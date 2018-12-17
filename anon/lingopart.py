"""
Re-model participant information from a CSV provided by lingoturk
This version created by Erwin R. Komen 
Date: 17/dec/2018 
"""
import sys, getopt, os.path, importlib
import os, sys
import util, csv, json
import openpyxl
from openpyxl import Workbook

errHandle = util.ErrHandle()

# ----------------------------------------------------------------------------------
# Name :    main
# Goal :    Main body of the function
# History:
# 15/sep/2017    ERK Created
# ----------------------------------------------------------------------------------
def main(prgName, argv) :
  flInput = ''        # input directory name
  flOutput = ''       # output directory name

  try:
    sSyntax = prgName + ' -i <input file>'
    # get all the arguments
    try:
      # Get arguments and options
      opts, args = getopt.getopt(argv, "hi:", ["-inputfile="])
    except getopt.GetoptError:
      print(sSyntax)
      sys.exit(2)
    # Walk all the arguments
    for opt, arg in opts:
      if opt in ("-h", "--help"):
        print(sSyntax)
        sys.exit(0)
      elif opt in ("-i", "--ifile"):
        flInput = arg
    # Check if all arguments are there
    if (flInput == ''):
      errHandle.DoError(sSyntax)

    # Determine what the output file will be
    flOutput = flInput.replace(".csv", ".xlsx")

    # Continue with the program
    errHandle.Status('Input is "' + flInput + '"')
    errHandle.Status('Output is "' + flOutput + '"')

    # Call the function that does the job
    oArgs = {'input': flInput,
             'output': flOutput}
    if (not process_lingopart(oArgs)) :
      errHandle.DoError("Could not complete")
      return False
    
      # All went fine  
    errHandle.Status("Ready")
  except:
    # act
    errHandle.DoError("main")
    return False

# ----------------------------------------------------------------------------------
# Name :    process_lingopart
# Goal :    Convert one lingopart file into an excel
# History:
# 17/dec/2019    ERK Created
# ----------------------------------------------------------------------------------
def process_lingopart(oArgs):
    """Process one lingopart file"""

    row_id = 0          # The ID within participantstatistics
    row_workerid = 1    # The string ID of the worker/participant
    row_expid = 3       # The Experiment ID row
    row_statistics = 4  # The field that contains the JSON list of statistics
    lHeadings = ['id', 'workerid', 'expid']

    try:
        # Recover the arguments
        if "input" in oArgs: flInput = oArgs["input"]
        if "output" in oArgs: flOutput = oArgs["output"]

        # Check input file
        if not os.path.isfile(flInput):
            errHandle.Status("Please specify an input FILE")
            return False

        # Create a list with the output we need
        lPartInfo = []
        # Read the input file as a CSV file with comma-separation
        with open(flInput, encoding='utf-8') as csv_file:
            csv_reader = csv.reader(csv_file, delimiter='\t')
            for row in csv_reader:
                oRow = {}
                oRow['id'] = row[row_id]
                oRow['workerid'] = row[row_workerid]
                oRow['expid'] = row[row_expid]
                sStatistics = row[row_statistics]
                lStatistics = json.loads(sStatistics)
                for oStats in lStatistics:
                    # One statistics object contains at least a "name" and a value in the form of "answer"
                    # Note that the value may be an integer or a string or a boolean
                    sName = oStats['name']
                    oRow[sName] = oStats['answer']
                    # Check for the heading
                    if sName not in lHeadings:
                        lHeadings.append(sName)

                # Add this object to the list
                lPartInfo.append(oRow)

        # Now create an Excel file and add the information into it
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()

        # Create a first row of headings
        row = 1
        for idx, name in enumerate(lHeadings):
            cell_this = ws.cell(row=row, column=idx+1)
            cell_this.value = name

        # Walk all rows in lPartInfo
        for oRow in lPartInfo:
            row += 1
            for idx, name in enumerate(lHeadings):
                cell_this = ws.cell(row=row, column=idx+1)
                if name in oRow:
                    cell_this.value = oRow[name]
                else:
                    cell_this.value = ""

        # Save it
        wb.save(flOutput)

        # Return correctly
        return True
    except:
        errHandle.DoError("process_lingopart")
        return False

# ----------------------------------------------------------------------------------
# Goal :  If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
  # Call the main function with two arguments: program name + remainder
  main(sys.argv[0], sys.argv[1:])
