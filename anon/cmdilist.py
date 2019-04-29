"""
Purpose: turn the CMDI files in one directory into a CSV file with path-value pairs

Original created by by Erwin R. Komen 
Date: 29/apr/2019
"""

# import lxml.etree as et
# from xml.dom import minidom
import xml.etree.ElementTree as et
import getopt, copy, util
import os, sys
import openpyxl
from openpyxl import Workbook

# ============ SETTINGS FOR THIS INSTANCE =================
bLocal = True          # Process locally 
bAllowCrashing = True   # Allow program to stop upon the first error
if bLocal:
    base_dir = "d:/etc/corpora/nld/ParsedSonar/WR-P-P-G_newspapers/079"
    #   Specify the result dir
    result_dir = "d:/etc/corpora/nld/ParsedSonar/WR-P-P-G_newspapers/079"
else:
    base_dir = "/vol/bigdata/corpora2/SoNaRfromTST/SoNaRCorpus_NC_1.2/SONAR500/CMDI/WR-P-P-G_newspapers/079"
    #   Specify the result dir
    result_dir = "/vol/tensusers/ekomen/corpora/"
XSI_CMD = "http://www.clarin.eu/cmd/"
XSI_XSD = "http://catalog.clarin.eu/ds/ComponentRegistry/rest/registry/profiles/clarin.eu:cr1:p_1328259700943/xsd"
# =========================================================

errHandle = util.ErrHandle()

def get_error_message():
    arInfo = sys.exc_info()
    if len(arInfo) == 3:
        sMsg = str(arInfo[1])
        if arInfo[2] != None:
            sMsg += " at line " + str(arInfo[2].tb_lineno)
        return sMsg
    else:
        return ""

def open_xml(filename):
    doc_string = b""
    sMethod = "traditional" # "onego"

    try:
        tree = et.parse(filename)
        root = tree.getroot()
        return root
    except:
        sMsg = get_error_message()
        print("Error in open_xml: {}".format(sMsg))

def add_metadate(oXml, sYear):
    """Add publication date"""
    
    back_xml = copy.deepcopy(oXml)
    lSubPub = ['Published', 'PublicationName', 'PublicationPlace', 'PublicationDate', 'PublicationTime', 'PublicationVolume',
               'PublicationNumber', 'PublicationNumber', 'PublicationIssue', 'PublicationSection', 'PublicationSubSection',
               'PublicationPage', 'PublicationId', 'PublicationLanguage', 'PublicationScope', 'PublicationGenre']
    # Local variables

    try:
        # Set the namespace
        ns_string = 'http://www.clarin.eu/cmd/'
        ns = {'df': ns_string}

        # Use the correct top components
        topattributes = {'xmlns': "http://www.clarin.eu/cmd/" ,
                    'xmlns:xsi': "http://www.w3.org/2001/XMLSchema-instance/",
                    'xsi:schemaLocation': XSI_CMD + " " + XSI_XSD,
                    'CMDVersion':'1.1'}

        # Find the <Source> tag
        # root = back_xml.getroot()
        for source in back_xml.findall(".//xmlns:Source", topattributes):
            # See if there is a Publication child
            pub_children = source.findall("./xmlns:Publication", topattributes)
            if len(pub_children) == 0:
                # Add a child
                pub_root = et.SubElement(source, "Publication")
                # Add children
                for child in lSubPub:
                    newChild = et.SubElement(pub_root, child)
                    # Check if this is the date
                    if child == "PublicationDate":
                        # Set the text value
                        newChild.text = sYear
                    # Add the child
                    #ch_this = pub_root.append(newChild)
    except:
        sMsg = errHandle.get_error_message()
        errHandle.DoError("add_metadate")

    return back_xml

def parse_node(node, ancestor_string = "", bFirst=False, ch_num=None):
    """parse one node in an XML tree recursively"""

    # Adapt the path we are at right now
    sNodeTag = node.tag
    sNodeTag = sNodeTag.replace("{http://www.clarin.eu/cmd/}", "")
    if ch_num != None:
        sNodeTag = "{}_{}".format(sNodeTag, ch_num)
    if ancestor_string:
        node_string = ".".join([ancestor_string, sNodeTag])
    else:
        node_string = sNodeTag

    # Adapt the path list
    tag_list = [node_string]

    # Get the text of this node
    text = node.text
    if text:
        text_list = [text.strip()]
    else:
        text_list = [""]

    # for child_node in list(node):
    for ch_num, child_node in enumerate(node):
        child_tag_list, child_text_list = parse_node(child_node, ancestor_string=node_string, ch_num=ch_num)
        tag_list.extend(child_tag_list)
        text_list.extend(child_text_list)

    return tag_list, text_list


def process_cmdi_directory(oArgs):
    """Process one directory of CMDI files into a CSV index of content"""

    # Defaults
    dirInput = ""
    flOutput = ""
    path_list = []
    text_list_list = []

    try:
        # Recover the arguments
        if "input" in oArgs: dirInput = oArgs["input"]
        if "output" in oArgs: flOutput = oArgs["output"]

        # Create an Excel file that should hold the information
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        column = 0

        # Check input directory
        if not os.path.exists(dirInput) or not os.path.isdir(dirInput):
            errHandle.Status("Please specify an existing input directory")
            return False

        # Signal the first file to be processed, which takes care of the list of paths
        bIsFirst = True
        # Process all cmdi.xml files in this directory
        for file in os.listdir(base_dir):
            # Check the ending of this file
            if ".cmdi.xml" in file:

                # Try to open the file
                try:
                    input_file = os.path.abspath( os.path.join(base_dir, file))

                    cmdi_xml = open_xml(input_file)

                    # Process this CMDI file into two lists
                    tags, texts = parse_node(cmdi_xml)

                    # DEBUG: show where we are
                    print("Processing metadata of file {} tags={}, texts={}".format(file, len(tags), len(texts)))

                    # If this is the first file, then set the path_list
                    if bIsFirst:
                        # Reset the flag
                        bIsFirst = False
                        # Copy the lists
                        path_list = copy.copy(tags)
                        text_list = copy.copy(texts)
                        # Fill the first column
                        column = 1
                        for idx, path in enumerate(path_list):
                            row = 2 + idx
                            cell_this = ws.cell(row=row, column=column)
                            cell_this.value = path

                    else:
                        # This is not the first file, so process all paths and values
                        text_list = []
                        for path in path_list:
                            # Get the path in the tags
                            idx = next((index for index, s in enumerate(tags) if s == path), None)
                            if idx == None:
                                # Failure: if a path is not found, then add the default value empty string
                                errHandle.DoError("Could not find path {}".format(path))
                                text_list.append("")
                            else:
                                # Add the corresponding text
                                text_list.append(texts[idx])
                    # In all instances: add the values of this file to a new column
                    column += 1
                    row = 1 
                    cell_this = ws.cell(row=row, column=column)
                    cell_this.value = os.path.splitext( os.path.basename(file))[0]
                    for idx, value in enumerate(text_list):
                        row = 2 + idx
                        cell_this = ws.cell(row=row, column=column)
                        cell_this.value = value
                    # Adapt the grand total
                    text_list_list.append(text_list)


                except:
                    print('failed: {}'.format(file))
                    print("error: {}".format(get_error_message()))
                    if bAllowCrashing:
                        print("The program has been aborted")
                        break

        # Save it
        wb.save(flOutput)
        ## Create the output CSV
        #lines = []
        #for idx, value in enumerate(path_list):
        #    line = []
        #    line.append(value)
        #    for item in text_list_list:
        #        line.append(item[idx])
        #    linestring = "{}\n".format( "\t".join(line))
        #    lines.append(linestring)

        ## Save the output XLSX file
        #with open(flOutput, 'wb') as output:
        #    for line in lines:
        #        output.write(line.encode('utf-8'))
        print("Ready")
        return True
    except:
        sMsg = errHandle.get_error_message()
        errHandle.DoError("process_cmdi_directory")
        return False

# ----------------------------------------------------------------------------------
# Name :    main
# Goal :    Main body of the function
# History:
# 29/apr/2019    ERK Created
# ----------------------------------------------------------------------------------
def main(prgName, argv) :
    dirInput = ''     # input directory name
    dirOutput = ""    # Output directory
    flOutput = ''     # output file name

    try:
        sSyntax = prgName + ' -i <input file>'
        # get all the arguments
        try:
            # Get arguments and options
            opts, args = getopt.getopt(argv, "hi:o:", ["-idir="])
        except getopt.GetoptError:
            print(sSyntax)
            sys.exit(2)
        # Walk all the arguments
        for opt, arg in opts:
            if opt in ("-h", "--help"):
                print(sSyntax)
                sys.exit(0)
            elif opt in ("-i", "--idir"):
                dirInput = arg
            elif opt in ("-o", "--odir"):
                dirOutput = arg
        # Check if all arguments are there
        if (dirInput == ''):
            errHandle.DoError(sSyntax, True)

        # Determine what the output directory will be
        if dirOutput == "":
            dirOutput = dirInput

        # Determine the output file name
        flOutput = os.path.abspath(os.path.join(dirOutput, "cmdilist.xlsx"))

        # Continue with the program
        errHandle.Status('Input directory is "' + dirInput + '"')
        errHandle.Status('Output file is "' + flOutput + '"')

        # Call the function that does the job
        oArgs = {'input': dirInput,
                    'output': flOutput}
        if (not process_cmdi_directory(oArgs)) :
            errHandle.DoError("Could not complete")
            return False
    
            # All went fine  
        errHandle.Status("Ready")
    except:
        # act
        errHandle.DoError("main")
        return False


# ----------------------------------------------------------------------------------
# Goal :  If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
  # Call the main function with two arguments: program name + remainder
  main(sys.argv[0], sys.argv[1:])
