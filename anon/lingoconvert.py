"""
Do Pinyin conversion of results from a CSV provided by lingoturk
This version created by Erwin R. Komen 
Date: 19/dec/2018 
"""
import sys, getopt, os.path, importlib
import os, sys
import util, csv, json
import openpyxl
from openpyxl import Workbook

errHandle = util.ErrHandle()

# ----------------------------------------------------------------------------------
# Name :    main
# Goal :    Main body of the function
# History:
# 19/dec/2019    ERK Created
# ----------------------------------------------------------------------------------
def main(prgName, argv) :
  flInput = ''        # input file name
  flOutput = ''       # output file name

  try:
    sSyntax = prgName + ' -i <input file>'
    # get all the arguments
    try:
      # Get arguments and options
      opts, args = getopt.getopt(argv, "hi:", ["-inputfile="])
    except getopt.GetoptError:
      print(sSyntax)
      sys.exit(2)
    # Walk all the arguments
    for opt, arg in opts:
      if opt in ("-h", "--help"):
        print(sSyntax)
        sys.exit(0)
      elif opt in ("-i", "--ifile"):
        flInput = arg
    # Check if all arguments are there
    if (flInput == ''):
      errHandle.DoError(sSyntax)

    # Determine what the output file will be
    flOutput = flInput.replace(".csv", ".xlsx")

    # Continue with the program
    errHandle.Status('Input is "' + flInput + '"')
    errHandle.Status('Output is "' + flOutput + '"')

    # Call the function that does the job
    oArgs = {'input': flInput,
             'output': flOutput}
    if (not process_lingo_convert_results(oArgs)) :
      errHandle.DoError("Could not complete")
      return False
    
      # All went fine  
    errHandle.Status("Ready")
  except:
    # act
    errHandle.DoError("main")
    return False

# ----------------------------------------------------------------------------------
# Name :    process_lingo_convert_results
# Goal :    Convert one CSV results file into an excel, converting the answers in Pinyin
# History:
# 19/dec/2019    ERK Created
# ----------------------------------------------------------------------------------
def process_lingo_convert_results(oArgs):
    """Process one lingo results CSV file"""

    row_answer = 9      # The column that contains the answer (in JSON)
    row_answer_roman = 0
    row_answer_semantics = 0
    lHeadings = []
    lConnectives = [
        {'chinese': '但是', 'roman': 'danshi', 'semantics': 'concessive'},
        {'chinese': '比如', 'roman': 'biru', 'semantics': 'instantiation'},
        {'chinese': '可见', 'roman': 'kejian', 'semantics': 'causal'},
        {'chinese': '所以', 'roman': 'suoyi', 'semantics': 'causal'},
        {'chinese': '因此', 'roman': 'yinci', 'semantics': 'causal'},
        {'chinese': '于是', 'roman': 'yushi', 'semantics': 'causal'}
        ]

    try:
        # Recover the arguments
        if "input" in oArgs: flInput = oArgs["input"]
        if "output" in oArgs: flOutput = oArgs["output"]

        # Check input file
        if not os.path.isfile(flInput):
            errHandle.Status("Please specify an input FILE")
            return False

        def get_connective(sChinese):
            for oItem in lConnectives:
                if sChinese == oItem['chinese']:
                    return oItem
            return None

        # Create a list with the output we need
        lPartInfo = []
        # Read the input file as a CSV file with comma-separation
        with open(flInput, encoding='utf-8-sig') as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            bFirst = True
            for row in csv_reader:
                # Check if this is the first row, which contains the headers
                if bFirst:
                    # Get the heading names
                    for sName in row:
                        lHeadings.append(sName)
                    # Add columns for the answer explanation
                    row_answer_roman = len(lHeadings)
                    lHeadings.append("answer.roman")
                    row_answer_semantics = len(lHeadings)
                    lHeadings.append("answer.semantics")
                    # Reset 
                    bFirst = False
                else:
                    oRow = {}
                    for idx, value in enumerate(row):
                        oRow[lHeadings[idx]] = value

                    sAnswer = row[row_answer]
                    oAnswer = json.loads( json.loads(sAnswer))
                    sRoman = "(none)"
                    sSemantics = "(none)"
                    if 'manualAnswer1' in oAnswer:
                        sRoman = "(manual)"
                        sSemantics = "(unknown)"
                    elif 'connective1' in oAnswer:
                        sChinese = oAnswer['connective1']
                        oConnective = get_connective(sChinese)
                        if oConnective != None:
                            sRoman = oConnective['roman']
                            sSemantics = oConnective['semantics']
                    oRow["answer.roman"] = sRoman
                    oRow["answer.semantics"] = sSemantics

                    # Add this object to the list
                    lPartInfo.append(oRow)

        # Now create an Excel file and add the information into it
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()

        # Create a first row of headings
        row = 1
        for idx, name in enumerate(lHeadings):
            cell_this = ws.cell(row=row, column=idx+1)
            cell_this.value = name

        # Walk all rows in lPartInfo
        for oRow in lPartInfo:
            row += 1
            for idx, name in enumerate(lHeadings):
                cell_this = ws.cell(row=row, column=idx+1)
                if name in oRow:
                    cell_this.value = oRow[name]
                else:
                    cell_this.value = ""

        # Save it
        wb.save(flOutput)

        # Return correctly
        return True
    except:
        errHandle.DoError("process_lingopart")
        return False

# ----------------------------------------------------------------------------------
# Goal :  If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
  # Call the main function with two arguments: program name + remainder
  main(sys.argv[0], sys.argv[1:])
