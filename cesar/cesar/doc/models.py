# Django imports
from django.db import models
from django.utils import timezone
from django.contrib.auth.models import User
from django.db import models, transaction

# Non-django imports
import os
import sys
import time
import json
import re
import copy
import pytz
from io import StringIO 
import openpyxl

# XML processing
from xml.dom import minidom

# FOlia handling: pynlpl
import pynlpl
from pynlpl.textprocessors import tokenize      # , split_sentences

# New folia: FOliaPy
import folia.main as folia

# ================== OWN procedures ==========================================

from cesar.utils import ErrHandle
from cesar.seeker.models import import_data_file
from cesar.settings import WRITABLE_DIR, TIME_ZONE, MEDIA_ROOT
from cesar.tsg.models import TsgInfo


## Attempt to input FROG
#try:
#    from frog import Frog, FrogOptions
#    # See: https://frognlp.readthedocs.io/en/latest/pythonfrog.html
#except:
#    # It is not there...
#    frogurl = "https://webservices-lst.science.ru.nl/frog"
#    from clam.common.client import *

# OLD: frogurl = "https://webservices-lst.science.ru.nl/frog"
# New (2020, december)
frogurl = "https://webservices.cls.ru.nl/frog"
uctourl = "https://webservices.cls.ru.nl/ucto"
from clam.common.client import *

MAXPARAMLEN = 100
MAXPATH = 256
FROGPORT = 8020

# ==================================== Auxiliary ===============================================

def get_crpp_date(dtThis, readable=False):
    """Convert datetime to string"""

    if readable:
        # Convert the computer-stored timezone...
        dtThis = dtThis.astimezone(pytz.timezone(TIME_ZONE))
        # Model: yyyy-MM-dd'T'HH:mm:ss
        sDate = dtThis.strftime("%d/%B/%Y (%H:%M)")
    else:
        # Model: yyyy-MM-dd'T'HH:mm:ss
        sDate = dtThis.strftime("%Y-%m-%dT%H:%M:%S")
    return sDate

def wordlist_path(instance, filename):
    """Upload file to the right place,and remove old file if existing
    
    This function is used within the model Wordlist
    NOTE: this must be the relative path w.r.t. MEDIA_ROOT
    """

    oErr = ErrHandle()
    sBack = ""
    sSubdir = "wordlist"
    try:
        # The stuff that we return
        sBack = os.path.join(sSubdir, filename)

        # Add the subdir [wordlist]
        fullsubdir = os.path.abspath(os.path.join(MEDIA_ROOT, sSubdir))
        if not os.path.exists(fullsubdir):
            os.mkdir(fullsubdir)

        # Add the actual filename to form an absolute path
        sAbsPath = os.path.abspath(os.path.join(fullsubdir, filename))
        if os.path.exists(sAbsPath):
            # Remove it
            os.remove(sAbsPath)

    except:
        msg = oErr.get_error_message()
        oErr.DoError("wordlist_path")
    return sBack


# ==================================== MODELS =================================================

class LocTimeInfo(models.Model):
    """Information about a particular location or time"""

    # [1] Example kind of text
    example = models.TextField("Example text")
    # [1] Concreteness score
    score = models.CharField("Concreteness score", default="0.0", max_length=MAXPARAMLEN)

    def __str__(self):
        return self.example


class Expression(models.Model):
    """Multi-word expression"""

    # [1] The multi-word expression
    full = models.TextField("Multi-word expression")
    # [0-1] The lemma parts of the expression in a stringified JSON
    lemmas = models.TextField("The lemma parts", blank=True, null=True)
    # [0-1] The lemma's as determined via FROG
    frogged = models.TextField("Lemma parts by FROG", blank=True, null=True)
    # [1] Concreteness score
    score = models.CharField("Concreteness score", default="0.0", max_length=MAXPARAMLEN)

    def __str__(self):
        return self.full

    def save(self, force_insert = False, force_update = False, using = None, update_fields = None): #, *args, **kwargs):
        oErr = ErrHandle()
        try:
            # username = kwargs.get("username")
            # Check if 'lemmas' is okay with 'full'
            sFull = json.dumps( dict(score=self.score,lemmas= re.split("\s+", self.full)))
            oErr.Status("Expression: {}".format(sFull))
            if self.lemmas != sFull and self.frogged != None and self.frogged != "":
                # The frogged value and the lemmas should coincide
                sFull = json.dumps( dict(score=self.score,lemmas= json.loads(self.frogged)))
                if self.lemmas != sFull:
                    self.lemmas = sFull
            # Perform the actual saving
            response = super(Expression, self).save(force_insert, force_update, using, update_fields)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("Expression/save")
        # Return the appropriate response
        return response

    def make_frogged(self, username, sFull):
        """Compare the existing .full with the new sFull, and adapt .frogged if needed"""

        sBack = ""
        oErr = ErrHandle()
        try:
            if not sFull is None and sFull != "":
                # Frogging is needed
                bFound, lst_lemmas = FrogLink.get_lemmas(username, self.full)
                if bFound:
                    if not lst_lemmas is None and len(lst_lemmas) > 0:
                        lst_line = lst_lemmas[0]
                        if not lst_line is None and len(lst_line) > 0:
                            sBack = json.dumps(lst_line)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("Expression/check_frogged")

        return sBack

    def get_frogged(self):
        """Show the frogged lemma's of this MWE"""

        sBack = ""
        oErr = ErrHandle()
        try:
            if not self.frogged is None:
                lst_frogged = json.loads(self.frogged)
                if not lst_frogged is None:
                    sBack = ", ".join(lst_frogged)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("Expression/get_frogged")

        return sBack

    def get_lemmas(self):
        """Show the lemma's of this MWE"""

        sBack = ""
        oErr = ErrHandle()
        try:
            if not self.lemmas is None:
                oMwe = json.loads(self.lemmas)
                lst_lemma = oMwe.get("lemmas")
                if not lst_lemma is None:
                    sBack = ", ".join(lst_lemma)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("Expression/get_lemmas")

        return sBack

    def get_mwe_list():
        """get a list of objects, where each object has the score + the list of lemma parts"""

        oErr = ErrHandle()
        lBack = []
        oFull = None
        try:
            lAll = Expression.objects.all().values('id', 'score', 'lemmas', 'full')
            for oItem in lAll:
                id = oItem['id']
                score = oItem['score']
                full = oItem['full']
                sLemmas = oItem.get('lemmas')
                if sLemmas is None:
                    # The lemma object must be calculated
                    bFound, lst_lemmas = FrogLink.get_lemmas("erwin", full)
                    if bFound and not lst_lemmas is None and len(lst_lemmas) > 0:
                        lst_line = lst_lemmas[0]
                        if not lst_line is None and len(lst_line) > 0:
                            sLemmas = json.dumps(lst_line)
                            oItem['lemmas'] = sLemmas
                            oFull = dict(score=score, lemmas=lst_line)
                            obj = Expression.objects.filter(id=id).first()
                            if not obj is None:
                                obj.frogged = sLemmas
                                obj.lemmas = json.dumps(oFull)
                                obj.save()
                else:
                    oFull = json.loads(sLemmas)
                # oFull = json.loads(oItem['lemmas'])
                lemmas = oFull['lemmas']
                lBack.append(dict(score=score, lemmas=lemmas, full=full))
        except:
            msg = oErr.get_error_message()
            oErr.DoError("doc/Expression/get_mwe_list")
        return lBack

    def get_fullmwe_list():
        """get a list of objects, where each object has the score + the full MWE"""

        oErr = ErrHandle()
        lBack = []
        try:
            lAll = Expression.objects.all().values('score', 'full')
            for oItem in lAll:
                score = oItem['score']
                full = oItem['full']
                lBack.append(dict(score=score, full=full))
        except:
            msg = oErr.get_error_message()
            oErr.DoError("doc/Expression/get_fullmwe_list")
        return lBack

    def get_fullmwe_fit(lst_word):
        """Find the first Expression, where all its words match those consecutively in [lst_word]

        NOTE: *all* of the Expression's words should match, but lst_word may have words left
        """

        oBack = dict(status="none")
        oErr = ErrHandle()
        try:
            lAll = Expression.objects.all().values('score', 'full')
            for oItem in lAll:
                score = oItem['score']
                full =  oItem['full']
                # Split the full into words
                mwe_words = [x.lower() for x in full.split()]
                mwe_size = len(mwe_words)

                # Check if this MWE would fit
                if mwe_size <= len(lst_word):
                    bMatches = True
                    for idx in range(len(mwe_words)):
                        # Compare case-insensitive
                        if mwe_words[idx] != lst_word[idx].lower():
                            bMatches = False
                            break
                    # Do we have a match?
                    if bMatches:
                        # Yes, we have a match!
                        oBack['status'] = "ok"
                        oBack['mwe'] = copy.copy(oItem)
                        oBack['size'] = mwe_size
                        # Return what we found
                        return oBack
            # Getting here means that we did not find any match
        except:
            msg = oErr.get_error_message()
            oErr.DoError("doc/Expression/get_fullmwe_fit")

        return oBack

    def process_item(full, score):
        """Update or add an expression with the indicated score"""

        bResult = True
        oErr = ErrHandle()
        try:
            # Check if this element already exists
            obj = Expression.objects.filter(full__iexact=full).first()
            if obj is None:
                obj = Expression.objects.create(full=full)
            # Note that internally the score is saved with a "."
            if isinstance(score, str):
                score = score.replace(",", ".")
            else:
                score = str(score)
            # Only now check if the score has actually changed
            if obj.score != score:
                # Yes it's changed, so save it
                obj.score = score
                obj.save()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("Expression/process_item")
        return bResult


class FoliaDocs(models.Model):
    """Set of folia-encoded documents"""
    
    # [1] These belong to a particular user
    owner = models.ForeignKey(User, editable=False, on_delete=models.CASCADE, related_name="owner_foliadocs")

    def __str__(self):
        return self.owner.username


class FrogLink(models.Model):
    """This provides the basic link with FROG

    Can be either frog local or frog remote, through the CLAM service
    """

    # [1] Each froglink centers around a file that is uploaded, processed and made available
    name = models.CharField("Name to be used for this file", max_length=MAXPARAMLEN)
    # [1] Each link belongs to a set of docs (that belong to an owner)
    fdocs = models.ForeignKey(FoliaDocs, related_name="documents", on_delete=models.CASCADE)
    # [0-1] Full name is the full path of the folia.xml document on the server
    fullname = models.CharField("Full path of this file", max_length=MAXPATH, null=True, blank=True)
    # [0-1] Concreteness as stringified JSON object
    concr = models.TextField("Concreteness scores", null=True, blank=True)
    # [1] Each Froglink has been created at one point in time
    created = models.DateTimeField(default=timezone.now)

    def __str__(self):
        return self.name

    def create(name, username):
        """Possibly create a new item [name] for user [username]"""

        errHandle = ErrHandle()

        try:
            # Get the correct user
            owner = User.objects.filter(username=username).first()

            # Find a FoliaDocs instance for this user
            fd = FoliaDocs.objects.filter(owner=owner).first()
            if fd == None:
                fd = FoliaDocs(owner=owner)
                fd.save()

            obj = FrogLink.objects.filter(name=name, fdocs=fd).first()
            if obj == None:
                # Create it
                obj = FrogLink(name=name, fdocs=fd)
                obj.save()
            # Return the result 
            return obj, ""
        except:
            errHandle.DoError("FrogLink/create")
            return None, errHandle.get_error_message()

    def get_created(self):
        sBack = self.created.strftime("%d/%B/%Y (%H:%M)")
        return sBack

    def get_lemmas(username, sText):
        """Use frog to get a list of lemma-lines"""

        bBack = False
        lst_line = []
        sName = "_froglink_lemmas.txt"
        sDir = os.path.abspath(os.path.join(WRITABLE_DIR, "../folia"))
        sFile = os.path.abspath(os.path.join(WRITABLE_DIR, "../folia", sName ))
        oErr = ErrHandle()
        arErr = []
        try:
            # Check the directory
            if not os.path.exists(sDir):
                os.mkdir(sDir)
            
            # Get the right Froglink
            fl, msg = FrogLink.create(sName, username)

            if not fl is None:
                # Put the text into a file
                with open(sFile, "w", encoding="utf-8") as f:
                    f.write(sText)

                lines = sText.split("\n")

                # Convert the text to Folia XML
                oFolia = fl.read_doc(username, lines, sFile, None, None, arErr)

                if not oFolia is None and oFolia.get("status", "") == "ok":
                    # Read the file into a doc
                    doc = folia.Document(file=fl.fullname)
                    # Read through paragraphs
                    for para in doc.paragraphs():
                        # Read sentences
                        for sent in para.sentences():
                            lst_linelemma = []
                            # Read the words in a sentence
                            for oWord in sent.words():
                                # word = oWord['word']
                                word_text = str(oWord)
                                # Get to the POS and the lemma of this word
                                pos = oWord.annotation(folia.PosAnnotation)
                                postag = pos.cls.split("(")[0]
                                lemma = oWord.annotation(folia.LemmaAnnotation)                     
                                lemmatag = lemma.cls
                                # Add to line
                                lst_linelemma.append(lemmatag)
                            lst_line.append(lst_linelemma)
                    bBack = True
        except:
            msg = oErr.get_error_message()
            oErr.DoError("Froglink/get_lemmas")

        return bBack, lst_line

    def get_name(self):
        """Either get the defined name or a unique name"""

        sBack = ""
        if self.name is None or self.name =="":
            sBack = "item_{}".format(self.id)
        else:
            sBack = self.name
        return sBack

    def get_owner(self):
        """Return the owner of this concreteness document"""

        sBack = ""
        if self.fdocs != None:
            if self.fdocs.owner != None:
                sBack = self.fdocs.owner.username
        return sBack

    def read_doc(self, username, data_file, filename, clamuser, clampw, arErr, xmldoc=None, sName = None, oStatus = None):
        """Import a text file, parse it through the frogger, and create a Folia.xml file"""

        oBack = {'status': 'ok', 'count': 0, 'msg': "", 'user': username}
        errHandle = ErrHandle()
        oDoc = None
        iCount = 0
        inputType = "folia"
        frogType = "remote"     # Fix to remote -- or put to None if automatically choosing

        def get_error_log(result):
            """Try to get the error.log file contents"""
            for item in result.output:
                name = str(item)
                if "error.log" in name:
                    lText = []
                    for part in item:
                        lText.append(part.decode("utf-8").strip())
                        # lText.append(str(part))
                    sText = "\n".join(lText)
            return sText

        def get_error_view(errorpoint, statusmsg, errormsg, sLog, sText):
            lhtml = []
            lhtml.append("<div>Error in {}: {}</div>".format(errorpoint, statusmsg))
            lhtml.append("<div>Frog error: {}</div>".format(errormsg))
            lhtml.append("<div class='panel panel-default'>")
            lhtml.append("  <div class='panel-heading' data-toggle='collapse' data-target='#error_log'>Log Info</div>")
            lhtml.append("  <div id='error_log' class='collapse'><pre style='font-size: smaller;'>{}</pre></div>".format(sLog))
            lhtml.append("</div>")
            # Get a (numbered) listing of the text file
            lhtml.append("<div class='panel panel-default'>")
            lhtml.append("  <div class='panel-heading' data-toggle='collapse' data-target='#error_text'>Text file</div>")
            lhtml.append("  <div id='error_text' class='collapse'><pre style='font-size: smaller;'>{}</pre></div>".format(sText))
            lhtml.append("</div>")
            return "\n".join(lhtml)


        folProc = FoliaProcessor(username)
        try:
            if clamuser is None:
                clamuser = TsgInfo.get_value("clam_user")
                clampw = TsgInfo.get_value("clam_pw")

            # Check our location
            frogLoc = frogType if frogType != None else folProc.location()

            # Directory: one directory for each user
            dir = folProc.dir
            
            # Read and create basis-folia
            bResult, sMsg = folProc.basis_folia(filename, data_file)
            if not bResult:
                # There was some kind of error
                oBack['status'] = 'error'
                oBack['msg'] = "basis_folia load error: {}".format(sMsg)
            else:
                # DEBUG: show the frog location
                errHandle.Status("DEBUG: frogLoc={}".format(frogLoc))


                if frogLoc == "remote":
                    # Think of a project name
                    project = folProc.docstr
                    basicauth = True
                    # Get access to the webservice
                    clamclient = CLAMClient(frogurl, clamuser, clampw, basicauth = basicauth)
                    # First delete any previous project, if it exists
                    try:
                        result = clamclient.delete(project)
                        errHandle.Status("Removed previous project {} = {}".format(project, result))
                    except:
                        # No problem: no project has been removed
                        pass
                    # Only now start creating it
                    result = clamclient.create(project)
                    errHandle.Status("Created new project {} = {}".format(project, result))
                    data = clamclient.get(project)
                    # If we have textinput, then it is 'maininput' (see /frog/info)
                    if inputType == "text":
                        it = data.inputtemplate('maininput')
                        for param in it.parameters:
                            if param.id == 'author':
                                param.value = "cesar"
                                param.hasvalue = True
                        it.filename = self.name
                    else:
                        it = data.inputtemplate("foliainput")

                    # Add the Basis-Folia XML file as input to the project
                    basicf = folProc.basicf
                    result = clamclient.addinputfile(project, it, basicf, language='nl')
                    # Opstarten
                    #   See explanation at https://webservices.cls.ru.nl/frog/info/
                    #   we are skipping some modules  
                    #	  t - Tokeniser
                    #	  m - Multi-Word Detector
                    #	  p - Parser
                    #	  c - Chunker / Shallow parser
                    #	  n - Named Entity Recognition
                    result = clamclient.start(project, skip= ['m', 'p', 'c', 'n'])
                    if result.errors:
                        oBack['status'] = "error"
                        # Handle errors
                        statusmsg = result.statusmessage
                        errormsg = result.errormsg
                        # Find the error log
                        sText = get_error_log(result)
                        oBack['msg'] = get_error_view("Pre-CLAM", statusmsg, errormsg, sText, sMsg)
                        return oBack
                    # Otherwise loop until ready
                    while result.status != clam.common.status.DONE:
                        time.sleep(1)			            # Wacht 4 seconden
                        result = clamclient.get(project)	# Refresh status
                        statusmsg = result.statusmessage
                        completion = result.completion
                        if oStatus != None:
                            msg = "{}: {}% completed".format(statusmsg, completion)
                            oStatus.set("CLAM", msg=msg)
                    if result.errors:
                        oBack['status'] = "error"
                        # Handle errors
                        statusmsg = result.statusmessage
                        errormsg = result.errormsg
                        # Find the error log
                        sText = get_error_log(result)
                        oBack['msg'] = get_error_view("Post-CLAM", statusmsg, errormsg, sText, sMsg)
                        return oBack
                    # Now we are ready
                    for outputfile in result.output:
                        name = str(outputfile)
                        if ".xml"  in name:
                            # Download the Folia XML file to the user's dir
                            fout = os.path.abspath(os.path.join(dir, os.path.basename(str(outputfile))))
                            fout = fout.replace(".basis", ".folia")
                            outputfile.copy(fout)
                            iCount += 1
                            
                            # Note where it is
                            self.fullname = fout
                            self.save()

                            # Bugfix because of new libfolia version
                            # self.folia_annotation_reduction()

                        else:
                            # Download the other files to a log dir
                            logdir = os.path.abspath(os.path.join(dir, "log"))
                            if not os.path.exists(logdir):
                                os.mkdir(logdir)
                            elif not os.path.isdir(logdir):
                                os.remove(logdir)
                                os.mkdir(logdir)
                            fout = os.path.abspath(os.path.join(logdir, os.path.basename(str(outputfile))))
                            fout = fout.replace(".basis", ".folia")
                            outputfile.copy(fout)
                            iCount += 1
                    # Delete project again
                    clamclient.delete(project)
  
                else:
                    # Don't know this location type
                    pass


            # Make sure the requester knows how many have been added
            oBack['count'] = iCount   # The number of sermans added
            oBack['filename'] = filename
            oBack['concreteness'] = 0

        except:
            sError = errHandle.get_error_message()
            oBack['status'] = 'error'
            oBack['msg'] = sError

        # Return the object that has been created
        return oBack

    def folia_annotation_reduction(self):
        """Remove annotation layers that are not recognized by the current FoLiA"""

        oErr = ErrHandle()
        bOkay = True
        try:
            # (1) Read the file as XML
            fullname = self.fullname
            xmldoc = minidom.parse(fullname)
            # (2) find the <annotations>
            ndAnnot_list = xmldoc.getElementsByTagName("annotations")
            remove_list = []
            if ndAnnot_list.length > 0:
                annotations = ndAnnot_list[0]
                for annotation in annotations.childNodes:
                    if annotation.nodeType == minidom.Node.ELEMENT_NODE:
                        if annotation.tagName == "quote-annotation" or annotation.tagName == "alternative-annotation":
                            # Mark this one as one to be removed
                            remove_list.append(annotation)
            for item in reversed(remove_list):
                annotations.removeChild(item)
            # Save the result
            with open(fullname, "w", encoding="utf-8") as f:
                xmldoc.writexml(f)
        except:
            sError = oErr.get_error_message()
            oErr.DoError("folia_annotation_reduction")
            bOkay = False

        # Return okay
        return bOkay

    def do_concreteness(self):
        rules = [   {'pos': 'ADJ'},
                    {'pos': 'N'},
                    {'pos': 'WW', 'lemma_excl': ['hebben', 'zijn', 'zullen', 'willen', 'worden', 'moeten', 'mogen', 'kunnen', 'laten', 'doen']} ]
        bResult = False
        sMsg = ""
        str_parts = ""
        oErr = ErrHandle()
        method = "hidde"        # Alternatives: 'erwin', 'alpino', 'simple', 'hidde'
        mwe_method = "anyorder" # Alternatives: 'lemmas', 'full', 'anyorder'


        # Make sure we have a list of MWEs available for process_mwe()
        lst_mwe = None
        if mwe_method == "full":
           lst_mwe = Expression.get_fullmwe_list()
        elif mwe_method in ["lemmas", "anyorder"]:
           lst_mwe = Expression.get_mwe_list()

        re_diminuative = re.compile(r'.*(m|n|l|ng|p|b|t|d|k|f|v|s|z|ch|g)je(s)?$')
        re_etje = re.compile(r'.*(nn|mm|ll)etje$')
        re_etjes = re.compile(r'.*(nn|mm|ll)etjes$')
        re_ngetje = re.compile(r'.*ngetje$')
        re_ngetjes = re.compile(r'.*ngetjes$')
        re_obs_je = re.compile(r'.*(p|b|t|d|k|f|v|s|z|ch|g)je$')
        re_obs_jes = re.compile(r'.*(p|b|t|d|k|f|v|s|z|ch|g)jes$')
        re_kje = re.compile(r'.*nkje$')
        re_kjes = re.compile(r'.*nkjes$')
        re_pje = re.compile(r'.*mpje$')
        re_pjes = re.compile(r'.*mpjes$')
        re_tje = re.compile(r'.*tje$')
        re_tjes = re.compile(r'.*tjes$')

        def treat_word(posonly, posfull, lemma):
            """Check if this category must be included"""
            bFound = False
            for rule in rules:
                if 'pos' in rule and rule['pos'] == posonly:
                    # Continue
                    if 'lemma_excl' in rule:
                        bFound = lemma not in rule['lemma_excl']
                    else:
                        bFound = True
                    # Whatever the outcome, we have found our result now
                    break
            return bFound

        def strip_diminuative(sWord):
            """Take the diminuative ending from the word"""

            sBack = sWord
            oErr = ErrHandle()
            try:
                # Sonorant -etje
                if re_etje.match(sWord):
                    sBack = sWord[:-5]
                elif re_etjes.match(sWord):
                    sBack = sWord[:-6]
                elif re_ngetje.match(sWord):
                    sBack = sWord[:-4]
                elif re_ngetjes.match(sWord):
                    sBack = sWord[:-5]
                elif re_tje.match(sWord):
                    # beentje >> been
                    sBack = sWord[:-3]
                elif re_tjes.match(sWord):
                    # leeuwtjes >> leeuw
                    sBack = sWord[:-4]
                elif re_obs_je.match(sWord):
                    sBack = sWord[:-2]
                elif re_obs_jes.match(sWord):
                    sBack = sWord[:-3]
                elif re_kje.match(sWord):
                    # koninkje >> koning
                    sBack = sWord[:-3] + "g"
                elif re_kjes.match(sWord):
                    # palinkjes >> paling
                    sBack = sWord[:-4] + "g"
                elif re_pje.match(sWord):
                    # duimpje >> duim
                    sBack = sWord[:-3]
                elif re_pjes.match(sWord):
                    # kruimpjes >> kruim
                    sBack = sWord[:-4]
            except:
                msg = oErr.get_error_message()
                oErr.DoError("strip_diminuative")
            return sBack

        def is_anyorder(sublemmas, lst, skiplist):
            """are the lemma's specified in a MWE available in the sentencein *ANY* order?"""

            lFound = []
            bResult = False
            oErr = ErrHandle()
            try:
                sub = copy.copy(sublemmas)
                ln = len(sub)
                for idx, word in enumerate(lst):
                    if not idx in skiplist and word in sub:
                        lFound.append(idx)
                        sub.remove(word)
                if len(lFound) == ln:
                    bResult = True
                else:
                    bResult = False
                    lFound = []
            except:
                msg = oErr.get_error_message()
                oErr.DoError("is_anyorder")
                lFound = []
                bResult = False
            return bResult, lFound

        def is_sub(sub, lst):
            ln = len(sub)
            for i in range(len(lst) - ln + 1):
                if all(sub[j] == lst[i+j] for j in range(ln)):
                    return True, i
            return False, -1

        def process_mwe(lst_sent):
            """Find any MWEs in the sentences list's lemma's and take them out"""
            lst_back = []
            lst_found_mwe = []

            oErr = ErrHandle()
            try:
                # Action depends on the method that we are following
                if mwe_method == "full":
                    # Convert the sentence into a list of words
                    lst_word = []
                    for oWord in lst_sent:
                        word = str(oWord)
                        lst_word.append(word)
                    # NOTE: the following doesn't work, apparently...
                    # lst_word = [str(word) for word in lst_sent]

                    # Look at the full MWE expressions
                    idx_full = 0
                    while idx_full < len(lst_word):
                        # Check if there is a fit for an MWE starting at this point
                        oResult = Expression.get_fullmwe_fit(lst_word[idx_full:])
                        if oResult['status'] == "ok":
                            # We have a match! Look at: score, full, size
                            mwe_size = oResult['size']
                            mwe = oResult['mwe']
                            mwe['size'] = mwe_size
                            mwe['idx'] = idx_full
                            # Return the MWE as one whole expression
                            lst_found_mwe.append(mwe)
                            # Skip to the next one
                            idx_full += mwe_size
                        else:
                            # Return the Folia Word object
                            oBack = dict(idx=idx_full, word=copy.copy(lst_sent[idx_full]))
                            lst_back.append(oBack)
                            idx_full += 1
                elif mwe_method == "lemmas":
                    # Look at a list of lemma's

                    # Create this list of lemma's
                    lst_lemma = []
                    for word in lst_sent:
                        lemma = word.annotation(folia.LemmaAnnotation)
                        lst_lemma.append(lemma.cls)
                    # Check all available MWEs
                    oFound = {}
                    for mwe in lst_mwe:
                        lemmas = mwe['lemmas']
                        score = mwe['score']
                        bFound, idx = is_sub(lemmas, lst_lemma)
                        if bFound:
                            # Note that this one needs to be taken out
                            oFound[idx] = mwe

                    idx = 0
                    count = len(lst_sent)
                    while idx < count:
                        if idx in oFound:
                            # Skip a number of items
                            mwe = oFound[idx]
                            idx += len(mwe['lemmas'])
                            lst_found_mwe.append(mwe)
                        else:
                            # Add item to list
                            lst_back.append(lst_sent[idx])
                            # Continue with next one
                            idx += 1
                elif mwe_method == "anyorder":
                    # Look at a list of lemma's
                    # Goal: are the lemma's specified in a MWE available in the sentencein *ANY* order?

                    # Create this list of lemma's
                    lst_lemma = []
                    for word in lst_sent:
                        lemma = word.annotation(folia.LemmaAnnotation)
                        lst_lemma.append(lemma.cls)
                    # Check all available MWEs
                    oFound = {}
                    lst_skip = []
                    for mwe in lst_mwe:
                        lemmas = mwe['lemmas']
                        score = mwe['score']
                        full = mwe['full']
                        # Beware of empty lemma lists!!
                        if len(lemmas) > 0:
                            # Find a list of indexes of the words in [lemmas] belonging to the MWE
                            bFound, lst_index = is_anyorder(lemmas, lst_lemma, lst_skip)
                            if bFound:
                                # This MWE is found inside our sentence [lst_lemma]
                                mwe['size'] = len(lemmas)
                                mwe['idx'] = lst_index[0]
                                lst_found_mwe.append(mwe)
                                # The [lst_index] contains the indices of the words in [lemmas] that are
                                #    part of the MWE, and may therefore be skipped from further processing
                                for idx in lst_index:
                                    # The list [lst_skip] will contain the indices of the words that may be skipped
                                    lst_skip.append(idx)

                    idx = 0
                    count = len(lst_sent)
                    while idx < count:
                        # Just make sure that we do not return words from the skippables [lst_skip]
                        if not idx in lst_skip:
                            # Return the Folia Word object
                            oBack = dict(idx=idx, word=copy.copy(lst_sent[idx]))
                            lst_back.append(oBack)
                        # Continue with next one
                        idx += 1
            except:
                msg = oErr.get_error_message()
                oErr.DoError("process_mwe")

            return lst_found_mwe, lst_back

        try:
            # Create a regular expression to detect a content word
            # re_content = re.compile(r'^(VNW|N)$')
            # Read the file into a doc
            doc = folia.Document(file=self.fullname)
            # Read through paragraphs
            para_scores = []
            word_id = 1
            for para in doc.paragraphs():
                # Read sentences
                sent_scores = []
                for sent in para.sentences():
                    # Read through the words
                    word_scores = []

                    # Get an initial list of words
                    sent_words = [x for x in sent.words()]

                    # Adapt it using MWE
                    sent_mwes, sent_words = process_mwe(sent_words)

                    # for word in sent.words():
                    # Note: these are the 'remaining' words, that are not part of the MWEs
                    for oWord in sent_words:
                        word = oWord['word']
                        word_text = str(word)
                        # Get to the POS and the lemma of this word
                        pos = word.annotation(folia.PosAnnotation)
                        postag = pos.cls.split("(")[0]
                        lemma = word.annotation(folia.LemmaAnnotation)                     
                        lemmatag = lemma.cls

                        # Check if we need to process this word (if it is a 'content' word)
                        if treat_word(postag, pos, lemmatag):
                            # Start out: we have not found it yet
                            obj_fixed = None
                            lst_homonyms = []       # Empty homonym list to start with

                            # (1) Attempt to find Quantity/measure
                            if obj_fixed is None:
                                # TODO: which issue is this??
                                pass

                            # (2) Attempt to find neologisms
                            if obj_fixed is None:
                                obj_fixed = Neologism.objects.filter(stimulus__iexact=lemmatag).first()
                                if obj_fixed is None:
                                    # Get the *first* neologism fitting
                                    obj_fixed = Neologism.objects.filter(stimulus__iexact=word_text).first()

                            # (3) Attempt to find homonyms
                            if obj_fixed is None:
                                # Search the homonym list
                                homonyms = Homonym.objects.filter(stimulus__iexact=lemmatag, postag=postag)
                                # Found any homonyms?
                                h_found = homonyms.count()
                                if h_found == 1:
                                    # There is one unique respons -- take it
                                    obj_fixed = homonyms.first()
                                    score = obj_fixed.get_concreteness()
                                elif h_found > 0:
                                    # There are multiple possibilities: Add details for the choice for a user
                                    for idx, obj_this in enumerate(homonyms):
                                        # Create a homonym object
                                        oHomonym = {}
                                        oHomonym['hnum'] = idx + 1
                                        oHomonym['pos'] = obj_this.postag
                                        oHomonym['meaning'] = obj_this.meaning
                                        oHomonym['score'] = obj_this.get_concreteness()
                                        # Add this object to the list
                                        lst_homonyms.append(oHomonym)
                                    # The default score will be the first one
                                    obj_fixed = homonyms.first()
                                    score = obj_fixed.get_concreteness()

                            # (4) Look in other lists supplied by Wordlist
                            if obj_fixed is None:
                                if postag == "":
                                    worddefs = Worddef.objects.filter(stimulus__iexact=lemmatag)
                                else:
                                    worddefs = Worddef.objects.filter(stimulus__iexact=lemmatag, postag__iexact=postag)
                                    if worddefs.count() == 0:
                                        worddefs = Worddef.objects.filter(stimulus__iexact=lemmatag)
                                o_found = worddefs.count()
                                if o_found > 0:
                                    # There are one or more responses: take the first
                                    obj_fixed = worddefs.first()
                                    score = obj_fixed.get_concreteness()

                            # (5) Brysbaert: check if a word is in the brysbaertlist
                            if obj_fixed is None:
                                obj_fixed = Brysbaert.objects.filter(stimulus=word_text).first()
                                if not obj_fixed is None:
                                    # The concreteness of this word is in obj_fixed
                                    score = obj_fixed.get_concreteness()
                                else:
                                    # Get the *FIRST* brysbaert equivalent if existing
                                    obj_fixed = Brysbaert.objects.filter(stimulus=lemmatag).first()

                            if obj_fixed is None:
                                # Check if there are multiple morph parts
                                # morph_parts = [m.text() for m in word.morphemes()]
                                morph_parts = []
                                for m in word.morphemes():
                                    try:
                                        m_text = m.text()
                                        # Do *NOT* include morpheme parts that are the same as the whole word
                                        if m_text != word_text.lower():
                                            morph_parts.append(m_text)
                                    except:
                                        ging_niet = 1
                                        pass
                                score = -1
                                if len(morph_parts)>0:
                                    # There are multiple morphemes
                                    brysb_parts = []
                                    score = 0
                                    bIgnore = False
                                    for idx, m in enumerate(morph_parts):
                                        obj_fixed = Brysbaert.objects.filter(stimulus=m).first()
                                        if obj_fixed == None:
                                            score = -1
                                            bIgnore = True
                                            #break
                                            brysb_parts.append(-1)
                                        else:
                                            # Add the score to the list
                                            brysb_parts.append(obj_fixed.get_concreteness())
                                    # debugging: get the parts as string
                                    sParts = json.dumps(brysb_parts)

                                    # TRYING
                                    if bIgnore:
                                        if method == "erwin":
                                            lst_parts, brysb_parts, str_parts = Brysbaert.best_fit(word.text())
                                            # Do we have something?
                                            if len(lst_parts) > 0:
                                                bIgnore = False
                                            # debugging: get the parts as string
                                            sParts = json.dumps(brysb_parts)
                                            lemmatag = "{} ({})".format(lemmatag,str_parts)
                                        elif method == "hidde":
                                            # Break up the word *right-to-left* in largest Brysbaert known parts:
                                            lst_parts, brysb_parts, str_parts = Brysbaert.best_fit(word.text(), right_to_left=True)
                                            # Do we have something?
                                            if len(lst_parts) > 0:
                                                bIgnore = False
                                                # Check the last part not being [je]
                                                if len(word.text()) > 2 and lst_parts[0].stimulus == "je":
                                                    bIgnore = True
                                                    # Check if this is a diminuative that needs separate treatment
                                                    sKernwoord = strip_diminuative(word.text())
                                                    if sKernwoord != word.text():
                                                        # Break up the word *right-to-left* in largest Brysbaert known parts:
                                                        lst_parts, brysb_parts, str_parts = Brysbaert.best_fit(sKernwoord, right_to_left=True)
                                                        # Do we have something?
                                                        if len(lst_parts) > 0:
                                                            bIgnore = False
                                            # debugging: get the parts as string
                                            sParts = json.dumps(brysb_parts)
                                            lemmatag = "{} ({})".format(lemmatag,str_parts)
                                        elif method == "simple":
                                            # DOn't do anything additional
                                            pass
                                        elif method == "alpino":
                                            # Make use of Alpino parsing
                                            pass

                                    if not bIgnore: # Was: score == 0:
                                        # Determine the average
                                        for m in brysb_parts:
                                            score += m
                                        score = score / len(brysb_parts)
                                    else:
                                        # change the 'lemmatag' to reflect the breaking up of this word into morphemes
                                        lemmatag = "{} (={})".format(lemmatag, "-".join(morph_parts))
                            else:
                                # The concreteness of this word is in obj_fixed
                                score = obj_fixed.get_concreteness()
                            # Process the score of this content word
                            oScore = {}
                            oScore['word'] = word_text # word.text()
                            oScore['word_id'] = word_id
                            oScore['pos'] = postag
                            oScore['pos_full'] = pos.cls
                            oScore['lemma'] = lemmatag
                            oScore['concr'] = "NiB" if score < 0 else str(score)
                            oScore['homonyms'] = lst_homonyms
                            if len(lst_homonyms) > 0:
                                # By default select the first homonym
                                oScore['hnum'] = 1
                            oScore['idx'] = oWord['idx']
                            word_id += 1
                            # Add it in all lists
                            word_scores.append(oScore)

                    # Add any MWEs that were found
                    for oMWE in sent_mwes:
                        oScore = dict(word="", word_id=word_id, pos="",pos_full="", lemma="", concr=0.0)
                        # Get the text of the MWE
                        if mwe_method == "full":
                            sText = oMWE['full']
                            score = oMWE['score']
                            oScore['word'] = sText
                            oScore['word_id'] = word_id
                            oScore['pos'] = "MWE"
                            oScore['pos_full'] = "MWE"
                            oScore['lemma'] = sText
                            oScore['concr'] = str(score)
                            oScore['idx'] = oMWE['idx']
                            word_id += oMWE['size']
                        elif mwe_method in [ "lemmas", "anyorder"]:
                            # sText = " ".join(oMWE['lemmas'])
                            sText = oMWE['full']
                            score = oMWE['score']
                            oScore['word'] = sText
                            oScore['word_id'] = word_id
                            oScore['pos'] = "MWE"
                            oScore['pos_full'] = "MWE"
                            oScore['lemma'] = sText
                            oScore['concr'] = str(score)
                            oScore['idx'] = oMWE['idx']
                            word_id += 1
                        # Add it in all lists
                        word_scores.append(oScore)

                    # Sort word scores, based on idx
                    word_scores = sorted(word_scores, key = lambda x: x['idx'])

                    # Process the results of this sentence
                    score = 0
                    total = 0
                    n = len(word_scores)
                    for obj in word_scores:
                        if obj['concr'] == "NiB":
                            n -= 1
                        else:
                            score += float(obj['concr'])
                            total += float(obj['concr'])
                    if n < 1:
                        n = 1
                    # The average of the sentence is the sum of the word-scores divided by the number of word scores
                    avg = score / n
                    sent_scores.append({'score': avg, 'n': n, 'total': total, 'sentence': sent.text(), 'list': word_scores})

                # Process the results of this paragraph
                score = 0
                n = 0
                total = 0
                for obj in sent_scores:
                    total += obj['total']
                    n += obj['n']
                if n == 0:
                    avg = 0.0
                else:
                    avg = total / n
                para_scores.append({'score': avg, 'n': n, 'total': total, 'paragraph': para.text(), 'list': sent_scores})

            # Process the results of this text
            score = 0
            n = 0
            total = 0
            for obj in para_scores:
                total += obj['total']
                n += obj['n']
            if n == 0:
                avg = 0.0
            else:
                avg = total / n
            oText = {'text': self.name, 'score': avg, 'n': n, 'total': total, 'list': para_scores}

            # Add the concreteness as string
            self.concr = json.dumps(oText)
            self.save()
            bResult = True
            return bResult, sMsg
        except:
            bResult = False
            sMsg = oErr.get_error_message()
            return bResult, sMsg

    def recalculate(oConcr):
        """Re-calculate the concreteness measures and return"""

        oErr = ErrHandle()
        try:
            for oPara in oConcr['list']:
                for oSent in oPara['list']:
                    # Process the results of this sentence
                    score = 0
                    total = 0
                    n = len(oSent['list'])
                    for obj in oSent['list']:
                        if obj['concr'] == "NiB":
                            n -= 1
                        else:
                            score += float(obj['concr'])
                            total += float(obj['concr'])
                    if n < 1:
                        n = 1
                    # The average of the sentence is the sum of the word-scores divided by the number of word scores
                    avg = score / n
                    # Adapt the sentence totals
                    oSent['score'] = avg
                    oSent['n'] = n
                    oSent['total'] = total
                # Process the results of this paragraph
                score = 0
                n = 0
                total = 0
                for obj in oPara['list']:
                    total += obj['total']
                    n += obj['n']
                if n == 0:
                    avg = 0.0
                else:
                    avg = total / n
                oPara['score'] = avg
                oPara['n'] = n
                oPara['total'] = total
            # Process the results of this text
            score = 0
            n = 0
            total = 0
            for obj in oConcr['list']:
                total += obj['total']
                n += obj['n']
            if n == 0:
                avg = 0.0
            else:
                avg = total / n
            oConcr['score'] = avg
            oConcr['n'] = n
            oConcr['total'] = total
        except:
            msg = oErr.get_error_message()
            oErr.DoError("FrogLink.recalculate")
        return oConcr

    def get_csv(self):
        """Convert [concr] to CSV-string with header row""" 

        def get_total(oThis):
            total = oThis.get("total")
            if total is None or total == "NiB":
                sBack = "NiB"
            else:
                sBack = float(total)
            return sBack

        lCsv = []
        oErr = ErrHandle()
        try:
            # Start with the header
            oLine = "{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}".format("par", "snt", "wrd", "text", "score", "n", "pos", "total")
            lCsv.append(oLine)
            # Get the concr object
            oConcr = json.loads(self.concr)
            # Process measures for the text as a whole
            score = "NiB" if oConcr['score'] == "NiB" else float(oConcr['score'])
            total = get_total(oConcr)
            oLine = "{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}".format(
                "(text)", "(text)", "(text)", "(the whole text)", score, oConcr['n'], "", total)
            lCsv.append(oLine)
            # Do paragraphs
            for idx_p, para in enumerate(oConcr['list']):
                # Output a line for the paragraph
                score = "NiB" if para['score'] == "NiB" else float(para['score'])
                total = get_total(para)
                oLine = "{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}".format(
                    idx_p+1, "(para)", "(para)", para['paragraph'], score, para['n'], "", total)
                lCsv.append(oLine)
                # Do sentences
                for idx_s, sent in enumerate(para['list']):
                    # Output a line for the sentence
                    score = "NiB" if sent['score'] == "NiB" else float(sent['score'])
                    total = get_total(sent)
                    oLine = "{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}".format(
                        idx_p+1, idx_s+1, "(sent)", sent['sentence'], score, sent['n'], "", total)
                    lCsv.append(oLine)
                    # Do words
                    for idx_w, word in enumerate(sent['list']):
                        # Output a line for the paragraph
                        score = "NiB" if word['concr'] == "NiB" else float(word['concr'])
                        total = score
                        oLine = "{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}".format(
                            idx_p+1, idx_s+1, idx_w+1, word['lemma'], score, 1, word['pos_full'], total)
                        lCsv.append(oLine)
            # REturn the whole
            return "\n".join(lCsv)
        except:
            sMsg = oErr.get_error_message()
            oErr.DoError("FrogLink get_csv")
            return ""


class FoliaProcessor():
    """Functions to help create or process a folia document"""

    docstr = ""         # The identifier of this document
    username = ""       # The owner of this document
    dir = ""            # Directory for the output
    basicf = ""         # If any: where the basis folis is
    frogClient = None   
    frog = None
    doc = None
    re_single = None
    re_double = None

    def __init__(self, username):
        # Check and/or create the appropriate directory for the user
        dir = os.path.abspath(os.path.join( WRITABLE_DIR, "../folia", username))
        if not os.path.exists(dir):
            os.mkdir(dir)
        # Set the dir locally
        self.dir = dir
        # Set regex
        self.re_single = re.compile(u"[‘’´]")
        self.re_double = re.compile(u"[“”]")

    def location(self):
        """Give the location we can work from: local, client or remote"""

        if self.frog != None:
            sBack = "local"
        elif self.frogClient != None:
            sBack = "client"
        else:
            sBack = "remote"
        return sBack

    def basis_folia(self, filename, data_contents):

        errHandle = ErrHandle()
        bReturn = False
        sMsg = ""
        try:
            # Read file into array
            lines = []
            bFirst = True
            for line in data_contents:
                if isinstance(line, str):
                    sLine = line.strip()
                else:
                    sLine = line.decode("utf-8").strip()
                if bFirst:
                    sLine = sLine.replace(u'\ufeff', '')
                    bFirst = False
                # Change curly quotes
                sLine = self.re_single.sub("'", sLine)
                sLine = self.re_double.sub('"', sLine)
                # Insert a space before ".." or "..."
                sLine = sLine.replace("..", " ..")
                # Remove some bad symbols
                sLine = sLine.replace(">", "")
                sLine = sLine.replace("<", "")
                # Put a space before a comma
                sLine = sLine.replace(",", " ,")
                lines.append(sLine)

            # Check and/or create the appropriate directory for the user
            dir = self.dir

            # create a folia document with a numbered id
            docstr = os.path.splitext( os.path.basename(filename))[0].replace(" ", "_").strip()
            # Make sure it starts with a letter
            if not re.match(r'^[a-zA-Z_]', docstr):
                docstr = "t_" + docstr
            # Make sure we remember the docstr
            self.docstr = docstr

            # Start creating a folia document
            doc = folia.Document(id=docstr)
            text = doc.add(folia.Text)

            # Walk through the JSON 
            lFolia = []
            counter = 0
            for sLine in lines:
                # Check for empty
                sLine = sLine.strip()
                if sLine != "":
                    counter += 1
                    lFolia.append("{}: {}".format(counter, sLine))
                    # Append a paragraph
                    para = text.add(folia.Paragraph)

                    # NOTE: do *NOT* Set the <t> of this paragraph -- it will differ from what Folia calculates itself
                    # para.settext(sLine)

                    # Split text into sentences
                    sf = StringIO(sLine)
                    for s_list in pynlpl.textprocessors.Tokenizer(sf, splitsentences=True):
                        for s in s_list:
                            sentence = para.add(folia.Sentence)
                            for token in s:
                                # Strip off anything from x0100 upwards (unicode)
                                token = re.sub(r'[^\x00-\xFF]', '', token)
                                # ANything left?
                                if token != "":
                                    sentence.add(folia.Word, token)
                            # Add text to sentence
                            sentence.settext(sentence.text())
                    # Add text to paragraph
                    para.settext(para.text())
            sMsg = "\n".join(lFolia)

            # THink of a correct name for Basic folia
            f =  os.path.abspath(os.path.join(dir, docstr) + ".basis.xml")
            self.basicf = f
            # Save the doc
            doc.save(f)
            # Make sure the doc is in this class instance
            self.doc = doc
            # Set positive return
            bReturn = True
        except:
            sMsg = errHandle.get_error_message()
            errHandle.DoError("basis_folia")
            bReturn = False
        # Return what has happened
        return bReturn, sMsg


class Homonym(models.Model):
    """Homonym list"""

    # [1] The lemma 
    stimulus = models.CharField("Lemma of word", max_length=MAXPARAMLEN)
    # [0-1] Possibly the POS tag
    postag = models.CharField("POS tag", blank=True, null=True, max_length=MAXPARAMLEN)
    # [0-1] The meaning of this variant
    meaning = models.TextField("Meaning", blank=True, null=True)
    # [1] Metric 1: concrete_m
    m = models.FloatField("Concrete m", default=0.0)

    def __str__(self):
        return self.stimulus

    def clear():
        Homonym.objects.all().delete()
        return True

    def find_or_create(stimulus, m, postag=None):
        """Find existing or create new item"""
        
        obj = None
        sMsg = ""
        oErr = ErrHandle()
        try:
            obj = Homonym.objects.filter(stimulus=stimulus).first()
            if obj == None:
                obj = Homonym.objects.create(stimulus=stimulus, m=m, postag=postag)
        except:
            sMsg = oErr.get_error_message()
            oErr.DoError("find_or_create")
        # Return the result
        return obj, sMsg

    def get_concreteness(self):
        return self.m


class Neologism(models.Model):
    """Neologism list"""

    # [1] The lemma 
    stimulus = models.CharField("Lemma of word", max_length=MAXPARAMLEN)
    # [1] Metric 1: concrete_m
    m = models.FloatField("Concrete m", default=0.0)
    # [0-1] Possibly the POS tag
    postag = models.CharField("POS tag", blank=True, null=True, max_length=MAXPARAMLEN)

    def __str__(self):
        return self.stimulus

    def clear():
        Neologism.objects.all().delete()
        return True

    def find_or_create(stimulus, m, postag=None):
        """Find existing or create new item"""
        
        obj = None
        sMsg = ""
        oErr = ErrHandle()
        try:
            obj = Neologism.objects.filter(stimulus=stimulus).first()
            if obj == None:
                obj = Neologism.objects.create(stimulus=stimulus, m=m, postag=postag)
        except:
            sMsg = oErr.get_error_message()
            oErr.DoError("find_or_create")
        # Return the result
        return obj, sMsg

    def get_concreteness(self):
        return self.m


class Wordlist(models.Model):
    """A wordlist has a name and provides a container for a number of [Worddef] items"""

    # [1] A wordlist must have a name
    name = models.CharField("Name", max_length=MAXPARAMLEN)

    # [0-1] Optional description
    description = models.TextField("Description", blank=True, null=True)

    # [0-1] File for uploading
    upload = models.FileField("Excel file", null=True, blank=True, upload_to=wordlist_path)
    # [0-1] Name of the worksheet from which info must be loaded
    sheet = models.CharField("Worksheet", blank=True, null=True, max_length=MAXPARAMLEN)

    # [1] Each Froglink has been created at one point in time
    created = models.DateTimeField(default=timezone.now)
    # [0-1] Time this module was last updated
    saved = models.DateTimeField(null=True, blank=True)

    def __str__(self):
        return self.name

    def save(self, force_insert = False, force_update = False, using = None, update_fields = None): #, *args, **kwargs):
        # Adapt the save date
        self.saved = timezone.now()
        # Now do the saving
        response = super(Wordlist, self).save(force_insert, force_update, using, update_fields)
        # Return the response
        return response

    def get_created(self):
        sBack = self.created.strftime("%d/%B/%Y (%H:%M)")
        return sBack

    def get_saved(self):
        sBack = "-"
        if self.saved != None:
            sBack = self.saved.strftime("%d/%B/%Y (%H:%M)")
        return sBack

    def get_upload(self):
        """If file has been filled in, get the file name"""

        sBack = "-"
        if not self.upload is None:
            sBack = self.upload
        return sBack

    def get_worddef(self, oItem):
        """Get (and possibly create) a worddef"""

        oErr = ErrHandle()
        obj = None
        try:
            woord = oItem.get("woord")
            score = oItem.get("score")
            postag = oItem.get("postag")

            if postag is None:
                obj = Worddef.objects.filter(wordlist=self, stimulus__iexact=woord).first()
            else:
                obj = Worddef.objects.filter(wordlist=self, stimulus__iexact=woord, postag__iexact=postag).first()
            if obj is None:
                # It is not yet there, so create it
                obj = Worddef.objects.create(wordlist=self, stimulus=woord, m=score)
                if not postag is None and postag != "":
                    obj.postag = postag
                    obj.save()
            else:
                # Double check the score
                if obj.m != score:
                    obj.m = score
                    obj.save()
        except:

            msg = oErr.get_error_message()
            oErr.DoError("Wordlist/get_worddef")
        return obj

    def read_upload(self):
        """Import or re-import the XLSX file's worksheet"""

        def get_postag(rCatch, postag_dict, woord, postag):
            """Either copy a postag or derive it from the word"""

            oErr = ErrHandle()
            try:
                # Just make sure we have no leading of trailing blanks
                woord = woord.strip()

                if postag is None or postag == "":
                    # Try to get the postag out of [woord]
                    if "(" in woord and ")" in woord:
                        # There should be a postag inside [woord]
                        result = rCatch.match(woord)
                        if result:
                            woord = result.group(1)
                            postag = result.group(2)
                # Possibly change POSTAG
                if not postag is None and postag in postag_dict:
                    postag = postag_dict[postag]
            except:
                msg = oErr.get_error_message()
                oErr.DoError("Wordlist/get_postag")
            return woord, postag

        oErr = ErrHandle()
        sContent = ""
        rCatch = re.compile( r'(.*)\s\(([A-Z]+)\)')
        postag_dict = dict(ZN="N", BNW="ADJ", WW="WW", TUSSENWERPSEL="BW")
        try:
            # Get the file and read it
            data_file = self.upload
            sheet = self.sheet

            # Check if it exists
            if not data_file is None and data_file != "":

                # Use openpyxl to read the correct worksheet
                wb = openpyxl.load_workbook(data_file, read_only=True)
                # Try to find the correct worksheet
                # worksheets = wb.worksheets
                ws = None
                if sheet is None:
                    ws = wb.active
                else:
                    for ws_this in wb:
                        if ws_this.title.lower() == sheet.lower():
                            ws = ws_this 
                            break
                # Found it?
                if not ws is None:
                    # Expectations: first  column is *WOORD*
                    #               second column is *SCORE* (concreetheid)
                    #  (optional)   third  column is *POS* (part of speech)
                    row_no = 2
                    while ws.cell(row=row_no, column=1).value != None:
                        # Get the woord and the score
                        woord = ws.cell(row=row_no, column=1).value
                        score = ws.cell(row=row_no, column=2).value
                        postag = ws.cell(row=row_no, column=3).value
                        if isinstance(score,float):
                            # Turn it into a string with a period decimal separator
                            score = str(score).replace(",", ".")
                        woord, postag = get_postag(rCatch, postag_dict, woord, postag)
                        oItem = dict(woord=woord, score=score, postag=postag)

                        # Add the item to the list of [Worddef] for this [Wordlist]
                        obj = self.get_worddef(oItem) 

                        # Go to the next row
                        row_no += 1


        except:
            msg = oErr.get_error_message()
            oErr.DoError("Wordlist/read_upload")
        return sContent


class Worddef(models.Model):
    """Generic word definition with concreteness score"""

    # [1] A generic word definition belongs to a particular wordlist
    wordlist = models.ForeignKey(Wordlist, on_delete=models.CASCADE, related_name="wordlist_worddefs")
    # [1] The lemma 
    stimulus = models.CharField("Lemma of word", max_length=MAXPARAMLEN)
    # [1] Metric 1: concrete_m
    m = models.FloatField("Concrete m", default=0.0)
    # [0-1] Possibly the POS tag
    postag = models.CharField("POS tag", blank=True, null=True, max_length=MAXPARAMLEN)

    def __str__(self):
        sBack = "{}: {}".format(self.wordlist.name, self.stimulus)
        return sBack

    def save(self, force_insert = False, force_update = False, using = None, update_fields = None): # , *args, **kwargs):
        # Now do the saving
        response = super(Worddef, self).save(force_insert, force_update, using, update_fields)
        # Return the response
        return response

    def clear(name):
        bResult = False
        oErr = ErrHandle()
        try:
            # Remove all the word definitions from the wordlist that is indicated
            if not name is None:
                wordlist = Wordlist.objects.filter(name__iexact=name).first()
                if not wordlist is None:
                    # Remove all the word definitions associated with this wordlist
                    wordlist.wordlist_worddefs.all().delete()
                    bResult = True
        except:
            msg = oErr.get_error_message()
            oErr.DoError("Worddef/clear")
        return bResult

    def find_or_create(name, stimulus, m, postag=None):
        """Find existing or create new item"""
        
        obj = None
        sMsg = ""
        oErr = ErrHandle()
        try:
            # Find the right wordlist
            if not name is None:
                wordlist = Wordlist.objects.filter(name__iexact=name).first()
                if not wordlist is None:
                    obj = Worddef.objects.filter(wordlist=wordlist, stimulus=stimulus).first()
                    if obj == None:
                        obj = Worddef.objects.create(wordlist=wordlist, stimulus=stimulus, m=m, postag=postag)
        except:
            sMsg = oErr.get_error_message()
            oErr.DoError("find_or_create")
        # Return the result
        return obj, sMsg

    def get_concreteness(self):
        return self.m

    def get_postag(self):
        """Get the POS tag, if that is specified"""

        sBack = "-"
        if not self.postag is None:
            sBack = self.postag
        return sBack

    def get_wordlist(self, html=False):
        """Get the name of the wordlist"""

        sBack = ""
        oErr = ErrHandle()
        try:
            if not self.wordlist is None:
                sBack = self.wordlist.name
                if html:
                    url = reverse("wordlist_details", kwargs={'pk': self.wordlist.id})
                    sBack = "<span class='badge signature gr'><a href='{}'>{}</a></span>".format(
                        url, sBack)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("Worddef/get_wordlist")
        return sBack


class Brysbaert(models.Model):
    """Information from the Brysbaert list of concreteness

    Fields in the Excel:
       stimulus, List, Concrete_m, Concrete_sd, Number_of_ratings, Number_of_N-respones, Number_of_subjects
    """

    # [1] The lemma 
    stimulus = models.CharField("Lemma of word", max_length=MAXPARAMLEN)
    # [1] The list number
    list = models.IntegerField("List number", default=0)
    # [1] Metric 1: concrete_m
    m = models.FloatField("Concrete m", default=0.0)
    # [1] Metric 2: concrete_sd
    sd = models.FloatField("Concrete sd", default=0.0)
    # [1] Metric 3: Number_of_ratings
    ratings = models.FloatField("Number_of_ratings", default=0.0)
    # [1] Metric 4: Number_of_N-responses
    responses = models.FloatField("Number_of_N-responses", default=0.0)
    # [1] Metric 5: Number_of_subjects
    subjects = models.FloatField("Number_of_subjects", default=0.0)

    def __str__(self):
        return self.stimulus

    def best_fit(woord, right_to_left=False):
        """Break up a word in parts, and find the best Brysbaert fit"""

        def get_best(sWord, lst_id, lst_m):
            """Left-to-right version of get_best()"""
            sBest = ""
            lExpression = []

            # Sanity check
            if sWord != "":
                for ln in range(len(sWord)):
                    # Add this to the expression
                    if ln > 0:
                        length = ln+1
                        if right_to_left:
                            lExpression.append(sWord[-length:])
                        else:
                            lExpression.append(sWord[0:length])
                # Now make a regular search expression
                sExpression = "^({})$".format( "|".join(lExpression))
                # Search for the longest entry in BrysBaert
                lst_qs = Brysbaert.objects.filter(stimulus__iregex=sExpression).values('stimulus', 'id')
                # Do we have a match?
                if len(lst_qs) > 0:
                    # Yes, we have a match - at least partial
                    lst_sorted = sorted( lst_qs, key=lambda o:len(o['stimulus']), reverse=True)
                    oMatch = lst_sorted[0]
                    sMatch = oMatch['stimulus']
                    idMatch = oMatch['id']
                    # Add this to the list we return
                    lst_id.append(idMatch)
                    lst_m.append(sMatch)
                    # COntinue searching and adding
                    if right_to_left:
                        # Right to left: the new string is *until* the match
                        end = len(sWord) - len(sMatch)
                        new_string = sWord[0:end]
                    else:
                        # Left to right: the new string starts just *after* the match
                        new_string = sWord[len(sMatch):]
                    return get_best(new_string, lst_id, lst_m)
                else:
                    # No match: try starting from the next character
                    lst_id.append(-1)
                    lst_m.append(sWord[1:1])
                    return get_best(sWord[1:], lst_id, lst_m)
            return lst_id, lst_m

        oErr = ErrHandle()
        sBack = ""
        lst_part = []
        lst_brysb = []
        ignore_bound_morphemes_pref = ['af', 'ver', 'in']   # Initial or medial
        ignore_bound_morphemes_post = ['en']                # Final only
        try:
            # Get a list of Brysbaert ID's
            lst_id, lst_m = get_best(woord, [], [])
            # Get the actual Brysbaert objects in a list
            if len(lst_id) > 0:
                lst_part = []
                lst_skip = []
                # Reverse the list if this is needed
                if right_to_left:
                    # Yes, reverse the lists
                    lst_id.reverse()
                    lst_m.reverse()
                with transaction.atomic():
                    for idx, id in enumerate(lst_id):
                        is_last = (idx == len(lst_id) -1)
                        if id < 0:
                            lst_skip.append(lst_m[idx])
                        else:
                            obj = Brysbaert.objects.filter(id=id).first()
                            bSkip = False
                            if is_last and obj.stimulus in ignore_bound_morphemes_post:
                                bSkip = True
                            elif (not is_last) and (obj.stimulus in ignore_bound_morphemes_pref):
                                bSkip = True
                            if not bSkip:
                                lst_part.append(obj)
                                lst_brysb.append(obj.get_concreteness())
                    # We now have all valid parts
                iStop = 1
            # Create a string of the stuf we found
            sBack = "-".join([x.stimulus for x in lst_part])
        except:
            msg = oErr.get_error_message()
            oErr.DoError("Brysbaert/best_fit")
            lst_part = []
            lst_brysb = []
        return lst_part, lst_brysb, sBack

    def clear():
        Brysbaert.objects.all().delete()
        return True

    def find_or_create(stimulus, listnum, m, sd, ratings, responses, subjects):
        """Find existing or create new item"""
        
        obj = None
        sMsg = ""
        oErr = ErrHandle()
        try:
            obj = Brysbaert.objects.filter(stimulus=stimulus).first()
            if obj == None:
                obj = Brysbaert(stimulus=stimulus, list=listnum, m=m, sd=sd, ratings=ratings, responses=responses, subjects=subjects)
                obj.save()
        except:
            sMsg = oErr.get_error_message()
            oErr.DoError("find_or_create")
        # Return the result
        return obj, sMsg

    def get_concreteness(self):
        return self.m


# ================= POS-TAGGING TWITTER ========================================


class TwitterMsg(models.Model):
    """Room for one twitter message, its set of tokens and its POS tagging"""

    # [1] Must have a message
    message = models.TextField("Message", blank=False, null=False)
    # [1] The coordinate / location of this message
    coordinate = models.CharField("Coordinate", max_length=MAXPARAMLEN)
    # [0-1] The row number within Excel
    row = models.IntegerField("Row", blank=True, null=True)

    # [0-1] Should be tokenized plain text string
    tokens = models.TextField("Tokens", blank=True, null=True)
    # [0-1] Should be POS-tagged JSON string
    postags = models.TextField("POS tags", blank=True, null=True)

    def __str__(self):
        return self.coordinate

    def check_files(self):
        """Check if all files that can be made on the basis of what we have are there"""

        bResult = True
        oErr = ErrHandle()
        try:
            bare_file = self.get_filename()
            # Check the text file
            text_file = "{}.txt".format(bare_file)
            if not os.path.exists(text_file):
                # Save it
                with open(text_file, "w", encoding="utf-8") as f:
                    f.write(self.message)

            # Check the tokens
            tokens_file = "{}.tok".format(bare_file)
            if not self.tokens is None:
                if not os.path.exists(tokens_file):
                    # Save it
                    with open(tokens_file, "w", encoding="utf-8") as f:
                        f.write(self.tokens)
            elif os.path.exists(tokens_file):
                # The file exists - do we need to read it?
                if self.tokens is None or self.tokens == "":
                    # Read it
                    sResult = None
                    with open(tokens_file, "r", encoding="utf-8") as f:
                        sResult = f.read()
                    if not sResult is None:
                        self.tokens = sResult.strip()
                        self.save()

            # Check the tokens
            postags_file = "{}.json".format(bare_file)
            if not self.postags is None:
                if not os.path.exists(postags_file):
                    # Save it
                    with open(postags_file, "w", encoding="utf-8") as f:
                        f.write(self.postags)
            elif os.path.exists(postags_file):
                # The file exists - do we need to read it?
                if self.postags is None or self.postags == "":
                    # Read it
                    sResult = None
                    with open(postags_file, "r", encoding="utf-8") as f:
                        sResult = f.read()
                    if not sResult is None:
                        self.postags = sResult.strip()
                        self.save()
        except:
            msg = oErr.get_error_message()
            oErr.DoError("check_files")

        return bResult

    def get_filename(self):
        """Construct and pass on the filename for this twitter message"""

        sBack = ""
        oErr = ErrHandle()
        try:
            # Check and/or create the appropriate directory for the user
            dir = os.path.abspath(os.path.join( WRITABLE_DIR, "../folia", "twitter"))
            if not os.path.exists(dir):
                os.mkdir(dir)

            # Check for row-number
            if not self.row is None:
                dir = os.path.abspath(os.path.join(dir, "row{:04}".format(self.row)))
                if not os.path.exists(dir):
                    os.mkdir(dir)

            # Combine into a filename
            if self.row is None:
                sBack = os.path.abspath(os.path.join(dir, "tw_{}".format(self.coordinate)))
            else:
                sBack = os.path.abspath(os.path.join(dir, "tw_{:04}_{}".format(self.row, self.coordinate)))
            # NOTE: 
            #   the calling program should append e.g. ".txt" or what is needed
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_filename")

        return sBack

    def twitter_ucto(self, coordinate, content, clamuser, clampw):
        """Perform UCTO on the 'content' and return the result"""
        oErr = ErrHandle()

        sBack = ""
        try:
            # Think of a project name
            project = "cesar_twitter_{}".format(coordinate)
            basicauth = True
            # Get access to the webservice
            clamclient = CLAMClient(uctourl, clamuser, clampw, basicauth = basicauth)
            # First delete any previous project, if it exists
            try:
                result = clamclient.delete(project)
                errHandle.Status("Removed previous UCTO project {} = {}".format(project, result))
            except:
                # No problem: no project has been removed
                pass
            # Only now start creating it
            result = clamclient.create(project)
            errHandle.Status("Created new UCTO project {} = {}".format(project, result))
            data = clamclient.get(project)

        except:
            sError = oErr.get_error_message()
            oErr.DoError("")
        return sBack

    def twitter_cell(self, coordinate, content, clamuser=None, clampw=None):
        """Process one twitter cell's contents with coordinate within Excel (e.g. W12)"""

        oBack = {'status': 'ok', 'count': 0, 'msg': "", 'user': username}
        frogType = "remote"     # Fix to remote -- or put to None if automatically choosing
        iCount = 0
        oErr = ErrHandle()

        try:
            if clamuser is None:
                clamuser = TsgInfo.get_value("clam_user")
                clampw = TsgInfo.get_value("clam_pw")

            # Check our location
            frogLoc = frogType if frogType != None else folProc.location()

            # Directory: one directory for each user
            dir = folProc.dir

            # Perform UCTO
            bResult, sMsg = self.twitter_ucto(coordinate, content, clamuser, clampw)
            if not bResult:
                # There was some kind of error
                oBack['status'] = 'error'
                oBack['msg'] = "ucto load error: {}".format(sMsg)
            else:
                pass

        except:
            sError = oErr.get_error_message()
            oBack['status'] = 'error'
            oBack['msg'] = sError

        # Return the object that has been created
        return oBack

    def twitter_excel(self, username, data_file, filename):
        """Prepare a TWITTER Excel and add POS tags"""

        oBack = {'status': 'ok', 'count': 0, 'msg': "", 'user': username}
        oErr = ErrHandle()
        try:
            # (1) Read the Excel and extract the cells with Twitter text

            # (2) Process all the twitter texts

            # (2.1) Tokenize using UCTO

            # (2.2) Find POS tags

            # (2.3) Re-combine tokens + POS tags

            # (3) Adapt the changed Excel cells

            # (4) Make the output available

            pass
        except:
            sError = oErr.get_error_message()
            oBack['status'] = 'error'
            oBack['msg'] = sError

        # Return the object that has been created
        return oBack



# ======================= NEXIS UNI =======================

class NexisDocs(models.Model):
    """Set of text files for Nexis research"""
    
    # [1] These belong to a particular user
    owner = models.ForeignKey(User, editable=False, on_delete=models.CASCADE, related_name="owner_nexisdocs")

    def __str__(self):
        return self.owner.username

    def get_obj(username):
        nd = None
        owner = User.objects.filter(username=username).first()
        if owner != None:
            nd = NexisDocs.objects.filter(owner=owner).first()
        return nd


class NexisBatch(models.Model):
    """A batch is a set of files that have been uploaded on a certain time"""

    # [1] Each Batch has been created at one point in time
    created = models.DateTimeField(default=timezone.now)
    # [0-1] Eatch batch has a number of files
    count = models.IntegerField("Number of files", default=0)
    # [1] Each batch belongs to a set of docs (that belong to an owner)
    ndocs = models.ForeignKey(NexisDocs, related_name="ndocsbatches", on_delete=models.CASCADE)

    def __str__(self):
        sBack = str(self.created)
        return sBack

    def create(username):
        """Possibly create a new batch for user [username]"""

        errHandle = ErrHandle()

        try:
            # Get the correct user
            owner = User.objects.filter(username=username).first()

            # Find a NexisDocs instance for this user
            fd = NexisDocs.objects.filter(owner=owner).first()
            if fd == None:
                fd = NexisDocs(owner=owner)
                fd.save()

            # Create it
            obj = NexisBatch(ndocs=fd)
            obj.save()
            # Return the result 
            return obj, ""
        except:
            errHandle.DoError("NexisBatch/create")
            return None, errHandle.get_error_message()


class NexisLink(models.Model):
    """Basic information from a Nexis text file: text + metadata"""

    # [1] Each froglink centers around a file that is uploaded, processed and made available
    name = models.CharField("Name to be used for this file", max_length=MAXPARAMLEN)
    # [0-1] Full name is the full path of the txt-file on the server
    fullname = models.CharField("Full path of this file", max_length=MAXPATH, null=True, blank=True)
    # [0-1] Text metadata as stringified JSON object
    nmeta = models.TextField("Nexis metadata", null=True, blank=True)
    # [0-1] Text body
    nbody = models.TextField("Nexis text", null=True, blank=True)
    # [1] Each Froglink has been created at one point in time
    created = models.DateTimeField(default=timezone.now)

    # [1] Each link belongs to a set of docs (that belong to an owner)
    ndocs = models.ForeignKey(NexisDocs, related_name="nexisdocuments", on_delete=models.CASCADE)
    # [1] Each nexislink should belong to a batch
    batch = models.ForeignKey(NexisBatch, null=True, on_delete=models.CASCADE, related_name="batchlinks")

    def __str__(self):
        return self.name

    def create(name, username, batch):
        """Possibly create a new item [name] for user [username]"""

        errHandle = ErrHandle()

        try:
            # Get the correct user
            owner = User.objects.filter(username=username).first()

            # Find a NexisDocs instance for this user
            fd = NexisDocs.objects.filter(owner=owner).first()
            if fd == None:
                fd = NexisDocs(owner=owner)
                fd.save()

            #obj = NexisLink.objects.filter(name=name).first()
            #if obj == None:
            # Just always Create it
            obj = NexisLink(name=name, ndocs=fd, batch=batch)
            obj.save()
            # Return the result 
            return obj, ""
        except:
            errHandle.DoError("NexisLink/create")
            return None, errHandle.get_error_message()

    def read_doc(self, username, data_file, filename, arErr, xmldoc=None, sName = None, oStatus = None):
        """Import a text file, split in text and metadata and store that in the NexisLink instance
        
        The syntax of a file:
            <nexisdoc> := <metadata> <body> <footer>

            <metadata> := <title> <newspaper> <date> <copyright> <meta>*

            <title> := $line NL
            <newspaper> := $line NL
            <date> := $line NL
            <copyright := $line NL

            <meta> := <keyword> ':' <metatext> NL
            <keyword> := 'Section' | 'Length' | 'Byline' | 'Highlight' |'Dateline'
            <metatext> := $line ( NL $line )*

            <body> := 'Body' NL ( $line NL )*

            <footer> := <loaddate> <eod>
            <loaddate> := ‘Load-Date:’ $line NL
            <eod> := ‘End of Document’ NL
        """

        def skip_empty_lines(lst, idx):
            while idx < len(lst) and lst[idx] == "": idx += 1
            return idx

        def get_anywhere(lst, must, start_from = 0, eow=False, year=False):
            bFound = False
            item_found = None
            index_found = -1
            for idx, item in enumerate(lst[start_from:]):
                if bFound:
                    # index_found = idx + start_from
                    break
                item_lower = item.lower()
                for m in must:
                    if m in item_lower: 
                        if eow:
                            pattern = r'.*{}\s*$'.format(m)
                            if re.match(pattern, item_lower):
                                bFound = True
                                item_found = item
                                index_found = idx + start_from
                                break
                        elif year:
                            pattern = r'.*\d\d\d\d\s*.*$'
                            if re.match(pattern, item_lower):
                                bFound = True
                                item_found = item
                                index_found = idx + start_from
                                break
                        else:
                            bFound = True
                            item_found = item
                            index_found = idx + start_from
                            break
            # Found anything?
            index_found += 1
            return bFound, item_found, index_found

        def get_line_item(lst, idx, inside = None, must = None):
            item = lst[idx]
            # Is there a 'must'?
            if must != None:
                bFound = False
                item_lower = item.lower()
                for m in must:
                    if m in item_lower: 
                        bFound = True
                        break
                if not bFound:
                    # Check if we can find it from here
                    idx_local = idx + 1
                    while idx_local < len(lst) and not bFound:
                        item = lst[idx_local]
                        item_lower = item.lower()
                        bFound = False
                        for m in must:
                            if m in item_lower: 
                                bFound = True
                                break
                        if not bFound:
                            idx_local += 1
                    if not bFound:
                        return idx, None
            elif inside != None:
                # Check if we can append to [item] from the next non-empty line
                bReady = False
                bFound = False
                while not bReady:
                    # GO to next non-empty line
                    idy = skip_empty_lines(lst, idx+1)
                    # Check if the contents is inside [inside]
                    if lst[idy] not in inside:
                        bReady = True
                    else:
                        idx = idy
                        bFound = True
                # Has something been found?
                if not bFound:
                    return idx, None

            # Skip any following empty lines
            idx = skip_empty_lines(lst, idx+1)
            # Return the index of the first non-empty line + the item we found
            return idx, item

        def get_line_meta(lst, idx):
            key = ""
            value = ""
            try:
                item = lst[idx]
                if ":" in item:
                    colon = item.index(":")
                    key = item[:colon]
                    value = item[colon+1:].strip()
                    # Go to the next line
                    idx += 1
                    # Check if the next line is not empty
                    while idx < len(lst) and lst[idx] != "" and ":" not in lst[idx]:
                        value = "{} {}".format(value, lst[idx])
                        idx += 1
                    while idx < len(lst) and lst[idx] == "": idx += 1

                return idx, key, value
            except:
                sError = errHandle.get_error_message()
                return -1, "", ""

        oBack = {'status': 'ok', 'count': 0, 'msg': "", 'user': username}
        errHandle = ErrHandle()
        oDoc = None
        iCount = 0
        inputType = "nexistext"
        nexisProc = NexisProcessor(username)
        day = ['maandag', 'dinsdag', 'woensdag', 'donderdag', 'vrijdag', 'zaterdag', 'zondag',
               'monday', 'tuesday', 'wednesday', 'thursday','friday', 'saturday', 'sunday']
        paper = ['telegraaf', 'dagblad', 'handelsblad', 'trouw', 'volkskrant', 'nrc.next', 'nrc']
        copyright = ['copyright']
        highlight = ['highlight:']
        byline = ['byline:']
        section = ['section:']

        nexis_filename = ""

        try:
            # Provide the right nexis filename
            nexis_filename = "nexislink_{}.txt".format(str(self.id).zfill(6))

            # Read and create basis-folia
            # Note: changed from [filename] to [nexis_filename]
            oResult = nexisProc.basis_text(nexis_filename, data_file)
            if not oResult['okay']:
                # There was some kind of error
                oBack['status'] = 'error'
                oBack['msg'] = "basis_text load error: {}".format( oResult['msg'])
            else:
                # DEBUG: show the frog location
                errHandle.Status("DEBUG: basis_text has been read")

                # Get the metadata and the body
                lst_meta = oResult['metadata']
                lst_body = oResult['body']

                # Process the metadata
                oMeta = {}
                bTitle = False
                # Add all the metadata 
                oMeta['raw'] = "\n".join(lst_meta)
                # Skip empty lines
                iLine = skip_empty_lines(lst_meta, 0)
                # Get the title, newspaper, date and copyright
                iLine, oMeta['title'] = get_line_item(lst_meta, iLine, inside=filename)
                if oMeta['title'] == None: 
                    oMeta['title'] = filename.replace(".txt", "")
                else:
                    bTitle = True
                
                bFound, oMeta['newspaper'], iLine = get_anywhere(lst_meta, must=paper, eow=True) # , start_from=iLine+1)
                if oMeta['newspaper'] == None:
                    oMeta['newspaper'] = "(Unknown newspaper)"
                elif not bTitle:
                    # See if he previous position has the title
                    if lst_meta[iLine-2] != "":
                        oMeta['title'] = lst_meta[iLine-2]

                bFound, oMeta['newsdate'], iLine = get_anywhere(lst_meta, must=day, year=True)
                # iLine, oMeta['newsdate'] = get_line_item(lst_meta, iLine, must=day)
                if oMeta['newsdate'] == None:
                    oMeta['newsdate'] = "(Cannot find the news date)"
                elif "Unknown" in oMeta['newspaper']:
                    # Take the newspaper as the line immediately preceding the newsdate
                    oMeta['newspaper'] = lst_meta[iLine-2]
                    if not bTitle and lst_meta[iLine-3] != "":
                        oMeta['title'] = lst_meta[iLine-3]

                bFound, oMeta['copyright'], iLine = get_anywhere(lst_meta, must=copyright)
                # iLine, oMeta['copyright'] = get_line_item(lst_meta, iLine, must=copyright)
                if oMeta['copyright'] == None:
                    oMeta['copyright'] = "(no copyright line)"

                # Get any more metadata, starting from the first line with colon
                iLine = 1
                while iLine < len(lst_meta) and not ":" in lst_meta[iLine]:
                    iLine += 1
                while iLine < len(lst_meta) and ":" in lst_meta[iLine]:
                    # Extract one meta element
                    iLine, key, value = get_line_meta(lst_meta, iLine)
                    # NOTE: the key must not contain a space
                    if key != "" and value != "" and not " " in key:
                        oMeta[key.lower()] = value
                # Any last line left?
                if iLine == len(lst_meta)-1 and ":" not in lst_meta[iLine]:
                    # There is still one final line -- not sure what that is though
                    oMeta['last'] = lst_meta[iLine]

                self.nmeta = json.dumps( oMeta)
                self.nbody = "\n".join(lst_body)
                self.save()
                iCount += 1
                oBack['name'] = self.name
                oBack['title'] = oMeta['title']

            # Make sure the requester knows how many have been added
            oBack['count'] = iCount   # The number of files added

        except:
            sError = errHandle.get_error_message()
            oBack['status'] = 'error'
            oBack['msg'] = sError

        # Return the object that has been created
        return oBack
    

class NexisProcessor():
    """Functions to help create or process a nexis uni document"""

    docstr = ""         # The identifier of this document
    username = ""       # The owner of this document
    dir = ""            # Directory for the output
    basicf = ""         # If any: where the basis folia is
    frogClient = None   
    frog = None
    doc = None
    re_single = None
    re_double = None
    bDebug = True

    def __init__(self, username):
        # Check and/or create the appropriate directory for the user
        dir = os.path.abspath(os.path.join( WRITABLE_DIR,  "../nexis", username))
        if not os.path.exists(dir):
            os.mkdir(dir)
        # Set the dir locally
        self.dir = dir
        # Set regex
        self.re_single = re.compile(u"[‘’´]")
        self.re_double = re.compile(u"[“”]")

    def basis_text(self, filename, data_contents):

        errHandle = ErrHandle()
        oBack = dict(okay=True)
        try:
            errHandle.Status("NexisProcessor: {}".format(filename))
            # Read file into array
            lines = []
            bFirst = True
            idx = 0
            for line in data_contents:
                idx += 1
                # ===================================================
                if self.bDebug:
                    errHandle.Status("[{}]: [{}]".format(idx, line))
                # ===================================================
                sLine = line.decode("utf-8").strip()
                if bFirst:
                    sLine = sLine.replace(u'\ufeff', '')
                    bFirst = False
                # Change curly quotes
                sLine = self.re_single.sub("'", sLine)
                sLine = self.re_double.sub('"', sLine)
                # Insert a space before ".." or "..."
                sLine = sLine.replace("..", " ..")
                lines.append(sLine)

            # Check and/or create the appropriate directory for the user
            dir = self.dir

            # ===================================================
            if self.bDebug:
                # Save the file just to be sure
                save_name = os.path.abspath(os.path.join(dir, filename))
                with open(save_name, "w", encoding = "utf-8") as fd:
                    fd.write("\n".join(lines))
            # ===================================================

            ## create a folia document with a numbered id
            #docstr = os.path.splitext( os.path.basename(filename))[0].replace(" ", "_").strip()
            ## Make sure we remember the docstr
            #self.docstr = docstr

            # Split into metadata and body
            iBodyStart = -1
            iBodyEnd = -1
            iLoadDate = -1
            iMetaEnd = -1
            loaddate = None
            end_of_text = []    # Array of possible end-of-text indexes
                                # (idea of Joost Grunwald)
            for idx, line in enumerate(lines):
                if line.startswith( "Body"):
                    iBodyStart = idx + 1
                    while lines[iBodyStart] == "":
                        iBodyStart += 1
                    iMetaEnd = idx - 1
                elif line.startswith("End of Document"):
                    # iBodyEnd = idx - 1
                    end_of_text.append(idx - 1)
                elif line.startswith("Bekijk de oorspronkelijke pagina"):
                    end_of_text.append(idx - 1)
                elif line.startswith("* Vervangende titel"):
                    # Issue #137: check for replacement title
                    end_of_text.append(idx - 1)
                elif line.startswith("Load-Date:"):
                    iLoadDate = idx - 1
                    colon = line.index(":") + 1
                    loaddate = line[colon:].strip()
                    end_of_text.append(idx - 1)
            # If there is a loaddate, then that is the end of the body
            # Determine the best candidate for the end-of-text
            if len(end_of_text) == 0:
                iBodyEnd = iBodyStart
            else:
                # Take the highest index in the list
                end_of_text.sort()
                iBodyEnd = end_of_text[0]
            # if iLoadDate >= 0: iBodyEnd = iLoadDate

            # Get the Meta and the body
            oBack['metadata'] = lines[:iMetaEnd]
            oBack['body'] = lines[iBodyStart:iBodyEnd]
            oBack['loaddate'] = loaddate

            # Issue #137: check for replacement title
            for idx, line in enumerate(oBack['body']):
                if "* vervangende titel" in line.lower():
                    # Get the previous line
                    alt_title = oBack['body'][idx-1]
                    oBack['body'].pop(idx)
                    oBack['body'].pop(idx-1)
                    oBack['metadata'].append('alt_title: {}'.format(alt_title))
                    break
            # Issuee #139: Look for Graphic
            if iBodyEnd < len(lines):
                for idx, line in enumerate(lines[iBodyEnd+1:]):
                    if line.startswith("Graphic"):
                        # Skip empty lines
                        bFound = False
                        start = iBodyEnd + 1 + idx 
                        for line_next in lines[start+1:]:
                            if line_next != "":
                                oBack['metadata'].append('graphic: {}'.format(line_next))
                                bFound = True
                                break
                        if bFound:
                            break
        except:
            oBack['msg'] = errHandle.get_error_message()
            errHandle.DoError("basis_text")
            oBack['okay'] = False
        # Return what has happened
        return oBack
