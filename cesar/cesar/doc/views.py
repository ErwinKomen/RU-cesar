"""
Definition of views for the DOC app.
"""

import sys
import json
import re
from django import template
from django.apps import apps
from django.contrib.auth.models import User, Group
from django.db import models, transaction
from django.db.models import Q
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse, HttpRequest
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.generic.detail import DetailView
from datetime import datetime
import os.path, io, shutil
import tarfile
import openpyxl
import docx
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook
from pypdf import PdfReader

import clam.common.client
import clam.common.data
import clam.common.status
import time

from cesar.settings import APP_PREFIX, WRITABLE_DIR
from cesar.basic.views import BasicList, BasicDetails, BasicPart, user_is_authenticated, add_rel_item
from cesar.browser.models import Status
from cesar.browser.views import nlogin
from cesar.seeker.views import csv_to_excel
from cesar.basic.models import Information
from cesar.tsg.models import TsgInfo
from cesar.doc.models import CLAMClient, FrogLink, FoliaDocs, Brysbaert, NexisDocs, NexisLink, NexisBatch, NexisProcessor, get_crpp_date, \
    LocTimeInfo, Expression, Neologism, Homonym, TwitterMsg, Worddef, Wordlist, Comparison
from cesar.doc.forms import UploadFilesForm, UploadNexisForm, UploadOneFileForm, NexisBatchForm, FrogLinkForm, \
    WordlistForm, LocTimeForm, ExpressionForm, UploadMwexForm, HomonymForm, UploadTwitterExcelForm, \
    WorddefForm, ConcreteForm
from cesar.doc.adaptations import listview_adaptations
from cesar.utils import ErrHandle

# Global debugging 
bDebug = False

TABLET_EDITOR = "tablet_editor"

def user_is_ingroup(request, sGroup):
    # Is this user part of the indicated group?
    username = request.user.username
    user = User.objects.filter(username=username).first()
    # glist = user.groups.values_list('name', flat=True)

    # Only look at group if the user is known
    if user == None:
        glist = []
    else:
        glist = [x.name for x in user.groups.all()]

        # Only needed for debugging
        if bDebug:
            ErrHandle().Status("User [{}] is in groups: {}".format(user, glist))
    # Evaluate the list
    bIsInGroup = (sGroup in glist)
    return bIsInGroup

def user_is_superuser(request):
    bFound = False
    # Is this user part of the indicated group?
    username = request.user.username
    if username != "":
        user = User.objects.filter(username=username).first()
        if user != None:
            bFound = user.is_superuser
    return bFound

def adapt_m2m(cls, instance, field1, qs, field2, extra = [], extrargs = {}, qfilter = {}, 
              related_is_through = False, userplus = None, added=None, deleted=None):
    """Adapt the 'field' of 'instance' to contain only the items in 'qs'
    
    The lists [added] and [deleted] (if specified) will contain links to the elements that have been added and deleted
    If [deleted] is specified, then the items will not be deleted by adapt_m2m(). Caller needs to do this.
    """

    errHandle = ErrHandle()
    try:
        # Get current associations
        lstQ = [Q(**{field1: instance})]
        for k,v in qfilter.items(): lstQ.append(Q(**{k: v}))
        through_qs = cls.objects.filter(*lstQ)
        if related_is_through:
            related_qs = through_qs
        else:
            related_qs = [getattr(x, field2) for x in through_qs]
        # make sure all items in [qs] are associated
        if userplus == None or userplus:
            for obj in qs:
                if obj not in related_qs:
                    # Add the association
                    args = {field1: instance}
                    if related_is_through:
                        args[field2] = getattr(obj, field2)
                    else:
                        args[field2] = obj
                    for item in extra:
                        # Copy the field with this name from [obj] to 
                        args[item] = getattr(obj, item)
                    for k,v in extrargs.items():
                        args[k] = v
                    # cls.objects.create(**{field1: instance, field2: obj})
                    new = cls.objects.create(**args)
                    if added != None:
                        added.append(new)

        # Remove from [cls] all associations that are not in [qs]
        # NOTE: do not allow userplus to delete
        for item in through_qs:
            if related_is_through:
                obj = item
            else:
                obj = getattr(item, field2)
            if obj not in qs:
                if deleted == None:
                    # Remove this item
                    item.delete()
                else:
                    deleted.append(item)
        # Return okay
        return True
    except:
        msg = errHandle.get_error_message()
        return False

def getText(data_file):
    """Given a Word .dox document, extract its text data"""

    oErr = ErrHandle()
    sBack = "-"
    try:
        doc = docx.Document(data_file)
        fullText = []
        for para in doc.paragraphs:
            fullText.append(para.text)
        sBack = '\n'.join(fullText)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("getText")
    return sBack

def getPdfText(data_file):
    """Given a PDF document, extract its text data"""
    oErr = ErrHandle()
    sBack = "-"
    try:
        reader = PdfReader(data_file)
        lst_text = []
        for page in reader.pages:
            sText = page.extract_text()
            lst_text.append(sText)
        sBack = '\n'.join(lst_text)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("getPdfText")
    return sBack

# Views belonging to the Cesar Document Processing app.

# ========== CONCRETENESS ==============================

def concrete_main(request):
    """The main page of working with documents for concreteness."""

    assert isinstance(request, HttpRequest)
    template = 'doc/concrete_main.html'
    frmUpload = UploadFilesForm()
    frmBrysb = UploadOneFileForm()
    frmTwitter = UploadTwitterExcelForm()
    superuser = request.user.is_superuser

    # Basic authentication
    if not user_is_authenticated(request):
        return nlogin(request)

    qd = request.GET
    clamdefine = False
    if 'clamdefine' in qd:
        v = qd.get('clamdefine')
        clamdefine = (v == "true" or v == "1")

    # Make sure the MWE list is okay
    Expression.get_mwe_list()

    # Get a list of already uploaded files too
    text_list = []
    for item in FrogLink.objects.filter(Q(fdocs__owner__username=request.user)).order_by('-created').values(
        'id', 'created', 'concr'):

        sConcr = item.get('concr')
        id = item.get('id')
        if sConcr is None:
            obj = dict(id=id, show=False)
            text_list.append(obj)
        else:
            sCreated = item.get("created")
            if not sCreated is None:
                sCreated = sCreated.strftime("%d/%B/%Y (%H:%M)")
            obj = json.loads(sConcr)
            obj['id'] = id
            obj['show'] = True
            obj['created'] = sCreated
            text_list.append(obj)

    context = {'title': 'Tablet process',
               'frmUpload': frmUpload,
               'frmBrysb': frmBrysb,
               'frmTwitter': frmTwitter,
               'clamdefine': clamdefine,
               'superuser': superuser,
               'message': 'Radboud University CESAR',
               'textlist': text_list,
               'intro_breadcrumb': 'Tablet',
               'year': datetime.now().year}

    if user_is_ingroup(request, TABLET_EDITOR) or  user_is_superuser(request):
        # Adapt the app editor status
        context['is_app_editor'] = True
        context['is_tablet_editor'] = context['is_app_editor']

    return render(request, template, context)

def import_brysbaert(request):
    """Ad-hoc procedure to allow importing Brysbaert tab-separated file into Model"""

    # Initialisations
    # NOTE: do ***not*** add a breakpoint until *AFTER* form.is_valid
    arErr = []
    error_list = []
    transactions = []
    data = {'status': 'ok', 'html': ''}
    template_name = 'doc/import_brys.html'
    obj = None
    data_file = ""
    bClean = False
    username = request.user.username
    statuscode = ""
    oImportName = {
        "bry": "Brysbaert",
        "neo": "Neologism",
        "hom": "Homonym"
        }
    oErr = ErrHandle()

    try:
        # Check if the user is authenticated and if it is POST
        if request.user.is_authenticated and request.method == 'POST':

            # Remove previous status object for this user
            Status.objects.filter(user=username).delete()
            # Create a status object
            oStatus = Status(user=username, type="brysb", status="preparing", msg="please wait")
            oStatus.save()
        
            form = UploadOneFileForm(request.POST, request.FILES)
            lResults = []
            if form.is_valid():
                # NOTE: from here a breakpoint may be inserted!
                print('import_brysb: valid form')

                # Check what the import type is
                qd = request.POST
                import_type = qd.get('import_type', 'bry')

                # Extinct: there can be a different type of import
                # import_name = "Brysbaert" if import_type == "bry" else "Neologism"

                # Issue #180: allow at least three import types, one of which is neologism
                import_name = oImportName.get(import_type, "Brysbaert")

                # Get the contents of the imported file
                file = request.FILES['file_field']

                if bClean:
                    # Clear whatever there was in either Brysbaert or Neologism
                    if import_type == "bry":
                        Brysbaert.clear()
                    elif import_type == "neo":
                        Neologism.clear()
                    elif import_type == "hom":
                        Homonym.clean()

                # The kind of file reading depends on the import_type
                if import_type == "bry":
                    # Expecting a tab-separated file

                    # Read the file into a structure
                    lLines = []
                    bFirst = True
                    for line in file:
                        if bFirst:
                            bFirst = False
                        else:
                            sLine = line.decode("utf-8-sig").strip()
                            if sLine != "":
                                lLines.append(sLine.split("\t"))
                    # Create an iterator for this list
                    lIter = iter(lLines)

                    # Now we have it all, so indicate that
                    oStatus.set("phase 2", msg="chunk-adding information")

                    # Iterate over the contents in chunks
                    iChunk = 0
                    iChunkSize = 100
                    iChunkLen = len(lLines) // iChunkSize + 1
                    iNum = 0

                    iCount = 0
                    iPass = 0
                    # Iterate over the chunks
                    arPart = next(lIter)
                    for chunk_this in range(iChunkLen):
                        iChunk += 1
                        # Show where we are
                        oStatus.set("chunking", msg="processing chunk {} of {}".format(iChunk, iChunkLen))
                        print("working Brysbaert File #{}".format(iCount), file=sys.stderr)
                        # Treat the items from 
                        with transaction.atomic():
                            for idx in range(iChunkSize):
                                # Check if there is some meat
                                if arPart != None:
                                    # Double check length
                                    if len(arPart) == 7:
                                        try:
                                            # get the different parts
                                            stimulus = arPart[0]
                                            listnum = arPart[1]
                                            m = float( arPart[2].replace(",", "."))
                                            sd = float(arPart[3].replace(",", "."))
                                            ratings = float(arPart[4].replace(",", "."))
                                            responses = float(arPart[5].replace(",", "."))
                                            subjects = float(arPart[6].replace(",", "."))

                                            # Just add in one go
                                            obj = Brysbaert(stimulus=stimulus, list=listnum, m=m, sd=sd, ratings=ratings, responses=responses, subjects=subjects)
                                            obj.save()

                                            # Keep track of where we are
                                            iCount += 1
                                        except:
                                            iPass += 1


                                    # Get to the next text
                                    try:
                                        arPart = next(lIter)
                                    except StopIteration as e:
                                        break

                    # We are ready
                    statuscode = "completed"

                elif import_type == "neo":
                    # Expecting an XLSX file

                    # Read the excel
                    wb = openpyxl.load_workbook(file)
                    # We expect the data to be in the first worksheet
                    ws = wb.worksheets[0]

                    # Skip the first row with headings
                    row_no = 2
                    iCount = 0
                    iPass = 0
                    # Walk all rows that have content in cell 1
                    while ws.cell(row=row_no, column=1).value != None:
                        # Get the woord and the score
                        woord = ws.cell(row=row_no, column=1).value
                        score = ws.cell(row=row_no, column=2).value
                        postag = None

                        # Adapt the score into a string
                        if isinstance(score,float):
                            # Turn it into a string with a period decimal separator
                            score = str(score).replace(",", ".")

                        # Make sure we are stripped of spaces and the like
                        woord = woord.strip()
                        # Adapt the word: take off anything starting with left bracket
                        arWoord = woord.split("(")
                        if len(arWoord) > 1:
                            woord = arWoord[0].strip()
                            postag = arWoord[1].strip()
                            if "ZN" in postag:
                                postag = "N"
                            elif "BNW" in postag:
                                postag = "ADJ"
                            elif "WW" in postag:
                                postag = "WW"
                            elif "TUSSENWERPSEL" in postag:
                                postag = "BW"
                            else:
                                postag = None

                        # find or create an item based on this information
                        obj = Neologism.objects.filter(stimulus=woord, postag=postag).first()
                        if obj is None:
                            # Create it
                            obj = Neologism.objects.create(stimulus=woord, m=score, postag=postag)
                            iCount += 1
                        else:
                            iPass += 1

                        # Go to the next row
                        row_no += 1
                    # We are ready
                    statuscode = "completed"

                elif import_type == "hom":
                    # This is homonyms

                    # Read the excel
                    wb = openpyxl.load_workbook(file)
                    # We expect the data to be in the first worksheet
                    ws = wb.worksheets[0]

                    # Skip the first row with headings
                    row_no = 2
                    iCount = 0
                    iPass = 0
                    # Walk all rows that have content in cell 1
                    while ws.cell(row=row_no, column=1).value != None:
                        # Get the woord and the score
                        woord = ws.cell(row=row_no, column=1).value
                        postag = ws.cell(row=row_no, column=2).value
                        meaning = ws.cell(row=row_no, column=3).value
                        score = ws.cell(row=row_no, column=5).value

                        # Adapt the score into a string
                        if isinstance(score,float):
                            # Turn it into a string with a period decimal separator
                            score = str(score).replace(",", ".")

                        # Make sure we are stripped of spaces and the like
                        woord = woord.strip()
                        postag = postag.strip()
                        meaning = meaning.strip()

                        # Possibly adapt postag
                        if "ZN" in postag:
                            postag = "N"
                        elif "BNW" in postag:
                            postag = "ADJ"
                        elif "WW" in postag:
                            postag = "WW"
                        elif "TUSSENWERPSEL" in postag:
                            postag = "BW"
                        else:
                            postag = None

                        # find or create an item based on this information
                        obj = Homonym.objects.filter(stimulus=woord, postag=postag, meaning=meaning).first()
                        if obj is None:
                            # Create it
                            obj = Homonym.objects.create(stimulus=woord, m=score, postag=postag, meaning=meaning)
                            iCount += 1
                        else:
                            iPass += 1

                        # Go to the next row
                        row_no += 1
                    # We are ready
                    statuscode = "completed"

                if statuscode == "error":
                    data['status'] = "error"
                    print("error import_brysbaert #1", file=sys.stderr)
                else:
                    oStatus.set("ready", msg="Read all of {}: {}, skipped {}".format(import_name, iCount, iPass))
                # Get a list of errors
                error_list = [str(item) for item in arErr]

                # Create the context
                context = dict(
                    statuscode=statuscode,
                    result_count=iCount,
                    result_skip =iPass,
                    error_list=error_list
                    )

                if len(arErr) == 0:
                    # Get the HTML response
                    data['html'] = render_to_string(template_name, context, request)
                else:
                    data['html'] = "There are errors in importing {}".format(import_name)


            else:
                data['html'] = 'invalid form: {}'.format(form.errors)
                data['status'] = "error"
        else:
            data['html'] = 'Only use POST and make sure you are logged in'
            data['status'] = "error"
    except:
        msg = oErr.get_error_message()
        data['status'] = "error"
        data['html'] = "Import Brysbaert/Neologism error: {}".format(msg)
 
    # Return the information
    return JsonResponse(data)

def import_concrete(request):
    """Import one or more TEXT (utf8) files that need to be transformed into FoLiA with FROG"""

    # Initialisations
    # NOTE: do ***not*** add a breakpoint until *AFTER* form.is_valid
    arErr = []
    error_list = []
    transactions = []
    data = {'status': 'ok', 'html': ''}
    impossible_extensions = ['doc', 'xml']
    template_name = 'doc/import_docs.html'
    obj = None
    data_file = ""
    bClean = False
    username = request.user.username
    oErr = ErrHandle()
    re_number = re.compile( r"^\d[.,\d]*$")

    # Check if the user is authenticated and if it is POST
    if request.user.is_authenticated and request.method == 'POST':

        # Remove previous status object for this user
        Status.objects.filter(user=username).delete()
        # Create a status object
        oStatus = Status(user=username, type="docs", status="preparing", msg="please wait")
        oStatus.save()

        # Other initialisations
        concretes = []
        
        form = UploadFilesForm(request.POST, request.FILES)
        lResults = []
        if form.is_valid():
            # NOTE: from here a breakpoint may be inserted!
            print('import_docs: valid form')

            # Get user name and password
            clamuser = None
            clampw = None
            qd = request.POST
            if 'clamuser' in qd:
                clamuser = request.POST.get("clamuser")
                clampw = request.POST.get("clampw")

            # Initialisations
            fd = None   # FoliaDocs

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                for data_file in files:
                    filename = data_file.name

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]
                        sBare = arFile[0].strip().replace(" ", "_").replace("-", "_")

                        # Check the bare file name
                        if re_number.match(sBare):
                            # Invalid filename
                            statuscode = "error"
                            msg = "Please change the filename. It should start with a character."
                            arErr.append(msg)
                            oResult = {'status': 'error', 'msg': msg}
                        else:

                            # Adaptation if this is a docx
                            if extension == "docx":
                                # Assuming this is a docx file
                                data_text = getText(data_file)
                                # The text should be split into lines
                                data_file = data_text.split("\n")
                                extension = "txt"

                            elif extension == "pdf":
                                data_text = getPdfText(data_file)
                                # The text should be split into lines
                                data_file = data_text.split("\n")
                                extension = "txt"

                            # Further processing depends on the extension
                            oResult = None
                            
                            if extension in impossible_extensions:  
                                # extension == "doc" or extension == "docx" or extension == "xml":
                                # Cannot process these
                                oResult = {'status': 'error', 'msg': 'cannot process non-text files'}
                            else:
                                # Assume this is a text file: create a froglink
                                fl, msg = FrogLink.create(name=sBare, username=username)
                                if fl == None:
                                    # Some error occurred
                                    statuscode = "error"
                                    oStatus.set("error", msg=msg)
                                    # Break out of the for-loop
                                    break
                                # Read and convert into folia.xml
                                oResult = fl.read_doc(username, data_file, filename, clamuser, clampw, arErr, oStatus=oStatus)
                                # Possibly get the link to the owner's FoliaDocs
                                if fd == None:
                                    # Get the foliadocs link
                                    fd = fl.fdocs

                            # Determine a status code
                            if oResult == None or oResult['status'] == "error" :
                                statuscode = "error"
                                msg = "" if oResult == None or 'msg' not in oResult else oResult['msg']
                                oStatus.set("error", msg=msg)
                                # Break out of the for-loop
                                break
                            else:
                                # Indicate that the folia.xml has been created
                                oStatus.set("working", msg="Created folia.xml file")
                                oErr.Status("Created folia.xml file")
                                # Next step: determine concreteness for this file
                                bResult, msg = fl.do_concreteness()
                                if bResult == False:
                                    arErr.append(msg)
                                    oStatus.set("error", msg=msg)
                                    statuscode = "error"
                                else:
                                    # Make sure we return the concreteness
                                    concretes.append(json.loads(fl.concr))
                                    # Show where we are
                                    statuscode = "completed"
                        if oResult == None:
                            arErr.append("There was an error. No manuscripts have been added")
                        else:
                            lResults.append(oResult)
            if statuscode == "error":
                data['status'] = "error"
            else:
                oStatus.set("ready", msg="Read all files")
            # Get a list of errors
            error_list = [str(item) for item in arErr if len(str(item)) > 0]

            # Create the context
            context = dict(
                statuscode=statuscode,
                results=lResults,
                object=fd,
                concretes=concretes,
                error_list=error_list
                )

            if len(arErr) == 0 or len(arErr[0]) == 0:
                # Get the HTML response
                data['html'] = render_to_string(template_name, context, request)
            else:
                lHtml = []
                lHtml.append("There are errors in importing this doc")
                for item in arErr:
                    lHtml.append("<br />- {}".format(str(item)))
                data['html'] = "\n".join(lHtml)


        else:
            data['html'] = 'invalid form: {}'.format(form.errors)
            data['status'] = "error"
    else:
        data['html'] = 'Only use POST and make sure you are logged in'
        data['status'] = "error"
 
    # Return the information
    return JsonResponse(data)


# =============== CONCRETE ===============================


class ConcreteDownload(DetailView):
    """Allow loading file that has been analyzed for concreteness"""

    model = FrogLink
    template_name = 'doc/foliadocs_view.html'

    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            # Do not allow to get a good response
            response = nlogin(request)
        else:
            self.object = self.get_object()
            context = self.get_context_data(object=self.object)
            response = self.render_to_response(context)
            #response.content = treat_bom(response.rendered_content)
        return response

    def post(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            # Do not allow to get a good response
            response = nlogin(request)
        else:
            self.object = self.get_object()
            # Initialize
            response = None
            context = self.get_context_data(object=self.object)
            # Get the download type
            self.qd = request.POST
            if 'downloadtype' in self.qd:
                # Get the download type and the data itself
                dtype = self.qd['downloadtype']
            
                if dtype == "json":
                    dext = ".json"
                    sContentType = "application/json"
                    # Get the ddata
                    ddata = json.dumps(json.loads( self.object.concr), indent=2)
                elif dtype == "tsv":
                    dext = ".tsv"
                    sContentType = "text/tab-separated-values"
                    # Get the data as a CSV string
                    ddata = self.object.get_csv()
                elif dtype == "excel":
                    dext = ".xlsx"
                    sContentType = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    # Get the data as a CSV string
                    ddata = self.object.get_csv()

                # Determine a file name
                sBase = self.object.get_name()
                sFileName = "{}_concreet{}".format(sBase, dext)

                # Excel needs additional conversion
                if dtype == "xlsx" or dtype == "excel":
                    # Convert 'compressed_content' to an Excel worksheet
                    response = HttpResponse(content_type=sContentType)
                    response['Content-Disposition'] = 'attachment; filename="{}"'.format(sFileName)    
                    response = csv_to_excel(ddata, response, delimiter="\t")
                else:
                    response = HttpResponse(ddata, content_type=sContentType)
                    response['Content-Disposition'] = 'attachment; filename="{}"'.format(sFileName)    

            else:
                response = self.render_to_response(context)
        # Return the response we have
        return response

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(ConcreteDownload, self).get_context_data(**kwargs)

        # Get parameters for the search (if it is GET)
        initial = self.request.GET

        status = ""
        if 'status' in initial:
            status = initial['status']
        context['status'] = status

        # Return what we have
        return context


class ConcreteEdit(BasicDetails):
    """Details view for one concreteness-treated file"""

    model = FrogLink
    mForm = FrogLinkForm
    prefix = "concr"
    title= "ConcreteEdit"
    has_select2 = True      # We are using Select2 in the FrogLinkForm
    mainitems = []

    def get_context_data(self, **kwargs):
        context = super(ConcreteEdit, self).get_context_data(**kwargs)

        # See if there is a manual afterurl
        if 'afterurl' in self.qd:
            afterurl = self.qd.get("afterurl")
            context['afterdelurl'] = afterurl
        return context

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:

            # Define the main items to show and edit
            context['mainitems'] = [
                {'type': 'plain', 'label': "Owner:",        'value': instance.get_owner()       },
                {'type': 'plain', 'label': "Date:",         'value': instance.get_created()     },
                {'type': 'plain', 'label': "Name:",         'value': instance.name              },
                {'type': 'plain', 'label': "Score:",        'value': instance.get_score_string()},
                {'type': 'plain', 'label': "Size:",         'value': instance.get_size()        },
                {'type': 'plain', 'label': "Compare with:", 'value': instance.get_compare()     }
                ]
            context['is_app_editor'] = user_is_ingroup(self.request, "seeker_user")
            context['is_tablet_editor'] = False
            # Double check: only the owner may edit his/her own tablet texts
            current_user = self.request.user.username
            if current_user == instance.get_owner():
                context['is_app_uploader'] = context['is_app_editor']
                if user_is_ingroup(self.request, TABLET_EDITOR) or  user_is_superuser(self.request):
                    # Adapt the app editor status
                    context['is_tablet_editor'] = True
                else:
                    # In fact: any legitimate USER may edit here!
                    context['is_tablet_editor'] = True

                # Allow the 'field_key' for: 'name', 'comparison'
                for oItem in context['mainitems']:
                    if oItem['label'] == "Name:":
                        oItem['field_key'] = "name"
                    elif oItem['label'] == "Compare with:":
                        oItem['field_list'] = "concrlist"

            else:
                context['is_app_editor'] = False

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ConcreteEdit/add_to_context")

        # Return the context we have made
        return context

    def after_save(self, form, instance):
        msg = ""
        bResult = True
        oErr = ErrHandle()
        
        try:
            # Process many-to-many changes: Add and remove relations in accordance with the new set passed on by the user
            # (1) 'collections'
            concrlist = form.cleaned_data['concrlist']
            current = [x.id for x in instance.comparisons.all().order_by('id')]
            after = [x.id for x in concrlist.order_by('id')]
            if current != after:
                # Perform adaptation
                adapt_m2m(Comparison, instance, "base", concrlist, "target")
                # Also make sure that details are re-loaded
                self.afterurl = reverse('froglink_details', kwargs={'pk': instance.id})

        except:
            msg = oErr.get_error_message()
            bResult = False
        return bResult, msg


class ConcreteDetails(ConcreteEdit):
    """Viewing concreteness file as HTML, including the text layout"""
    rtype = "html"

    def add_to_context(self, context, instance):
        # Call the base implementation first to get a context
        context = super(ConcreteDetails, self).add_to_context(context, instance)

        oErr = ErrHandle()
        try:
            # Do we have a JSON response in self.concr?
            if not instance.concr is None and instance.concr != "":
                # Make sure to add [otext] and[tnumber]
                context['tnumber'] = 1
                if instance.concr is None or instance.concr == "":
                    obj = dict(id=instance.id, show=False)
                else:
                    obj = json.loads(instance.concr)

                    # Double check if changes are needed in the object
                    bNeedSaving = False
                    bRecalculate = False
                    word_id = 1
                    for oPara in obj['list']:
                        for oSent in oPara['list']:
                            lst_sent = []
                            for oWord in oSent['list']:
                                concr = str(oWord.get("concr", ""))
                                id = oWord.get('word_id')
                                if id is None:
                                    oWord['word_id'] = word_id
                                    word_id += 1
                                    bNeedSaving = True
                                    lst_sent.append(oWord)
                                elif concr == "-1":
                                    # This word should be removed
                                    bNeedSaving = True
                                else:
                                    # Add to sentence
                                    lst_sent.append(oWord)
                            # Replace the current list
                            if bNeedSaving:
                                oSent['list'] = lst_sent

                    # Do we need to save the (recalculated) results?
                    if bNeedSaving:
                        instance.concr = json.dumps(obj)
                        instance.save()

                    obj['id'] = instance.id
                    obj['show'] = True
                context['otext'] = obj

                # Also provide the loctime info
                qs = LocTimeInfo.objects.all().order_by('example')
                context['loctimes'] = qs

                # Provide a <form> with a field that allows selecting multiple (other) texts (from this user)
                concForm = ConcreteForm(instance=instance)
                context['concForm'] = concForm

                sAfter = render_to_string('doc/foliadocs_view.html', context, self.request)
                context['after_details'] = sAfter


                # Lists of related objects
                related_objects = []
                resizable = True
                index = 1
                sort_start = '<span class="sortable"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_start_int = '<span class="sortable integer"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_start_mix = '<span class="sortable mixed"><span class="fa fa-sort sortshow"></span>&nbsp;'
                sort_end = '</span>'

                # List of Worddefs part of the Wordlist
                compares = dict(title="Comparison with other texts", prefix="dcomp")
                if resizable: compares['gridclass'] = "resizable"

                rel_list =[]
                qs = instance.comparisons.all().order_by('name')
                # qs = Worddef.objects.filter(wordlist=instance).order_by('stimulus')
                for item in qs:
                    # Fields: name, size, score

                    url = reverse('froglink_details', kwargs={'pk': item.id})
                    rel_item = []

                    # Order number for this FrogLink
                    add_rel_item(rel_item, index, False, align="right")
                    index += 1

                    # Name
                    name_txt = item.name
                    add_rel_item(rel_item, name_txt, False, main=True, link=url)

                    # Size
                    size_txt = item.get_size()
                    add_rel_item(rel_item, size_txt, False, main=False, align="right", link=url)

                    # Score
                    score_txt = "{0:.3f}".format(item.get_score())
                    add_rel_item(rel_item, score_txt, False, main=False, align="right", link=url, nowrap=False)


                    # Add this line to the list
                    rel_list.append(dict(id=item.id, cols=rel_item))

                compares['rel_list'] = rel_list

                compares['columns'] = [
                    '{}<span>#</span>{}'.format(sort_start_int, sort_end), 
                    '{}<span>Name</span>{}'.format(sort_start, sort_end), 
                    '{}<span>Size</span>{}'.format(sort_start_mix, sort_end), 
                    '{}<span>Score</span>{}'.format(sort_start_int, sort_end)
                    ]
                related_objects.append(compares)

                # Add all related objects to the context
                context['related_objects'] = related_objects
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ConcreteDetails/add_to_context")

        # Return the adapted context
        return context


class ConcreteUpdate(BasicDetails):
    """Update concreteness information"""

    model = FrogLink
    prefix = "concr"

    def custom_init(self, instance):
        """Just handle the default situation"""

        # Do we have a JSON response in self.concr?
        if instance.concr == None or instance.concr == "":
            obj = dict(id=instance.id, show=False)
        else:
            obj = json.loads(instance.concr)

            # Check if there is any concretedata
            sConcretedata = self.qd.get("concretedata")
            concretedata = {}
            if not sConcretedata is None:
                concretedata = json.loads(sConcretedata)
            concrete_scores = concretedata.get("scores", {})
            concrete_loctime = concretedata.get("loctime", {})

            # Double check if changes are needed in the object
            bNeedSaving = False
            bRecalculate = False

            # If there are any LocTime elements, at least add them to the appropriate list
            bLocTime = False
            if len(concrete_loctime) > 0:
                bLocTime = True
                for k,v in concrete_loctime.items():
                    Expression.process_item(k, v)
                bNeedSaving = True
                bRecalculate = True

            # Walk through the data with the adapted list
            word_id = 1
            for oPara in obj['list']:
                for oSent in oPara['list']:
                    # First try to tackle the added LocTime expressions (if any)
                    if bLocTime:
                        for k,v in concrete_loctime.items():
                            if k in oSent['sentence']:
                                # add this as MWE
                                oWord = dict(lemma=k, concr=v, word_id=word_id)
                                word_id += 1
                                oSent['list'].insert(0, oWord)

                    # Now continue to treat the separate words
                    for oWord in oSent['list']:
                        id = oWord.get('word_id')
                        if id is None or bLocTime:
                            oWord['word_id'] = word_id
                            word_id += 1
                            bNeedSaving = True
                        else:
                            word_id = id
                        if str(word_id) in concrete_scores:
                            oWord['concr'] = concrete_scores[str(word_id)]
                            # Check if this change is due to a homonym
                            lst_homonyms = oWord.get("homonyms")
                            if not lst_homonyms is None:
                                for oItem in lst_homonyms:
                                    if oItem['score'] == oWord['concr']:
                                        # Found it!
                                        oWord['hnum'] = oItem['hnum']
                                        break
                            bNeedSaving = True
                            bRecalculate = True
            # Is recalculation needed?
            if bRecalculate:
                obj = FrogLink.recalculate(obj)

            # Do we need to save the (recalculated) results?
            if bNeedSaving:
                instance.concr = json.dumps(obj)
                if isinstance(obj['score'], float):
                    instance.score = obj['score']
                instance.save()
        return None


class ConcreteListView(BasicList):
    """Search and list nexis batches"""

    model = FrogLink
    listform = FrogLinkForm
    prefix = "concr"
    new_button = False      # Don't show a new button, because new items can only be added by downloading
    plural_name = "Concreteness text files"
    sg_name = "Concreteness file"
    downloadname = "ConcreteList"
    has_select2 = True      # We are using Select2 in the FrogLinkForm
    delete_line = True      # Allow deleting a line
    bUseFilter = True
    superuser = False
    order_cols = ['created', 'score', 'size', 'name', '']
    order_default = ['-created', 'score', 'size', 'name']
    order_heads = [
        {'name': 'Date',  'order': 'o=1', 'type': 'str', 'custom': 'created',    'linkdetails': True},
        {'name': 'Score', 'order': 'o=2', 'type': 'str', 'custom': 'score',      'linkdetails': True},
        {'name': 'Size',  'order': 'o=3', 'type': 'int', 'custom': 'size',       'linkdetails': True, 'align': "right"},
        {'name': 'Name',  'order': 'o=4', 'type': 'str', 'field':  'name',       'linkdetails': True, 'main': True},
        {'name': '',      'order': '',    'type': 'str', 'options': 'delete', 'classes': 'tdnowrap'}]
    filters = [
        {'name': 'Name',  'id': 'filter_name',  'enabled': False}
        ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'name',  'dbfield':  'name',     'keyS': 'name'}
            ]},
        {'section': 'other', 'filterlist': [
            {'filter': 'owner', 'fkfield':  'fdocs__owner', 'keyS': 'owner', 'keyFk': 'id', 'keyList': 'ownlist', 'infield': 'id'},
            {'filter': 'fdocs',     'fkfield': 'fdocs',  'keyFk': 'fdocs'}]}
        ]

    def initializations(self):
        oErr = ErrHandle()
        try:
            # Check if I am superuser or not
            self.superuser = self.request.user.is_superuser
            if self.superuser:
                self.order_cols = ['created', 'fdocs__owner__username', 'score', 'size', 'name', '']
                self.order_default = ['-created', 'fdocs__owner__username', 'score', 'size', 'name']
                self.order_heads = [
                    {'name': 'Date',  'order': 'o=1', 'type': 'str', 'custom': 'created',    'linkdetails': True},
                    {'name': 'Owner', 'order': 'o=2', 'type': 'str', 'custom': 'owner',      'linkdetails': True},
                    {'name': 'Score', 'order': 'o=3', 'type': 'str', 'custom': 'score',      'linkdetails': True},
                    {'name': 'Size',  'order': 'o=4', 'type': 'int', 'custom': 'size',       'linkdetails': True, 'align': "right"},
                    {'name': 'Name',  'order': 'o=5', 'type': 'str', 'field':  'name',       'linkdetails': True, 'main': True},
                    {'name': '',      'order': '',    'type': 'str', 'options': 'delete', 'classes': 'tdnowrap'}]
                self.filters = [
                    {'name': 'Name',  'id': 'filter_name',  'enabled': False},
                    {'name': 'Owner', 'id': 'filter_owner', 'enabled': False}
                    ]
                self.searches = [
                    {'section': '', 
                     'filterlist': [
                        {'filter': 'name',  'dbfield':  'name',     'keyS': 'name'},
                        {'filter': 'owner', 'fkfield':  'fdocs__owner', 'keyS': 'owner', 'keyFk': 'id', 'keyList': 'ownlist', 'infield': 'id'}
                        ]
                     },
                    {'section': 'other', 
                     'filterlist': [
                        {'filter': 'fdocs', 'fkfield': 'fdocs', 'keyFk': 'fdocs'}
                        ]
                     }
                    ]
                sValue = Information.get_kvalue("doc_score")
                if not sValue in ['ok', 'done']:
                    # Walk all the FrogLink items
                    for obj in FrogLink.objects.all():
                        if obj.score is None or obj.score == 0.0:
                            score = obj.get_score(True)
                            obj.score = score
                            obj.save()

                    Information.set_kvalue("doc_score", "done")

                sValue = Information.get_kvalue("doc_size")
                if not sValue in ['ok', 'done']:
                    # Walk all the FrogLink items
                    for obj in FrogLink.objects.all():
                        if obj.size is None or obj.size <= 0:
                            size = obj.get_size(True)
                            obj.size = size
                            obj.save()

                    Information.set_kvalue("doc_size", "done")

            self.downloadname = "ConcreteList_{}".format(timezone.now().strftime("%Y%m%d"))
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ConcreteListView/initializations")

        return None

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []
        oErr = ErrHandle()
        try:

            # Figure out what to show
            if custom == "created":
                sBack = instance.created.strftime("%d/%B/%Y (%H:%M)")
            elif custom == "owner":
                sBack = instance.fdocs.owner.username
            elif custom == "score":
                # Get the score
                score = instance.get_score()
                sBack = "{0:.3f}".format(score)
            elif custom == "size":
                # Get the score
                size = instance.get_size()
                sBack = "{}".format(size)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ConcreteListView/get_field_value")
        # Retourneer wat kan
        return sBack, sTitle

    def adapt_search(self, fields):
        # Initialisations
        lstExclude=None
        qAlternative = None
        oErr = ErrHandle()
        try:
            if not self.superuser:
                # Make sure only batches are shown for which this user is the owner
                username = self.request.user.username
                fields['ownlist'] = User.objects.filter(username = username)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("ConcreteListView/adapt_search")

        # Return standard
        return fields, lstExclude, qAlternative

    def add_to_context(self, context, initial):
        # Allow simple seeker_user to work with this
        context['is_app_editor'] = user_is_ingroup(self.request, "seeker_user")
        context['is_app_uploader'] = context['is_app_editor']
        context['is_tablet_editor'] = user_is_ingroup(self.request, "tablet_editor")
        return context


# =============== LOCTIME ===============================


class LocTimeEdit(BasicDetails):
    """Edit a loctime information element"""

    model = LocTimeInfo
    mForm = LocTimeForm
    prefix = "loct"
    title = "Location and time Edit"
    mainitems = []

    def add_to_context(self, context, instance):
        """Add to the existing context"""
        # Only moderators are to be allowed
        if user_is_ingroup(self.request, TABLET_EDITOR) or  user_is_superuser(self.request): 
            # Define the main items to show and edit
            context['mainitems'] = [
                {'type': 'plain', 'label': "Example:", 'value': instance.example, 'field_key': "example"},
                {'type': 'plain', 'label': "Score:",   'value': instance.score,   'field_key': "score"}
                ]       
            # Adapt the app editor status
            context['is_app_editor'] = user_is_superuser(self.request) or user_is_ingroup(self.request, TABLET_EDITOR)
            context['is_tablet_editor'] = context['is_app_editor']
        # Return the context we have made
        return context
    

class LocTimeDetails(LocTimeEdit):
    """Viewing LocTime information items"""
    rtype = "html"


class LocTimeList(BasicList):
    """List loctime elements"""

    model = LocTimeInfo
    listform = LocTimeForm
    prefix = "loct"
    sg_name = "Location and time item"
    plural_name = "Location and time items"
    new_button = True      # Do show a new button
    view_only = False
    order_cols = ['example', 'score']
    order_default = order_cols
    order_heads = [
        {'name': 'Example', 'order': 'o=1', 'type': 'str', 'field': 'example', 'main': True, 'linkdetails': True},
        {'name': 'Score',   'order': 'o=2', 'type': 'str', 'field': 'score',   'linkdetails': True},
        ]
    filters = [ {"name": "Example", "id": "filter_example", "enabled": False}
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'example',   'dbfield': 'example', 'keyS': 'example'}]} 
        ] 
    
    def initializations(self):
        oErr = ErrHandle()
        try:
            # Check if this is view_only or editable
            is_superuser = user_is_superuser(self.request)
            is_tablet_editor = user_is_ingroup(request, "tablet_editor")

            if not is_superuser and not is_tablet_editor:
                #if self.view_only:
                for oItem in self.order_heads:
                    if 'linkdetails' in oItem:
                        oItem.pop("linkdetails")
        except:
            msg = oErr.get_error_message()
            oErr.DoError("LocTimeTable/init")
        return None

    def add_to_context(self, context, initial):
        # Only moderators are to be allowed
        allow_editing = False
        if not self.view_only:
            allow_editing = user_is_ingroup(self.request, TABLET_EDITOR) or  user_is_superuser(self.request)

        if allow_editing:
            # Adapt the app editor status
            context['is_app_editor'] = True
            context['is_tablet_editor'] = context['is_app_editor']
        else:
            # View only
            context['is_app_editor'] = False
            context['is_tablet_editor'] = context['is_app_editor']
        return context



# =============== EXPRESSION ===============================


class ExpressionEdit(BasicDetails):
    """Edit a Expression information element"""

    model = Expression
    mForm = ExpressionForm
    prefix = "expr"
    title = "Multi-word Expression"
    mainitems = []

    def add_to_context(self, context, instance):
        """Add to the existing context"""
        # Only moderators are to be allowed
        if user_is_ingroup(self.request, TABLET_EDITOR) or  user_is_superuser(self.request): 
            # Define the main items to show and edit
            context['mainitems'] = [
                {'type': 'plain', 'label': "Expression:", 'value': instance.full,   'field_key': "full"},
                {'type': 'plain', 'label': "Score:",      'value': instance.score,  'field_key': "score"},
                {'type': 'plain', 'label': "Lemma's:",    'value': instance.get_lemmas()                },
                {'type': 'plain', 'label': "Frogged:",    'value': instance.get_frogged()               },
                ]       
            # Adapt the app editor status
            context['is_app_editor'] = user_is_superuser(self.request) or user_is_ingroup(self.request, TABLET_EDITOR)
            context['is_tablet_editor'] = context['is_app_editor']
        # Return the context we have made
        return context

    def before_save(self, form, instance):
        """Before a new or changed expression is saved, its [frogged] attribute may need to be re-calculated"""

        bResult = True
        msg = ""
        bRecalculate = False
        oErr = ErrHandle()
        try:
            # Is this a new expression?
            if instance is None:
                pass
            elif instance.id is None:
                if not instance.full is None and len(instance.full) > 0:
                    bRecalculate = True
            else:
                # Get the previous instance
                previous = Expression.objects.filter(id=instance.id).first()
                # Check if the previous list differs
                if previous.full != instance.full:
                    bRecalculate = True

            # Do we need to recalculate?
            if bRecalculate:
                # Get a possible new frogged line
                sFrogged = instance.make_frogged(self.request.user.username, instance.full)
                if sFrogged != "":
                    # It will get saved with the form
                    form.instance.frogged = sFrogged

        except:
            msg = oErr.get_error_message()
            oErr.DoError("ExpressionEdit/before_save")
        return bResult, msg
    

class ExpressionDetails(ExpressionEdit):
    """Viewing Expression information items"""
    rtype = "html"


class ExpressionList(BasicList):
    """List Expression elements"""

    model = Expression
    listform = ExpressionForm
    prefix = "expr"
    sg_name = "Multi-word Expression"
    plural_name = "Multi-word Expressions"
    new_button = True      # Do show a new button
    order_cols = ['full', 'score']
    order_default = order_cols
    order_heads = [
        {'name': 'Expression', 'order': 'o=1', 'type': 'str', 'field': 'full', 'main': True, 'linkdetails': True},
        {'name': 'Score',      'order': 'o=2', 'type': 'str', 'field': 'score',   'linkdetails': True},
        ]
    filters = [ {"name": "Expression", "id": "filter_full", "enabled": False}
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'full',   'dbfield': 'full', 'keyS': 'full'}]} 
        ] 
    uploads = [
        {"title": "mwex", 
         "label": "MWE Excel", 
         "url": "import_mwex", 
         "msg": "Upload MWE Excel file", 
         "type": "single"}]

    def initializations(self):
        # ======== One-time adaptations ==============
        kwargs = {'username': self.request.user.username}
        # username = self.request.user.username
        listview_adaptations("doc_list", **kwargs)

        return None

    def add_to_context(self, context, initial):
        # Only moderators are to be allowed
        if user_is_ingroup(self.request, TABLET_EDITOR) or  user_is_superuser(self.request): 
            # Adapt the app editor status
            context['is_app_editor'] = user_is_superuser(self.request) or user_is_ingroup(self.request, TABLET_EDITOR)
            context['is_tablet_editor'] = context['is_app_editor']
        return context


# =============== HOMONYM ===============================


class HomonymEdit(BasicDetails):
    """Edit a Homonym information element"""

    model = Homonym
    mForm = HomonymForm
    prefix = "hnym"
    title = "Homonym"
    mainitems = []

    def add_to_context(self, context, instance):
        """Add to the existing context"""
        # Only moderators are to be allowed
        if user_is_ingroup(self.request, TABLET_EDITOR) or  user_is_superuser(self.request): 
            # Define the main items to show and edit
            context['mainitems'] = [
                {'type': 'plain', 'label': "Lemma:",            'value': instance.stimulus, 'field_key': "stimulus"},
                {'type': 'plain', 'label': "Part-of-speech:",   'value': instance.postag,   'field_key': "postag"},
                {'type': 'plain', 'label': "Sense of meaning:", 'value': instance.meaning,  'field_key': "meaning"},
                {'type': 'plain', 'label': "Metric:",           'value': instance.m,        'field_key': "m"    },
                ]       
            # Adapt the app editor status
            context['is_app_editor'] = user_is_superuser(self.request) or user_is_ingroup(self.request, TABLET_EDITOR)
            context['is_tablet_editor'] = context['is_app_editor']
        # Return the context we have made
        return context
    

class HomonymDetails(HomonymEdit):
    """Viewing Homonym information items"""
    rtype = "html"


class HomonymList(BasicList):
    """List Homonym elements"""

    model = Homonym
    listform = HomonymForm
    prefix = "hnym"
    sg_name = "Homonym"
    plural_name = "Homonyms"
    new_button = True      # Do show a new button
    order_cols = ['stimulus', 'meaning', 'postag', 'm']
    order_default = order_cols
    order_heads = [
        {'name': 'Lemma',               'order': 'o=1', 'type': 'str', 'field': 'stimulus', 'linkdetails': True},
        {'name': 'Sense of meaning',    'order': 'o=3', 'type': 'str', 'field': 'meaning',  'linkdetails': True, 'main': True},
        {'name': 'POS',                 'order': 'o=2', 'type': 'str', 'field': 'postag',   'linkdetails': True},
        {'name': 'Metric',              'order': 'o=4', 'type': 'int', 'field': 'm',        'linkdetails': True, 'align': "right"},
        ]
    filters = [ 
        {"name": "Lemma", "id": "filter_stimulus", "enabled": False},
        {"name": "Sense", "id": "filter_sense", "enabled": False}
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'stimulus',   'dbfield': 'stimulus', 'keyS': 'stimulus'},
            {'filter': 'sense',      'dbfield': 'meaning',  'keyS': 'meaning'}
            ]
         } 
        ] 

    def add_to_context(self, context, initial):
        # Only moderators are to be allowed
        if user_is_ingroup(self.request, TABLET_EDITOR) or  user_is_superuser(self.request): 
            # Adapt the app editor status
            context['is_app_editor'] = user_is_superuser(self.request) or user_is_ingroup(self.request, TABLET_EDITOR)
            context['is_tablet_editor'] = context['is_app_editor']
        return context


# =============== WORDLIST ===============================


class WordlistEdit(BasicDetails):
    """Edit a Wordlist information element"""

    model = Wordlist
    mForm = WordlistForm
    prefix = "wlst"
    title = "Wordlist"
    mainitems = []

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Only tablet editors or superusers are to be allowed
            if user_is_ingroup(self.request, TABLET_EDITOR) or  user_is_superuser(self.request): 
                # Define the main items to show and edit
                context['mainitems'] = [
                    {'type': 'plain', 'label': "Name:",         'value': instance.name,         'field_key': "name"},
                    {'type': 'plain', 'label': "Description:",  'value': instance.description,  'field_key': "description"},
                    {'type': 'plain', 'label': "Excel file:",   'value': instance.get_upload(), 'field_key': "upload"},
                    {'type': 'plain', 'label': "Worksheet",     'value': instance.sheet,        'field_key': "sheet"},
                    {'type': 'safe',  'label': "",              'value': self.get_button(instance)                  }
                    ]       
                # Adapt the app editor status
                context['is_app_editor'] = user_is_superuser(self.request) or user_is_ingroup(self.request, TABLET_EDITOR)
                context['is_tablet_editor'] = context['is_app_editor']

                # context['after_details'] = self.get_button(instance)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("WordlistEdit/add_to_context")

        # Return the context we have made
        return context

    def get_button(self, instance):
        """Get a HTML button to process the Excel file with the wordlist"""

        sBack = ""
        oErr = ErrHandle()
        try:
            # Only provide a button, if all the necessary details have been provided
            if not instance is None and not instance.upload is None and instance.upload != "" and not instance.sheet is None:
                html = []
                url = reverse('wordlist_upload', kwargs={'pk': instance.id})
                html.append("<a role='button' class='btn btn-xs jumbo-3' href='{}' ".format(url))
                html.append("title='Upload the file into the system'>{}<a>".format("Upload..."))

                sBack = "\n".join(html)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("WordlistEdit/get_button")

        return sBack
    

class WordlistDetails(WordlistEdit):
    """Viewing Wordlist information items"""
    rtype = "html"

    def add_to_context(self, context, instance):
        # First get the 'standard' context from AuworkEdit
        context = super(WordlistDetails, self).add_to_context(context, instance)

        context['sections'] = []

        oErr = ErrHandle()
        try:

            # Lists of related objects
            related_objects = []
            resizable = True
            index = 1
            sort_start = '<span class="sortable"><span class="fa fa-sort sortshow"></span>&nbsp;'
            sort_start_int = '<span class="sortable integer"><span class="fa fa-sort sortshow"></span>&nbsp;'
            sort_start_mix = '<span class="sortable mixed"><span class="fa fa-sort sortshow"></span>&nbsp;'
            sort_end = '</span>'

            # List of Worddefs part of the Wordlist
            worddefs = dict(title="Word definitions with this wordlist", prefix="wdef")
            if resizable: worddefs['gridclass'] = "resizable"

            rel_list =[]
            qs = Worddef.objects.filter(wordlist=instance).order_by('stimulus')
            for item in qs:
                # Fields: stimulus, postag, score (=m)

                url = reverse('worddef_details', kwargs={'pk': item.id})
                rel_item = []

                # Order number for this worddef
                add_rel_item(rel_item, index, False, align="right")
                index += 1

                # Stimulus
                stimulus_txt = item.stimulus
                add_rel_item(rel_item, stimulus_txt, False, main=True, link=url)

                # POS tag (if any)
                postag_txt = item.get_postag()
                add_rel_item(rel_item, postag_txt, False, main=False, link=url)

                # Score
                score_txt = "{}".format(item.get_concreteness())
                add_rel_item(rel_item, score_txt, False, main=False, align="right", link=url, nowrap=False)


                # Add this line to the list
                rel_list.append(dict(id=item.id, cols=rel_item))

            worddefs['rel_list'] = rel_list

            worddefs['columns'] = [
                '{}<span>#</span>{}'.format(sort_start_int, sort_end), 
                '{}<span>Stimulus</span>{}'.format(sort_start, sort_end), 
                '{}<span>POS tag</span>{}'.format(sort_start_mix, sort_end), 
                '{}<span>Score</span>{}'.format(sort_start_int, sort_end)
                ]
            related_objects.append(worddefs)

            # Add all related objects to the context
            context['related_objects'] = related_objects
        except:
            msg = oErr.get_error_message()
            oErr.DoError("WordlistDetails/add_to_context")

        # Return the context we have made
        return context


class WordlistUpload(WordlistDetails):

    initRedirect = True

    def initializations(self, request, pk):
        # Perform the original initialization
        super(WordlistUpload, self).initializations(request, pk)

        oErr = ErrHandle()
        try:
            # Get the parameters
            self.qd = request.POST

            # Get the instance
            instance = self.object

            if not instance is None:
                # Actually try to read this Excel file
                instance.read_upload()

            # The default redirectpage is just this manuscript
            self.redirectpage = reverse("wordlist_details", kwargs = {'pk': instance.id})

            # Getting here means all went well
        except:
            msg = oErr.get_error_message()
            oErr.DoError("TemplateImport/initializations")
        return None


class WordlistList(BasicList):
    """List Wordlist elements"""

    model = Wordlist
    listform = WordlistForm
    prefix = "wlst"
    sg_name = "Wordlist"
    plural_name = "Wordlists"
    new_button = True      # Do show a new button
    order_cols = ['name', '']
    order_default = order_cols
    order_heads = [
        {'name': 'Name',    'order': 'o=1', 'type': 'str', 'field': 'name',  'linkdetails': True, 'main': True},
        {'name': 'Saved',   'order': '',    'type': 'str', 'field': 'saved', 'linkdetails': True, 'align': "right"},
        ]
    filters = [ 
        {"name": "Name", "id": "filter_name", "enabled": False},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'name',   'dbfield': 'name', 'keyS': 'name'},
            ]
         } 
        ] 

    def add_to_context(self, context, initial):
        # Only moderators are to be allowed
        if user_is_ingroup(self.request, TABLET_EDITOR) or  user_is_superuser(self.request): 
            # Adapt the app editor status
            context['is_app_editor'] = user_is_superuser(self.request) or user_is_ingroup(self.request, TABLET_EDITOR)
            context['is_tablet_editor'] = context['is_app_editor']
        return context


# =============== WORDLIST ===============================


class WorddefEdit(BasicDetails):
    """Edit a Worddef information element"""

    model = Worddef
    mForm = WorddefForm
    prefix = "wdef"
    title = "Worddef"
    mainitems = []

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        oErr = ErrHandle()
        try:
            # Only tablet editors or superusers are to be allowed
            if user_is_ingroup(self.request, TABLET_EDITOR) or  user_is_superuser(self.request): 
                # Define the main items to show and edit
                context['mainitems'] = [
                    {'type': 'plain', 'label': "Wordlist:", 'value': instance.get_wordlist() },
                    {'type': 'plain', 'label': "Stimulus:", 'value': instance.stimulus,             'field_key': "stimulus"},
                    {'type': 'plain', 'label': "POS tag:",  'value': instance.get_postag(),         'field_key': "postag"},
                    {'type': 'plain', 'label': "Score:",    'value': instance.get_concreteness(),   'field_key': "m"},
                    ]       
                # Adapt the app editor status
                context['is_app_editor'] = user_is_superuser(self.request) or user_is_ingroup(self.request, TABLET_EDITOR)
                context['is_tablet_editor'] = context['is_app_editor']

            # Add a button back to the Wordlist
            topleftlist = []
            if not instance.wordlist is None:
                wl = instance.wordlist
                buttonspecs = {'label': "W", 
                     'title': "Go to wordlist {}".format(wl.name), 
                     'url': reverse('wordlist_details', kwargs={'pk': wl.id})}
                topleftlist.append(buttonspecs)
            context['topleftbuttons'] = topleftlist

        except:
            msg = oErr.get_error_message()
            oErr.DoError("WorddefEdit/add_to_context")

        # Return the context we have made
        return context
    

class WorddefDetails(WorddefEdit):
    """Viewing Worddef information items"""
    rtype = "html"


class WorddefList(BasicList):
    """List Worddef elements"""

    model = Worddef
    listform = WorddefForm
    prefix = "wdef"
    sg_name = "Word definition"
    plural_name = "Word definitions"
    new_button = False      # Do *NOT* show a new button
    order_cols = ['wordlist__name', 'stimulus', 'postag', 'm']
    order_default = order_cols
    order_heads = [
        {'name': 'Word list', 'order': 'o=1', 'type': 'str', 'custom': 'wordlist','linkdetails': True},
        {'name': 'Stimulus',  'order': 'o=2', 'type': 'str', 'field': 'stimulus', 'linkdetails': True, 'main': True},
        {'name': 'POS tag',   'order': 'o=3', 'type': 'str', 'field': 'postag',   'linkdetails': True},
        {'name': 'Score',     'order': 'o=4', 'type': 'int', 'field': 'm',        'linkdetails': True, 'align': "right"},
        ]
    filters = [ 
        {"name": "Stimulus", "id": "filter_stimulus", "enabled": False},
        {"name": "POS tag",  "id": "filter_postag", "enabled": False},
               ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'stimulus',   'dbfield': 'stimulus', 'keyS': 'stimulus'},
            {'filter': 'postag',     'dbfield': 'postag',   'keyS': 'postag'},
            ]
         } 
        ] 

    def add_to_context(self, context, initial):
        # Only moderators are to be allowed
        if user_is_ingroup(self.request, TABLET_EDITOR) or  user_is_superuser(self.request): 
            # Adapt the app editor status
            context['is_app_editor'] = user_is_superuser(self.request) or user_is_ingroup(self.request, TABLET_EDITOR)
            context['is_tablet_editor'] = context['is_app_editor']
        return context

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []

        # Figure out what to show
        if custom == "wordlist":
            # Figure out the name of the wordlist
            sBack = instance.get_wordlist()

        # Retourneer wat kan
        return sBack, sTitle



# ================ TWITTER =============================

def twitter_main(request):
    """The main page of working with documents for concreteness."""

    assert isinstance(request, HttpRequest)
    template = 'doc/twitter_main.html'
    frmUpload = UploadFilesForm()
    frmBrysb = UploadOneFileForm()
    frmTwitter = UploadTwitterExcelForm()
    superuser = request.user.is_superuser

    # Basic authentication
    if not user_is_authenticated(request):
        return nlogin(request)

    qd = request.GET
    clamdefine = False
    if 'clamdefine' in qd:
        v = qd.get('clamdefine')
        clamdefine = (v == "true" or v == "1")

    ## Get a list of already uploaded files too
    #text_list = []
    #for item in FrogLink.objects.filter(Q(fdocs__owner__username=request.user)).order_by('-created'):
    #    if item.concr == None or item.concr == "":
    #        obj = dict(id=item.id, show=False)
    #        text_list.append(obj)
    #    else:
    #        obj = json.loads(item.concr)
    #        obj['id'] = item.id
    #        obj['show'] = True
    #        obj['created'] = item.get_created()
    #        text_list.append(obj)

    tweetcount = TwitterMsg.objects.count()

    context = {'title': 'Twitter process',
               'frmTwitter': frmTwitter,
               'clamdefine': clamdefine,
               'superuser': superuser,
               'tweetcount': tweetcount,
               'message': 'Radboud University CESAR',
               'intro_breadcrumb': 'Twitter',
               'year': datetime.now().year}

    if user_is_ingroup(request, TABLET_EDITOR) or  user_is_superuser(request):
        # Adapt the app editor status
        context['is_app_editor'] = True
        context['is_tablet_editor'] = context['is_app_editor']

    return render(request, template, context)


# ======================= DUTCH SPEECH ===================


def transcribe_dutch(request):
    """Transcribe Dutch MP3 speech
    
    """

    asr_url = "https://webservices.cls.ru.nl/asr_nl"
    inputtemplate = "InputMP3File"
    localfilename = None
    oErr = ErrHandle()
    try:
        assert isinstance(request, HttpRequest)
        template = 'doc/twitter_main.html'
        frmUpload = UploadFilesForm()

        superuser = request.user.is_superuser

        # Basic authentication
        if not user_is_authenticated(request):
            return nlogin(request)

        qd = request.GET
        context = {'title': 'Dutch transcription',
                   'frmSoundFile': frmUpload,
                   'superuser': superuser,
                   'message': 'Radboud University CESAR',
                   'intro_breadcrumb': 'Transcribe',
                   'year': datetime.now().year}

        if user_is_superuser(request):
            # Adapt the app editor status
            context['is_app_editor'] = True
            context['is_tablet_editor'] = context['is_app_editor']

            # Get the file location
            inputfile = Information.get_kvalue("asrnl_file")

            # Regular CLAM stuff
            clamuser = TsgInfo.get_value("clam_user")
            clampw = TsgInfo.get_value("clam_pw")
            # Think of a project name
            project = "asrnl"
            basicauth = True
            # Get access to the webservice
            clamclient = CLAMClient(asr_url, clamuser, clampw, basicauth = basicauth)
            # First delete any previous project, if it exists
            try:
                result = clamclient.delete(project)
                oErr.Status("Removed previous project {} = {}".format(project, result))
            except:
                # No problem: no project has been removed
                pass
            # Only now start creating it
            result = clamclient.create(project)
            oErr.Status("Created new project {} = {}".format(project, result))
            data = clamclient.get(project)

            # Set the input file
            clamclient.addinputfile(project, data.inputtemplate(inputtemplate), inputfile)

            # Start the project
            data = clamclient.start(project)

            # Check for possible error(s)
            if data.errors:
                print("An error occured: " + data.errormsg, file=sys.stderr)
                for parametergroup, paramlist in data.parameters:
                    for parameter in paramlist:
                        if parameter.error:
                            print("Error in parameter " + parameter.id + ": " + parameter.error, file=sys.stderr)
                # Remove the project, since an error ha occured
                clamclient.delete(project) 
                sys.exit(1)

            #If everything went well, the system is now running, we simply wait until it is done and retrieve the status in the meantime
            while data.status != clam.common.status.DONE:
                time.sleep(5)                       # wait 5 seconds before polling status
                data = clamclient.get(project)      # get status again
                statusmsg = data.statusmessage
                completion = data.completion
                msg = "Running. Status={}: {}% completed".format(statusmsg, completion)
                oErr.Status(msg)
                # x = data.status
                # print("\tRunning: " + str(data.completion) + '% -- ' + data.statusmessage, file=sys.stderr)

            #Iterate over output files
            for outputfile in data.output:
                try:    
                    # Metadata contains information on output template
                    outputfile.loadmetadata() 
                except:
                    continue

                # Get the output template
                outputtemplate = outputfile.metadata.provenance.outputtemplate_id

                # Think of a local filename
                localfilename = os.path.basename(str(outputfile))
                outputfile.copy(localfilename)

                # Or: iterate ove rthe textual contents
                for line in outputfile.readlines():
                    print(line)

            # Create the appropriate response
            # response = render(request, template, context)
            response = "READY"
    except:
        msg = oErr.get_error_message()
        oErr.DoError("transcribe_dutch")
        response = "<html><body><h3>Error</h3><div>{}</div></body></html>".format(msg)

    return response




# ================ NEXIS UNI ===========================
def nexis_main(request):
    """The main page of working with Nexis Uni documents."""

    # Initializations
    template = 'doc/nexis_main.html'
    frmUpload = UploadNexisForm()

    # Make sure this is just a HTTP request
    assert isinstance(request, HttpRequest)

    # Basic authentication
    if not user_is_authenticated(request):
        return nlogin(request)

    # Other initializations
    superuser = request.user.is_superuser
    username = request.user.username

    # Get a list of processed batches
    nd = NexisDocs.get_obj(username)
    batch_list = []
    if nd != None:
        batch_list = nd.ndocsbatches.all().order_by("-created")



    # Create the context
    context = {'title': 'Document processing',
               'frmUpload': frmUpload,
               'superuser': superuser,
               'message': 'Radboud University CESAR',
               'batchlist': batch_list,
               'intro_breadcrumb': 'Nexis Uni',
               'year': datetime.now().year}
    
    # Render the template as HTML
    return render(request, template, context)

def import_nexis(request):
    """Import one or more TEXT (utf8) files that need to be transformed into FoLiA with FROG"""

    # Initialisations
    # NOTE: do ***not*** add a breakpoint until *AFTER* form.is_valid
    arErr = []
    error_list = []
    transactions = []
    data = {'status': 'ok', 'html': ''}
    template_name = 'doc/import_nexis.html'
    obj = None
    data_file = ""
    bClean = False
    statuscode = ""
    username = request.user.username
    oErr = ErrHandle()
    re_number = re.compile( r"^\d[.,\d]*$")

    # Check if the user is authenticated and if it is POST
    if request.user.is_authenticated and request.method == 'POST':

        # Remove previous status object for this user
        Status.objects.filter(user=username).delete()
        # Create a status object
        oStatus = Status(user=username, type="docs", status="preparing", msg="please wait")
        oStatus.save()

        # Other initialisations
        lmeta = []
        
        form = UploadNexisForm(request.POST, request.FILES)
        lResults = []
        if form.is_valid():
            # NOTE: from here a breakpoint may be inserted!
            print('import_nexis: valid form')

            # Initialisations
            nd = None   # NexisDocs

            # Get the contents of the imported file
            files = request.FILES.getlist('files_field')
            if files != None:
                # Create a NexisBatch
                batch, msg = NexisBatch.create(username=username)
                for idx, data_file in enumerate(files):
                    filename = data_file.name

                    # Set the status
                    oStatus.set("reading", msg="{}: file={}".format(idx, filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        arErr.append("No source file specified for the selected project")
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]
                        # sBare = arFile[0].strip().replace(" ", "_")
                        sBare = arFile[0].strip()

                        # Check the bare file name
                        if False: # re_number.match(sBare):
                            # Invalid filename
                            statuscode = "error"
                            msg = "Please change the filename. It should start with a character."
                            arErr.append(msg)
                            oResult = {'status': 'error', 'msg': msg}
                        else:

                            # Further processing depends on the extension
                            oResult = None
                            if extension == "doc" or extension == "docx" or extension == "xml":
                                # Cannot process these
                                oResult = {'status': 'error', 'msg': 'cannot process non-text files'}
                            else:
                                # Assume this is a text file: create a NexisLink and parse it
                                fl, msg = NexisLink.create(name=sBare, username=username, batch=batch)
                                if fl == None:
                                    # Some error occurred
                                    statuscode = "error"
                                    oStatus.set("error", msg=msg)
                                    # Break out of the for-loop
                                    break
                                # Read and convert into folia.xml
                                oResult = fl.read_doc(username, data_file, filename, arErr, oStatus=oStatus)
                                # Possibly get the link to the owner's NexisDocs
                                if nd == None:
                                    # Get the foliadocs link
                                    nd = fl.ndocs
                                # Add the meta data
                                oResult['metadata'] = json.loads(fl.nmeta)
                                oResult['name'] = fl.name

                                # Show where we are
                                statuscode = "completed"

                        if oResult == None:
                            arErr.append("There was an error. No manuscripts have been added")
                        else:
                            lResults.append(oResult)
            if statuscode == "error":
                data['status'] = "error"
            else:
                # Set the number of files in this batch
                iCount = batch.batchlinks.all().count()
                batch.count = iCount
                batch.save()
                # Show we are ready
                oStatus.set("ready", msg="Read all files")
            # Get a list of errors
            error_list = [str(item) for item in arErr if len(str(item)) > 0]

            # Create the context
            context = dict(
                statuscode=statuscode,
                results=lResults,
                object=nd,
                error_list=error_list
                )

            if len(arErr) == 0 or len(arErr[0]) == 0:
                # Get the HTML response
                data['html'] = render_to_string(template_name, context, request)
            else:
                lHtml = []
                lHtml.append("There are errors in importing this doc")
                for item in arErr:
                    lHtml.append("<br />- {}".format(str(item)))
                data['html'] = "\n".join(lHtml)


        else:
            data['html'] = 'invalid form: {}'.format(form.errors)
            data['status'] = "error"
    else:
        data['html'] = 'Only use POST and make sure you are logged in'
        data['status'] = "error"
 
    # Return the information
    return JsonResponse(data)

def import_mwex(request):
    """Import one MWE Excel file"""

    # Initialisations
    # NOTE: do ***not*** add a breakpoint until *AFTER* form.is_valid
    arErr = []
    error_list = []
    statuscode = ""
    data = {'status': 'ok', 'html': ''}
    template_name = 'doc/import_mwex.html'
    username = request.user.username
    oErr = ErrHandle()

    # Check if the user is authenticated and if it is POST
    if request.user.is_authenticated and request.method == 'POST':

        # Remove previous status object for this user
        Status.objects.filter(user=username).delete()
        # Create a status object
        oStatus = Status(user=username, type="mwex", status="preparing", msg="please wait")
        oStatus.save()
        
        form = UploadMwexForm(request.POST, request.FILES)
        lResults = []
        if form.is_valid():
            # NOTE: from here a breakpoint may be inserted!
            print('import_mwex: valid form')

            # Get the contents of the imported file
            files = request.FILES.getlist('file_source')
            if files != None:
                if len(files) > 0:
                    # Only read the first
                    data_file = files[0]
                    filename = data_file.name

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        arErr.append("No source file specified for the MWE Excel list")
                        statuscode = "error"
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]
                        sBare = arFile[0].strip()

                        # Further processing depends on the extension
                        oResult = None
                        if extension == "xlsx":
                            # Read the excel
                            wb = openpyxl.load_workbook(data_file)
                            ws = None
                            sheetnames = wb.sheetnames
                            # Check the number of worksheets
                            if len(sheetnames) > 0:
                                # walk through the worksheets and check their name
                                for sname in sheetnames:
                                    if "mwe" in sname.lower():
                                        ws = wb[sname]
                                        break
                            else:
                                # We expect the data to be in the first worksheet
                                ws = wb.active
                            # Did we find it?
                            if not ws is None:
                                # Expectations: first  column is *WOORD*
                                #               second column is *SCORE* (concreetheid)
                                row_no = 2
                                lAdded = []
                                lSkipped = []
                                while ws.cell(row=row_no, column=1).value != None:
                                    # Get the woord and the score
                                    woord = ws.cell(row=row_no, column=1).value
                                    score = ws.cell(row=row_no, column=2).value
                                    if isinstance(score,float):
                                        # Turn it into a string with a period decimal separator
                                        score = str(score).replace(",", ".")
                                    oItem = dict(woord=woord, score=score)
                                    # Check if these are already processed in the Epression table
                                    obj = Expression.objects.filter(full=woord).first()
                                    if obj is None:
                                        # Add this expression
                                        obj = Expression.objects.create(full=woord, score=score)
                                        lAdded.append(oItem)
                                    else:
                                        lSkipped.append(oItem)
                                        # Double check the score value
                                        if score != obj.score:
                                            obj.score = score
                                            obj.save()
                                    # Go to the next row
                                    row_no += 1
                                # Add results
                                oResult = {}
                                oResult['added'] = lAdded
                                oResult['skipped'] = lSkipped
                                oResult['added_count'] = len(lAdded)
                                oResult['skipped_count'] = len(lSkipped)
                            # Show where we are
                            statuscode = "completed"
                        else:
                            # Cannot process these
                            oResult = {'status': 'error', 'msg': 'cannot process non Excel (xlsx) files'}
                            statuscode = "error"

                        if oResult == None:
                            arErr.append("There was an error. No Excel file has been loaded")
                        else:
                            lResults.append(oResult)


            if statuscode == "error":
                data['status'] = "error"
            else:
                # Show we are ready
                oStatus.set("ready", msg="Read all data")
            # Get a list of errors
            error_list = [str(item) for item in arErr if len(str(item)) > 0]

            # Create the context
            context = dict(
                statuscode=statuscode,
                filename=filename,
                results=lResults,
                error_list=error_list
                )

            if len(arErr) == 0 or len(arErr[0]) == 0:
                # Get the HTML response
                data['html'] = render_to_string(template_name, context, request)
            else:
                lHtml = []
                lHtml.append("There are errors in importing this mwex")
                for item in arErr:
                    lHtml.append("<br />- {}".format(str(item)))
                data['html'] = "\n".join(lHtml)

        else:
            data['html'] = 'invalid form: {}'.format(form.errors)
            data['status'] = "error"

    else:
        data['html'] = 'Only use POST and make sure you are logged in'
        data['status'] = "error"
 
    # Return the information
    return JsonResponse(data)

def import_twitter_excel(request):
    """Import one Twitter Excel file into TwitterMsg objects"""

    # Initialisations
    # NOTE: do ***not*** add a breakpoint until *AFTER* form.is_valid
    arErr = []
    error_list = []
    statuscode = ""
    sDoFiles = "false"            # Whether to create server-side files for the Twitter messages
    data = {'status': 'ok', 'html': ''}
    template_name = 'doc/import_twitter.html'
    username = request.user.username
    oErr = ErrHandle()

    # Check if the user is authenticated and if it is POST
    if request.user.is_authenticated and request.method == 'POST':

        # Remove previous status object for this user
        Status.objects.filter(user=username).delete()
        # Create a status object
        oStatus = Status.objects.create(user=username, type="twitter", status="preparing", msg="please wait")
        
        form = UploadTwitterExcelForm(request.POST, request.FILES)
        lResults = []
        if form.is_valid():
            # NOTE: ===================================================
            # ====== starting *HERE* a breakpoint may be inserted! ====
            print('import_twitter: valid form')

            # Check for server-side file creation
            cleaned_data = form.cleaned_data
            sDoFiles = cleaned_data.get("dofiles", sDoFiles)
            dofiles = (sDoFiles.lower() == "true")

            # Get the contents of the imported file
            files = request.FILES.getlist('file_source')
            if files != None:
                if len(files) > 0:
                    # Only read the first
                    data_file = files[0]
                    filename = data_file.name

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        arErr.append("No source file specified for the Twitter Excel")
                        statuscode = "error"
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]
                        sBare = arFile[0].strip()

                        # Create room for all the Twitter cells to be read
                        lst_twitter = []

                        # Further processing depends on the extension
                        oResult = None
                        if extension == "xlsx":
                            # Read the excel
                            wb = openpyxl.load_workbook(data_file)
                            # We expect the data to be in the first worksheet
                            ws = wb.worksheets[0]

                            oStatus.set("processing", msg="file={}".format(filename))
                            # Start reading the excel row-by-row
                            bFirstRow = True
                            bHasChanged = False
                            text_columns = []
                            row_no = 1
                            for row in ws.iter_rows():
                                # Show where we are
                                oStatus.set("processing", msg="row={}".format(row_no))
                                # Check contents of the first cell in this row
                                v = row[0].value
                                if v is None or v == "":
                                    # Break from the loop
                                    break
                                # Check for first forw
                                if bFirstRow:
                                    # Treat the columns in here
                                    for cell in row:
                                        value = cell.value
                                        if not value is None and isinstance(value, str):
                                            if "text_tweet_" in value.lower():
                                                # Need to add this column
                                                text_columns.append(cell.col_idx)
                                    # We have treated the first row
                                    bFirstRow = False
                                else:
                                    # Consider all other rows
                                    for column in text_columns:
                                        cell = row[column-1]
                                        if not cell.value is None and cell.value != "":
                                            # Add this cell and its text value
                                            lst_twitter.append(dict(row=row_no, coordinate=cell.coordinate, message=cell.value))

                                            # Add this combination as a TwitterMsg object
                                            obj = TwitterMsg.objects.filter(row=row_no, coordinate=cell.coordinate, message=cell.value).first()
                                            if obj is None:
                                                obj = TwitterMsg.objects.create(row=row_no, coordinate=cell.coordinate, message=cell.value)

                                            # Check and possibly save file
                                            if dofiles:
                                                # Check files and possibly update [obj] with information
                                                obj.check_files()
                                                # If any updates have been made
                                                if not obj.postags is None and obj.postags != "":
                                                    # Start creating a list of token + postag
                                                    lst_postags = []
                                                    # Read the postags string
                                                    arPosTags = obj.postags.split("]\n[")
                                                    for idx, sent in enumerate(arPosTags):
                                                        if len(sent) > 0:
                                                            # Build the actual sentence, based on the position in arPosTags
                                                            if idx == 0:
                                                                # This is the initial
                                                                if idx + 1 == len(arPosTags):
                                                                    # This is the final
                                                                    pass
                                                                else:
                                                                    # Non-final: add ]
                                                                    sSent = "{}]".format(sent)
                                                            elif idx + 1 == len(arPosTags):
                                                                # This is the final, but not the initial
                                                                sSent = "[{}".format(sent)
                                                            else:
                                                                # Non-initial and Non-final
                                                                sSent = "[{}]".format(sent)

                                                            lst_sent = json.loads(sSent)
                                                            for oItem in lst_sent:
                                                                pos = oItem.get("pos")
                                                                word = oItem.get("word")
                                                                if not pos is None:
                                                                    tag = pos.get("tag")
                                                                    lst_postags.append("{}_{}".format(word, tag))
                                                    # Replace the Excel cell
                                                    cell.value = " ".join(lst_postags)
                                                    bHasChanged = True
                                # Go to the next row in counting
                                row_no += 1

                            # Possibly save file
                            if bHasChanged:
                                # Yes, save file
                                outfile = os.path.abspath(os.path.join(WRITABLE_DIR, "..", "folia", filename.replace(".xlsx", "-out.xlsx")))
                                wb.save(outfile)

                            # Add results
                            oResult = {}
                            oResult['tweets'] = lst_twitter
                            oResult['count'] = len(lst_twitter)
                            oResult['rows'] = row_no
                            # Show where we are
                            statuscode = "completed"
                        else:
                            # Cannot process these
                            oResult = {'status': 'error', 'msg': 'cannot process non Excel (xlsx) files'}
                            statuscode = "error"

                        if oResult == None:
                            arErr.append("There was an error. No Excel file has been loaded")
                        else:
                            lResults.append(oResult)


            if statuscode == "error":
                data['status'] = "error"
            else:
                # Show we are ready
                oStatus.set("ready", msg="Read all data")
            # Get a list of errors
            error_list = [str(item) for item in arErr if len(str(item)) > 0]

            # Create the context
            context = dict(
                statuscode=statuscode,
                filename=filename,
                results=lResults,
                error_list=error_list
                )

            if len(arErr) == 0 or len(arErr[0]) == 0:
                # Get the HTML response
                data['html'] = render_to_string(template_name, context, request)
            else:
                lHtml = []
                lHtml.append("There are errors in importing this mwex")
                for item in arErr:
                    lHtml.append("<br />- {}".format(str(item)))
                data['html'] = "\n".join(lHtml)

        else:
            data['html'] = 'invalid form: {}'.format(form.errors)
            data['status'] = "error"

    else:
        data['html'] = 'Only use POST and make sure you are logged in'
        data['status'] = "error"
 
    # Return the information
    return JsonResponse(data)


def import_twitada_excel(request):
    """Import one tagged Twitter Excel file and process it into another"""

    # Initialisations
    # NOTE: do ***not*** add a breakpoint until *AFTER* form.is_valid
    arErr = []
    error_list = []
    statuscode = ""
    sDoFiles = "false"            # Whether to create server-side files for the Twitter messages
    data = {'status': 'ok', 'html': ''}
    template_name = 'doc/import_twitter.html'
    username = request.user.username
    oErr = ErrHandle()

    def iseven(number):
        return "even" if number % 2 == 0 else "odd"

    # Check if the user is authenticated and if it is POST
    if request.user.is_authenticated and request.method == 'POST':

        # Remove previous status object for this user
        Status.objects.filter(user=username).delete()
        # Create a status object
        oStatus = Status.objects.create(user=username, type="twitada", status="preparing", msg="please wait")
        
        form = UploadTwitterExcelForm(request.POST, request.FILES)
        lResults = []
        if form.is_valid():
            # NOTE: ===================================================
            # ====== starting *HERE* a breakpoint may be inserted! ====
            print('import_twitada: valid form')

            # Check for server-side file creation
            cleaned_data = form.cleaned_data

            # Get the contents of the imported file
            files = request.FILES.getlist('file_source')
            if files != None:
                if len(files) > 0:
                    # Only read the first
                    data_file = files[0]
                    filename = data_file.name

                    # Set the status
                    oStatus.set("reading", msg="file={}".format(filename))

                    # Get the source file
                    if data_file == None or data_file == "":
                        arErr.append("No source file specified for the Twitter Excel")
                        statuscode = "error"
                    else:
                        # Check the extension
                        arFile = filename.split(".")
                        extension = arFile[len(arFile)-1]
                        sBare = arFile[0].strip()

                        # Create room for all the Twitter cells to be read
                        lst_twitter = []

                        # Further processing depends on the extension
                        oResult = None
                        if extension == "xlsx":
                            # Read the excel
                            wb = openpyxl.load_workbook(data_file)
                            # We expect the data to be in the first worksheet
                            ws = wb.worksheets[0]

                            oStatus.set("processing", msg="file={}".format(filename))
                            # Start reading the excel row-by-row
                            bFirstRow = True
                            bHasChanged = False
                            text_columns = []
                            row_no = 1
                            for row in ws.iter_rows():
                                # Show where we are
                                oStatus.set("processing", msg="row={}".format(row_no))
                                # Check contents of the first cell in this row
                                v = row[0].value
                                if v is None or v == "":
                                    # Break from the loop
                                    break
                                oStatus.set("Reading row={}".format(row_no))
                                # Check for first forw
                                if bFirstRow:
                                    # Treat the columns in here
                                    for cell in row:
                                        value = cell.value
                                        if not value is None and isinstance(value, str):
                                            if "text_tweet_" in value.lower():
                                                # Need to add this column
                                                text_columns.append(cell.col_idx)
                                    # We have treated the first row
                                    bFirstRow = False
                                else:
                                    # Consider all other rows
                                    for column in text_columns:
                                        cell = row[column-1]
                                        if not cell.value is None and cell.value != "":
                                            # Add this cell and its text value
                                            lst_twitter.append(dict(row=row_no, column=column, coordinate=cell.coordinate, message=cell.value))
                                            bHasChanged = True

                                # Go to the next row in counting
                                row_no += 1

                            # Close the worksheet and workbook
                            wb.close()

                            # Possibly save file
                            if bHasChanged:
                                # Yes, think of save file
                                outfile = os.path.abspath(os.path.join(WRITABLE_DIR, "..", "folia", filename.replace(".xlsx", "-list.xlsx")))

                                # Open workbook for this
                                wb_out = openpyxl.Workbook()
                                ws_out = wb_out.worksheets[0]

                                # Create first row with headers
                                row_no = 1
                                headers = ["row", "column", "coordinate", "word", "coltype"]
                                for idx, header in enumerate(headers):
                                    ws_out.cell(row=row_no, column=idx+1).value = header

                                # Process all other cells
                                for oTwitter in lst_twitter:
                                    row_this = oTwitter.get("row")
                                    col_this = oTwitter.get("column")
                                    coo_this = oTwitter.get("coordinate")
                                    sentence = oTwitter.get("message")

                                    oStatus.set("Converting", msg="main row={} word row={}".format(row_this, row_no))

                                    if not sentence is None and sentence != "":
                                        for word in sentence.split(" "):
                                            row_no += 1
                                            ws_out.cell(row=row_no, column=1).value = row_this
                                            ws_out.cell(row=row_no, column=2).value = col_this
                                            ws_out.cell(row=row_no, column=3).value = coo_this
                                            ws_out.cell(row=row_no, column=4).value = word
                                            ws_out.cell(row=row_no, column=5).value = iseven(col_this)

                                # Save the new workbook
                                wb_out.save(outfile)

                            # Add results
                            oResult = {}
                            oResult['tweets'] = lst_twitter
                            oResult['count'] = len(lst_twitter)
                            oResult['rows'] = row_no
                            # Show where we are
                            statuscode = "completed"
                        else:
                            # Cannot process these
                            oResult = {'status': 'error', 'msg': 'cannot process non Excel (xlsx) files'}
                            statuscode = "error"

                        if oResult == None:
                            arErr.append("There was an error. No Excel file has been loaded")
                        else:
                            lResults.append(oResult)


            if statuscode == "error":
                data['status'] = "error"
            else:
                # Show we are ready
                oStatus.set("ready", msg="Read all data")
            # Get a list of errors
            error_list = [str(item) for item in arErr if len(str(item)) > 0]

            # Create the context
            context = dict(
                statuscode=statuscode,
                filename=filename,
                results=lResults,
                error_list=error_list
                )

            if len(arErr) == 0 or len(arErr[0]) == 0:
                # Get the HTML response
                data['html'] = render_to_string(template_name, context, request)
            else:
                lHtml = []
                lHtml.append("There are errors in importing this mwex")
                for item in arErr:
                    lHtml.append("<br />- {}".format(str(item)))
                data['html'] = "\n".join(lHtml)

        else:
            data['html'] = 'invalid form: {}'.format(form.errors)
            data['status'] = "error"

    else:
        data['html'] = 'Only use POST and make sure you are logged in'
        data['status'] = "error"
 
    # Return the information
    return JsonResponse(data)


class NexisBatchEdit(BasicDetails):
    """Details view for one batch"""

    model = NexisBatch
    mForm = NexisBatchForm
    prefix = "nbatch"
    title = "BatchEdit"
    rtype = "json"
    mainitems = []

    def add_to_context(self, context, instance):
        """Add to the existing context"""

        # Define the main items to show and edit
        context['mainitems'] = [
            {'type': 'plain', 'label': "Date:",             'value': instance.created,  'field_key': 'created'},
            {'type': 'plain', 'label': "Number of texts:",  'value': instance.count,    'field_key': 'count'}
            ]
        context['is_app_editor'] = user_is_ingroup(self.request, "nexis_editor")
        context['is_app_uploader'] = context['is_app_editor']
        # Return the context we have made
        return context


class NexisBatchDetails(NexisBatchEdit):
    """Viewing nexis as HTML"""
    rtype = "html"


class NexisBatchDownload(BasicPart):
    MainModel = NexisBatch
    templat_name = "seeker/download_status.html"
    dtype = "tar.gz"
    action = "download"

    def get_data(self, prefix, dtype):
        gzdata = None
        oErr = ErrHandle()
        try:
            # Who am I?
            username = self.request.user.username
            nexisProc = NexisProcessor(username)
            # What is the directory to use?
            dir = nexisProc.dir
            # Get the NexisBatch object
            batch = self.obj
            if batch != None:
                # Determine file name
                srcdir = os.path.join(dir, "nexisbatch_{}".format(batch.id))
                # Create dir if not existing
                if not os.path.exists(srcdir):
                    os.mkdir(srcdir)

                # Create a list of key names and of meta info
                key_names = ['file_id']
                meta_lines = []
                csv_lines = []

                # Get all the link objects as files in [srcdir]
                qs = batch.batchlinks.all()
                for obj in qs:
                    # Figure out where to put them
                    sbare_name = "nexislink_{}".format(str(obj.id).zfill(6))
                    sbare = os.path.join(srcdir,sbare_name)
                    fmeta = "{}.meta".format(sbare)
                    ftext = "{}.txt".format(sbare)
                    # Write the contents of metadata
                    oMeta = json.loads(obj.nmeta)
                    with open(fmeta, "w") as fp:
                        json.dump(oMeta, fp, indent=2)
                    # Write the text as UTF8
                    with io.open(ftext, 'w', encoding='utf8') as fp:
                        fp.write(obj.nbody)

                    meta = dict(file_id=sbare_name)
                    for k,v in oMeta.items():
                        if not k in key_names: key_names.append(k)
                        meta[k] = v
                    meta_lines.append(meta)

                # Create CSV output
                csv_lines.append(key_names)
                for meta in meta_lines:
                    oCsvLine = []
                    for k in key_names:
                        oCsvLine.append(meta.get(k, ""))
                    csv_lines.append(oCsvLine)
                fexcel = os.path.join(srcdir, "nexisbatch_{}.xlsx".format(str(batch.id).zfill(6)))
                self.csv_to_excel(csv_lines, fexcel)

                # Copy from [srcdir] into tar.gz
                ofname = "{}.tar.gz".format(srcdir)
                with tarfile.open(ofname, "w:gz") as tar:
                    tar.add(srcdir, arcname=os.path.basename(srcdir))

                # Now remove the source directory
                shutil.rmtree(srcdir)

                # Now read the file as binary data
                with io.open(ofname, "rb") as fp:
                    gzdata = fp.read()
        except:
            msg = oErr.get_error_message()

        # Return the data
        return gzdata

    def csv_to_excel(self, csv_lines, filename):
        """Convert CSV data to an Excel worksheet"""

        # Start workbook
        wb = openpyxl.Workbook()
        # ws = wb.get_active_sheet()
        ws = wb.active
        ws.title="Data"

        # Read the header cells and make a header row in the worksheet
        headers = csv_lines[0]
        for col_num in range(len(headers)):
            c = ws.cell(row=1, column=col_num+1)
            c.value = headers[col_num]
            c.font = openpyxl.styles.Font(bold=True)
            # Set width to a fixed size
            ws.column_dimensions[get_column_letter(col_num+1)].width = 8.0        

        row_num = 1
        lCsv = []
        for row in csv_lines[1:]:
            # Keep track of the EXCEL row we are in
            row_num += 1
            # Walk the elements in the data row
            # oRow = {}
            for idx, cell in enumerate(row):
                c = ws.cell(row=row_num, column=idx+1)
                # attempt to see this as a float
                cell_value = row[idx]
                try:
                    cell_value = float(cell_value)
                except ValueError:
                    pass
                c.value = cell_value
                c.alignment = openpyxl.styles.Alignment(wrap_text=False)

        # Save the result in the response
        wb.save(filename)
        return True
    

class NexisListView(BasicList):
    """Search and list nexis batches"""

    model = NexisBatch
    listform = NexisBatchForm
    prefix = "nbatch"
    new_button = False      # Don't show a new button, because new items can only be added by downloading
    plural_name = "Batches of Nexis text files"
    sg_name = "Nexis Batch"
    has_select2 = False     # Don't use Select2 here
    delete_line = True      # Allow deleting a line
    bUseFilter = True
    superuser = False
    order_cols = ['created', 'count', '']
    order_default = ['-created', 'count']
    order_heads = [{'name': 'Date',  'order': 'o=1', 'type': 'str', 'custom': 'created', 'linkdetails': True},
                   {'name': 'Texts', 'order': 'o=2', 'type': 'int', 'field':  'count', 'linkdetails': True, 'main': True},
                   {'name': '',      'order': '',    'type': 'str', 'custom': 'links', 'align': 'right'},
                   {'name': '',      'order': '',    'type': 'str', 'options': 'delete', 'classes': 'tdnowrap'}]
    filters = [
        {'name': 'Date', 'id': 'filter_created', 'enabled': False}
        ]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'created',   'dbfield':  'created',  'keyS': 'created'}
            ]},
        {'section': 'other', 'filterlist': [
            {'filter': 'ndocs',     'fkfield': 'ndocs',  'keyFk': 'ndocs'}]}
        ]
    uploads = [{"title": "nbatch", "label": "Batch", "url": "import_nexis", "msg": "Upload Nexis text files", "type": "multiple"}]

    def initializations(self):
        # Check if I am superuser or not
        self.superuser = self.request.user.is_superuser
        if self.superuser:
            self.order_cols = ['created', 'count', 'ndocs__owner__username', '']
            self.order_heads = [
                {'name': 'Date',  'order': 'o=1', 'type': 'str', 'custom': 'created', 'linkdetails': True},
                {'name': 'Texts', 'order': 'o=2', 'type': 'int', 'field':  'count', 'linkdetails': True, 'main': True},
                {'name': 'User',  'order': 'o=3', 'type': 'str', 'custom':  'user', 'linkdetails': True},
                {'name': '',      'order': '',    'type': 'str', 'custom': 'links', 'align': 'right'},
                {'name': '',      'order': '',    'type': 'str', 'options': 'delete', 'classes': 'tdnowrap'}]

        return None

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []

        # Figure out what to show
        if custom == "created":
            # sBack = instance.created.strftime("%d/%B/%Y (%H:%M)")
            sBack = get_crpp_date(instance.created, True)
        elif custom == "links":
            # Show the download links
            url = reverse('nexisbatch_download', kwargs={'pk': instance.id})
            html.append('<a href="#" title="Download compressed tar.gz" downloadtype="tar.gz" ajaxurl="{}" onclick="ru.basic.post_download(this);">'.format(url))
            html.append('<span class="glyphicon glyphicon-download"></span></a>')
            
            # COmbineer
            sBack = "\n".join(html)
        elif custom == "user":
            sBack = instance.ndocs.owner.username

        # Retourneer wat kan
        return sBack, sTitle

    def adapt_search(self, fields):
        # Initialisations
        lstExclude=None
        qAlternative = None
        if not self.superuser:
            # Make sure only batches are shown for which this user is the owner
            username = self.request.user.username
            owner = User.objects.filter(username = username).first()
            ndocs = NexisDocs.objects.filter(owner=owner)
            if ndocs != None:
                fields['ndocs'] = ndocs

        # Return standard
        return fields, lstExclude, qAlternative

    def add_to_context(self, context, initial):
        context['is_app_editor'] = user_is_ingroup(self.request, "nexis_editor")
        context['is_app_uploader'] = context['is_app_editor']
        return context

