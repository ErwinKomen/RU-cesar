"""
Definition of views for the BASIC app.
"""

from django.apps import apps
from django.contrib.auth.models import User, Group
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.db.models import Q, Prefetch, Count, F
from django.db.models.functions import Lower
from django.db.models.query import QuerySet 
from django.forms.models import model_to_dict
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse, FileResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.generic.detail import DetailView
from django.views.generic.base import RedirectView
from django.views.generic import ListView, View

import json
import fnmatch
import os
import base64
from datetime import datetime
import openpyxl
from openpyxl.utils.cell import get_column_letter
from io import StringIO
import csv

# provide error handling
from .utils import ErrHandle


# Some constants that can be used
paginateSize = 20
paginateSelect = 15
paginateValues = (100, 50, 20, 10, 5, 2, 1, )

# Global debugging 
bDebug = False

# General functions serving the list and details views

def get_application_name():
    """Try to get the name of this application"""

    # Walk through all the installed apps
    for app in apps.get_app_configs():
        # Check if this is a site-package
        if "site-package" not in app.path:
            # Get the name of this app
            name = app.name
            # Take the first part before the dot
            project_name = name.split(".")[0]
            return project_name
    return "unknown"
# Provide application-specific information
PROJECT_NAME = get_application_name()
APPLICATION_NAME = "Cesar"
app_uploader = "{}_uploader".format(PROJECT_NAME.lower())
app_editor = "{}_editor".format(PROJECT_NAME.lower())
app_userplus = "{}_userplus".format(PROJECT_NAME.lower())

def user_is_authenticated(request):
    # Is this user authenticated?
    username = request.user.username
    user = User.objects.filter(username=username).first()
    response = False if user == None else user.is_authenticated
    return response

def user_is_ingroup(request, sGroup):
    # Is this user part of the indicated group?
    username = request.user.username
    user = User.objects.filter(username=username).first()
    # glist = user.groups.values_list('name', flat=True)

    # Only look at group if the user is known
    if user == None:
        glist = []
    else:
        glist = [x.name for x in user.groups.all()]

        # Only needed for debugging
        if bDebug:
            ErrHandle().Status("basic: User [{}] is in groups: {}".format(user, glist))
    # Evaluate the list
    bIsInGroup = (sGroup in glist)
    return bIsInGroup

def get_breadcrumbs(request, name, is_menu, lst_crumb=[], **kwargs):
    """Process one visit and return updated breadcrumbs"""

    # Initialisations
    p_list = []
    p_list.append({'name': 'Home', 'url': reverse('home')})
    # Find out who this is
    username = "anonymous" if request.user == None else request.user.username
    if username != "anonymous" and request.user.username != "":
        # Add the visit
        currenturl = request.get_full_path()
        # Visit.add(username, name, currenturl, is_menu, **kwargs)
        # Set the full path, dependent on the arguments we get
        for crumb in lst_crumb:
            if len(crumb) == 2:
                p_list.append(dict(name=crumb[0], url=crumb[1]))
            else:
                pass
        # Also add the final one
        p_list.append(dict(name=name, url=currenturl))
    # Return the breadcrumbs
    return p_list

def get_current_datetime():
    return timezone.now()

def action_model_changes(form, instance):
    field_values = model_to_dict(instance)
    changed_fields = form.changed_data
    changes = {}
    exclude = []
    if hasattr(form, 'exclude'):
        exclude = form.exclude
    for item in changed_fields: 
        if item in field_values:
            changes[item] = field_values[item]
        elif item not in exclude:
            # It is a form field
            try:
                representation = form.cleaned_data[item]
                if isinstance(representation, QuerySet):
                    # This is a list
                    rep_list = []
                    for rep in representation:
                        rep_str = str(rep)
                        rep_list.append(rep_str)
                    representation = json.dumps(rep_list)
                elif isinstance(representation, str) or isinstance(representation, int):
                    representation = representation
                elif isinstance(representation, object):
                    try:
                        representation = representation.__str__()
                    except:
                        representation = str(representation)
                changes[item] = representation
            except:
                changes[item] = "(unavailable)"
    return changes

def has_string_value(field, obj):
    response = (field != None and field in obj and obj[field] != None and obj[field] != "")
    return response

def has_list_value(field, obj):
    response = (field != None and field in obj and obj[field] != None and len(obj[field]) > 0)
    return response

def isempty(value):
    response = True
    if value != None:
        if isinstance(value, str):
            response = (value == "")
        elif isinstance(value, int):
            response = False
        else:
            response = (len(value) == 0)
    return response

def has_obj_value(field, obj):
    response = (field != None and field in obj and obj[field] != None)
    return response

def adapt_search(val):
    # First trim
    val = val.strip()
    # Then add start and en matter 
    val = '^' + fnmatch.translate(val) + '$'
    return val

def treat_bom(sHtml):
    """REmove the BOM marker except at the beginning of the string"""

    # Check if it is in the beginning
    bStartsWithBom = sHtml.startswith(u'\ufeff')
    # Remove everywhere
    sHtml = sHtml.replace(u'\ufeff', '')
    # Return what we have
    return sHtml

def make_search_list(filters, oFields, search_list, qd):
    """Using the information in oFields and search_list, produce a revised filters array and a lstQ for a Queryset"""

    def enable_filter(filter_id, head_id=None):
        for item in filters:
            if filter_id in item['id']:
                item['enabled'] = True
                # Break from my loop
                break
        # Check if this one has a head
        if head_id != None and head_id != "":
            for item in filters:
                if head_id in item['id']:
                    item['enabled'] = True
                    # Break from this sub-loop
                    break
        return True

    def get_value(obj, field, default=None):
        if field in obj:
            sBack = obj[field]
        else:
            sBack = default
        return sBack

    oErr = ErrHandle()

    try:
        # (1) Create default lstQ
        lstQ = []

        # (2) Reset the filters in the list we get
        for item in filters: item['enabled'] = False
    
        # (3) Walk all sections
        for part in search_list:
            head_id = get_value(part, 'section')

            # (4) Walk the list of defined searches
            for search_item in part['filterlist']:
                keyS = get_value(search_item, "keyS")
                keyId = get_value(search_item, "keyId")
                keyFk = get_value(search_item, "keyFk")
                keyList = get_value(search_item, "keyList")
                infield = get_value(search_item, "infield")
                dbfield = get_value(search_item, "dbfield")
                fkfield = get_value(search_item, "fkfield")
                keyType = get_value(search_item, "keyType")
                filter_type = get_value(search_item, "filter")
                s_q = ""
                arFkField = []
                if fkfield != None:
                    arFkField = fkfield.split("|")
               
                # Main differentiation: fkfield or dbfield
                if fkfield:
                    # Check for keyS
                    if has_string_value(keyS, oFields):
                        # Check for ID field
                        if has_string_value(keyId, oFields):
                            val = oFields[keyId]
                            if not isinstance(val, int): 
                                try:
                                    val = val.id
                                except:
                                    pass
                            enable_filter(filter_type, head_id)
                            s_q = Q(**{"{}__id".format(fkfield): val})
                        elif has_obj_value(fkfield, oFields):
                            val = oFields[fkfield]
                            enable_filter(filter_type, head_id)
                            s_q = Q(**{fkfield: val})
                        else:
                            val = oFields[keyS]
                            enable_filter(filter_type, head_id)
                            # We are dealing with a foreign key (or multiple)
                            if len(arFkField) > 1:
                                iStop = 1
                            # we are dealing with a foreign key, so we should use keyFk
                            s_q = None
                            for fkfield in arFkField:
                                if "*" in val:
                                    val = adapt_search(val)
                                    s_q_add = Q(**{"{}__{}__iregex".format(fkfield, keyFk): val})
                                else:
                                    s_q_add = Q(**{"{}__{}__iexact".format(fkfield, keyFk): val})
                                if s_q == None:
                                    s_q = s_q_add
                                else:
                                    s_q |= s_q_add
                    elif has_obj_value(fkfield, oFields):
                        val = oFields[fkfield]
                        enable_filter(filter_type, head_id)
                        if isinstance(val, Q):
                            s_q = val
                        else:
                            s_q = Q(**{fkfield: val})
                            external = get_value(search_item, "external")
                            if has_string_value(external, oFields):
                                qd[external] = getattr(val, "name")
                elif dbfield:
                    # We are dealing with a plain direct field for the model
                    # OR: it is also possible we are dealing with a m2m field -- that gets the same treatment
                    # Check for keyS
                    if has_string_value(keyS, oFields):
                        # Check for ID field
                        if has_string_value(keyId, oFields):
                            val = oFields[keyId]
                            enable_filter(filter_type, head_id)
                            s_q = Q(**{"{}__id".format(dbfield): val})
                        elif has_obj_value(keyFk, oFields):
                            val = oFields[keyFk]
                            enable_filter(filter_type, head_id)
                            s_q = Q(**{dbfield: val})
                        else:
                            val = oFields[keyS]
                            enable_filter(filter_type, head_id)
                            if isinstance(val, int):
                                s_q = Q(**{"{}".format(dbfield): val})
                            elif "*" in val:
                                val = adapt_search(val)
                                s_q = Q(**{"{}__iregex".format(dbfield): val})
                            else:
                                s_q = Q(**{"{}__iexact".format(dbfield): val})
                    elif keyType == "has":
                        # Check the count for the db field
                        val = oFields[filter_type]
                        if val == "yes" or val == "no":
                            enable_filter(filter_type, head_id)
                            if val == "yes":
                                s_q = Q(**{"{}__gt".format(dbfield): 0})
                            else:
                                s_q = Q(**{"{}".format(dbfield): 0})

                # Check for list of specific signatures
                if has_list_value(keyList, oFields):
                    s_q_lst = ""
                    enable_filter(filter_type, head_id)
                    code_list = [getattr(x, infield) for x in oFields[keyList]]
                    if fkfield:
                        # Now we need to look at the id's
                        if len(arFkField) > 1:
                            # THere are more foreign keys: combine in logical or
                            s_q_lst = ""
                            for fkfield in arFkField:
                                if s_q_lst == "":
                                    s_q_lst = Q(**{"{}__{}__in".format(fkfield, infield): code_list})
                                else:
                                    s_q_lst |= Q(**{"{}__{}__in".format(fkfield, infield): code_list})
                        else:
                            # Just one foreign key
                            s_q_lst = Q(**{"{}__{}__in".format(fkfield, infield): code_list})
                    elif dbfield:
                        s_q_lst = Q(**{"{}__in".format(infield): code_list})
                    s_q = s_q_lst if s_q == "" else s_q | s_q_lst

                # Possibly add the result to the list
                if s_q != "": lstQ.append(s_q)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("make_search_list")
        lstQ = []

    # Return what we have created
    return filters, lstQ, qd

def make_ordering(qs, qd, order_default, order_cols, order_heads):

    oErr = ErrHandle()

    try:
        bAscending = True
        sType = 'str'
        order = []
        colnum = ""
        # reset 'used' feature for all heads
        for item in order_heads: item['used'] = None
        if 'o' in qd and qd['o'] != "":
            colnum = qd['o']
            if '=' in colnum:
                colnum = colnum.split('=')[1]
            if colnum != "":
                order = []
                iOrderCol = int(colnum)
                bAscending = (iOrderCol>0)
                iOrderCol = abs(iOrderCol)

                # Set the column that it is in use
                order_heads[iOrderCol-1]['used'] = 1
                # Get the type
                sType = order_heads[iOrderCol-1]['type']
                for order_item in order_cols[iOrderCol-1].split(";"):
                    if order_item != "":
                        if sType == 'str':
                            order.append(Lower(order_item).asc(nulls_last=True))
                        else:
                            order.append(F(order_item).asc(nulls_last=True))
                if bAscending:
                    order_heads[iOrderCol-1]['order'] = 'o=-{}'.format(iOrderCol)
                else:
                    # order = "-" + order
                    order_heads[iOrderCol-1]['order'] = 'o={}'.format(iOrderCol)

                # Reset the sort order to ascending for all others
                for idx, item in enumerate(order_heads):
                    if idx != iOrderCol - 1:
                        # Reset this sort order
                        order_heads[idx]['order'] = order_heads[idx]['order'].replace("-", "")
        else:
            orderings = []
            for order_item in order_default:
                if ";" in order_item:
                    for sub_item in order_item.split(";"):
                        orderings.append(sub_item)
                else:
                    orderings.append(order_item)
            for order_item in orderings:
                if order_item != "":
                    if "-" in order_item:
                        order.append(order_item)
                    else:
                        order.append(Lower(order_item))

           #  order.append(Lower(order_cols[0]))
        if sType == 'str':
            if len(order) > 0:
                qs = qs.order_by(*order)
        else:
            qs = qs.order_by(*order)
        # Possibly reverse the order
        if not bAscending:
            qs = qs.reverse()
    except:
        msg = oErr.get_error_message()
        oErr.DoError("make_ordering")
        lstQ = []

    return qs, order_heads, colnum

def add_rel_item(rel_item, value, resizable=False, title=None, align=None, link=None, nowrap=True, main=None, draggable=None):
    oAdd = dict(value=value)
    if resizable: oAdd['initial'] = 'small'
    if title != None: oAdd['title'] = title
    if align != None: oAdd['align'] = align
    if link != None: oAdd['link'] = link
    if main != None: oAdd['main'] = main
    if nowrap != None: oAdd['nowrap'] = nowrap
    if draggable != None: oAdd['draggable'] = draggable
    rel_item.append(oAdd)
    return True

def base64_encode(sInput):
    message_bytes = sInput.encode("utf8")
    base64_bytes = base64.b64encode(message_bytes)
    sOutput = base64_bytes.decode("utf8")
    return sOutput

def base64_decode(sInput):
    base64_bytes = sInput.encode('utf8')
    message_bytes = base64.b64decode(base64_bytes)
    sOutput = message_bytes.decode('utf8')
    return sOutput

def csv_to_excel(sCsvData, response, delimiter=","):
    """Convert CSV data to an Excel worksheet"""

    oErr = ErrHandle()
    try:
        # Start workbook
        wb = openpyxl.Workbook()
        ws = wb.active  # wb.get_active_sheet()
        ws.title="Data"

        # Start accessing the string data 
        f = StringIO(sCsvData)
        reader = csv.reader(f, delimiter=delimiter)

        # Read the header cells and make a header row in the worksheet
        headers = next(reader)
        for col_num in range(len(headers)):
            c = ws.cell(row=1, column=col_num+1)
            c.value = headers[col_num]
            c.font = openpyxl.styles.Font(bold=True)
            # Set width to a fixed size
            ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0        

        row_num = 1
        lCsv = []
        for row in reader:
            # Keep track of the EXCEL row we are in
            row_num += 1
            # Walk the elements in the data row
            # oRow = {}
            for idx, cell in enumerate(row):
                c = ws.cell(row=row_num, column=idx+1)
                c.value = row[idx]
                c.alignment = openpyxl.styles.Alignment(wrap_text=False)
        # Save the result in the response
        wb.save(response)
    except:
        msg = oErr.get_error_message()
        oErr.DoError("csv_to_excel")

    return response



# The views that are defined by 'basic'

class BasicList(ListView):
    """Basic listview
    
    This listview inherits the standard listview and adds a few automatic matters
    """

    paginate_by = 15
    entrycount = 0
    qd = None
    bFilter = False
    basketview = False
    template_name = 'basic/basic_list.html'
    bHasParameters = False
    bUseFilter = False
    new_button = True
    initial = None
    listform = None
    has_select2 = False
    plural_name = ""
    sg_name = ""
    basic_name = ""
    basic_name_prefix = ""
    basic_edit = ""
    basic_details = ""
    basic_add = ""
    add_text = "Add a new"
    prefix = ""
    order_default = []
    order_cols = []
    order_heads = []
    filters = []
    searches = []
    downloads = []
    custombuttons = []
    list_fields = []
    uploads = []
    delete_line = False
    none_on_empty = False
    use_team_group = False
    redirectpage = ""
    downloadname = None
    lst_typeaheads = []
    sort_order = ""
    param_list = []
    extend_template = "layout.html"
    qs = None
    page_function = "ru.basic.search_paged_start"

    def initializations(self):
        return None

    def get_app_access(self, context):
        return True

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(BasicList, self).get_context_data(**kwargs)

        oErr = ErrHandle()

        #self.initializations()

        # Get parameters for the search
        if self.initial == None:
            initial = self.request.POST if self.request.POST else self.request.GET
        else:
            initial = self.initial

        # Need to load the correct form
        if self.listform:
            prefix = "" if self.prefix == "any" else self.prefix
            if self.use_team_group:
                frm = self.listform(initial, prefix=prefix, username=self.request.user.username, team_group=app_editor, userplus=app_userplus)
            else:
                frm = self.listform(initial, prefix=prefix)
            if frm.is_valid():
                context['{}Form'.format(self.prefix)] = frm
                # Get any possible typeahead parameters
                lst_form_ta = getattr(frm, "typeaheads", None)
                if lst_form_ta != None:
                    for item in lst_form_ta:
                        self.lst_typeaheads.append(item)

            if self.has_select2:
                context['has_select2'] = True
            context['listForm'] = frm

        # Determine the count 
        context['entrycount'] = self.entrycount # self.get_queryset().count()

        # Make sure the paginate-values are available
        context['paginateValues'] = paginateValues

        if 'paginate_by' in initial:
            context['paginateSize'] = int(initial['paginate_by'])
        else:
            context['paginateSize'] = paginateSize

        # Need to pass on a pagination function
        if self.page_function:
            context['page_function'] = self.page_function

        # Set the page number if needed
        if 'page_obj' in context and 'page' in initial and initial['page'] != "":
            # context['page_obj'].number = initial['page']
            page_num = int(initial['page'])
            context['page_obj'] = context['paginator'].page( page_num)
            # Make sure to adapt the object_list
            context['object_list'] = context['page_obj']
            self.param_list.append("page={}".format(page_num))

        # Make sure the parameter list becomes available
        params = ""
        if len(self.param_list) > 0:
            params = base64_encode( "&".join(self.param_list))
        context['params'] = params

        # Set the title of the application
        if self.plural_name =="":
            self.plural_name = str(self.model._meta.verbose_name_plural)
        context['title'] = self.plural_name
        if self.basic_name == "":
            if self.basic_name_prefix == "":
                self.basic_name = str(self.model._meta.model_name)
            else:
                self.basic_name = "{}{}".format(self.basic_name_prefix, self.prefix)
        context['titlesg'] = self.sg_name if self.sg_name != "" else self.basic_name.capitalize()
        context['basic_name'] = self.basic_name
        if self.basic_add:
            basic_add = reverse(self.basic_add)
        else:
            basic_add = reverse("{}_details".format(self.basic_name))
        context['basic_add'] = basic_add
        context['basic_list'] = reverse("{}_list".format(self.basic_name))
        context['basic_edit'] = self.basic_edit if self.basic_edit != "" else "{}_edit".format(self.basic_name)
        context['basic_details'] = self.basic_details if self.basic_details != "" else "{}_details".format(self.basic_name)

        # Make sure to transform the 'object_list'  into a 'result_list'
        context['result_list'] = self.get_result_list(context['object_list'])

        context['sortOrder'] = self.sort_order

        context['new_button'] = self.new_button
        context['add_text'] = self.add_text
        context['extend_template'] = self.extend_template

        # Possible generic EXCEL download...
        context['download_excel'] = None
        if not self.qs is None and hasattr(self.model, 'specification'):
            list_spec = self.model.specification
            if isinstance(list_spec, list):
                # There is a specification list: make Excel download available by providing the right URL for downloading
                context['download_excel'] = context['basic_list']

        # Adapt possible downloads
        if len(self.downloads) > 0:
            for item in self.downloads:
                if 'url' in item and item['url'] != "" and "/" not in item['url']:
                    item['url'] = reverse(item['url'])
            context['downloads'] = self.downloads

        # Specify possible upload
        if len(self.uploads) > 0:
            for item in self.uploads:
                if 'url' in item and item['url'] != "" and "/" not in item['url']:
                    item['url'] = reverse(item['url'])
            context['uploads'] = self.uploads

        # Custom buttons
        if len(self.custombuttons) > 0:
            for item in self.custombuttons:
                if 'template_name' in item:
                    # get the code of the template
                    pass
            context['custombuttons'] = self.custombuttons

        # Delete button per line?
        if self.delete_line:
            context['delete_line'] = True

        # Make sure we pass on the ordered heads
        context['order_heads'] = self.order_heads
        context['has_filter'] = self.bFilter
        fsections = []
        # Adapt filters with the information from searches
        for section in self.searches:
            oFsection = {}
            bHasValue = False
            # Add filter section name
            section_name = section['section']
            if section_name != "" and section_name not in fsections:
                oFsection = dict(name=section_name, has_value=False)
                # fsections.append(dict(name=section_name))
            # Copy the relevant search filter
            for item in section['filterlist']:
                # Find the corresponding item in the filters
                id = "filter_{}".format(item['filter'])
                for fitem in self.filters:
                    if id == fitem['id']:
                        try:
                            fitem['search'] = item
                            fitem['has_keylist'] = False
                            # Add possible fields
                            if 'keyS' in item and item['keyS'] in frm.cleaned_data: 
                                fitem['keyS'] = frm[item['keyS']]
                                if fitem['keyS'].value(): 
                                    bHasValue = True
                            if 'keyList' in item and item['keyList'] in frm.cleaned_data: 
                                if frm.fields[item['keyList']].initial or frm.cleaned_data[item['keyList']].count() > 0: 
                                    bHasValue = True
                                fitem['keyList'] = frm[item['keyList']]
                                fitem['has_keylist'] = True
                            if 'keyS' in item and item['keyS'] in frm.cleaned_data: 
                                if 'dbfield' in item and item['dbfield'] in frm.cleaned_data:
                                    fitem['dbfield'] = frm[item['dbfield']]
                                    if fitem['dbfield'].value(): 
                                        bHasValue = True
                                elif 'fkfield' in item and item['fkfield'] in frm.cleaned_data:
                                    fitem['fkfield'] = frm[item['fkfield']]                                    
                                    if fitem['fkfield'].value(): bHasValue = True
                                else:
                                    # There is a keyS without corresponding fkfield or dbfield
                                    pass
                            break
                        except:
                            sMsg = oErr.get_error_message()
                            break
            if bHasValue: 
                oFsection['has_value'] = True
            if oFsection != None: fsections.append(oFsection)

        # Make it available
        context['filters'] = self.filters
        context['fsections'] = fsections
        context['list_fields'] = self.list_fields

        # Add any typeaheads that should be initialized
        context['typeaheads'] = json.dumps( self.lst_typeaheads)

        # Check if user may upload
        context['is_authenticated'] = user_is_authenticated(self.request)
        context['authenticated'] = context['is_authenticated'] 
        context['is_app_uploader'] = user_is_ingroup(self.request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(self.request, app_editor)
        context['is_app_userplus'] = user_is_ingroup(self.request, app_userplus)

        # Possible adapt the access stuff
        self.get_app_access(context)

        # Process this visit and get the new breadcrumbs object
        prevpage = reverse('home')
        context['prevpage'] = prevpage
        context['breadcrumbs'] = get_breadcrumbs(self.request, self.plural_name, True)

        context['usebasket'] = self.basketview

        # Allow others to add to context
        context = self.add_to_context(context, initial)

        # Return the calculated context
        return context

    def add_to_context(self, context, initial):
        return context

    def get_data(self, prefix, dtype, response=None):
        """Gather the data as CSV, including a header line and comma-separated"""

        # Initialize
        lData = []
        sData = ""
        oErr = ErrHandle()
        try:
            # Sanity check - should have been done already
            if not hasattr(self.model, 'specification'):
                return sData

            # Get the specification
            specification = getattr(self.model, 'specification')

            # Need to know who this user (profile) is
            user = self.request.user
            username = user.username
            #profile = user.user_profiles.first()
            team_group = app_editor
            # kwargs = {'profile': profile, 'username': username, 'team_group': team_group}
            kwargs = {'username': username, 'team_group': team_group}

            # Get the queryset, which is based on the listview parameters
            qs = self.get_queryset()

            # Start workbook
            wb = openpyxl.Workbook()
            # Create worksheet with data
            ws = wb.active
            ws.title = "Data"

            # Create header cells
            headers = [x['name'] for x in specification if x['type'] != "" ]
            headers.insert(0, "Id")
            for col_num in range(len(headers)):
                c = ws.cell(row=1, column=col_num+1)
                c.value = headers[col_num]
                c.font = openpyxl.styles.Font(bold=True)
                # Set width to a fixed size
                ws.column_dimensions[get_column_letter(col_num+1)].width = 5.0        

            row_num = 1
            # Walk the items in the queryset
            for obj in qs:
                row_num += 1
                # Add the object id
                ws.cell(row=row_num, column=1).value = obj.id
                col_num = 2
                # Walk the items
                for item in specification:
                    if item['type'] != "":
                        key, value = obj.custom_getkv(item, kwargs=kwargs)
                        ws.cell(row=row_num, column=col_num).value = value
                        col_num += 1

            # Save it
            wb.save(response)
            sData = response
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_data")
        return sData

    def download_excel(self, dtype):
        """Create a generic Excel download based on [specification]"""

        response = None
        oErr = ErrHandle()
        try:

            # Make a download name
            downloadname = self.model.__name__
            if not self.downloadname is None and self.downloadname != "":
                downloadname = self.downloadname
            appl_name = APPLICATION_NAME
            sDbName = "{}_{}.xlsx".format(appl_name, downloadname)

            # Convert 'compressed_content' to an Excel worksheet
            sContentType = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            response = HttpResponse(content_type=sContentType)
            response['Content-Disposition'] = 'attachment; filename="{}"'.format(sDbName)    
            # Get the Data
            sData = self.get_data('', dtype, response)

        except:
            msg = oErr.get_error_message()
            oErr.DoError("process_download")

        return response

    def get_result_list(self, obj_list):
        result_list = []
        # Walk all items in the object list
        for obj in obj_list:
            # Transform this object into a list of objects that can be shown
            result = dict(id=obj.id)
            fields = []
            for head in self.order_heads:
                fobj = dict(value="")
                fname = None
                if 'field' in head:
                    # This is a field that can be shown
                    fname = head['field']
                    default = "" if 'default' not in head else head['default']
                    value = getattr(obj, fname, default)
                    if not value is None:
                        fobj['value'] = value
                elif 'custom' in head:
                    # The user should provide a way to determine the value for this field
                    fvalue, ftitle = self.get_field_value(obj, head['custom'])
                    if not fvalue is None:
                        fobj['value']= fvalue
                    if ftitle != None:
                        fobj['title'] = ftitle
                    fname = head['custom']
                classes = []
                if fname != None: classes.append("{}-{}".format(self.basic_name, fname))
                if 'linkdetails' in head and head['linkdetails']: fobj['linkdetails'] = True
                if 'main' in head and head['main']:
                    fobj['styles'] = "width: 100%;"
                    fobj['main'] = True
                    if self.delete_line:
                        classes.append("ms editable")
                elif 'options' in head and len(head['options']) > 0:
                    options = head['options']
                    if 'delete' in options:
                        fobj['delete'] = True
                    fobj['styles'] = "width: {}px;".format(50 * len(options))
                    classes.append("tdnowrap")
                else:
                    fobj['styles'] = "width: 100px;"
                    classes.append("tdnowrap")
                if 'align' in head and head['align'] != "":
                    fobj['align'] = head['align'] 
                fobj['classes'] = " ".join(classes)
                fields.append(fobj)
            # Make the list of field-values available
            result['fields'] = fields
            # Add to the list of results
            result_list.append(result)
        return result_list

    def get_template_names(self):
        names = [ self.template_name ]
        return names

    def get_field_value(self, instance, custom):
        return "", ""

    def get_paginate_by(self, queryset):
        """
        Paginate by specified value in default class property value.
        """
        return self.paginate_by

    def get_basketqueryset(self):
        """User-specific function to get a queryset based on a basket"""
        return None

    def adapt_search(self, fields):
        return fields, None, None
  
    def get_queryset(self):
        oErr = ErrHandle()
        try:
            self.initializations()

            # Get the parameters passed on with the GET or the POST request
            get = self.request.GET if self.request.method == "GET" else self.request.POST
            get = get.copy()
            self.qd = get

            username=self.request.user.username
            team_group=app_editor

            self.bFilter = False
            self.bHasParameters = (len(get) > 0)
            bHasListFilters = False
            if self.bHasParameters:
                # y = [x for x in get ]
                bHasListFilters = len([x for x in get if self.prefix in x and get[x] != ""]) > 0
                if not bHasListFilters:
                    self.basketview = ("usebasket" in get and get['usebasket'] == "True")

            # Initial setting of qs
            qs = self.model.objects.none()

            # Get the queryset and the filters
            if self.basketview:
                self.basketview = True
                # We should show the contents of the basket
                # (1) Reset the filters
                for item in self.filters: item['enabled'] = False
                # (2) Indicate we have no filters
                self.bFilter = False
                # (3) Set the queryset -- this is listview-specific
                qs = self.get_basketqueryset()

                # Do the ordering of the results
                order = self.order_default
                qs, self.order_heads, colnum = make_ordering(qs, self.qd, order, self.order_cols, self.order_heads)
            elif self.bHasParameters or self.bUseFilter:
                self.basketview = False
                lstQ = []
                # Indicate we have no filters
                self.bFilter = False

                # Read the form with the information
                prefix = self.prefix
                if prefix == "any": prefix = ""
                if self.use_team_group:
                    thisForm = self.listform(self.qd, prefix=prefix, username=username, team_group=team_group)
                else:
                    thisForm = self.listform(self.qd, prefix=prefix)

                if thisForm.is_valid():
                    # Process the criteria for this form
                    oFields = thisForm.cleaned_data

                    # Set the param_list variable
                    self.param_list = []
                    lookfor = "{}-".format(prefix)
                    for k,v in self.qd.items():
                        if lookfor in k and not isempty(v):
                            self.param_list.append("{}={}".format(k,v))

                    # Allow user to adapt the list of search fields
                    oFields, lstExclude, qAlternative = self.adapt_search(oFields)
                
                    self.filters, lstQ, self.initial = make_search_list(self.filters, oFields, self.searches, self.qd)
                
                    # Calculate the final qs
                    if len(lstQ) == 0 and not self.none_on_empty:
                        if lstExclude:
                            qs = self.model.exclude(*lstExclude)
                        else:
                            # Just show everything
                            qs = self.model.objects.all()
                    else:
                        # There is a filter, so build it up
                        filter = lstQ[0]
                        for item in lstQ[1:]:
                            filter = filter & item
                        if qAlternative:
                            filter = ( filter ) | ( ( qAlternative ) & filter )

                        # DEBUG
                        #qs = self.model.objects.filter(Q(fdocs__owner_id=64))

                        # Check if excluding is needed
                        if lstExclude:
                            # qs = self.model.objects.filter(*lstQ).exclude(*lstExclude).distinct()
                            qs = self.model.objects.filter(filter).exclude(*lstExclude).distinct()
                        else:
                            # qs = self.model.objects.filter(*lstQ).distinct()
                            qs = self.model.objects.filter(filter).distinct()
                        # Only set the [bFilter] value if there is an overt specified filter
                        for filter in self.filters:
                            if filter['enabled'] and ('head_id' not in filter or filter['head_id'] != 'filter_other'):
                                self.bFilter = True
                                break
                        # OLD self.bFilter = True
                elif not self.none_on_empty:
                    # Just show everything
                    qs = self.model.objects.all().distinct()

                    # x = self.model.objects.filter(filter).count()
                # Do the ordering of the results
                order = self.order_default
                qs, self.order_heads, colnum = make_ordering(qs, self.qd, order, self.order_cols, self.order_heads)
            else:
                # No filter and no basked: show all
                self.basketview = False
                qs = self.model.objects.all().distinct()
                order = self.order_default
                qs, tmp_heads, colnum = make_ordering(qs, self.qd, order, self.order_cols, self.order_heads)
            self.sort_order = colnum

            # Determine the length
            # self.entrycount = len(qs)
            self.entrycount = qs.count()

            # Return the resulting filtered and sorted queryset
            self.qs = qs
        except:
            msg = oErr.get_error_message()
            oErr.DoError("get_queryset")
        return qs

    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            # Do not allow to get a good response
            response = redirect(reverse('nlogin'))
        else:
            # FIrst do my own initializations
            self.initializations()

            # Get the parameters passed on with the GET or the POST request
            get = request.GET if request.method == "GET" else request.POST
            get = get.copy()

            ## If this is a 'usersearch' then replace the parameters
            #usersearch_id = get.get("usersearch")
            #if usersearch_id != None:
            #    get = UserSearch.load_parameters(usersearch_id, get)
            self.qd = get

            # Check for downloading
            if request.method == "POST" and self.qd.get("action") == "download":
                dtype = self.qd.get("dtype", "")
                if dtype != "":
                    # This is not the regular listview, but just a downloading action
                    response = self.download_excel(dtype)
            else:
                # Then check if we have a redirect or not
                if self.redirectpage == "":
                    # We can continue with the normal 'get()'
                    response = super(BasicList, self).get(request, *args, **kwargs)
                else:
                    response = redirect(self.redirectpage)
        # REturn the appropriate response
        return response

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)
    

class BasicDetails(DetailView):
    """Extension of the normal DetailView class for PASSIM"""

    template_name = ""      # Template for GET
    template_post = ""      # Template for POST
    formset_objects = []    # List of formsets to be processed
    form_objects = []       # List of form objects
    afternewurl = ""        # URL to move to after adding a new item
    prefix = ""             # The prefix for the one (!) form we use
    previous = None         # Start with empty previous page
    title = ""              # The title to be passed on with the context
    titlesg = None          # Alternative title in singular
    rtype = "json"          # JSON response (alternative: html)
    dname = None            # If rtype == download, this may be the name of the downloaded document
    prefix_type = ""        # Whether the adapt the prefix or not ('simple')
    mForm = None            # Model form
    basic_name = None
    basic_name_prefix = ""
    basic_add = ""
    delimiter = ","
    extend_template = "layout.html"
    add_text = "Add a new"
    permission = "write"    # Permission can be: (nothing), "read" and "write"
    new_button = False
    do_not_save = False
    listview = None
    listviewtitle = None
    has_select2 = False
    newRedirect = False     # Redirect the page name to a correct one after creating
    use_team_group = False
    redirectpage = ""       # Where to redirect to
    afterurl = None
    add = False             # Are we adding a new record or editing an existing one?
    is_basic = True         # Is this a basic details/edit view?
    history_button = False  # Show history button for this view
    lst_typeahead = []

    def get(self, request, pk=None, *args, **kwargs):
        # Initialisation
        data = {'status': 'ok', 'html': '', 'statuscode': ''}
        # always do this initialisation to get the object
        self.initializations(request, pk)
        if not request.user.is_authenticated:
            # Do not allow to get a good response
            if self.rtype == "json":
                data['html'] = "(No authorization)"
                data['status'] = "error"

                # Set any possible typeaheads
                data['typeaheads'] = self.lst_typeahead

                response = JsonResponse(data)
            else:
                response = redirect(reverse('nlogin'))
        else:
            context = self.get_context_data(object=self.object)

            if self.is_basic and self.template_name == "":
                if self.rtype == "json":
                    self.template_name = "basic/basic_edit.html"
                else:
                    self.template_name = "basic/basic_details.html"
            # Possibly indicate form errors
            # NOTE: errors is a dictionary itself...
            if 'errors' in context and len(context['errors']) > 0:
                data['status'] = "error"
                data['msg'] = context['errors']

            if self.rtype == "json":
                # We render to the _name 
                sHtml = render_to_string(self.template_name, context, request)
                sHtml = sHtml.replace("\ufeff", "")
                data['html'] = sHtml
                # Set any possible typeaheads
                data['typeaheads'] = self.lst_typeahead
                response = JsonResponse(data)
            elif self.redirectpage != "":
                return redirect(self.redirectpage)
            else:
                # Set any possible typeaheads
                context['typeaheads'] = json.dumps(self.lst_typeahead)
                # This takes self.template_name...
                oErr = ErrHandle()
                try:
                    sHtml = render_to_string(self.template_name, context, request)
                    sHtml = sHtml.replace("\ufeff", "")
                    response = HttpResponse(sHtml)
                except:
                    msg = oErr.get_error_message()
                    response = None
                # response = self.render_to_response(context)

        # Return the response
        return response

    def post(self, request, pk=None, *args, **kwargs):
        # Initialisation
        data = {'status': 'ok', 'html': '', 'statuscode': ''}
        # always do this initialisation to get the object
        self.initializations(request, pk)
        # Make sure only POSTS get through that are authorized
        if request.user.is_authenticated:
            context = self.get_context_data(object=self.object)
            # Check if 'afternewurl' needs adding
            if 'afternewurl' in context:
                data['afternewurl'] = context['afternewurl']
            # Check if 'afterdelurl' needs adding
            if 'afterdelurl' in context:
                data['afterdelurl'] = context['afterdelurl']
            # Possibly indicate form errors
            # NOTE: errors is a dictionary itself...
            if 'errors' in context and len(context['errors']) > 0:
                data['status'] = "error"
                data['msg'] = context['errors']

            if self.is_basic and self.template_name == "":
                if self.rtype == "json":
                    self.template_name = "basic/basic_edit.html"
                else:
                    self.template_name = "basic/basic_details.html"

            if self.rtype == "json":
                if self.template_post == "": self.template_post = self.template_name
                response = render_to_string(self.template_post, context, request)
                response = response.replace("\ufeff", "")
                data['html'] = response
                # Set any possible typeaheads
                data['typeaheads'] = self.lst_typeahead
                if not self.afterurl is None:
                    data['afterurl'] = self.afterurl
                    self.afterurl = None
                response = JsonResponse(data)
            elif self.rtype == "download":
                if self.template_post == "": self.template_post = self.template_name
                # We are being asked to download something
                if self.dtype != "":
                    # Initialise return status
                    oBack = {'status': 'ok'}
                    sType = "csv" if (self.dtype == "xlsx") else self.dtype

                    # Get the data
                    if self.dtype == "htmldoc":
                        sData = render_to_string(self.template_post, context, request)
                        sData = sData.replace("\ufeff", "")
                    else:
                        sData = self.get_data('', self.dtype)
                    # Decode the data and compress it using gzip
                    bUtf8 = (self.dtype != "db")
                    bUsePlain = (self.dtype == "xlsx" or self.dtype == "csv")

                    # Create name for download
                    modelname = self.model.__name__
                    obj_id = "n" if self.object == None else self.object.id
                    if self.dname == None:
                        sDbName = "cesar_{}_{}.{}".format(modelname, obj_id, self.dtype)
                    else:
                        sDbName = self.dname
                    sContentType = ""
                    if self.dtype == "csv":
                        sContentType = "text/tab-separated-values"
                    elif self.dtype == "json":
                        sContentType = "application/json"
                    elif self.dtype == "xlsx":
                        sContentType = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    elif self.dtype == "tar.gz":
                        sContentType = "application/gzip"
                    elif self.dtype == "htmldoc":
                        sContentType = "application/msword"

                    # Excel needs additional conversion
                    if self.dtype == "xlsx":
                        # Convert 'compressed_content' to an Excel worksheet
                        response = HttpResponse(content_type=sContentType)
                        response['Content-Disposition'] = 'attachment; filename="{}"'.format(sDbName)    
                        response = csv_to_excel(sData, response, self.delimiter)
                    else:
                        response = HttpResponse(sData, content_type=sContentType)
                        response['Content-Disposition'] = 'attachment; filename="{}"'.format(sDbName)    

                    # Continue for all formats
                        
                    # return gzip_middleware.process_response(request, response)
                    return response
            elif self.newRedirect and self.redirectpage != "":
                # Redirect to this page
                return redirect(self.redirectpage)
            else:
                # Set any possible typeaheads
                context['typeaheads'] = json.dumps(self.lst_typeahead)
                # This takes self.template_name...
                response = self.render_to_response(context)
        else:
            data['html'] = "(No authorization)"
            data['status'] = "error"
            response = JsonResponse(data)

        # Return the response
        return response

    def get_data(self, prefix, dtype, response=None):
        return ""

    def initializations(self, request, pk):
        # Store the previous page
        # self.previous = get_previous_page(request)

        self.lst_typeahead = []

        # Copy any pk
        self.pk = pk
        self.add = pk is None
        # Get the parameters
        if request.POST:
            self.qd = request.POST
        else:
            self.qd = request.GET

        # Check for action
        if 'action' in self.qd:
            self.action = self.qd['action']

        # Find out what the Main Model instance is, if any
        if self.add:
            self.object = None
        else:
            # Get the instance of the Main Model object
            # NOTE: if the object doesn't exist, we will NOT get an error here
            self.object = self.get_object()

        # Possibly perform custom initializations
        self.custom_init(self.object)
        
    def custom_init(self, instance):
        pass

    def before_delete(self, instance):
        """Anything that needs doing before deleting [instance] """
        return True, "" 

    def after_new(self, form, instance):
        """Action to be performed after adding a new item"""
        return True, "" 

    def before_save(self, form, instance):
        """Action to be performed after saving an item preliminarily, and before saving completely"""
        return True, "" 

    def after_save(self, form, instance):
        """Actions to be performed after saving"""
        return True, "" 

    def add_to_context(self, context, instance):
        """Add to the existing context"""
        return context

    def process_formset(self, prefix, request, formset):
        return None

    def get_formset_queryset(self, prefix):
        return None

    def get_form_kwargs(self, prefix):
        return None

    def get_history(self, instance):
        """Get the history of this element"""
        return ""

    def get_app_access(self, context):
        return True

    def get_context_data(self, **kwargs):
        # Get the current context
        context = super(BasicDetails, self).get_context_data(**kwargs)

        # Check this user: is he allowed to UPLOAD data?
        context['authenticated'] = user_is_authenticated(self.request)
        context['is_app_uploader'] = user_is_ingroup(self.request, app_uploader)
        context['is_app_editor'] = user_is_ingroup(self.request, app_editor)
        context['is_app_userplus'] = user_is_ingroup(self.request, app_userplus)

        self.get_app_access(context)

        # context['prevpage'] = get_previous_page(self.request) # self.previous
        context['afternewurl'] = ""

        context['topleftbuttons'] = ""
        context['history_button'] = self.history_button

        if context['authenticated']:
            self.permission = "read"
            if context['is_app_editor']:
                self.permission = "write"
        context['permission'] = self.permission
        context['extend_template'] = self.extend_template

        if self.has_select2:
            context['has_select2'] = True

        # Possibly define where a listview is
        classname = self.model._meta.model_name
        if self.basic_name == None or self.basic_name == "":
            if self.basic_name_prefix == "":
                self.basic_name = classname
            else:
                self.basic_name = "{}{}".format(self.basic_name_prefix, self.prefix)
        basic_name = self.basic_name
        if self.listview != None:
            context['listview'] = self.listview
        else:
            listviewname = "{}_list".format(basic_name)
            try:
                context['listview'] = reverse(listviewname)
            except:
                context['listview'] = reverse('home')

        if self.basic_add:
            basic_add = reverse(self.basic_add)
        else:
            basic_add = reverse("{}_details".format(basic_name))
        context['basic_add'] = basic_add

        context['new_button'] = self.new_button
        context['add_text'] = self.add_text

        if self.is_basic:
            context['afterdelurl'] = context['listview']

        # Get the parameters passed on with the GET or the POST request
        get = self.request.GET if self.request.method == "GET" else self.request.POST
        initial = get.copy()
        self.qd = initial

        # Possibly first get params
        params = ""
        if "params" in dict(self.qd):
            params = base64_decode( "".join(self.qd.pop("params")))
        context['params'] = params

        # Now see if anything is left
        self.bHasFormInfo = (len(self.qd) > 0)

        # Set the title of the application
        context['title'] = self.title

        # Get the instance
        instance = self.object

        # Prepare form
        frm = self.prepare_form(instance, context, initial)

        if frm:

            if instance == None:
                instance = frm.instance
                self.object = instance

            # Walk all the formset objects
            bFormsetChanged = False
            for formsetObj in self.formset_objects:
                formsetClass = formsetObj['formsetClass']
                prefix  = formsetObj['prefix']
                formset = None
                form_kwargs = self.get_form_kwargs(prefix)
                if 'noinit' in formsetObj and formsetObj['noinit'] and not self.add:
                    # Only process actual changes!!
                    if self.request.method == "POST" and self.request.POST and not self.do_not_save:

                        #if self.add:
                        #    # Saving a NEW item
                        #    if 'initial' in formsetObj:
                        #        formset = formsetClass(self.request.POST, self.request.FILES, prefix=prefix, initial=formsetObj['initial'], form_kwargs = form_kwargs)
                        #    else:
                        #        formset = formsetClass(self.request.POST, self.request.FILES, prefix=prefix, form_kwargs = form_kwargs)
                        #else:
                        #    # Get a formset including any stuff from POST
                        #    formset = formsetClass(self.request.POST, prefix=prefix, instance=instance)

                        # Get a formset including any stuff from POST
                        formset = formsetClass(self.request.POST, prefix=prefix, instance=instance)
                        # Process this formset
                        self.process_formset(prefix, self.request, formset)
                        
                        # Process all the correct forms in the formset
                        for subform in formset:
                            if subform.is_valid():
                                # DO the actual saving
                                subform.save()

                                # Log the SAVE action
                                details = {'id': instance.id}
                                details["savetype"] = "add" # if bNew else "change"
                                details['model'] = subform.instance.__class__.__name__
                                if subform.changed_data != None and len(subform.changed_data) > 0:
                                    details['changes'] = action_model_changes(subform, subform.instance)
                                self.action_add(instance, details, "add")

                                # Signal that the *FORM* needs refreshing, because the formset changed
                                bFormsetChanged = True

                        if formset.is_valid():
                            # Load an explicitly empty formset
                            formset = formsetClass(initial=[], prefix=prefix, form_kwargs=form_kwargs)
                        else:
                            # Retain the original formset, that now contains the error specifications per form
                            # But: do *NOT* add an additional form to it
                            pass

                    else:
                        # All other cases: Load an explicitly empty formset
                        formset = formsetClass(initial=[], prefix=prefix, form_kwargs=form_kwargs)
                else:
                    # show the data belonging to the current [obj]
                    qs = self.get_formset_queryset(prefix)
                    if qs == None:
                        formset = formsetClass(prefix=prefix, instance=instance, form_kwargs=form_kwargs)
                    else:
                        formset = formsetClass(prefix=prefix, instance=instance, queryset=qs, form_kwargs=form_kwargs)
                # Process all the forms in the formset
                ordered_forms = self.process_formset(prefix, self.request, formset)
                if ordered_forms:
                    context[prefix + "_ordered"] = ordered_forms
                # Store the instance
                formsetObj['formsetinstance'] = formset
                # Add the formset to the context
                context[prefix + "_formset"] = formset
                # Get any possible typeahead parameters
                lst_formset_ta = getattr(formset.form, "typeaheads", None)
                if lst_formset_ta != None:
                    for item in lst_formset_ta:
                        self.lst_typeahead.append(item)

            # Essential formset information
            for formsetObj in self.formset_objects:
                prefix = formsetObj['prefix']
                if 'fields' in formsetObj: context["{}_fields".format(prefix)] = formsetObj['fields']
                if 'linkfield' in formsetObj: context["{}_linkfield".format(prefix)] = formsetObj['linkfield']

            # Check if the formset made any changes to the form
            if bFormsetChanged:
                # OLD: 
                frm = self.prepare_form(instance, context)

            # Put the form and the formset in the context
            context['{}Form'.format(self.prefix)] = frm
            context['basic_form'] = frm
            context['instance'] = instance
            context['options'] = json.dumps({"isnew": (instance == None)})

            # Possibly define the admin detailsview
            if instance:
                admindetails = "admin:seeker_{}_change".format(classname)
                try:
                    context['admindetails'] = reverse(admindetails, args=[instance.id])
                except:
                    pass
            context['modelname'] = self.model._meta.object_name
            context['titlesg'] = self.titlesg if self.titlesg else self.title if self.title != "" else basic_name.capitalize()

            # Make sure we have a url for editing
            if instance and instance.id:
                # There is a details and edit url
                context['editview'] = reverse("{}_edit".format(basic_name), kwargs={'pk': instance.id})
                context['detailsview'] = reverse("{}_details".format(basic_name), kwargs={'pk': instance.id})
            # Make sure we have an url for new
            context['addview'] = reverse("{}_details".format(basic_name))

        # Determine breadcrumbs and previous page
        if self.is_basic:
            title = self.title if self.title != "" else basic_name
            if self.rtype == "json":
                # This is the EditView
                context['breadcrumbs'] = get_breadcrumbs(self.request, "{} edit".format(title), False)
                prevpage = reverse('home')
                context['prevpage'] = prevpage
            else:
                # This is DetailsView
                # Process this visit and get the new breadcrumbs object
                prevpage = context['listview']
                context['prevpage'] = prevpage
                crumbs = []
                if self.listviewtitle == None:
                    crumbs.append([title + " list", prevpage])
                else:
                    crumbs.append([ self.listviewtitle, prevpage])
                current_name = title if instance else "{} (new)".format(title)
                context['breadcrumbs'] = get_breadcrumbs(self.request, current_name, True, crumbs)

        # Possibly add to context by the calling function
        if instance.id:
            context = self.add_to_context(context, instance)
            # Double check access
            if context['is_app_editor']:
                self.permission = "write"
                context['permission'] = self.permission
            if self.history_button:
                # Retrieve history
                context['history_contents'] = self.get_history(instance)


        # fill in the form values
        if frm and 'mainitems' in context:
            for mobj in context['mainitems']:
                # Check for possible form field information
                if 'field_key' in mobj: 
                    mobj['field_abbr'] = "{}-{}".format(frm.prefix, mobj['field_key'])
                    mobj['field_key'] = frm[mobj['field_key']]
                if 'field_view' in mobj: mobj['field_view'] = frm[mobj['field_view']]
                if 'field_ta' in mobj: mobj['field_ta'] = frm[mobj['field_ta']]
                if 'field_list' in mobj: mobj['field_list'] = frm[mobj['field_list']]

                # Calculate view-mode versus any-mode
                #  'field_key' in mainitem or 'field_list' in mainitem and permission == "write"  or  is_app_userplus and mainitem.maywrite
                if self.permission == "write":       # or app_userplus and 'maywrite' in mobj and mobj['maywrite']:
                    mobj['allowing'] = "edit"
                else:
                    mobj['allowing'] = "view"
                if ('field_key' in mobj or 'field_list' in mobj) and (mobj['allowing'] == "edit"):
                    mobj['allowing_key_list'] = "edit"
                else:
                    mobj['allowing_key_list'] = "view"

        # Define where to go to after deletion
        if 'afterdelurl' not in context or context['afterdelurl'] == "":
            context['afterdelurl'] = get_previous_page(self.request)

        # Return the calculated context
        return context

    def action_add(self, instance, actiontype, details):
        """User can fill this in to his/her liking"""

        # Example: 
        #   Action.add(self.request.user.username, instance.__class__.__name__, "delete", json.dumps(details))
        pass

    def prepare_form(self, instance, context, initial=[]):
        # Initialisations
        bNew = False
        mForm = self.mForm
        frm = None
        method = "plain"    # "username" would be the other
        oErr = ErrHandle()
        username=self.request.user.username
        team_group=app_editor
        userplus = app_userplus

        try:
            # If there is no mForm, leave
            if mForm is None:
                return None

            # Determine the prefix
            if self.prefix_type == "":
                id = "n" if instance == None else instance.id
                prefix = "{}-{}".format(self.prefix, id)
            else:
                prefix = self.prefix

            # Check if this is a POST or a GET request
            if self.request.method == "POST" and not self.do_not_save:
                # Determine what the action is (if specified)
                action = ""
                if 'action' in initial: action = initial['action']
                if action == "delete":
                    # The user wants to delete this item
                    try:
                        bResult, msg = self.before_delete(instance)
                        if bResult:
                            # Log the DELETE action
                            details = {'id': instance.id}
                            self.action_add(instance, details, "delete")
                        
                            # Remove this sermongold instance
                            instance.delete()
                        else:
                            # Removing is not possible
                            context['errors'] = {'delete': msg }
                    except:
                        msg = oErr.get_error_message()
                        # Create an errors object
                        context['errors'] = {'delete':  msg }

                    if 'afterdelurl' not in context or context['afterdelurl'] == "":
                        context['afterdelurl'] = get_previous_page(self.request, True)

                    # Make sure we are returning JSON
                    self.rtype = "json"

                    # Possibly add to context by the calling function
                    if instance.id:
                        context = self.add_to_context(context, instance)

                    # No need to retern a form anymore - we have been deleting
                    return None
            
                # All other actions just mean: edit or new and send back
                # Make instance available
                context['object'] = instance
                self.object = instance

                # Do we have an existing object or are we creating?
                if instance == None:
                    # Saving a new item
                    if self.use_team_group:
                        frm = mForm(initial, prefix=prefix, username=username, team_group=team_group, userplus=userplus)
                    else:
                        frm = mForm(initial, prefix=prefix)
                    bNew = True
                    self.add = True
                elif len(initial) == 0:
                    # Create a completely new form, on the basis of the [instance] only
                    if self.use_team_group:
                        frm = mForm(prefix=prefix, instance=instance, username=username, team_group=team_group, userplus=userplus)
                    else:
                        frm = mForm(prefix=prefix, instance=instance)
                else:
                    # Editing an existing one
                    if self.use_team_group:
                        frm = mForm(initial, self.request.FILES, prefix=prefix, instance=instance, username=username, team_group=team_group, userplus=userplus)
                    else:
                        frm = mForm(initial, self.request.FILES, prefix=prefix, instance=instance)
                # Both cases: validation and saving
                if frm.is_valid():
                    # The form is valid - do a preliminary saving
                    obj = frm.save(commit=False)
                    # Any checks go here...
                    bResult, msg = self.before_save(form=frm, instance=obj)
                    if bResult:
                        # Now save it for real
                        if method == "username":
                            kwargs={'username': self.request.user.username}
                            obj.save(**kwargs)
                        else:
                            obj.save()
                        # Log the SAVE action
                        details = {'id': obj.id}
                        details["savetype"] = "new" if bNew else "change"
                        if frm.changed_data != None and len(frm.changed_data) > 0:
                            details['changes'] = action_model_changes(frm, obj)
                        self.action_add(obj, details, "save")

                        # Make sure the form is actually saved completely
                        frm.save()
                        instance = obj
                    
                        # Any action(s) after saving
                        bResult, msg = self.after_save(frm, obj)
                    else:
                        context['errors'] = {'save': msg }
                else:
                    # We need to pass on to the user that there are errors
                    context['errors'] = frm.errors

                # Check if this is a new one
                if bNew:
                    if self.is_basic:
                        self.afternewurl = context['listview']
                        if self.rtype == "html":
                            # Make sure we do a page redirect
                            self.newRedirect = True
                            self.redirectpage = reverse("{}_details".format(self.basic_name), kwargs={'pk': instance.id})
                    # Any code that should be added when creating a new [SermonGold] instance
                    bResult, msg = self.after_new(frm, instance)
                    if not bResult:
                        # Removing is not possible
                        context['errors'] = {'new': msg }
                    # Check if an 'afternewurl' is specified
                    if self.afternewurl != "":
                        context['afternewurl'] = self.afternewurl
                
            else:
                # Check if this is asking for a new form
                if instance == None:
                    # Get the form for the sermon
                    if self.use_team_group:
                        frm = mForm(prefix=prefix, username=username, team_group=team_group, userplus=userplus)
                    else:
                        frm = mForm(prefix=prefix)
                else:
                    # Get the form for the sermon
                    if self.use_team_group:
                        frm = mForm(instance=instance, prefix=prefix, username=username, team_group=team_group, userplus=userplus)
                    else:
                        frm = mForm(instance=instance, prefix=prefix)
                if frm.is_valid():
                    iOkay = 1
                # Walk all the form objects
                for formObj in self.form_objects:
                    formClass = formObj['form']
                    prefix = formObj['prefix']
                    # This is only for *NEW* forms (right now)
                    form = formClass(prefix=prefix)
                    context[prefix + "Form"] = form
                    # Get any possible typeahead parameters
                    lst_form_ta = getattr(formObj['forminstance'], "typeaheads", None)
                    if lst_form_ta != None:
                        for item in lst_form_ta:
                            self.lst_typeahead.append(item)

            # Get any possible typeahead parameters
            if frm != None:
                lst_form_ta = getattr(frm, "typeaheads", None)
                if lst_form_ta != None:
                    for item in lst_form_ta:
                        self.lst_typeahead.append(item)
        except:
            msg = oErr.get_error_message()
            oErr.DoError("BasicDetails/prepare_form")
        # Return the form we made
        return frm


class BasicPart(View):
    """This is my own versatile handling view.

    Note: this version works with <pk> and not with <object_id>
    """

    # Initialisations
    arErr = []              # errors   
    template_name = None    # The template to be used
    template_err_view = None
    form_validated = True   # Used for POST form validation
    savedate = None         # When saving information, the savedate is returned in the context
    add = False             # Are we adding a new record or editing an existing one?
    obj = None              # The instance of the MainModel
    action = ""             # The action to be undertaken
    delimiter = ","         # for excel/csv downloading
    dtype = ""
    MainModel = None        # The model that is mainly used for this form
    form_objects = []       # List of forms to be processed
    formset_objects = []    # List of formsets to be processed
    previous = None         # Return to this
    bDebug = False          # Debugging information
    redirectpage = ""       # Where to redirect to
    data = {'status': 'ok', 'html': ''}       # Create data to be returned    
    
    def post(self, request, pk=None):
        # A POST request means we are trying to SAVE something
        self.initializations(request, pk)
        # Initialize typeahead list
        lst_typeahead = []

        # Explicitly set the status to OK
        self.data['status'] = "ok"
        
        if self.checkAuthentication(request):
            # Build the context
            context = dict(object_id = pk, savedate=None)
            context['authenticated'] = user_is_authenticated(request)
            context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
            context['is_app_editor'] = user_is_ingroup(request, app_editor)

            # Action depends on 'action' value
            if self.action == "":
                if self.bDebug: self.oErr.Status("ResearchPart: action=(empty)")
                # Walk all the forms for preparation of the formObj contents
                for formObj in self.form_objects:
                    # Are we SAVING a NEW item?
                    if self.add:
                        # We are saving a NEW item
                        formObj['forminstance'] = formObj['form'](request.POST, prefix=formObj['prefix'])
                        formObj['action'] = "new"
                    else:
                        # We are saving an EXISTING item
                        # Determine the instance to be passed on
                        instance = self.get_instance(formObj['prefix'])
                        # Make the instance available in the form-object
                        formObj['instance'] = instance
                        # Get an instance of the form
                        formObj['forminstance'] = formObj['form'](request.POST, prefix=formObj['prefix'], instance=instance)
                        formObj['action'] = "change"

                # Initially we are assuming this just is a review
                context['savedate']="reviewed at {}".format(get_current_datetime().strftime("%X"))

                # Iterate again
                for formObj in self.form_objects:
                    prefix = formObj['prefix']
                    # Adapt if it is not readonly
                    if not formObj['readonly']:
                        # Check validity of form
                        if formObj['forminstance'].is_valid() and self.is_custom_valid(prefix, formObj['forminstance']):
                            # Save it preliminarily
                            instance = formObj['forminstance'].save(commit=False)
                            # The instance must be made available (even though it is only 'preliminary')
                            formObj['instance'] = instance
                            # Perform actions to this form BEFORE FINAL saving
                            bNeedSaving = formObj['forminstance'].has_changed()
                            if self.before_save(prefix, request, instance=instance, form=formObj['forminstance']): bNeedSaving = True
                            if formObj['forminstance'].instance.id == None: bNeedSaving = True
                            if bNeedSaving:
                                # Perform the saving
                                instance.save()
                                # Log the SAVE action
                                details = {'id': instance.id}
                                if formObj['forminstance'].changed_data != None:
                                    details['changes'] = action_model_changes(formObj['forminstance'], instance)
                                if 'action' in formObj: details['savetype'] = formObj['action']
                                Action.add(request.user.username, self.MainModel.__name__, instance.id, "save", json.dumps(details))
                                # Set the context
                                context['savedate']="saved at {}".format(get_current_datetime().strftime("%X"))
                                # Put the instance in the form object
                                formObj['instance'] = instance
                                # Store the instance id in the data
                                self.data[prefix + '_instanceid'] = instance.id
                                # Any action after saving this form
                                self.after_save(prefix, instance=instance, form=formObj['forminstance'])
                            # Also get the cleaned data from the form
                            formObj['cleaned_data'] = formObj['forminstance'].cleaned_data
                        else:
                            self.arErr.append(formObj['forminstance'].errors)
                            self.form_validated = False
                            formObj['cleaned_data'] = None
                    else:
                        # Form is readonly

                        # Check validity of form
                        if formObj['forminstance'].is_valid() and self.is_custom_valid(prefix, formObj['forminstance']):
                            # At least get the cleaned data from the form
                            formObj['cleaned_data'] = formObj['forminstance'].cleaned_data

                            # x = json.dumps(sorted(self.qd.items(), key=lambda kv: kv[0]), indent=2)
                    # Add instance to the context object
                    context[prefix + "Form"] = formObj['forminstance']
                    # Get any possible typeahead parameters
                    lst_form_ta = getattr(formObj['forminstance'], "typeaheads", None)
                    if lst_form_ta != None:
                        for item in lst_form_ta:
                            lst_typeahead.append(item)
                # Walk all the formset objects
                for formsetObj in self.formset_objects:
                    prefix  = formsetObj['prefix']
                    if self.can_process_formset(prefix):
                        formsetClass = formsetObj['formsetClass']
                        form_kwargs = self.get_form_kwargs(prefix)
                        if self.add:
                            # Saving a NEW item
                            if 'initial' in formsetObj:
                                formset = formsetClass(request.POST, request.FILES, prefix=prefix, initial=formsetObj['initial'], form_kwargs = form_kwargs)
                            else:
                                formset = formsetClass(request.POST, request.FILES, prefix=prefix, form_kwargs = form_kwargs)
                        else:
                            # Saving an EXISTING item
                            instance = self.get_instance(prefix)
                            qs = self.get_queryset(prefix)
                            if qs == None:
                                formset = formsetClass(request.POST, request.FILES, prefix=prefix, instance=instance, form_kwargs = form_kwargs)
                            else:
                                formset = formsetClass(request.POST, request.FILES, prefix=prefix, instance=instance, queryset=qs, form_kwargs = form_kwargs)
                        # Process all the forms in the formset
                        self.process_formset(prefix, request, formset)
                        # Store the instance
                        formsetObj['formsetinstance'] = formset
                        # Make sure we know what we are dealing with
                        itemtype = "form_{}".format(prefix)
                        # Adapt the formset contents only, when it is NOT READONLY
                        if not formsetObj['readonly']:
                            # Is the formset valid?
                            if formset.is_valid():
                                # Possibly handle the clean() for this formset
                                if 'clean' in formsetObj:
                                    # Call the clean function
                                    self.clean(formset, prefix)
                                has_deletions = False
                                if len(self.arErr) == 0:
                                    # Make sure all changes are saved in one database-go
                                    with transaction.atomic():
                                        # Walk all the forms in the formset
                                        for form in formset:
                                            # At least check for validity
                                            if form.is_valid() and self.is_custom_valid(prefix, form):
                                                # Should we delete?
                                                if 'DELETE' in form.cleaned_data and form.cleaned_data['DELETE']:
                                                    # Check if deletion should be done
                                                    if self.before_delete(prefix, form.instance):
                                                        # Log the delete action
                                                        details = {'id': form.instance.id}
                                                        Action.add(request.user.username, itemtype, form.instance.id, "delete", json.dumps(details))
                                                        # Delete this one
                                                        form.instance.delete()
                                                        # NOTE: the template knows this one is deleted by looking at form.DELETE
                                                        has_deletions = True
                                                else:
                                                    # Check if anything has changed so far
                                                    has_changed = form.has_changed()
                                                    # Save it preliminarily
                                                    sub_instance = form.save(commit=False)
                                                    # Any actions before saving
                                                    if self.before_save(prefix, request, sub_instance, form):
                                                        has_changed = True
                                                    # Save this construction
                                                    if has_changed and len(self.arErr) == 0: 
                                                        # Save the instance
                                                        sub_instance.save()
                                                        # Adapt the last save time
                                                        context['savedate']="saved at {}".format(get_current_datetime().strftime("%X"))
                                                        # Log the delete action
                                                        details = {'id': sub_instance.id}
                                                        if form.changed_data != None:
                                                            details['changes'] = action_model_changes(form, sub_instance)
                                                        Action.add(request.user.username, itemtype,sub_instance.id, "save", json.dumps(details))
                                                        # Store the instance id in the data
                                                        self.data[prefix + '_instanceid'] = sub_instance.id
                                                        # Any action after saving this form
                                                        self.after_save(prefix, sub_instance)
                                            else:
                                                if len(form.errors) > 0:
                                                    self.arErr.append(form.errors)
                                
                                    # Rebuild the formset if it contains deleted forms
                                    if has_deletions or not has_deletions:
                                        # Or: ALWAYS
                                        if qs == None:
                                            formset = formsetClass(prefix=prefix, instance=instance, form_kwargs=form_kwargs)
                                        else:
                                            formset = formsetClass(prefix=prefix, instance=instance, queryset=qs, form_kwargs=form_kwargs)
                                        formsetObj['formsetinstance'] = formset
                            else:
                                # Iterate over all errors
                                for idx, err_this in enumerate(formset.errors):
                                    if '__all__' in err_this:
                                        self.arErr.append(err_this['__all__'][0])
                                    elif err_this != {}:
                                        # There is an error in item # [idx+1], field 
                                        problem = err_this 
                                        for k,v in err_this.items():
                                            fieldName = k
                                            errmsg = "Item #{} has an error at field [{}]: {}".format(idx+1, k, v[0])
                                            self.arErr.append(errmsg)

                            # self.arErr.append(formset.errors)
                    else:
                        formset = []
                    # Get any possible typeahead parameters
                    lst_formset_ta = getattr(formset.form, "typeaheads", None)
                    if lst_formset_ta != None:
                        for item in lst_formset_ta:
                            lst_typeahead.append(item)
                    # Add the formset to the context
                    context[prefix + "_formset"] = formset
            elif self.action == "download":
                # We are being asked to download something
                if self.dtype != "":
                    # Initialise return status
                    oBack = {'status': 'ok'}
                    sType = "csv" if (self.dtype == "xlsx") else self.dtype

                    # Get the data
                    sData = self.get_data('', self.dtype)
                    # Decode the data and compress it using gzip
                    bUtf8 = (self.dtype != "db")
                    bUsePlain = (self.dtype == "xlsx" or self.dtype == "csv")

                    # Create name for download
                    modelname = self.MainModel.__name__
                    obj_id = "n" if self.obj == None else self.obj.id
                    sDbName = "cesar_{}_{}.{}".format(modelname, obj_id, self.dtype)
                    sContentType = ""
                    if self.dtype == "csv":
                        sContentType = "text/tab-separated-values"
                    elif self.dtype == "json":
                        sContentType = "application/json"
                    elif self.dtype == "xlsx":
                        sContentType = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    elif self.dtype == "tar.gz":
                        sContentType = "application/gzip"

                    # Excel needs additional conversion
                    if self.dtype == "xlsx":
                        # Convert 'compressed_content' to an Excel worksheet
                        response = HttpResponse(content_type=sContentType)
                        response['Content-Disposition'] = 'attachment; filename="{}"'.format(sDbName)    
                        response = csv_to_excel(sData, response, self.delimiter)
                    else:
                        response = HttpResponse(sData, content_type=sContentType)
                        response['Content-Disposition'] = 'attachment; filename="{}"'.format(sDbName)    

                    # Continue for all formats
                        
                    # return gzip_middleware.process_response(request, response)
                    return response
            elif self.action == "delete":
                # The user requests this to be deleted
                if self.before_delete():
                    # Log the delete action
                    details = {'id': self.obj.id}
                    Action.add(request.user.username, self.MainModel.__name__, self.obj.id, "delete", json.dumps(details))
                    # We have permission to delete the instance
                    self.obj.delete()
                    context['deleted'] = True

            # Allow user to add to the context
            context = self.add_to_context(context)
            
            # Possibly add data from [context.data]
            if 'data' in context:
                for k,v in context['data'].items():
                    self.data[k] = v

            # First look at redirect page
            self.data['redirecturl'] = ""
            if self.redirectpage != "":
                self.data['redirecturl'] = self.redirectpage
            # Check if 'afternewurl' needs adding
            # NOTE: this should only be used after a *NEW* instance has been made -hence the self.add check
            if 'afternewurl' in context and self.add:
                self.data['afternewurl'] = context['afternewurl']
            else:
                self.data['afternewurl'] = ""
            if 'afterdelurl' in context:
                self.data['afterdelurl'] = context['afterdelurl']

            # Make sure we have a list of any errors
            error_list = [str(item) for item in self.arErr]
            context['error_list'] = error_list
            context['errors'] = json.dumps( self.arErr)
            if len(self.arErr) > 0:
                # Indicate that we have errors
                self.data['has_errors'] = True
                self.data['status'] = "error"
            else:
                self.data['has_errors'] = False
            # Standard: add request user to context
            context['requestuser'] = request.user

            # Set any possible typeaheads
            self.data['typeaheads'] = lst_typeahead

            # Get the HTML response
            if len(self.arErr) > 0:
                if self.template_err_view != None:
                     # Create a list of errors
                    self.data['err_view'] = render_to_string(self.template_err_view, context, request)
                else:
                    self.data['error_list'] = error_list
                    self.data['errors'] = self.arErr
                self.data['html'] = ''
                # We may not redirect if there is an error!
                self.data['redirecturl'] = ''
            elif self.action == "delete":
                self.data['html'] = "deleted" 
            elif self.template_name != None:
                # In this case reset the errors - they should be shown within the template
                sHtml = render_to_string(self.template_name, context, request)
                sHtml = treat_bom(sHtml)
                self.data['html'] = sHtml
            else:
                self.data['html'] = "no template_name specified" 

            # At any rate: empty the error basket
            self.arErr = []
            error_list = []

        else:
            self.data['html'] = "Please log in before continuing"

        # Return the information
        return JsonResponse(self.data)
        
    def get(self, request, pk=None): 
        self.data['status'] = 'ok'
        # Perform the initializations that need to be made anyway
        self.initializations(request, pk)
        # Initialize typeahead list
        lst_typeahead = []

        # Continue if authorized
        if self.checkAuthentication(request):
            context = dict(object_id = pk, savedate=None)
            context['prevpage'] = self.previous
            context['authenticated'] = user_is_authenticated(request)
            context['is_app_uploader'] = user_is_ingroup(request, app_uploader)
            context['is_app_editor'] = user_is_ingroup(request, app_editor)
            # Walk all the form objects
            for formObj in self.form_objects:        
                # Used to populate a NEW research project
                # - CREATE a NEW research form, populating it with any initial data in the request
                initial = dict(request.GET.items())
                if self.add:
                    # Create a new form
                    formObj['forminstance'] = formObj['form'](initial=initial, prefix=formObj['prefix'])
                else:
                    # Used to show EXISTING information
                    instance = self.get_instance(formObj['prefix'])
                    # We should show the data belonging to the current Research [obj]
                    formObj['forminstance'] = formObj['form'](instance=instance, prefix=formObj['prefix'])
                # Add instance to the context object
                context[formObj['prefix'] + "Form"] = formObj['forminstance']
                # Get any possible typeahead parameters
                lst_form_ta = getattr(formObj['forminstance'], "typeaheads", None)
                if lst_form_ta != None:
                    for item in lst_form_ta:
                        lst_typeahead.append(item)
            # Walk all the formset objects
            for formsetObj in self.formset_objects:
                formsetClass = formsetObj['formsetClass']
                prefix  = formsetObj['prefix']
                form_kwargs = self.get_form_kwargs(prefix)
                if self.add:
                    # - CREATE a NEW formset, populating it with any initial data in the request
                    initial = dict(request.GET.items())
                    # Saving a NEW item
                    formset = formsetClass(initial=initial, prefix=prefix, form_kwargs=form_kwargs)
                else:
                    # Possibly initial (default) values
                    if 'initial' in formsetObj:
                        initial = formsetObj['initial']
                    else:
                        initial = None
                    # show the data belonging to the current [obj]
                    instance = self.get_instance(prefix)
                    qs = self.get_queryset(prefix)
                    if qs == None:
                        formset = formsetClass(prefix=prefix, instance=instance, form_kwargs=form_kwargs)
                    else:
                        formset = formsetClass(prefix=prefix, instance=instance, queryset=qs, initial=initial, form_kwargs=form_kwargs)
                # Get any possible typeahead parameters
                lst_formset_ta = getattr(formset.form, "typeaheads", None)
                if lst_formset_ta != None:
                    for item in lst_formset_ta:
                        lst_typeahead.append(item)
                # Process all the forms in the formset
                ordered_forms = self.process_formset(prefix, request, formset)
                if ordered_forms:
                    context[prefix + "_ordered"] = ordered_forms
                # Store the instance
                formsetObj['formsetinstance'] = formset
                # Add the formset to the context
                context[prefix + "_formset"] = formset
            # Allow user to add to the context
            context = self.add_to_context(context)
            # Make sure we have a list of any errors
            error_list = [str(item) for item in self.arErr]
            context['error_list'] = error_list
            context['errors'] = self.arErr
            # Standard: add request user to context
            context['requestuser'] = request.user

            # Set any possible typeaheads
            self.data['typeaheads'] = json.dumps(lst_typeahead)
            
            # Get the HTML response
            sHtml = render_to_string(self.template_name, context, request)
            sHtml = treat_bom(sHtml)
            self.data['html'] = sHtml
        else:
            self.data['html'] = "Please log in before continuing"

        # Return the information
        return JsonResponse(self.data)
      
    def checkAuthentication(self,request):
        # first check for authentication
        if not request.user.is_authenticated:
            # Simply redirect to the home page
            self.data['html'] = "Please log in to work on this project"
            return False
        else:
            return True

    def rebuild_formset(self, prefix, formset):
        return formset

    def initializations(self, request, object_id):
        # Store the previous page
        #self.previous = get_previous_page(request)
        # Clear errors
        self.arErr = []
        # COpy the request
        self.request = request
        # Copy any object id
        self.object_id = object_id
        self.add = object_id is None
        # Get the parameters
        if request.POST:
            self.qd = request.POST
        else:
            self.qd = request.GET

        # Check for action
        if 'action' in self.qd:
            self.action = self.qd['action']

        # Find out what the Main Model instance is, if any
        if self.add:
            self.obj = None
        else:
            # Get the instance of the Main Model object
            self.obj =  self.MainModel.objects.filter(pk=object_id).first()
            # NOTE: if the object doesn't exist, we will NOT get an error here
        # ALWAYS: perform some custom initialisations
        self.custom_init()

    def get_instance(self, prefix):
        return self.obj

    def is_custom_valid(self, prefix, form):
        return True

    def get_queryset(self, prefix):
        return None

    def get_form_kwargs(self, prefix):
        return None

    def get_data(self, prefix, dtype):
        return ""

    def before_save(self, prefix, request, instance=None, form=None):
        return False

    def before_delete(self, prefix=None, instance=None):
        return True

    def after_save(self, prefix, instance=None, form=None):
        return True

    def add_to_context(self, context):
        return context

    def process_formset(self, prefix, request, formset):
        return None

    def can_process_formset(self, prefix):
        return True

    def custom_init(self):
        pass    

    

