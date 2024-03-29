"""
Definition of views for the SEEKER app.
"""

from django.core.exceptions import FieldDoesNotExist
from django.core.paginator import Paginator
from django.db.models import Q
from django.db.models.functions import Lower
from django.forms import formset_factory
from django.forms import inlineformset_factory, BaseInlineFormSet, modelformset_factory
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.views.decorators.csrf import csrf_exempt
from django.views.generic.detail import DetailView
from django.views.generic.list import MultipleObjectMixin
from django.views.generic.edit import CreateView, DeleteView
from django.views.generic.base import RedirectView
from django.views.generic import ListView, View
from django import template
from functools import reduce
from operator import __or__ as OR
import zipfile
import csv
import base64

import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook
from io import StringIO

# My own application
# from basic.views import BasicList
from cesar.basic.views import BasicList

from cesar.seeker.forms import *
from cesar.seeker.models import *
from cesar.seeker.models import SIMPLENAME
from cesar.seeker.convert import decompressSafe, get_crpp_date
from cesar.browser.models import Part, Corpus, Sentence
from cesar.browser.views import get_item_list, adapt_search, user_is_ingroup
from cesar.browser.services import get_crpp_sent_info
from cesar.seeker.services import crpp_dbget
from cesar.settings import APP_PREFIX

paginateEntries = 20

# DEFINE a 'related' formset
# RelatedFormset = formset_factory(RelatedForm, formset=BaseRelatedFormset, can_delete=True, extra=0, min_num=0)
RelatedFormset = formset_factory(RelatedForm, can_delete=True, extra=0, min_num=0)


def check_arguments(arg_formset, functiondef, qs_gvar, qs_cvar, qs_dvar, target):

    oErr = ErrHandle()
    # Take the functiondef as available in this argument
    arg_defs = ArgumentDef.objects.filter(function=functiondef)

    for index, arg_form in enumerate(arg_formset):
        # Access the argument definition of this one
        argdef_this = arg_defs[index]
        # Check out the obltype parameter
        obltype = argdef_this.obltype

        # Initialise the querysets
        if obltype != 'str' and obltype != 'int':
            # Global variables can only be strings, and only strings can reduce to integers
            arg_form.fields['gvar'].queryset = GlobalVariable.objects.none()
        else:
            arg_form.fields['gvar'].queryset = qs_gvar

        # Look at construction variables
        if qs_cvar == None:
            arg_form.fields['cvar'].queryset = ConstructionVariable.objects.none()
        else:
            arg_form.fields['cvar'].queryset = qs_cvar

        # Look at data-dependant variables: only accept those that have the correct output type
        arg_form.fields['dvar'].queryset = VarDef.get_restricted_vardef_list( qs_dvar, obltype)

        # If we have an obligatory type, then limit the possibilities for the user
        if obltype != None and obltype != '':
            # Get the list of choices
            choice_list = arg_form.fields['argtype'].choices
            choices = []
            bShowFunctions = False
            for choice in choice_list:
                # Each choice is a tuple -- look at the first part of the tuple
                if keep_argtype(choice[0], obltype, arg_form):
                    choices.append(choice)
                    if choice[0] == "func" or choice[0] == "calc":
                        bShowFunctions = True
            # Take over the choices for argtype
            arg_form.fields['argtype'].choices = choices
            # Are functions going to be included?
            if bShowFunctions:
                # Since we are restricting choice, make sure that only those functions are offered
                #   at this level, that return a matching type
                arg_form.fields['functiondef'].queryset = FunctionDef.get_functions_with_type(obltype)

        # Add the information required by 'seeker/function_args.html' for each argument
        arg_form.target = target
        arg_form.targetid = 'research_part_' + arg_form.target
        if arg_form.instance != None and arg_form.instance.id != None:
            arg_form.url_edit = reverse(arg_form.targetid, kwargs={'object_id': arg_form.instance.id})
        arg_form.url_new = reverse(arg_form.targetid)
        # Get the instance from this form
        try:
            arg = arg_form.save(commit=False)
        except:
            oErr.DoError("check_arguments error")
        # Check if the argument definition is set

        if arg.id == None or arg.argumentdef_id == None:
            # Get the argument definition for this particular argument
            arg.argumentdef = arg_defs[index]
            arg_form.initial['argumentdef'] = arg_defs[index]
            arg_form.initial['argtype'] = arg_defs[index].argtype
            # Preliminarily save
            arg_form.save(commit=False)

        # Initialize the value of argval, if this can be done
        if obltype == 'str' or obltype == 'int' or obltype == 'bool':
            if arg.argval == '[]':
                arg_form.initial['argval'] = argdef_this.argval
                arg = arg_form.save(commit=False)
            else:
                oErr.Status("check_arguments: arg{} val={}".format(index, arg_form.fields['argval']))



    # Return the adatpted formset
    return arg_formset


def keep_argtype(sChoice, sOblType, arg_form):
    """Check if argument of type [sChoice] is okay for [sOblType]"""

    bBack = True
    if sChoice == 'fixed':
        # Can be string or integer
        bBack = (sOblType == 'str' or sOblType == 'int' or sOblType == 'bool')
    elif sChoice == 'gvar':
        # Can be string or integer
        bBack = (sOblType == 'str' or sOblType == 'int')
    elif sChoice == 'cvar':
        # Don't know
        # bBack = (arg_form.fields['cvar'].queryset.count() > 0)
        bBack = False
    elif sChoice == 'func':
        # Don't know
        pass
    elif sChoice == 'cnst':
        # This thing should be prevented anyway
        bBack = False
        # bBack = (sOblType == 'cnst' or sOblType == 'clist')
    elif sChoice == 'raxis':
        bBack = (sOblType == 'raxis')
    elif sChoice == 'hit':
        bBack = (sOblType == 'cnst' or sOblType == 'clist')
    elif sChoice == 'dvar':
        bBack = (arg_form.fields['dvar'].queryset.count() > 0)
    elif sChoice == 'rcond':
        bBack = (sOblType == 'rcond')
    elif sChoice == 'rcnst':
        bBack = (sOblType == 'cnst' or sOblType == 'clist')
    return bBack


def cleanup_functions():
    """Remove superfluous functions"""

    has_changed = False
    # Find all functions that are not pointed to from any of the construction variables
    lstCvarId = [item.id for item in ConstructionVariable.objects.exclude(function=None)]
    lstCondId = [item.id for item in Condition.objects.exclude(function=None)]
    lstFeatId = [item.id for item in Feature.objects.exclude(function=None)]
    lstFunDel = Function.objects.exclude(root__in=lstCvarId).exclude(rootcond__in=lstCondId)
    # Now delete those that need deleting (cascading is done in the function model itself)
    with transaction.atomic():
        for fun_this in lstFunDel:
            fun_this.delete()
            # Make sure that deletions get saved
            has_changed = True
    return has_changed


def treat_bom(sHtml):
    """REmove the BOM marker except at the beginning of the string"""

    # Check if it is in the beginning
    bStartsWithBom = sHtml.startswith(u'\ufeff')
    # Remove everywhere
    sHtml = sHtml.replace(u'\ufeff', '')
    # Return what we have
    return sHtml


def csv_to_excel(sCsvData, response, delimiter=","):
    """Convert CSV data to an Excel worksheet"""

    # Start workbook
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title="Data"

    # Start accessing the string data 
    f = StringIO(sCsvData)
    reader = csv.reader(f, delimiter=delimiter)

    # Read the header cells and make a header row in the worksheet
    headers = next(reader)
    for col_num in range(len(headers)):
        c = ws.cell(row=1, column=col_num+1)
        c.value = headers[col_num]
        c.font = openpyxl.styles.Font(bold=True)
        # Set width to a fixed size
        ws.column_dimensions[get_column_letter(col_num+1)].width = 8.0        

    row_num = 1
    lCsv = []
    for row in reader:
        # Keep track of the EXCEL row we are in
        row_num += 1
        # Walk the elements in the data row
        # oRow = {}
        for idx, cell in enumerate(row):
            c = ws.cell(row=row_num, column=idx+1)
            # attempt to see this as a float
            cell_value = row[idx]
            try:
                cell_value = float(cell_value)
            except ValueError:
                pass
            c.value = cell_value
            c.alignment = openpyxl.styles.Alignment(wrap_text=False)

    # Save the result in the response
    wb.save(response)
    return response


def get_best_format(part, sFormat):
    """Check if the combination part/sFormat is okay, otherwise try another format"""

    format = choice_value(CORPUS_FORMAT, sFormat)
    alt_format = 0 if format == 1 else 1

    # Check if 'part'is a string or not
    if isinstance(part, str):
        part = Part.objects.filter(id=part).first()

    num = Text.objects.filter(part=part, format=format).count()
    alt_num = Text.objects.filter(part=part, format=alt_format).count()
    if num == 0 or (num + 10) < alt_num:
        sFormat = get_format_name(alt_format)
    return sFormat



class CustomInlineFormset(BaseInlineFormSet):
    def clean(self):
        super(CustomInlineFormset, self).clean()


class GatewayDetailView(DetailView):
    model = Gateway
    form_class = GatewayForm
    template_name = 'seeker/gateway_view.html'


class GatewayCreateView(CreateView):
    model = Gateway


class FunctionListView(ListView):
    """List all the functions that are available"""

    model = FunctionDef
    template_name = 'seeker/function_list.html'
    paginate_by = paginateEntries
    entrycount = 0
    qs = None

    def render_to_response(self, context, **response_kwargs):

        currentuser = self.request.user

        context['object_list'] = FunctionDef.objects.all().order_by('name')
        context['authenticated'] = currentuser.is_authenticated
        # Make sure the correct URL is being displayed
        return super(FunctionListView, self).render_to_response(context, **response_kwargs)


class ResultListView(ListView):
    """List all the search-result-baskets that are available"""

    model = Basket
    template_name = 'seeker/result_list.html'
    paginate_by = paginateEntries
    entrycount = 0
    qs = None
    order_cols = ['research__name', '', 'part__name', '', 'saved']
    order_heads = [{'name': 'Project', 'order': 'o=1', 'type': 'str'}, 
                   {'name': 'Language', 'order': '', 'type': 'str'}, 
                   {'name': 'Corpus', 'order': 'o=3', 'type': 'str'},
                   {'name': 'Hits', 'order': '', 'type': 'int'},
                   {'name': 'Date', 'order': 'o=5', 'type': 'str'}]

    def render_to_response(self, context, **response_kwargs):
        # Make sure we get the user and check the authentication
        currentuser = self.request.user
        context['authenticated'] = currentuser.is_authenticated

        # Make sure the correct URL is being displayed
        return super(ResultListView, self).render_to_response(context, **response_kwargs)

    def get_context_data(self, **kwargs):

        # Get the base implementation first of the context
        context = super(ResultListView, self).get_context_data(**kwargs)

        # Make sure we get the user and check the authentication
        currentuser = self.request.user
        context['authenticated'] = currentuser.is_authenticated

        # Initialise the list conditions: restrict results to current user
        lCombi = []
        lstQ = []
        lstQ.append(Q(research__owner = currentuser))

        # Add our own element(s)
        order = [Lower('research__name'), Lower('part__name')]
        initial = self.request.GET
        bAscending = True
        sType = 'str'
        if 'o' in initial:
            order = []
            iOrderCol = int(initial['o'])
            bAscending = (iOrderCol>0)
            iOrderCol = abs(iOrderCol)
            order.append(Lower( self.order_cols[iOrderCol-1]))
            sType = self.order_heads[iOrderCol-1]['type']
            if bAscending:
                self.order_heads[iOrderCol-1]['order'] = 'o=-{}'.format(iOrderCol)
            else:
                # order = "-" + order
                self.order_heads[iOrderCol-1]['order'] = 'o={}'.format(iOrderCol)
        if sType == 'str':
            qs = Basket.objects.filter(*lstQ).order_by(*order)
        else:
            qs = Basket.objects.filter(*lstQ).order_by(*order)
        # Possibly reverse the order
        if not bAscending:
            qs = qs.reverse()

        # Combine list with Basket/Quantor information
        for itm in  qs:
            # Get any quantor 
            quantor = Quantor.objects.filter(basket=itm).first()
            # Add the elements to the list -- only if the quantor is not none
            if quantor != None:
                lCombi.append({'basket': itm, 'quantor': quantor})            
        context['object_list'] = lCombi

        context['order_heads'] = self.order_heads

        # Check if there are any projects still running for this user
        oCrpp = crpp_status(currentuser.username, "list")
        running = None
        if oCrpp != None and 'commandstatus' in oCrpp and oCrpp['commandstatus'] == "ok":
            job_list = oCrpp['list']
            for job in job_list:
                if job['user'] == currentuser.username and job['status'] != 'interrupt'  and job['status'] != 'completed':
                    # Present this user-job
                    basket = Basket.objects.filter(jobid=job['id'],research__owner__username=job['user']).order_by('-saved').first()
                    if basket != None:
                        running = {}
                        running['basket'] = basket
                        running['research_id'] = basket.research.id
                        running['research_name'] = basket.research.name
                        running['corpus_name'] = basket.part.name
                        break
        context['running'] = running

        # Return the calculated context
        return context


class SeekerListView(ListView):
    """List all the research projects available"""

    model = Research
    template_name = 'seeker/research_list.html'
    paginate_by = paginateEntries
    entrycount = 0
    qs = None
    order_cols = ['owner', '', 'name', 'saved', 'purpose']
    order_heads = [{'name': 'Owner', 'order': 'o=1', 'type': 'str'}, 
                   {'name': 'Type', 'order': '', 'type': 'str'}, 
                   {'name': 'Name', 'order': 'o=3', 'type': 'str'},
                   {'name': 'Saved', 'order': 'o=4', 'type': 'str'},
                   {'name': 'Purpose', 'order': 'o=5', 'type': 'str'}]

    def render_to_response(self, context, **response_kwargs):
        # helper function...
        def get_parent_nodeid(glist, instance, groupid, user):
            """Find the parent-group [instance] from glist"""
            # Obvious case
            if glist == [] or instance == None: return groupid
            # Get the parent of the group
            grp = instance.name
            # Yes: look for the correct parent
            for g in glist:
                if g['group'] == grp and g['user'] == user:
                    return g['nodeid']
            # Getting here means we haven't found the group
            return groupid

        def path_exists(glist, p):
            """Check if glist has the group_path p"""
            for g in glist:
                if g['group_path'] == p:
                    return True
            return False

        sType = self.request.GET.get('listtype', '')
        if 'copyerror' not in context:
            if 'copyerror' in response_kwargs:
                context['copyerror'] = response_kwargs['copyerror']
            else:
                context['copyerror'] = None
        if sType == 'copy':
            # Can we change the path?
            self.request.path = reverse('seeker_list')
            # Need to make a copy
            if 'object_id' in self.kwargs:
                object_id = self.kwargs['object_id']
                sMsg = research_copy(object_id)
            else:
                sMsg = "The [object_id] is not passed along in the HTML call"
            # If needed, adapt the context
            if sMsg != None and sMsg != '':
                context['copyerror'] = sMsg
            ## Redirect appropriately
            #return HttpResponseRedirect(reverse('seeker_list', kwargs={'copyerror': sMsg}))
        elif sType == 'del':
            # We can delete this one
            if 'object_id' in self.kwargs:
                object_id = self.kwargs['object_id']
                # Try to delete it
                sMsg = research_del(object_id)
            else:
                sMsg = "The [object_id] is not passed along in the HTML call"

        currentuser = self.request.user
        # Is this user logged in?
        if currentuser.is_authenticated:
            # make sure we pass this on
            authenticated = True
            # Get the correct list of research projects:
            # - All my own projects
            # - All projects shared with me in the groups I belong to
            my_projects = Research.objects.filter(Q(owner=currentuser))
            lstQ = []
            lstQ.append(~Q(owner=currentuser))
            lstQ.append(~Q(sharegroups__group__in=currentuser.groups.all()))
    
            # This is the correct queryset - don't touch it!!!
            qs = Research.objects.filter(Q(owner=currentuser)|Q(sharegroups__group__in=currentuser.groups.all()))

            # Perform the sorting
            order = [Lower('owner'), '-saved']
            initial = self.request.GET
            bAscending = True
            sType = 'str'
            if 'o' in initial:
                order = []
                iOrderCol = int(initial['o'])
                bAscending = (iOrderCol>0)
                iOrderCol = abs(iOrderCol)
                order.append(Lower( self.order_cols[iOrderCol-1]))
                sType = self.order_heads[iOrderCol-1]['type']
                if bAscending:
                    self.order_heads[iOrderCol-1]['order'] = 'o=-{}'.format(iOrderCol)
                else:
                    # order = "-" + order
                    self.order_heads[iOrderCol-1]['order'] = 'o={}'.format(iOrderCol)
            if sType == 'str':
                qs = qs.order_by(*order)
            else:
                qs = qs.order_by(*order)
            # Possibly reverse the order
            if not bAscending:
                qs = qs.reverse()
            # We now have it
            research_list = qs

            # Combine into a combi-list
            combi_list = []
            for item in research_list:
                may_read = item.has_permission(currentuser, 'r')
                may_write = item.has_permission(currentuser, 'w')
                may_delete = item.has_permission(currentuser, 'd')
                combi_list.append({"project": item, 
                                   "may_read":may_read,
                                   "may_write": may_write,
                                   "may_delete": may_delete})

            # Provide a list of projects divided into ResGroup items
            resgroup_list = []
            # Check who I am
            sCurrentUser = currentuser.username
            # Get a list of the ID-values of other users 
            # -- not equal to me
            # -- only those that share one or more projects with me
            qs_users = Research.objects\
                         .exclude(Q(owner=currentuser))\
                         .filter(Q(sharegroups__group__in=currentuser.groups.all()))\
                         .order_by(Lower('owner__username')).distinct().values('owner')
            # Create a list of user-ids, starting from the currentuser
            currentuser_id = User.objects.filter(Q(username=sCurrentUser)).first().id
            user_list = []
            user_list.append(currentuser_id)
            # Now append the other user id's
            for usr in qs_users:
                user_list.append(usr['owner'])

            sUser = ""
            sGroup = ""
            max_depth = 0
            parent = 1      # Parent '1' means the root of all
            nodeid = 1
            prj_list = []
            grp_list = []   # List of groups that have already been shown

            # Process in the order of users
            for user_id in user_list:
                grp_list = []   # List of groups that have already been shown
                # Get projects of this user in an ordered way:
                # -- only those that are shared with me
                lstQ = []
                lstQ.append(Q(owner__id=user_id))
                if user_id != currentuser_id:
                    lstQ.append(Q(sharegroups__group__in=currentuser.groups.all()))
                # qs = Research.objects.filter(*lstQ).order_by(Lower('group__parent'), Lower('group'), Lower( 'name'))
                qs = Research.objects.filter(*lstQ)
                # Sort them appropriately
                qs_sorted = sorted(qs, key=lambda x: x.group_path() + "/" + x.name.lower())
                # Get the name of the user
                sUser = User.objects.filter(Q(id=user_id)).first().username
                # Set parameters for this user
                nodeid += 1
                groupid = nodeid
                sGroup = ""
                # Create a group for this user
                oGrp = {'group': sUser, 
                        'group_path': "/" + sUser,
                        'user': sUser,
                        'prj': None,
                        'nodeid': nodeid,
                        'childof': 1,           # Each user is child of the highest level
                        'depth': 1 }
                oGrp['minwidth'] = (oGrp['depth']-1) * 20
                # Add USER-level as group to list
                resgroup_list.append(oGrp)
                
                # Set the new parent: the user-group
                parent = groupid

                # Walk through the projects of this user
                for res_item in qs_sorted:
                    # Set the new parent: the user-group
                    parent = groupid
                    # Each research project is a new item
                    nodeid += 1
                    grp_path = "/" + sUser + res_item.group_path()
                    # Get and add the depth
                    depth = res_item.group_depth() + 2
                    if depth > max_depth:
                        max_depth = depth
                    # What group does this research project belong to?
                    sResItemGroup = "" if res_item.group == None else res_item.group.name
                    # Is the group staying the same?
                    if sResItemGroup != sGroup and not path_exists(resgroup_list, grp_path):
                        # Keep track of the current group
                        sGroup = sResItemGroup
                        # Determine what to show
                        sGroupOrUser = sGroup if sGroup != "" else sUser
                        # Show all groups that have not been shown yet
                        if sGroup == "":
                            # Now show this newly started search-containing group
                            parent = get_parent_nodeid(resgroup_list, None, groupid, sUser)
                            oGrp = {'group': sGroupOrUser, 'group_path': grp_path, 'prj': None, 
                                    'user': sUser, 'nodeid': nodeid, 'childof': parent, 'depth': depth-1}
                            oGrp['minwidth'] = (oGrp['depth']-2) * 20
                            # Add group to list
                            resgroup_list.append(oGrp)
                            # Add group to the list of already-shown groups
                            grp_list.append(this_group)
                            # Adapt values
                            parent = nodeid
                            nodeid += 1
                        else:
                            # Look for not-yet-shown-ancestor groups and show them
                            this_group = res_item.group
                            show_list = []
                            while this_group != None and this_group not in grp_list:
                                show_list.append(this_group)
                                this_group = this_group.parent
                            # Now walk this list in reversed order
                            for this_group in reversed(show_list):
                                # Show this group
                                sGroupOrUser = this_group.name
                                grp_depth = this_group.group_depth() + 2
                                if grp_depth > max_depth:
                                    max_depth = grp_depth
                                # Now show this newly started search-containing group
                                parent = get_parent_nodeid(resgroup_list, this_group.parent, groupid, sUser)
                                oGrp = {'group': sGroupOrUser, 'group_path': grp_path, 'prj': None, 
                                        'user': sUser, 'nodeid': nodeid, 'childof': parent, 'depth': grp_depth-1}
                                oGrp['minwidth'] = (oGrp['depth']-2) * 20
                                # Add group to list
                                resgroup_list.append(oGrp)
                                # Add group to the list of already-shown groups
                                grp_list.append(this_group)
                                # Adapt values: all items that follow will be part of this group
                                parent = nodeid
                                nodeid += 1

                    # Determine what to show
                    sGroupOrUser = sGroup if sGroup != "" else sUser
                    # Create a new item 
                    parent = get_parent_nodeid(resgroup_list, res_item.group, groupid, sUser)
                    oGrp = {'group': sGroupOrUser, 'group_path': grp_path, 'prj': res_item, 'nodeid': nodeid, 
                            'user': sUser, 'childof': parent, 'depth': depth,
                            'may_read': res_item.has_permission(currentuser, 'r'),
                            'may_write': res_item.has_permission(currentuser, 'w'),
                            'may_delete': res_item.has_permission(currentuser, 'd')}
                    oGrp['minwidth'] = (oGrp['depth']-1) * 20
                    # Add group to list
                    resgroup_list.append(oGrp)
            # Add the Remainder into the elements
            for oGrp in resgroup_list:
                oGrp['remainder'] = max_depth - oGrp['depth'] + 1

            # Create a well ordered list of search groups
            sgroup_list = ResGroup.objects.filter(Q(owner=currentuser)).order_by(Lower('parent__name'), Lower('name'), )
        else:
            combi_list = []
            research_list = []
            resgroup_list = []
            sgroup_list = []
            max_depth = 0
            authenticated = False

        # DEBUGGING
        x = json.dumps([{'group': x['group'], 
                         'nodeid': x['nodeid'], 
                         'depth': x['depth'], 
                         'childof': x['childof'], 
                         'prj': '-' if x['prj'] == None else x['prj'].name} for x in resgroup_list])

        # Wrap up the context
        context['combi_list'] = combi_list
        context['order_heads'] = self.order_heads
        context['object_list'] = research_list
        context['import_form'] = UploadFileForm()
        context['authenticated'] = authenticated
        context['resgroups'] = resgroup_list
        context['sgroup_list'] = sgroup_list
        context['currentuser'] = currentuser.username
        context['max_depth'] = max_depth        # The number of columns needed

        # Make sure the correct URL is being displayed
        return super(SeekerListView, self).render_to_response(context, **response_kwargs)


class SimpleListView(BasicList):
    """List all the currently defined 'simple' searches"""

    model = Research
    listform = SimpleListForm
    prefix = "simple"
    basic_name = "simple"
    plural_name = "Simple searches"
    sg_name = "Simple search"
    basic_add = "simple_details"
    add_text = "Goto" 
    has_select2 = False
    paginate_by = 20
    bUseFilter = True
    delete_line = True
    new_button = False
    none_on_empty = True
    page_function = "ru.basic.search_paged_start"
    order_cols = ['owner__username', 'name', '', '', 'created', '']
    order_default = ['owner__username', 'created', 'name', '', '', '']
    order_heads = [{'name': 'Owner',        'order': 'o=1', 'type': 'str', 'custom': 'owner'},
                   {'name': 'Name',         'order': 'o=2', 'type': 'str', 'field': 'name', 'linkdetails': True},
                   {'name': 'Description',  'order': '',    'type': 'str', 'custom': 'description', 'main': True, 'linkdetails': True},
                   {'name': 'Status',       'order': '',    'type': 'str', 'custom': 'status'},
                   {'name': 'Created',      'order': 'o=4', 'type': 'str', 'custom': 'created'},
                   {'name': '', 'order': '', 'type': 'str', 'options': ['delete']}
                   ]
    filters = [ {"name": "Name", "id": "filter_name", "enabled": False},]
    searches = [
        {'section': '', 'filterlist': [
            {'filter': 'name', 'dbfield': 'name', 'keyS': 'name'}]},
        {'section': 'other', 'filterlist': [
            {'filter': 'owner', 'fkfield': 'owner',  'keyS': 'owner', 'keyFk': 'id', 'keyList': 'ownlist', 'infield': 'id' },
            {'filter': 'stype', 'dbfield': 'stype',  'keyS': 'stype'}]}
        ]
    custombuttons = []
    uploads = [
        {'title': 'simple', 'label': 'simple searches', 'url': 'import_file', 'msg': 'Specify the JSON file that contains the CESAR Simple project'}
        ]

    def adapt_search(self, fields):
        # Adapt the search to the keywords that *may* be shown
        lstExclude=None
        qAlternative = None

        # Make sure the owner is set to myself
        if fields['ownlist'] == None or len(fields['ownlist']) == 0:
            # Restrict to myself
            fields['ownlist'] = User.objects.filter(username=self.request.user.username)

            # Only show simple 
            fields['stype'] = "s"

        # Make sure SimplName is excluded
        lstExclude = [ Q(name=SIMPLENAME) ]
        
        # Return all that needs be
        return fields, lstExclude, qAlternative

    def get_field_value(self, instance, custom):
        sBack = ""
        sTitle = ""
        html = []

        if custom == "owner":
            sTitle = instance.owner.username
            html.append("<span style='color: darkgreen; font-size: small;'>{}</span>".format(instance.owner.username[:20]))
        elif custom == "description":
            sDescr = instance.purpose
            sCompact = instance.compact
            if sCompact:
                oCompact = json.loads(sCompact)
                if 'description' in oCompact:
                    sDescr = oCompact['description']
            sTitle = sDescr

            html.append("<span style='color: darkblue; font-size: small;'>{}</span>".format(sDescr))
        elif custom == "created":
            sTime = instance.created.strftime("%d/%b/%Y %H:%M")
            html.append("<span style='color: darkgreen; font-size: smaller;'>{}</span>".format(sTime))
        elif custom == "status":
            sStatus = "unknown"
            if instance.compact == None or instance.compact == "" or instance.compact[0] != "{":
                sStatus = "empty?"
            else:
                sStatus = "ok"
            html.append(sStatus)
        # Combine the HTML code
        sBack = "\n".join(html)
        return sBack, sTitle

    def add_to_context(self, context, initial):
        # Add a button for the default simple list view
        context['user_button'] = "<a class='btn btn-xs jumbo-1' role='button' href='{}' title='Go to the default simple search view'>Go Simple</a>".format(reverse('simple_details'))
        # Adapt for CESAR
        context['is_app_editor'] = user_is_ingroup(self.request, "seeker_user")
        context['is_app_uploader'] = user_is_ingroup(self.request, "radboud-tsg")
        return context


def get_changeform_initial_data(model, request):
    """
    Get the initial form data.
    Unless overridden, this populates from the GET params.
    """
    initial = dict(request.GET.items())
    for k in initial:
        try:
            f = model._meta.get_field(k)
        except FieldDoesNotExist:
            continue
        # We have to special-case M2Ms as a list of comma-separated PKs.
        if isinstance(f, models.ManyToManyField):
            initial[k] = initial[k].split(",")
    return initial


class SeekerForm():
    """Forms to work with the Seeker App"""

    formset_list = []

    def add_formset(self, ModelFrom, ModelTo, FormModel):
        FactSet = inlineformset_factory(ModelFrom, ModelTo,form=FormModel, min_num=1, extra=0, can_delete=True, can_order=True)
        oFormset = {'factory': FactSet, 'formset': None}
        self.formset_list.append(oFormset)


class ResearchExe(View):
    """General class to guide the execution of a research project"""

    # Initialisations:     
    arErr = []              # errors   
    template_name = None    # The template to be used
    action = ""             # The action to be undertaken
    data = {'status': 'ok', 'html': '', 'statuscode': ''}       # Create data to be returned   
    progress = ['start', 'finish', 'count', 'total', 'ready'] 
    completed = ['searchTime', 'searchDone', 'taskid', 'table', 'total']
    oErr = ErrHandle()
    bDebug = True

    def post(self, request, object_id=None):
        try:
            # A POST request means we are trying to SAVE something
            self.initializations(request, object_id)
            if self.checkAuthentication(request):
                # Define the context
                sStatusCode = "none"
                context = dict(object_id = object_id, savedate=None, statuscode=sStatusCode)
                # Action depends on 'action' value
                if self.action != "" and self.obj != None:
                    if self.bDebug: 
                        sNowTime = get_crpp_date(timezone.now())
                        self.oErr.Status("ResearchExe: action={} at [{}]".format(self.action, sNowTime))
                    if self.action == "prepare":
                        # Make sure we have the right parameters to work with
                        if "select_part" in self.qd and self.qd.get("select_part") != "":
                            research = self.obj
                            # Check if this object is not currently being executed
                            if research.get_status() != "":
                                # Stop it
                                research.stop()

                            # Check if this is a simple search
                            if "is_simple_search" in self.qd:
                                # It is a simple search: modify the research based on the information we receive
                                modify_simple_search(research, self.qd)

                            # Find out which corpus/part has been chosen
                            select_part = self.qd.get("select_part")
                            select_format = self.qd.get("searchFormat")

                            # Possibly get searchtype and search count
                            search_type = self.qd.get("search_type")
                            search_count = self.qd.get("search_count")
                            oOptions = { 'search_type': search_type, 
                                         'search_count': search_count}
                            # Get the partId
                            part = None
                            if select_part != "":
                                part = Part.objects.filter(Q(id=select_part)).first()

                            # Check if format adaptation is needed, should there be a better match
                            select_format = get_best_format(part, select_format)

                            # Check if the necessary ingredient(s) are there
                            basket = self.get_basket(select_format, part)
                            # Set the status and the jobid
                            basket.status = "prepared"
                            basket.jobid = ""
                            # Set options
                            basket.options = json.dumps(oOptions)
                            # Save the basket
                            basket.save()
                            context['statuscode'] = "prepared"
                            self.data['status'] = "prepared"
                            self.data['basket_id'] = basket.id
                            # Also provide all relevent command urls for the caller
                            self.data['basket_start'] = reverse("search_start", kwargs={'object_id': basket.id})
                            self.data['basket_progress'] = reverse("search_progress", kwargs={'object_id': basket.id})
                            self.data['basket_stop'] = reverse("search_stop", kwargs={'object_id': basket.id})
                            self.data['basket_watch'] = reverse("search_watch", kwargs={'object_id': basket.id})
                            self.data['basket_result'] = reverse("result_details", kwargs={'pk': basket.id})
                            # Also indicate the statuscode
                            sStatusCode = "preparing" 
                        else:
                            self.arErr.append("No corpus part to be searched has been selected")
                    elif self.action == "watch":
                        # Get the basket
                        basket = self.obj
                        self.data['basket_id'] = basket.id
                        # Also provide all relevent command urls for the caller
                        self.data['basket_start'] = reverse("search_start", kwargs={'object_id': basket.id})
                        self.data['basket_progress'] = reverse("search_progress", kwargs={'object_id': basket.id})
                        self.data['basket_stop'] = reverse("search_stop", kwargs={'object_id': basket.id})
                        self.data['basket_watch'] = reverse("search_watch", kwargs={'object_id': basket.id})
                        self.data['basket_result'] = reverse("result_details", kwargs={'pk': basket.id})
                        sStatusCode = "watching"
                        basket.set_status("caught the search")
                    elif self.action == "start":
                        # the basket is the object
                        basket = self.obj
                        # Start translation + execution
                        oBack = basket.research.execute(basket)
                        if self.bDebug: 
                            self.oErr.Status("ResearchExe: start...")
                            for key in oBack:
                                self.oErr.Status("   {}={}".format(key, str(oBack[key])[:20]))
                        # Check for errors
                        if oBack['status'] == "error":
                            # Make sure we set the basket status
                            self.obj.set_status("error")
                            # Add error to the error array
                            self.arErr.append(oBack['msg'])
                            sStatusCode = "error"
                            self.data['statuscode'] = "error"
                            self.oErr.Status("ResearchExe error: " + oBack['msg'])
                            if 'msg_list' in oBack:
                                context['msg_list'] = oBack['msg_list']
                            context['msg'] = oBack['msg']
                            # If we have an instance, then provide that in the context
                            if 'id' in oBack and 'type' in oBack:
                                # Get the type and the instance
                                inst_type = oBack['type']
                                inst_id = oBack['id']
                                context['error_jumptype'] = inst_type
                                context['error_jumpid'] = inst_id
                                # Provide a jump to the correction of the function
                                if inst_type == "feat":
                                    # Jump to 72 with the correct object_id
                                    context['error_jumpto'] = reverse("research_part_72", kwargs={'object_id': inst_id})
                                elif inst_type == "cond":
                                    # Jump to a condition
                                    context['error_jumpto'] = reverse("research_part_62", kwargs={'object_id': inst_id})
                                elif inst_type == "cvar":
                                    # Jump to a construction variable
                                    context['error_jumpto'] = reverse("research_part_43", kwargs={'object_id': inst_id})

                    elif self.action == "stop":
                        # Need to stop the execution of the project
                        # Find out which basket is treating this
                        basket = self.obj
                        sStatusCode = "stopping"
                        # Stop the basket
                        basket.research.stop_execute(basket)
                    elif self.action == "progress":
                        ## First check: if we are in error, then do not continue
                        #if self.obj.get_status() == "error":
                        #    sStatusCode = "error"
                        #    self.data['status'] = "error"
                        #else:

                        # Need to get the status of the project
                        # NOTE: the self.obj now is the BASKET!!
                        self.obj.set_status("loading current status")
                        oBack = self.obj.get_progress()
                        if oBack['commandstatus'] == "error":
                            self.arErr.append(oBack['msg'])
                            context['statuscode'] = "error"
                        else:
                            # All went well: provide information for the context
                            sStatusCode = ""
                            if 'code' in oBack['status']:
                                sStatusCode = oBack['status']['code']
                            elif 'code' in oBack:
                                sStatusCode = oBack['code']
                            else:
                                # This may be a problem?
                                iStop = True
                            # Pass on the status code in the context
                            context['statuscode'] = sStatusCode

                            if self.bDebug: self.oErr.Status("ResearchExe: statuscode=[{}]".format(sStatusCode))

                            # Action depends on the status code
                            if sStatusCode == "working":
                                # Now we have one set of feedback
                                for item in self.progress:
                                    context[item] = oBack[item]
                                # Add percentages
                                if oBack['total'] == 0:
                                    context['ptc_count'] = 0
                                    context['ptc_ready'] = 0
                                    context['pipecount'] = 0
                                    context['ptc_done'] = 0
                                else:
                                    context['ptc_count'] = 100 * oBack['count'] / oBack['total']
                                    context['ptc_ready'] = 100 * oBack['ready'] / oBack['total']
                                    context['pipecount'] = oBack['count'] - oBack['ready']
                                    context['ptc_done'] = int(context['ptc_ready'])
                                context['found'] = oBack['found']
                            elif sStatusCode == "preparing":
                                # Show where we are in preparing
                                sStatus = self.obj.get_status()
                                context['prep_status'] = sStatus
                                context['prep_job'] = "no job running yet" if self.obj.jobid == "" else self.obj.jobid
                                if sStatus == "error":
                                    sStatusCode = "stop"
                            elif sStatusCode == "error":
                                # Add the error description to the list
                                if 'message' in oBack:
                                    # Try to decypher
                                    sMsg = oBack['message']
                                    try:
                                        if sMsg.startswith("{"):
                                            # Find the first closing }
                                            sJson = sMsg.split('}', 1)[0] + "}"
                                            oJson = json.loads(sJson)
                                            lErr = []
                                            for key,value in oJson.items():
                                                self.arErr.append("{}: {}".format(key,value))
                                        else:
                                            # Not structured
                                            self.arErr.append(sMsg)
                                    except:
                                        # Could not decypher
                                        self.arErr.append(sMsg)
                                else:
                                    self.arErr.append("The /crpp engine returns an unknown error")
                            elif sStatusCode == "completed":
                                # Adapt the table we receive
                                oTable = oBack["table"]
                                for qcItem in oTable:
                                    sub = []
                                    for idx in range(0, len(qcItem['subcats'])):
                                        subcat = qcItem['subcats'][idx]
                                        subcount = qcItem['counts'][idx]
                                        sub.append({"subcat": subcat, "subcount": subcount})
                                    qcItem['sub'] = sub
                                # Copy all the items from [oBack] to [context]
                                for item in self.completed:
                                    context[item] = oBack[item]

                                # Make sure the searchTime is provided in seconds
                                context['searchTime'] = context['searchTime'] / 1000

                                # Final action: make all the information available into a Quantor for this basket
                                oQset = self.obj.set_quantor(oBack)
                                if oQset['status'] != 'ok':
                                    msg = "Could not set quantor"
                                    if 'msg' in oQset: msg = oQset['msg']
                                    self.arErr.append(msg)
                    elif self.action == "download":
                        if "select_part" in self.qd:
                            # Find out which corpus/part has been chosen
                            select_part = self.qd.get("select_part")
                            select_format = self.qd.get("searchFormat")

                            # Check if format adaptation is needed, should there be a better match
                            select_format = get_best_format(select_part, select_format)

                            # Determine which basket this is in
                            basket = self.get_basket(select_format, select_part)
                            # Translate project to Xquery
                            oBack = self.obj.to_crpx(select_part, select_format, basket)
                            # Check for errors
                            if oBack['status'] == "error":
                                # Add error to the error array
                                self.arErr.append(oBack['msg'])
                            else:
                                # Get the name adn the contents
                                sCrpxName = oBack['crpx_name']
                                sCrpxText = oBack['crpx_text']
                                response = HttpResponse(sCrpxText, content_type="application/xml")
                                response['Content-Disposition'] = 'attachment; filename="'+sCrpxName+'.crpx"'     
                                return response
                    elif self.action == "downloadjson":
                        # Prepare to download the search project as json
                        oBack = self.obj.get_json()
                        # Check for errors
                        if oBack['status'] == "error":
                            # Add error to the error array
                            self.arErr.append(oBack['msg'])
                        else:
                            # Get the name adn the contents
                            sJsonName = oBack['json_name']
                            sJsonText = oBack['json_data']
                            response = HttpResponse(sJsonText, content_type="application/json; charset=utf-8")
                            response['Content-Disposition'] = 'attachment; filename="'+sJsonName+'.json"'     
                            return response
                    elif self.action == "downloadfunction":
                        # Prepare to download the FUNCTION as json
                        # THIS WOULD ONLY DOWNLOAD THE TOP FUNCTION: oBack = self.obj.get_output()
                        if self.obj == None:
                            oBack['status'] = "error"
                            if not 'msg' in oBack:
                                oBack['msg'] = "ResearchExe - downloadfunction - no OBJ defined"
                        else:
                            if self.obj.root != None:
                                oBack = self.obj.root.get_json()
                                oBack['name'] = "{}-{}".format(self.obj.root.construction.name, self.obj.root.variable.name)
                            elif self.obj.rootcond != None:
                                oBack = self.obj.rootcond.get_json()
                                oBack['name'] = "{}".format(self.obj.rootcond.name)
                            elif self.obj.rootfeat != None:
                                oBack = self.obj.rootfeat.get_json()
                                oBack['name'] = "{}".format(self.obj.rootfeat.name)
                            else:
                                oBack['status'] = "error"
                                if not 'msg' in oBack:
                                    oBack['msg'] = "ResearchExe - downloadfunction - no root (cvar/cond/feat) defined"
                        # Check for errors
                        if 'status' in oBack and oBack['status'] == "error":
                            # Add error to the error array
                            self.arErr.append(oBack['msg'])
                        else:
                            # Get the name and the contents
                            sProjectName = "anonymous"
                            research = self.obj.get_research()
                            if research != None:
                                sProjectName = "{}_{}".format(research.name, research.owner.username)
                            oCode = oBack
                            sJsonName = "{}_{}".format(sProjectName, oCode['name'])
                            sJsonText = json.dumps(oCode, 2)
                            response = HttpResponse(sJsonText, content_type="application/json; charset=utf-8")
                            response['Content-Disposition'] = 'attachment; filename="'+sJsonName+'.json"'     
                            return response

                # Make sure we have a list of any errors
                error_list = [str(item) for item in self.arErr]
                self.data['error_list'] = error_list
                self.data['errors'] = self.arErr
                if len(self.arErr) >0:
                    # debugging error stuff
                    self.oErr.Status("ResearchExe error loc=1")
                    # Pass on the statuscode to the context for the template rendering
                    context['status'] = "error"
                    # Pass on the status to the calling function through [self.data]
                    self.data['status'] = "error"
                    self.data['statuscode'] = "error"
                    context['statuscode'] = "error"
                else:
                    # Pass on the statuscode
                    self.data['statuscode'] = sStatusCode
                    context['status'] = self.data['status']

                # Add items to context
                context['error_list'] = error_list
            
                # Get the HTML response
                self.data['html'] = render_to_string(self.template_name, context, request)
            
            else:
                self.data['html'] = "Please log in before continuing"

        except:
            # If there is a real error, then the user should know too
            self.data['status'] = "error"
            self.data['statuscode'] = "error"
            sMsg = self.oErr.get_error_message()
            self.data['html'] = sMsg
            self.oErr.Status("ResearchExe: point #8 gives message: {}".format(sMsg))

        # Return the information
        return JsonResponse(self.data)

    def get_basket(self, sFormat, part):

        # Get the research object
        research = self.obj
        # Check if 'part'is a string or not
        if isinstance(part, str):
            part = Part.objects.filter(id=part).first()
        # Prepare query
        lstQ = []
        lstQ.append(Q(research=research))
        lstQ.append(Q(format=sFormat))
        lstQ.append(Q(part=part))
        basket = Basket.objects.filter(*lstQ).first()
        if basket == None:
            # Create a new basket
            basket = Basket(research=research, part=part, format=sFormat, status="created", jobid="")
            basket.save()
        # Return the basket
        return basket

    def get(self, request, object_id=None): 
        self.initializations(request, object_id)
        if self.checkAuthentication(request):
            # Build the context
            context = dict(object_id = object_id, savedate=None)

            # Make sure we have a list of any errors
            error_list = [str(item) for item in self.arErr]
            self.data['error_list'] = error_list
            self.data['errors'] = self.arErr
            if len(self.arErr) >0:
                self.data['status'] = "error"

            # Add items to context
            context['error_list'] = error_list
            context['status'] = self.data['status']

            # Get the HTML response
            self.data['html'] = render_to_string(self.template_name, context, request)
        else:
            self.data['html'] = "Please log in before continuing"

        # Return the information
        return JsonResponse(self.data)

    def checkAuthentication(self,request):
        # first check for authentication
        if not request.user.is_authenticated:
            # Simply redirect to the home page
            self.data['html'] = "Please log in to work on a research project"
            return False
        else:
            return True

    def initializations(self, request, object_id):
        # Clear errors
        self.arErr = []
        self.data['status'] = "ok"
        self.data['html'] = ""
        self.data['statuscode'] = ""
        # COpy the request
        self.request = request
        # Copy any object id
        self.object_id = object_id
        # if self.action == "start" or self.action == "download":
        if self.action == "prepare" or self.action == "download" or \
            self.action == "downloadjson" or self.action == "downloadfunction":
            # Get the instance of the Main Model object
            self.obj =  self.MainModel.objects.get(pk=object_id)
        else:
            # In all other cases we are looking at a Basket object
            self.obj = Basket.objects.get(pk=object_id)
        # Get the parameters
        if request.POST:
            self.qd = request.POST
        else:
            self.qd = request.GET
        # Perform some custom initialisations
        self.custom_init()

    def custom_init(self):
        pass


class ResearchPrepare(ResearchExe):
    MainModel = Research
    action = "prepare"
    template_name = "seeker/exe_status.html"


class ResearchWatch(ResearchExe):
    MainModel = Basket
    action = "watch"
    template_name = "seeker/exe_status.html"


class ResearchStart(ResearchExe):
    MainModel = Basket
    action = "start"
    template_name = "seeker/exe_status.html"
    

class ResearchStop(ResearchExe):
    # MainModel = Research
    action = "stop"
    template_name = "seeker/exe_status.html"


class ResearchProgress(ResearchExe):
    action = "progress"
    template_name = "seeker/exe_status.html"


class ResearchDownload(ResearchExe):
    MainModel = Research
    template_name = "seeker/exe_status.html"
    action = "download"


class ResearchDownloadJson(ResearchExe):
    MainModel = Research
    template_name = "seeker/exe_status.html"
    action = "downloadjson"


class ResearchDownloadFunction(ResearchExe):
    MainModel = Function
    template_name = "seeker/exe_status.html"
    action = "downloadfunction"


class ResearchPart(View):
    # Initialisations:     
    arErr = []              # errors   
    template_name = None    # The template to be used
    template_err_view = None
    form_validated = True   # Used for POST form validation
    savedate = None         # When saving information, the savedate is returned in the context
    add = False             # Are we adding a new record or editing an existing one?
    obj = None              # The instance of the MainModel
    action = ""             # The action to be undertaken
    MainModel = None        # The model that is mainly used for this form
    form_objects = []       # List of forms to be processed
    formset_objects = []    # List of formsets to be processed
    bDebug = False          # Debugging information
    data = {'status': 'ok', 'html': ''}       # Create data to be returned    
    
    def post(self, request, object_id=None):
        # A POST request means we are trying to SAVE something
        self.initializations(request, object_id)

        # Explicitly set the status to OK
        self.data['status'] = "ok"

        if self.checkAuthentication(request):
            # Build the context
            context = dict(object_id = object_id, savedate=None)
            # Action depends on 'action' value
            if self.action == "":
                if self.bDebug: self.oErr.Status("ResearchPart: action=(empty)")
                # Walk all the forms for preparation of the formObj contents
                for formObj in self.form_objects:
                    # Are we SAVING a NEW item?
                    if self.add:
                        # We are saving a NEW item
                        formObj['forminstance'] = formObj['form'](request.POST, prefix=formObj['prefix'])
                    else:
                        # We are saving an EXISTING item
                        # Determine the instance to be passed on
                        instance = self.get_instance(formObj['prefix'])
                        # Make the instance available in the form-object
                        formObj['instance'] = instance
                        # Get an instance of the form
                        formObj['forminstance'] = formObj['form'](request.POST, prefix=formObj['prefix'], instance=instance)

                # Initially we are assuming this just is a review
                context['savedate']="reviewed at {}".format(datetime.now().strftime("%X"))

                # Iterate again
                for formObj in self.form_objects:
                    prefix = formObj['prefix']
                    # Adapt if it is not readonly
                    if not formObj['readonly']:
                        # Check validity of form
                        if formObj['forminstance'].is_valid() and self.is_custom_valid(prefix, formObj['forminstance']):
                            # Save it preliminarily
                            instance = formObj['forminstance'].save(commit=False)
                            # The instance must be made available (even though it is only 'preliminary')
                            formObj['instance'] = instance
                            # Perform actions to this form BEFORE FINAL saving
                            bNeedSaving = formObj['forminstance'].has_changed()
                            if self.before_save(prefix, request, instance=instance): bNeedSaving = True
                            if formObj['forminstance'].instance.id == None: bNeedSaving = True
                            if bNeedSaving:
                                # Perform the saving
                                instance.save()
                                # Set the context
                                context['savedate']="saved at {}".format(datetime.now().strftime("%X"))
                                # Put the instance in the form object
                                formObj['instance'] = instance
                                # Store the instance id in the data
                                self.data[prefix + '_instanceid'] = instance.id
                                # Any action after saving this form
                                self.after_save(prefix, instance)
                        else:
                            self.arErr.append(formObj['forminstance'].errors)
                            self.form_validated = False

                    # Add instance to the context object
                    context[prefix + "Form"] = formObj['forminstance']
                # Walk all the formset objects
                for formsetObj in self.formset_objects:
                    prefix  = formsetObj['prefix']
                    if self.can_process_formset(prefix):
                        formsetClass = formsetObj['formsetClass']
                        form_kwargs = self.get_form_kwargs(prefix)
                        if self.add:
                            # Saving a NEW item
                            formset = formsetClass(request.POST, request.FILES, prefix=prefix, form_kwargs = form_kwargs)
                        else:
                            # Saving an EXISTING item
                            instance = self.get_instance(prefix)
                            qs = self.get_queryset(prefix)
                            if qs == None:
                                formset = formsetClass(request.POST, request.FILES, prefix=prefix, instance=instance, form_kwargs = form_kwargs)
                            else:
                                formset = formsetClass(request.POST, request.FILES, prefix=prefix, instance=instance, queryset=qs, form_kwargs = form_kwargs)
                        # Process all the forms in the formset
                        self.process_formset(prefix, request, formset)
                        # Store the instance
                        formsetObj['formsetinstance'] = formset
                        # Adapt the formset contents only, when it is NOT READONLY
                        if not formsetObj['readonly']:
                            # Is the formset valid?
                            if formset.is_valid():
                                # Make sure all changes are saved in one database-go
                                with transaction.atomic():
                                    # Walk all the forms in the formset
                                    for form in formset:
                                        # At least check for validity
                                        if form.is_valid() and self.is_custom_valid(prefix, form):
                                            # Should we delete?
                                            if form.cleaned_data['DELETE']:
                                                # Delete this one
                                                form.instance.delete()
                                                # NOTE: the template knows this one is deleted by looking at form.DELETE
                                                # form.delete()
                                            else:
                                                # Check if anything has changed so far
                                                has_changed = form.has_changed()
                                                # Save it preliminarily
                                                instance = form.save(commit=False)
                                                # Any actions before saving
                                                if self.before_save(prefix, request, instance, form):
                                                    has_changed = True
                                                # Save this construction
                                                if has_changed: 
                                                    # Save the instance
                                                    instance.save()
                                                    # Adapt the last save time
                                                    context['savedate']="saved at {}".format(datetime.now().strftime("%X"))
                                                    # Store the instance id in the data
                                                    self.data[prefix + '_instanceid'] = instance.id
                                        else:
                                            if len(form.errors) > 0:
                                                self.arErr.append(form.errors)
                            else:
                                # Iterate over all errors
                                for idx, err_this in enumerate(formset.errors):
                                    if '__all__' in err_this:
                                        self.arErr.append(err_this['__all__'][0])
                                    elif err_this != {}:
                                        # There is an error in item # [idx+1], field 
                                        problem = err_this 
                                        for k,v in err_this.items():
                                            fieldName = k
                                            errmsg = "Item #{} has an error at field [{}]: {}".format(idx+1, k, v[0])
                                            self.arErr.append(errmsg)

                            # self.arErr.append(formset.errors)
                    else:
                        formset = []
                    # Add the formset to the context
                    context[prefix + "_formset"] = formset
            elif self.action == "download":
                # We are being asked to download something
                if self.dtype != "" and self.qcTarget and self.qcTarget > 0:
                    # Get the name of the CRP, starting from basket
                    sCrpName = self.basket.research.name
                    sUserName = self.basket.research.owner.username
                    sPartDir = self.basket.part.dir
                    oBack = {'status': 'ok'}
                    # Download the requested information
                    sType = "csv" if (self.dtype == "xlsx") else self.dtype
                    oData = crpp_dbget(sUserName, sCrpName, self.qcTarget, sType=sType, sPart=sPartDir)
                    if oData == None or oData['commandstatus'] != 'ok' or not 'db' in oData:
                        # Allow user to add to the context
                        context = self.add_to_context(context)
                        # Something went wrong
                        sMsg = ""
                        oBack['status'] = 'error'
                        oBack['commandstatus'] = sMsg
                        if 'code' in oData:
                            oBack['code'] = oData['code']
                            sMsg += " code: " + oBack['code']   
                        if 'message' in oData:
                            sMsg += "<br>" + oData['message']                     
                        self.arErr.append(sMsg)
                        context['error_list'] = [str(item) for item in self.arErr]
                        context['errors'] = self.arErr
                        return render(request, self.template_name, context)
                    else:
                        # All should be fine...
                        sData = oData['db']
                        # Decode the data and compress it using gzip
                        bUtf8 = (self.dtype != "db")
                        bUsePlain = (self.dtype == "xlsx" or self.dtype == "csv")
                        compressed_content = decompressSafe(sData, True, bUtf8, bUsePlain)
                        # Get the name and the contents
                        sLng = self.basket.part.corpus.get_lng_display()
                        sGz = "" if (bUsePlain) else ".gz"
                        sDbName = "{}_{}_{}_QC{}_Dbase.{}{}".format(sCrpName, sLng, sPartDir, self.qcTarget, self.dtype, sGz)
                        sContentType = "application/gzip"
                        if self.dtype == "csv":
                            sContentType = "text/tab-separated-values"
                        elif self.dtype == "xlsx":
                            sContentType = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

                        # Excel needs additional conversion
                        if self.dtype == "xlsx":
                            # Convert 'compressed_content' to an Excel worksheet
                            response = HttpResponse(content_type=sContentType)
                            response['Content-Disposition'] = 'attachment; filename="{}"'.format(sDbName)    
                            response = csv_to_excel(compressed_content, response)
                        else:
                            response = HttpResponse(compressed_content, content_type=sContentType)
                            response['Content-Disposition'] = 'attachment; filename="{}"'.format(sDbName)    

                        # Continue for all formats
                        
                        # return gzip_middleware.process_response(request, response)
                        return response
            elif self.action == "downloadhit":
                if 'downloadtype' in self.qd and 'downloaddata' in self.qd:
                    # Get the download type and the data itself
                    dtype = self.qd['downloadtype']
                    ddata = self.qd['downloaddata']
            
                    if dtype == "tree":
                        dext = ".svg"
                        sContentType = "application/svg"
                    elif dtype == "htable":
                        dext = ".html"
                        sContentType = "application/html"
                    elif (dtype == "htable-png" or dtype == "tree-png"):
                        dext = ".png"
                        # sContentType = "application/octet-stream"
                        sContentType = "image/png"
                        # Read base64 encoded part
                        arPart = ddata.split(";")
                        dSecond = arPart[1]
                        # Strip off the preceding "base64," part
                        ddata = dSecond.replace("base64,", "")
                        # Convert string to bytestring
                        ddata = ddata.encode()
                        # Decode base64 into binary
                        ddata = base64.decodestring(ddata)
                        # Strip -png off
                        dtype = dtype.replace("-png", "")


                    # Determine a file name from the information we have...
                    iResId = int(self.qd['resid'])
                    oResult = self.obj.get_result(iResId)
                    sBase = Text.strip_ext( oResult['File'])
                    sIdt = oResult['Locs']
                    if not sBase in sIdt:
                        sIdt = "{}_{}".format( sBase, sIdt)
                    sFileName = "{}_{}{}".format(sIdt, dtype, dext)

                    response = HttpResponse(ddata, content_type=sContentType)
                    response['Content-Disposition'] = 'attachment; filename="{}"'.format(sFileName)    

                    # For downloading: resturn this response
                    return response

            # Allow user to add to the context
            context = self.add_to_context(context)

            # Make sure we have a list of any errors
            error_list = [str(item) for item in self.arErr]
            context['error_list'] = error_list
            context['errors'] = self.arErr
            if len(self.arErr) > 0:
                # Indicate that we have errors
                self.data['has_errors'] = True
                self.data['status'] = "error"
            else:
                self.data['has_errors'] = False
            # Standard: add request user to context
            context['requestuser'] = request.user

            # Get the HTML response
            if len(self.arErr) > 0:
                if self.template_err_view != None:
                     # Create a list of errors
                    self.data['err_view'] = render_to_string(self.template_err_view, context, request)
                else:
                    self.data['error_list'] = error_list
                self.data['html'] = ''
            else:
                # In this case reset the errors - they should be shown within the template
                self.data['html'] = render_to_string(self.template_name, context, request)
            # At any rate: empty the error basket
            self.arErr = []
            error_list = []

        else:
            self.data['html'] = "Please log in before continuing"

        # Return the information
        return JsonResponse(self.data)
        
    def get(self, request, object_id=None): # , *args, **kwargs): # ): 
        self.data['status'] = 'ok'
        # Perform the initializations that need to be made anyway
        self.initializations(request, object_id)
        if self.checkAuthentication(request):
            context = dict(object_id = object_id, savedate=None)
            # Walk all the form objects
            for formObj in self.form_objects:        
                # Used to populate a NEW research project
                # - CREATE a NEW research form, populating it with any initial data in the request
                initial = dict(request.GET.items())
                if self.add:
                    # Create a new form
                    formObj['forminstance'] = formObj['form'](initial=initial, prefix=formObj['prefix'])
                else:
                    # Used to show EXISTING information
                    instance = self.get_instance(formObj['prefix'])
                    # We should show the data belonging to the current Research [obj]
                    formObj['forminstance'] = formObj['form'](instance=instance, prefix=formObj['prefix'])
                # Add instance to the context object
                context[formObj['prefix'] + "Form"] = formObj['forminstance']
            # Walk all the formset objects
            for formsetObj in self.formset_objects:
                formsetClass = formsetObj['formsetClass']
                prefix  = formsetObj['prefix']
                form_kwargs = self.get_form_kwargs(prefix)
                if self.add:
                    # - CREATE a NEW formset, populating it with any initial data in the request
                    initial = dict(request.GET.items())
                    # Saving a NEW item
                    formset = formsetClass(initial=initial, prefix=prefix, form_kwargs=form_kwargs)
                else:
                    # show the data belonging to the current [obj]
                    instance = self.get_instance(prefix)
                    qs = self.get_queryset(prefix)
                    if qs == None:
                        formset = formsetClass(prefix=prefix, instance=instance, form_kwargs=form_kwargs)
                    else:
                        formset = formsetClass(prefix=prefix, instance=instance, queryset=qs, form_kwargs=form_kwargs)
                # Process all the forms in the formset
                ordered_forms = self.process_formset(prefix, request, formset)
                if ordered_forms:
                    context[prefix + "_ordered"] = ordered_forms
                # Store the instance
                formsetObj['formsetinstance'] = formset
                # Add the formset to the context
                context[prefix + "_formset"] = formset
            # Allow user to add to the context
            context = self.add_to_context(context)
            # Make sure we have a list of any errors
            error_list = [str(item) for item in self.arErr]
            context['error_list'] = error_list
            context['errors'] = self.arErr
            # Standard: add request user to context
            context['requestuser'] = request.user
            
            # Get the HTML response
            sHtml = render_to_string(self.template_name, context, request)
            sHtml = treat_bom(sHtml)
            self.data['html'] = sHtml
        else:
            self.data['html'] = "Please log in before continuing"

        # Return the information
        return JsonResponse(self.data)
      
    def checkAuthentication(self,request):
        # first check for authentication
        if not request.user.is_authenticated:
            # Simply redirect to the home page
            self.data['html'] = "Please log in to work on a research project"
            return False
        else:
            return True

    def initializations(self, request, object_id):
        # Clear errors
        self.arErr = []
        # COpy the request
        self.request = request
        # Copy any object id
        self.object_id = object_id
        self.add = object_id is None
        # Get the parameters
        if request.POST:
            self.qd = request.POST
        else:
            self.qd = request.GET

        # Check for action
        if 'action' in self.qd:
            self.action = self.qd['action']

        # Find out what the Main Model instance is, if any
        if self.add:
            self.obj = None
        else:
            # Get the instance of the Main Model object
            self.obj =  self.MainModel.objects.filter(pk=object_id).first()
            # NOTE: if the object doesn't exist, we will NOT get an error here
            # Perform some custom initialisations
            self.custom_init()

    def get_instance(self, prefix):
        return self.obj

    def is_custom_valid(self, prefix, form):
        return True

    def get_queryset(self, prefix):
        return None

    def get_form_kwargs(self, prefix):
        return None

    def before_save(self, prefix, request, instance=None, form=None):
        return False

    def after_save(self, prefix, instance=None):
        return True

    def add_to_context(self, context):
        return context

    def process_formset(self, prefix, request, formset):
        return None

    def can_process_formset(self, prefix):
        return True

    def custom_init(self):
        pass


class ResGroupList(ListView):
    model = ResGroup
    template_name = 'seeker/resgroup_list.html'

    def render_to_response(self, context, **response_kwargs):
        currentuser = self.request.user
        context['authenticated'] = currentuser.is_authenticated

        # Create a well ordered list of search groups
        sgroup_list = ResGroup.objects.filter(Q(owner=currentuser)).order_by(Lower('parent__name'), Lower('name'), )
        context['sgroup_list'] = sgroup_list
        # Make sure the correct URL is being displayed
        return super(ResGroupList, self).render_to_response(context, **response_kwargs)


class ResGroupDetails(ResearchPart):
    template_name = 'seeker/resgroup_details.html'
    MainModel = ResGroup
    form_objects = [{'form': SeekerResGroupForm, 'prefix': 'resgroup', 'readonly': False}]
           
    def before_save(self, prefix, request, instance=None, form=None):
        has_changed = False
        if prefix == 'resgroup':
            resgroup = None
            for formObj in self.form_objects:
                if formObj['prefix'] == 'resgroup': resgroup = formObj['instance']
            if resgroup != None:
                # Check for the owner
                if resgroup.owner_id == None:
                    resgroup.owner = request.user
                    has_changed = True
        # Set the 'saved' one
        if has_changed and self.obj != None:
            self.obj.save()
        return has_changed

    def add_to_context(self, context):

        # Look at the object_id
        if context['object_id'] == None:
            if self.obj != None:
                context['object_id'] = "{}".format(self.obj.id)

        # Figure out the owner
        if self.obj == None:
            currentowner = self.request.user.username
        else:
            currentowner = self.obj.owner
        context['currentowner'] = currentowner

        # Set the list of groups *EXCLUDING* myself
        if self.obj == None:
            parent_list = []
        else:
            parent_list = ResGroup.objects.exclude(id=self.obj.id).order_by('name')
        context['parent_list'] = parent_list

        # Return the context we have adapted
        return context


class ResearchField(ResearchPart):
    """Change, set or delete one field of a Research instance"""

    template_name = 'seeker/research_field.html'
    MainModel = Research

    def add_to_context(self, context):

        # Initiali values for what we return
        context['field-msg'] = "(no changes)"
        context['field-status'] = ""
        # Find out which field is being processed
        objid = self.obj.id
        field_name = "prj-fname-{}".format(objid)
        field_value = "prj-fvalue-{}".format(objid)

        # Find out what action is needed
        if self.action == "save":
            # Get the field name and value
            if field_name in self.qd and field_value in self.qd:
                fname = self.qd[field_name]
                fvalue = self.qd[field_value]
                # Save the new value
                if fvalue == "(none)":
                    setattr(self.obj, fname, None)
                else:
                    setattr(self.obj, fname, fvalue)
                self.obj.save()
                context['field_msg'] = "research group setting has been updated"
                context['field_status'] = "saved"
            else:
                context['field_msg'] = "please provide {} and {}".format(field_name, field_value)
                context['field_status'] = "error"
        elif self.action == "delete":
            # Get the field name and value
            if field_name in self.qd:
                # Set this value to null
                fname = self.qd[field_name]
                # Remove the group from the research project
                setattr(self.obj, fname, None)
                self.obj.save()
                context['field_msg'] = "research is no longer under this group"
                context['field_status'] = "deleted"
            else:
                context['field_msg'] = "please provide {}".format(field_name)
                context['field_status'] = "error"

        return context


class ResearchPart1(ResearchPart):
    template_name = 'seeker/research_part_1.html'
    MainModel = Research
    form_objects = [{'form': GatewayForm, 'prefix': 'gateway', 'readonly': False},
                    {'form': SeekerResearchForm, 'prefix': 'research', 'readonly': False}]
    SharegFormSet = inlineformset_factory(Research, ShareGroup, 
                                        form=SharegForm, min_num=0, 
                                        extra=0, can_delete=True, can_order=False)
    formset_objects = [{'formsetClass': SharegFormSet, 'prefix': 'shareg', 'readonly': False}]
             
    def get_instance(self, prefix):
        if prefix == 'research':
            # Return the Research instance
            return self.obj
        elif prefix == 'shareg':
            # A sharegroup is mainly interested in the Research
            return self.obj
        else:
            # The other option is 'gateway': return the gateway instance
            return self.obj.gateway

    def before_save(self, prefix, request, instance=None, form=None):
        has_changed = False
        if prefix == 'research':
            research = None
            gateway = None
            for formObj in self.form_objects:
                if formObj['prefix'] == 'gateway': gateway = formObj['instance']
                if formObj['prefix'] == 'research': research = formObj['instance']
            if research != None:
                if research.gateway_id == None or research.gateway == None:
                    research.gateway = gateway
                    has_changed = True
                # Check for the owner
                if research.owner_id == None:
                    research.owner = request.user
                    has_changed = True
        # Set the 'saved' one
        if has_changed and self.obj != None:
            self.obj.save()
        return has_changed

    def after_save(self, prefix, instance=None):
        if prefix == 'research' and self.obj == None:
            # Capture the object
            self.obj = instance
            # Adapt the instanceid and the ajaxurl
            object_id = "{}".format(instance.id)
            self.data['instanceid'] = object_id

    def add_to_context(self, context):
        if context['object_id'] == None:
            if self.obj != None:
                context['object_id'] = "{}".format(self.obj.id)
        if self.obj == None:
            currentowner = None
        else:
            currentowner = self.obj.owner
        context['currentowner'] = currentowner
        return context

    def custom_init(self):
        if self.obj:
            gw = self.obj.gateway
            if gw:
                # Check and repair CVAR instances
                gw.check_cvar()
                # Make sure DVAR instances are ordered
                gw.order_dvar()
                # Make sure GVAR instances are ordered
                gw.order_gvar()
        return True

    def process_formset(self, prefix, request, formset):
        if prefix == 'shareg':
            currentuser = request.user
            # Need to process all the forms here
            for form in formset:
                x = form.fields['permission'].help_text
        return None


class ResearchPart2(ResearchPart):
    template_name = 'seeker/research_part_2.html'
    MainModel = Research
    WrdConstructionFormSet = inlineformset_factory(Gateway, Construction, 
                                                form=ConstructionWrdForm, min_num=1, 
                                                extra=0, can_delete=True, can_order=True)
    CnsConstructionFormSet = inlineformset_factory(Gateway, Construction, 
                                                form=ConstructionCnsForm, min_num=1, 
                                                extra=0, can_delete=True, can_order=True)
    formset_objects = [{'formsetClass': WrdConstructionFormSet, 'prefix': 'wrdconstruction', 'readonly': False},
                       {'formsetClass': CnsConstructionFormSet, 'prefix': 'cnsconstruction', 'readonly': False}]
    # TODO: possibly add ExtConstructionFormSet...
             
    def get_instance(self, prefix):
        if prefix == 'wrdconstruction' or prefix == 'cnsconstruction' or prefix == 'extconstruction':
            return self.obj.gateway

    def before_save(self, prefix, request, instance=None, form=None):
        has_changed = False
        if prefix == 'wrdconstruction':
            # Add the correct search item
            instance.search = SearchMain.create_item("word-group", form.cleaned_data['value'], 'groupmatches')
            has_changed = True
        elif prefix == 'cnsconstruction':
            # Add the correct search item
            instance.search = SearchMain.create_item("const-group", form.cleaned_data['cat_incl'], 
                                                     'groupmatches', form.cleaned_data['cat_excl'])
            has_changed = True
        elif prefix == 'extconstruction':
            # TODO: How to process this one??
            pass
        # Set the 'saved' one
        self.obj.save()
        return has_changed

    def add_to_context(self, context):
        if self.obj == None:
            targettype = None
            currentowner = None
        else:
            targettype = self.obj.targetType
            currentowner = self.obj.owner
        context['targettype'] = targettype
        context['currentowner'] = currentowner
        return context

    def process_formset(self, prefix, request, formset):
        # Get the owner of the research project
        if self.obj == None:
            owner = None
        else:
            owner = self.obj.owner
        currentuser = request.user
        if prefix == 'wrdconstruction':
            # Need to process all the forms here
            for form in formset:
                # Compare the owner with the current user
                if owner != None and owner != currentuser:
                    form.fields['name'].disabled = True
                    form.fields['value'].disabled = True
        elif prefix == 'cnsconstruction':
            # Need to process all the forms here
            for form in formset:
                # Compare the owner with the current user
                if owner != None and owner != currentuser:
                    # TODO: add intelligence here
                    form.fields['name'].disabled = True
                    pass
        return None

    def can_process_formset(self, prefix):
        if self.obj == None:
            return True
        else:
            return ( (prefix == 'wrdconstruction' and self.obj.targetType == "w") or \
                     (prefix == 'cnsconstruction' and self.obj.targetType == "c") or \
                     (prefix == 'extconstruction' and self.obj.targetType == "e") )

        
class ResearchPart3(ResearchPart):
    template_name = 'seeker/research_part_3.html'
    MainModel = Research
    GvarFormSet = inlineformset_factory(Gateway, GlobalVariable, 
                                        form=GvarForm, min_num=1, 
                                        extra=0, can_delete=True, can_order=True)
    formset_objects = [{'formsetClass': GvarFormSet, 'prefix': 'gvar', 'readonly': False}]
                
    def get_instance(self, prefix):
        if prefix == 'gvar':
            return self.obj.gateway

    def before_save(self, prefix, request, instance=None, form=None):
        # Set the 'saved' one
        self.obj.save()
        return False

    def add_to_context(self, context):
        if self.obj == None:
            currentowner = None
            targettype = ""
        else:
            currentowner = self.obj.owner
            targettype = self.obj.targetType
        context['currentowner'] = currentowner
        context['targettype'] = targettype
        return context


class ResearchPart4(ResearchPart):
    template_name = 'seeker/research_part_4.html'
    template_err_view = 'seeker/err_view.html'
    MainModel = Research
    VardefFormSet = inlineformset_factory(Gateway, VarDef, 
                                          form=VarDefForm, min_num=1, extra=0, 
                                          can_delete=True, can_order=True)
    formset_objects = [{'formsetClass': VardefFormSet, 'prefix': 'vardef', 'readonly': False}]
                
    def get_instance(self, prefix):
        if prefix == 'vardef':
            return self.obj.gateway

    def add_to_context(self, context):
        if self.obj == None:
            currentowner = None
            context['research_id'] = None
            targettype = ""
        else:
            currentowner = self.obj.owner
            context['research_id'] = self.obj.gateway.research.id
            targettype = self.obj.targetType
        context['currentowner'] = currentowner
        context['targettype'] = targettype
        # We also need to make the object_id available
        context['object_id'] = self.object_id
        # Provide a number for the newest item 
        context['new_order_number'] = len(self.obj.gateway.definitionvariables.all())
        return context

    def get_queryset(self, prefix):
        if prefix == "vardef":
            qs = self.obj.gateway.definitionvariables.all().order_by('order')
        else:
            qs = super(ResearchPart4, self).get_queryset(prefix)
        return qs

    def process_formset(self, prefix, request, formset):
        if prefix == 'vardef':
            # Sorting: see https://wiki.python.org/moin/HowTo/Sorting
            ordered_forms = sorted(formset.forms, key=lambda item: item.instance.order)
            # Make sure the initial values of the 'order' in the forms are set correctly
            for form in ordered_forms:
                form.fields['ORDER'].initial = form.instance.order
            return ordered_forms
        else:
            return None

    def is_custom_valid(self, prefix, form):
        if prefix == "vardef":
            valid, sMsg = form.check_order()
            if not valid:
                # Add the message to the error log
                self.arErr.append(sMsg)
        else:
            valid = super(ResearchPart4, self).is_custom_valid(prefix, form)
        return valid

    def before_save(self, prefix, request, instance=None, form=None):
        has_changed = False
        if prefix == 'vardef':
            # Retrieve the 'order' field
            if instance.order != form.cleaned_data['ORDER']:
                if form.cleaned_data['ORDER'] == None:
                    instance.order = len(instance.gateway.definitionvariables.all())+1
                else:
                    instance.order = form.cleaned_data['ORDER']
                has_changed = True
        # Set the 'saved' one
        self.obj.save()
        return has_changed
    

class ResearchPart42(ResearchPart):
    template_name = 'seeker/research_part_42.html'
    MainModel = VarDef
    CvarFormSet = inlineformset_factory(VarDef, ConstructionVariable, 
                                          form=CvarForm, min_num=1, extra=0)
    formset_objects = [{'formsetClass': CvarFormSet, 'prefix': 'cvar', 'readonly': False}]
                
    def get_instance(self, prefix):
        if prefix == 'cvar':
            # Note: the actual object returned for 'cvar' is a VARDEF...
            return self.obj

    def add_to_context(self, context):
        if self.obj == None:
            currentowner = None
            researchid = None
            targettype = ""
        else:
            currentowner = self.obj.gateway.research.owner
            researchid = self.obj.gateway.research.id
            targettype = self.obj.gateway.research.targetType

            # Check and adapt the cvar_formset
            cvar_formset = context['cvar_formset']
            lstQ = []
            lstQ.append(Q(gateway=self.obj.gateway))
            qsCns = Construction.objects.filter(*lstQ)
            for index, cvar_form in enumerate(cvar_formset):
                # Get the list of constructions
                cns_id = cvar_form.instance.construction.id
                qs = qsCns.exclude(Q(id=cns_id))
                # Adapt this cvar_form
                cvar_form.copyto = qs
            # Copy the adapted formset back
            context['cvar_formset'] = cvar_formset

            # Possibly execute post-saving task(s)
            self.process_task()

        context['currentowner'] = currentowner
        # We also need to make the object_id available
        context['object_id'] = self.object_id
        context['research_id'] = researchid
        context['vardef_this'] = self.obj
        context['targettype'] = targettype
        return context

    def process_task(self):
        """If the dictionary contains a 'task' element, it should be processed"""

        oErr = ErrHandle()
        try:
            # Get the task
            sTask = "" if (not '_task' in self.qd) else self.qd['_task']
            if sTask == "copy_cvar_function":
                # There are some more parameters we expect
                functionid = self.qd['_functionid']
                functionid = -1 if functionid == "" else int(functionid)
                constructionid = self.qd['_constructionid']
                constructionid = -1 if constructionid == "" else int(constructionid)
                if functionid>=0 and constructionid>=0:
                    # Find the function
                    function = Function.objects.filter(id=functionid).first()
                    # Find out what variable we are processing
                    cvar_src = function.root
                    variable = cvar_src.variable
                    construction = Construction.objects.filter(id=constructionid).first()
                    if construction != None and variable != None:
                        cvar_dst = ConstructionVariable.objects.filter(variable=variable, construction=construction).first()
                        if function != None and cvar_dst != None:
                            # Delayed committing doesn't work because of the recursivity
                            # with transaction.atomic():
                            if cvar_dst.function != None: 
                                # Now we are going to delete the old one
                                # Ideally the user should confirm this...
                                cvar_dst.function.delete()
                            cvar_dst.function = function.get_copy(root=cvar_dst)
                            cvar_dst.functiondef = function.functiondef
                            cvar_dst.type = function.root.type
                            response = cvar_dst.save()

            # REturn positively
            return True
        except:
            oErr.DoError("custom_init error")
            return False

    def before_save(self, prefix, request, instance=None, form=None):
        # NOTE: the 'instance' here is the CVAR instance

        has_changed = False
        arErr = []
        # When saving a CVAR, we need to check that the functions are okay
        if prefix == 'cvar':
            # Find the function attached to me - only if applicable!!
            if instance.type == "calc" and instance.functiondef != None:
                # Two situations:
                # - THere is no function yet
                # - There is a function, but with the wrong functiondef
                if instance.function == None or (instance.function != None and instance.functiondef != instance.function.functiondef):
                    # Does a previous function exist?
                    if instance.function:
                        # Remove the existing function
                        instance.function.delete()
                    # Create a new (obligatory) 'Function' instance, with accompanying Argument instances
                    instance.function = Function.create(instance.functiondef, instance, None, None)
                    # Indicate that changes have been made
                    has_changed = True
            elif instance.type == "json":
                # We are uploading a function: get the JSON contents
                if 'file_source' in form.cleaned_data:
                    data_file = form.cleaned_data["file_source"]
                    oData = import_data_file(data_file, arErr)
                    # Check the top layer
                    if oData != None and 'type' in oData and oData['type'] == 'calc':
                        # Make sure we change the type
                        instance.type = "calc"
                        if instance.function != None:
                            # Remove the old function
                            instance.function.delete()
                        # Be sure to get the correct gateway
                        gateway = self.obj.gateway
                        # Create the functions for this CVAR
                        func_main = Function.create_from_list(oData['value'], gateway, "cvar", instance, None, None)
                        func_main.save()
                        instance.function = func_main
                        instance.functiondef = func_main.functiondef
                        # Indicate that changes have been made
                        has_changed = True
                        # x = json.dumps( instance.function.root.get_json())
        # Set the 'saved' one, but ONLY if any changes have been made
        if has_changed: 
            self.obj.gateway.research.save()
        # Return the changed flag
        return has_changed


class Variable43(ResearchPart):
    """The definition of one particular variable"""

    MainModel = ConstructionVariable
    template_name = 'seeker/variable_view.html'

    def add_to_context(self, context):
        # Provide the top function
        context['function'] = self.obj.function
        return context


class Variable43t(ResearchPart):
    """The definition of one particular variable"""

    MainModel = ConstructionVariable
    template_name = 'seeker/variable_details.html'

    def add_to_context(self, context):
        # Add a list of functions to the context
        context['function_list'] = self.obj.get_functions()
        return context


class Condition63(ResearchPart):
    """The definition of one particular condition"""

    MainModel = Condition
    template_name = 'seeker/variable_view.html'

    def add_to_context(self, context):
        # Provide the top function
        context['function'] = self.obj.function
        return context


class Condition63t(ResearchPart):
    """The definition of one particular condition"""

    MainModel = Condition
    template_name = 'seeker/variable_details.html'

    def add_to_context(self, context):
        # Add a list of functions to the context
        context['function_list'] = self.obj.get_functions()
        return context


class Feature73(ResearchPart):
    """The definition of one particular condition"""

    MainModel = Feature
    template_name = 'seeker/variable_view.html'

    def add_to_context(self, context):
        # Provide the top function
        context['function'] = self.obj.function
        return context


class Feature73t(ResearchPart):
    """The definition of one particular condition"""

    MainModel = Feature
    template_name = 'seeker/variable_details.html'

    def add_to_context(self, context):
        # Add a list of functions to the context
        context['function_list'] = self.obj.get_functions()
        return context


class ResearchPart43(ResearchPart):
    """Starting from a CVAR of type 'function', allow defining that function"""

    MainModel = ConstructionVariable
    template_name = 'seeker/research_part_43.html'
    form_objects = [{'form': FunctionForm, 'prefix': 'function', 'readonly': False}]
    ArgFormSet = inlineformset_factory(Function, Argument, 
                                          form=ArgumentForm, min_num=1, extra=0)
    formset_objects = [{'formsetClass': ArgFormSet, 'prefix': 'arg', 'readonly': False}]
                
    def get_instance(self, prefix):
        if prefix == 'function' or prefix == 'arg':
            # This returns the FUNCTION object we are linked to
            cvar = self.obj
            if cvar.function_id == None or cvar.function == None:
                # Check the function definition
                if cvar.functiondef_id == None:
                    # There is an error: we need to have a function definition here
                    return None
                # Make sure both changes are saved in one database-go
                with transaction.atomic():
                    # Create a new function 
                    # function = Function(functiondef = cvar.functiondef, root = cvar)
                    function = Function.create(cvar.functiondef, cvar, None, None)
                    # Make sure the function instance gets saved
                    # function.save()
                    # Acc a link to this function from the CVAR object
                    cvar.function = function
                    # Make sure we save the CVAR object
                    cvar.save()
            return cvar.function

    def custom_init(self):
        """Make sure the arg-formset gets the correct number of arguments"""

        # Check if we have a CVAR object
        if self.obj:
            # Check if the object type is Calculate
            if self.obj.type == "calc":
                # Get the function definition
                functiondef = self.obj.functiondef
                if functiondef != None:
                    # Get the number of arguments
                    argnum = functiondef.argnum
                    # Adapt the minimum number of items in the argument formset
                    self.ArgFormSet = inlineformset_factory(Function, Argument, 
                                          form=ArgumentForm, min_num=argnum, extra=0)
                    self.formset_objects[0]['formsetClass'] = self.ArgFormSet

        return True

    def add_to_context(self, context):
        if self.obj == None:
            currentowner = None
            context['research_id'] = None
            context['vardef_this'] = None
            context['construction_this'] = None
            targettype = ""
        else:
            currentowner = self.obj.variable.gateway.research.owner
            context['research_id'] = self.obj.variable.gateway.research.id
            context['vardef_this'] = self.obj.variable
            context['construction_this'] = self.obj.construction
            targettype = self.obj.variable.gateway.research.targetType
            # Further action if this is a calculation
            if self.obj.type == "calc":
                # Need to specify the template for the function
                functiondef = self.obj.functiondef
                if functiondef == None:
                    # Provide a break point for debugging
                    iStop = 1

                # Adapt the arguments for this form
                # arg_formset = context['arg_formset']
                arg_defs = ArgumentDef.objects.filter(function=functiondef)

                # Calculate the initial queryset for 'gvar'
                qs_gvar = GlobalVariable.objects.filter(gateway=self.obj.construction.gateway)

                # Calculate the initial queryset for 'cvar'
                # NOTE: we take into account the 'order' field, which must have been defined properly...
                lstQ = []
                lstQ.append(Q(construction=self.obj.construction))
                lstQ.append(Q(variable__order__lt=self.obj.variable.order))
                qs_cvar = ConstructionVariable.objects.filter(*lstQ)

                # Calculate the initial queryset for 'dvar'
                lstQ = []
                lstQ.append(Q(gateway=self.obj.construction.gateway))
                lstQ.append(Q(order__lt=self.obj.variable.order))
                qs_dvar = VarDef.objects.filter(*lstQ).order_by('order')

                # - adapting the 'arg_formset' for the 'function' form (editable)
                fun_this = self.get_instance('function')
                if fun_this != None:
                    context['arg_formset'] = check_arguments(context['arg_formset'], functiondef, qs_gvar, qs_cvar, qs_dvar, '44')
                else:
                    context['arg_formset'] = None

                # Add a list of functions to the context
                context['function_list'] = self.obj.get_functions()

        context['currentowner'] = currentowner
        # We also need to make the object_id available
        context['object_id'] = self.object_id
        context['targettype'] = targettype
        return context
    
    def before_save(self, prefix, request, instance=None, form=None):
        # NOTE: the 'instance' is the CVAR instance

        has_changed = False
        # When we save an ARG, we need to add a link to the Function it belongs to
        if prefix == 'arg':
            # Get the 'function' instance
            function = None
            for formObj in self.form_objects:
                if formObj['prefix'] == 'function':
                    # Get the function object 
                    function = formObj['instance']
            # Link to this function
            if function != None and instance.function != function:
                instance.function = function
                has_changed = True
            # Check the argtype of this argument: is it NOT a function any more?
            if instance.argtype != "func":
                # Then REmove all functions pointing to me
                func_child = instance.functionparent.first()
                if func_child != None:
                    func_child.delete()
                    has_changed = True
            else:
                # THis is a function: check if the function definition has not changed
                func_child = instance.functionparent.first()
                # NOTE: the argument functiondef must equal the argument's function-child-functiondef
                if func_child == None or instance.functiondef != func_child.functiondef:
                    # The function definition changed >> replace the child with a completely NEW version
                    # [1] remove the child
                    if func_child != None:
                        func_child.delete()
                    # [2] Create a new version
                    func_child = Function.create(instance.functiondef, instance.function.root, None, instance)
                    # [3] Save it
                    func_child.save()
                    # Indicate changes
                    has_changed = True
        elif prefix == 'function':
            if instance != None:
                if self.obj.function != instance:
                    # Link the function-instance to the  CVAR instance
                    self.obj.function = instance
                    # Save the adapted CVAR instance
                    self.obj.save()
                    # We have already saved the above, so 'has_changed' does not need to be touched
        # Save the research object
        self.obj.variable.gateway.research.save()
        # Return the change-indicator to trigger saving
        return has_changed


class ResearchPart44(ResearchPart):
    """Starting from an Argument of type 'function', allow defining that function"""

    MainModel = Argument
    template_name = 'seeker/research_part_44.html'
    ArgFormSet = inlineformset_factory(Function, Argument, 
                                          form=ArgumentForm, min_num=1, extra=0)
    # Two forms:
    # - the 'parent' form is view-only and contains the argument we are supplying a function for
    # - the 'function' form is editable and contains the function for the argument 
    form_objects = [{'form': FunctionForm, 'prefix': 'parent', 'readonly': True},
                    {'form': FunctionForm, 'prefix': 'function', 'readonly': False}]
    # Two formsets:
    # - the 'arg'  formset belongs to the 'function' (see above)
    # - the 'parg' formset belongs to the 'parent'
    formset_objects = [{'formsetClass': ArgFormSet, 'prefix': 'arg', 'readonly': False},
                       {'formsetClass': ArgFormSet, 'prefix': 'parg', 'readonly': True}]
                
    def get_instance(self, prefix):
        # NOTE: For '44' the self.obj is an Argument!!

        if prefix == 'function' or prefix == 'arg':
            # This returns the EXISTING or NEW function object belonging to the argument
            qs = Function.objects.filter(parent=self.obj)
            # Does it exist?
            if qs.count() == 0:
                # It does not yet exist: create it
                # Make sure both changes are saved in one database-go
                with transaction.atomic():
                    # Create a new function 
                    function = Function(functiondef = self.obj.functiondef, 
                                        root = self.obj.function.root,
                                        parent = self.obj)
                    # Make sure the function instance gets saved
                    function.save()
            else:
                # It exists, so assign it
                function = qs[0]
            return function
        elif prefix == 'parent' or prefix == 'parg':
            # This returns the PARENT function object the argument belongs to
            return self.obj.function

    def custom_init(self):
        """Make sure the ARG formset gets the correct number of arguments"""

        # Check if we have a FUNCTION object
        fun_this = self.get_instance('function')
        if fun_this:
            # Get the function definition
            functiondef = fun_this.functiondef
            if functiondef != None:
                # Get the number of arguments
                argnum = functiondef.argnum
                # Adapt the minimum number of items in the argument formset
                self.ArgFormSet = inlineformset_factory(Function, Argument, 
                                        form=ArgumentForm, min_num=argnum, extra=0)
                # The 'arg' object is the one with index '0'
                self.formset_objects[0]['formsetClass'] = self.ArgFormSet

        return True

    def add_to_context(self, context):
        # NOTE: the instance (self.obj) for the '44' form is an ARGUMENT
        pfun_this = self.get_instance('parent')
        if pfun_this == None:
            currentowner = None
            context['research_id'] = None
            context['vardef_this'] = None
            context['construction_this'] = None
            context['cvar_this'] = None
            targettype = ""
        else:
            # Get to the CVAR instance
            cvar = pfun_this.root
            gateway = cvar.construction.gateway
            currentowner = gateway.research.owner
            context['research_id'] = gateway.research.id
            context['vardef_this'] = cvar.variable
            context['construction_this'] = cvar.construction
            context['cvar_this'] = cvar
            targettype = gateway.research.targetType

            # Since this is a '44' form, we know this is a calculation

            # Calculate the initial queryset for 'gvar'
            #   (These are the variables available for this GATEWAY)
            qs_gvar = GlobalVariable.objects.filter(gateway=gateway)

            # Calculate the initial queryset for 'cvar'
            lstQ = []
            lstQ.append(Q(construction=cvar.construction))
            lstQ.append(Q(variable__order__lt=cvar.variable.order))
            qs_cvar = ConstructionVariable.objects.filter(*lstQ)

            # Calculate the initial queryset for 'dvar'
            lstQ = []
            lstQ.append(Q(gateway=gateway))
            lstQ.append(Q(order__lt=cvar.variable.order))
            qs_dvar = VarDef.objects.filter(*lstQ).order_by('order')

            # - adapting the 'parg_formset' for the 'parent' form (view-only)
            context['parg_formset'] = check_arguments(context['parg_formset'], pfun_this.functiondef, qs_gvar, qs_cvar, qs_dvar, '44')

            # - adapting the 'arg_formset' for the 'function' form (editable)
            fun_this = self.get_instance('function')
            if fun_this != None:
                context['arg_formset'] = check_arguments(context['arg_formset'], fun_this.functiondef, qs_gvar, qs_cvar, qs_dvar, '44')
            else:
                context['arg_formset'] = None

            # Get a list of all ancestors
            context['anc_list'] = fun_this.get_ancestors()

        context['currentowner'] = currentowner
        # We also need to make the object_id available
        context['object_id'] = self.object_id
        context['targettype'] = targettype
        return context

    def before_save(self, prefix, request, instance=None, form=None):
        has_changed = False
        # When we save an ARG, we need to add a link to the Function it belongs to
        if prefix == 'arg':
            # The instance is the argument instance

            # Check the argtype of this argument: is it NOT a function any more?
            if instance.argtype != "func":
                # Then REmove all functions pointing to me
                func_child = instance.functionparent.first()
                if func_child != None:
                    func_child.delete()
                    has_changed = True
            else:
                # THis is a function: check if the function definition has not changed
                func_child = instance.functionparent.first()
                if func_child == None or instance.functiondef != func_child.functiondef:
                    # The function definition changed >> replace the child with a completely NEW version
                    # [1] remove the child
                    if func_child != None:
                        func_child.delete()
                    # [2] Create a new version
                    func_child = Function.create(instance.functiondef, instance.function.root, None, instance)
                    # [3] Save it here (or that is done one level up)
                    func_child.save()
                    # Indicate changes have been made
                    has_changed = True
        # Save the related RESEARCH object
        self.obj.function.root.construction.gateway.research.save()
        # Return the change-indicator to trigger saving
        return has_changed


class ResearchPart6(ResearchPart):
    template_name = 'seeker/research_part_6.html'
    MainModel = Research
    CondFormSet = inlineformset_factory(Gateway, Condition, 
                                        form=ConditionForm, min_num=1, 
                                        extra=0, can_delete=True, can_order=True)
    formset_objects = [{'formsetClass': CondFormSet, 'prefix': 'cond', 'readonly': False}]

    def custom_init(self):
        """If the dictionary contains a 'task' element, it should be processed"""

        oErr = ErrHandle()
        try:
            # Get the task
            sTask = "" if (not '_task' in self.qd) else self.qd['_task']
            if sTask == "copy_condition_function":
                # There are some more parameters we expect
                functionid = self.qd['_functionid']
                functionid = -1 if functionid == "" else int(functionid)
                conditionid = self.qd['_conditionid']
                conditionid = -1 if conditionid == "" else int(conditionid)
                if functionid>=0 and conditionid>=0:
                    # Find the function
                    function = Function.objects.filter(id=functionid).first()
                    condition = Condition.objects.filter(id=conditionid).first()
                    if function != None and condition != None:
                        # Copy the function to the construction
                        with transaction.atomic():
                            condition.function.delete()
                            condition.function = function.get_copy()
                            response = condition.save()

            # REturn positively
            return True
        except:
            oErr.DoError("custom_init error")
            return False

    def get_instance(self, prefix):
        if prefix == 'cond':
            # We need to have the gateway
            return self.obj.gateway

    def get_form_kwargs(self, prefix):
        if prefix == 'cond' and self.obj != None and self.obj.gateway != None:
            return {"gateway":  self.obj.gateway}
        else:
            return None

    def process_formset(self, prefix, request, formset):
        if prefix == 'cond':
            # Sorting: see ResearchPart4
            ordered_forms = sorted(formset.forms, key=lambda item: item.instance.order)
            # Make sure the initial values of the 'order' in the forms are set correctly
            for form in ordered_forms:
                form.fields['ORDER'].initial = form.instance.order
            return ordered_forms
        else:
            return None

    def add_to_context(self, context):
        # Note: the self.obj is the Research project
        if self.obj == None:
            currentowner = None
            targettype = ""
        else:
            # gateway = self.obj.gateway
            currentowner = self.obj.owner
            targettype = self.obj.targetType
            context['currentowner'] = currentowner
            context['targettype'] = targettype
        # Provide a number for the newest item 
        context['new_order_number'] = len(self.obj.gateway.conditions.all())
        # Return the context
        return context

    def before_save(self, prefix, request, instance=None, form=None):
        has_changed = False
        if prefix == 'cond':
            # Find the function attached to me - only if applicable!!
            if instance.condtype == "func" and instance.functiondef != None:
                # Two situations:
                # - THere is no function yet
                # - There is a function, but with the wrong functiondef
                if instance.function == None or (instance.function != None and instance.functiondef != instance.function.functiondef):
                    # If the instance is not yet made, it needs to be saved before we continue
                    if instance.id == None:
                        instance.save()
                    # Does a previous function exist?
                    if instance.function:
                        # Remove the existing function
                        instance.function.delete()
                    
                    # Create a new (obligatory) 'Function' instance, with accompanying Argument instances
                    instance.function = Function.create(instance.functiondef, None, instance, None)
                    # Indicate that changes have been made and saving of 'instance' is needed
                    has_changed = True
            elif instance.condtype == "json":
                # We are uploading a function: get the JSON contents
                if 'file_source' in form.cleaned_data:
                    data_file = form.cleaned_data["file_source"]
                    oData = import_data_file(data_file, self.arErr)
                    # Check the top layer
                    if oData != None and 'type' in oData and oData['type'] == 'func':
                        # If the instance is not yet made, it needs to be saved before we continue
                        if instance.id == None:
                            instance.save()
                        # Make sure we change the type
                        instance.condtype = "func"
                        if instance.function != None:
                            # Remove the old function
                            instance.function.delete()
                        # Be sure to get the correct gateway
                        gateway = self.obj.gateway
                        # Create the functions for this CVAR
                        func_main = Function.create_from_list(oData['value'], gateway, "cvar", None, instance, None)
                        # Double check if a function has been made
                        if func_main == None:
                            # No function made - return error
                            self.arErr.append("For some reason no condition-function was created")
                            return False                              
                        else:
                            # At least save the function
                            func_main.save()
                            instance.function = func_main
                            instance.functiondef = func_main.functiondef
                            # Indicate that changes have been made
                            has_changed = True
                            # Check this function: its return type should be BOOLEAN
                            func_out_type = func_main.get_outputtype()
                            if not arg_matches('bool', func_out_type):
                                # Not a boolean output type -- pass on this message
                                self.arErr.append("The selected function has a <b>{}</b> output. What is needed for a condition is a <b>Boolean</b> output.".format(func_out_type))  

            # Retrieve the 'order' field
            if instance.order != form.cleaned_data['ORDER']:
                if form.cleaned_data['ORDER'] == None:
                    instance.order = len(instance.gateway.conditions.all())+1
                else:
                    instance.order = form.cleaned_data['ORDER']
                has_changed = True
        # Return the change-indicator to trigger saving
        return has_changed
          

class ResearchPart62(ResearchPart):
    MainModel = Condition
    template_name = 'seeker/research_part_62.html'
    # Provide a form that allows filling in the specifics of a function
    form_objects = [{'form': FunctionForm, 'prefix': 'function', 'readonly': False}]
    # The function that is provided contains a particular number of arguments
    ArgFormSet = inlineformset_factory(Function, Argument, 
                                          form=ArgumentForm, min_num=1, extra=0)
    formset_objects = [{'formsetClass': ArgFormSet,  'prefix': 'arg',  'readonly': False}]

    def get_instance(self, prefix):
        if prefix == 'function' or prefix == 'arg':
            # This returns the FUNCTION object we are linked to
            condition = self.obj
            if condition.function_id == None or condition.function == None:
                # Check the function definition
                if condition.functiondef_id == None:
                    # There is an error: we need to have a function definition here
                    return None
                # Make sure both changes are saved in one database-go
                with transaction.atomic():
                    # Create a new function 
                    function = Function.create(condition.functiondef, None, condition, None)
                    # Add a link to this function from the condition object
                    condition.function = function
                    # Make sure we save the condition object
                    condition.save()
            return condition.function

    def custom_init(self):
        """Make sure the ARGUMENT formset gets the correct number of arguments"""

        # Check if we have a Condition object
        if self.obj:
            # Check if the condition-object type is 'func' - functiewaarde
            if self.obj.condtype == "func":
                # Get the function definition
                functiondef = self.obj.functiondef
                if functiondef != None:
                    # Get the number of arguments
                    argnum = functiondef.argnum
                    # Adapt the minimum number of items in the argument formset
                    self.ArgFormSet = inlineformset_factory(Function, Argument, 
                                          form=ArgumentForm, min_num=argnum, extra=0)
                    self.formset_objects[0]['formsetClass'] = self.ArgFormSet

        return True

    def add_to_context(self, context):
        # Note: the self.obj is a Condition
        condition = self.obj

        if condition == None:
            currentowner = None
            context['research_id'] = None
            context['condition_this'] = None
            targettype = ""
        else:
            context['condition_this'] = condition
            gateway = condition.gateway
            currentowner = gateway.research.owner
            context['research_id'] = gateway.research.id
            targettype = gateway.research.targetType

            # Further action if this condition is of the 'func' type
            if condition.condtype == "func":

                # Need to specify the template for the function
                functiondef = condition.functiondef
                # Adapt the arguments for this form
                arg_defs = ArgumentDef.objects.filter(function=functiondef)

                # Calculate the initial queryset for 'gvar'
                qs_gvar = GlobalVariable.objects.filter(gateway=gateway)


                # Calculate the initial queryset for 'dvar'
                lstQ = []
                lstQ.append(Q(gateway=gateway))
                qs_dvar = VarDef.objects.filter(*lstQ).order_by('order')

                # - adapting the 'arg_formset' for the 'function' form (editable)
                fun_this = self.get_instance('function')
                if fun_this != None:
                    context['arg_formset'] = check_arguments(context['arg_formset'], functiondef, qs_gvar, None, qs_dvar, '63')
                else:
                    context['arg_formset'] = None

                # Get the list of functions for this condition here
                context['function_list'] = self.obj.get_functions()

        context['currentowner'] = currentowner
        # We also need to make the object_id available
        context['object_id'] = self.object_id
        context['targettype'] = targettype
        return context

    #def check_arguments(self, arg_formset, functiondef, qs_gvar, qs_cvar, qs_dvar, target):
    #    # Take the functiondef as available in this argument
    #    arg_defs = ArgumentDef.objects.filter(function=functiondef)

    #    for index, arg_form in enumerate(arg_formset):
    #        # Initialise the querysets
    #        arg_form.fields['gvar'].queryset = qs_gvar
    #        if qs_cvar != None: arg_form.fields['cvar'].queryset = qs_cvar
    #        arg_form.fields['dvar'].queryset = qs_dvar
    #        # Add the information required by 'seeker/function_args.html' for each argument
    #        arg_form.target = target
    #        arg_form.targetid = 'research_part_' + arg_form.target
    #        if arg_form.instance != None:
    #            arg_form.url_edit = reverse(arg_form.targetid, kwargs={'object_id': arg_form.instance.id})
    #        arg_form.url_new = reverse(arg_form.targetid)
    #        # Get the instance from this form
    #        arg = arg_form.save(commit=False)
    #        # Check if the argument definition is set
    #        if arg.id == None or arg.argumentdef_id == None or arg.argumentdef_id != arg_defs[index].id:
    #            # Get the argument definition for this particular argument
    #            arg.argumentdef = arg_defs[index]
    #            arg_form.initial['argumentdef'] = arg_defs[index]
    #            arg_form.initial['argtype'] = arg_defs[index].argtype
    #            # Preliminarily save
    #            arg_form.save(commit=False)

    #    # Return the adatpted formset
    #    return arg_formset

    def before_save(self, prefix, request, instance=None, form=None):
        has_changed = False
        # When we save an ARG, we need to add a link to the Function it belongs to
        if prefix == 'arg':
            # The instance is the argument instance

            # Check the argtype of this argument: is it NOT a function any more?
            if instance.argtype != "func":
                # Then REmove all functions pointing to me
                func_child = instance.functionparent.first()
                if func_child != None:
                    func_child.delete()
                    has_changed = True
            else:
                # THis is a function: check if the function definition has not changed
                func_child = instance.functionparent.first()
                if func_child == None or instance.functiondef != func_child.functiondef:
                    # The function definition changed >> replace the child with a completely NEW version
                    # [1] remove the child (as well as anything attached to it)
                    if func_child != None:
                        func_child.delete()
                    # [2] Create a new version
                    func_child = Function.create(instance.functiondef, None, instance.function.rootcond, instance)
                    # [3] Save it here (or that is done one level up)
                    func_child.save()
                    # Indicate changes have been made
                    has_changed = True
        elif prefix == 'function':
            if instance != None:
                if self.obj.function != instance:
                    # Link the function-instance to the  CVAR instance
                    self.obj.function = instance
                    # Save the adapted CVAR instance
                    self.obj.save()
                    # We have already saved the above, so 'has_changed' does not need to be touched
        # Save the related RESEARCH object
        # NOTE: should this be function.rootcond.gateway.research???
        self.obj.function.condition.gateway.research.save()
        # Return the change-indicator to trigger saving
        return has_changed


class ResearchPart63(ResearchPart):
    MainModel = Argument
    template_name = 'seeker/research_part_63.html'
    ArgFormSet = inlineformset_factory(Function, Argument, 
                                          form=ArgumentForm, min_num=1, extra=0)
    # Two forms:
    # - the 'parent' form is view-only and contains the argument we are supplying a function for
    # - the 'function' form is editable and contains the function for the argument 
    form_objects = [{'form': FunctionForm, 'prefix': 'parent', 'readonly': True},
                    {'form': FunctionForm, 'prefix': 'function', 'readonly': False}]
    # Two formsets:
    # - the 'arg'  formset belongs to the 'function' (see above)
    # - the 'parg' formset belongs to the 'parent'
    formset_objects = [{'formsetClass': ArgFormSet, 'prefix': 'arg', 'readonly': False},
                       {'formsetClass': ArgFormSet, 'prefix': 'parg', 'readonly': True}]

    def get_instance(self, prefix):
        # NOTE: For '63' the self.obj is an Argument!!

        if prefix == 'function' or prefix == 'arg':
            # This returns the EXISTING or NEW function object belonging to the argument
            qs = Function.objects.filter(parent=self.obj)
            # Does it exist?
            if qs.count() == 0:
                # It does not yet exist: create it
                # Make sure both changes are saved in one database-go
                with transaction.atomic():
                    # Create a new function 
                    function = Function(functiondef = self.obj.functiondef, 
                                        rootcond = self.obj.function.rootcond,
                                        parent = self.obj)
                    # Make sure the function instance gets saved
                    function.save()
            else:
                # It exists, so assign it
                function = qs[0]
            return function
        elif prefix == 'parent' or prefix == 'parg':
            # This returns the PARENT function object the argument belongs to
            return self.obj.function

    def custom_init(self):
        """Make sure the ARG formset gets the correct number of arguments"""

        # Check if we have a FUNCTION object
        fun_this = self.get_instance('function')
        if fun_this:
            # Get the function definition
            functiondef = fun_this.functiondef
            if functiondef != None:
                # Get the number of arguments
                argnum = functiondef.argnum
                # Adapt the minimum number of items in the argument formset
                self.ArgFormSet = inlineformset_factory(Function, Argument, 
                                        form=ArgumentForm, min_num=argnum, extra=0)
                # The 'arg' object is the one with index '0'
                self.formset_objects[0]['formsetClass'] = self.ArgFormSet

        return True

    def add_to_context(self, context):
        # NOTE: the instance (self.obj) for the '44' form is an ARGUMENT
        pfun_this = self.get_instance('parent')
        if pfun_this == None:
            currentowner = None
            context['research_id'] = None
            context['vardef_this'] = None
            targettype = ""
        else:
            # Get to the CONDITION instance
            cond = pfun_this.rootcond
            gateway = cond.gateway
            currentowner = gateway.research.owner
            context['research_id'] = gateway.research.id
            context['vardef_this'] = cond.variable
            context['cond_this'] = cond
            targettype = gateway.research.targetType

            # Since this is a '63' form, we know this is a calculation

            # Calculate the initial queryset for 'gvar'
            #   (These are the variables available for this GATEWAY)
            qs_gvar = GlobalVariable.objects.filter(gateway=gateway)

            # Calculate the initial queryset for 'dvar'
            lstQ = []
            lstQ.append(Q(gateway=gateway))
            qs_dvar = VarDef.objects.filter(*lstQ).order_by('order')

            # - adapting the 'parg_formset' for the 'parent' form (view-only)
            context['parg_formset'] = check_arguments(context['parg_formset'], pfun_this.functiondef, qs_gvar, None, qs_dvar, '63')

            # - adapting the 'arg_formset' for the 'function' form (editable)
            fun_this = self.get_instance('function')
            if fun_this != None:
                context['arg_formset'] = check_arguments(context['arg_formset'], fun_this.functiondef, qs_gvar, None, qs_dvar, '63')
            else:
                context['arg_formset'] = None

            # Get a list of all ancestors
            context['anc_list'] = fun_this.get_ancestors()

        context['currentowner'] = currentowner
        # We also need to make the object_id available -- but is this the correct one?
        context['object_id'] = self.object_id
        context['targettype'] = targettype
        return context

    def before_save(self, prefix, request, instance=None, form=None):
        has_changed = False
        # When we save an ARG, we need to add a link to the Function it belongs to
        if prefix == 'arg':
            # The instance is the argument instance

            # Check the argtype of this argument: is it NOT a function any more?
            if instance.argtype != "func":
                # Then REmove all functions pointing to me
                func_child = instance.functionparent.first()
                if func_child != None:
                    func_child.delete()
                    has_changed = True
            else:
                # THis is a function: check if the function definition has not changed
                func_child = instance.functionparent.first()
                if func_child == None or instance.functiondef != func_child.functiondef:
                    # The function definition changed >> replace the child with a completely NEW version
                    # [1] remove the child
                    if func_child != None:
                        func_child.delete()
                    # [2] Create a new version
                    func_child = Function.create(instance.functiondef, None, instance.function.rootcond, instance)
                    # [3] Save it here (or that is done one level up)
                    func_child.save()
                    # Indicate changes have been made
                    has_changed = True
        # Save the related RESEARCH object
        if has_changed:
            self.obj.function.rootcond.gateway.research.save()
        # Return the change-indicator to trigger saving
        return has_changed


class ResearchPart7(ResearchPart):
    template_name = 'seeker/research_part_7.html'
    MainModel = Research
    FeatFormSet = inlineformset_factory(Gateway, Feature, 
                                        form=FeatureForm, min_num=1, 
                                        extra=0, can_delete=True, can_order=True)
    formset_objects = [{'formsetClass': FeatFormSet, 'prefix': 'feat', 'readonly': False}]

    def custom_init(self):
        """If the dictionary contains a 'task' element, it should be processed"""

        oErr = ErrHandle()
        try:
            # Get the task
            sTask = "" if (not '_task' in self.qd) else self.qd['_task']
            if sTask == "copy_feature_function":
                # There are some more parameters we expect
                functionid = self.qd['_functionid']
                functionid = -1 if functionid == "" else int(functionid)
                featureid = self.qd['_featureid']
                featureid = -1 if featureid == "" else int(featureid)
                if functionid>=0 and featureid>=0:
                    # Find the function
                    function = Function.objects.filter(id=functionid).first()
                    feature = feature.objects.filter(id=featureid).first()
                    if function != None and feature != None:
                        # Copy the function to the construction
                        with transaction.atomic():
                            feature.function.delete()
                            feature.function = function.get_copy()
                            response = feature.save()
            else:
                # Perform some pre-fetching
                self.gateway = self.obj.gateway
                vardef_list = self.gateway.get_vardef_list()
                self.vardef_str_list = VarDef.get_restricted_vardef_list( vardef_list, 'str')
                self.fundef_str_list = FunctionDef.get_functions_with_type('str')
            # REturn positively
            return True
        except:
            oErr.DoError("custom_init error")
            return False

    def get_instance(self, prefix):
        if prefix == 'feat':
            # We need to have the gateway
            return self.obj.gateway

    def get_form_kwargs(self, prefix):
        if prefix == 'feat' and self.obj != None and self.obj.gateway != None:
            form_kwargs = {"gateway":  self.obj.gateway,
                           "vardef_str_list": self.vardef_str_list,
                           "fundef_str_list": self.fundef_str_list  }
            return form_kwargs
        else:
            return None

    def process_formset(self, prefix, request, formset):
        if prefix == 'feat':
            # Sorting: see ResearchPart4
            ordered_forms = sorted(formset.forms, key=lambda item: item.instance.order)
            # Make sure the initial values of the 'order' in the forms are set correctly
            for form in ordered_forms:
                form.fields['ORDER'].initial = form.instance.order
            return ordered_forms
        else:
            return None

    def add_to_context(self, context):
        # Note: the self.obj is the Research project
        if self.obj == None:
            currentowner = None
            targettype = ""
        else:
            # gateway = self.obj.gateway
            currentowner = self.obj.owner
            targettype = self.obj.targetType
            context['currentowner'] = currentowner
            context['targettype'] = targettype
        # Provide a number for the newest item 
        context['new_order_number'] = len(self.obj.gateway.features.all())
        # Return the context
        return context

    def before_save(self, prefix, request, instance=None, form=None):
        has_changed = False
        if prefix == 'feat':
            # Find the function attached to me - only if applicable!!
            if instance.feattype == "func" and instance.functiondef != None:
                # Two situations:
                # - THere is no function yet
                # - There is a function, but with the wrong functiondef
                if instance.function == None or (instance.function != None and instance.functiondef != instance.function.functiondef):
                    # If the instance is not yet made, it needs to be saved before we continue
                    if instance.id == None:
                        instance.save()
                    # Does a previous function exist?
                    if instance.function:
                        # Remove the existing function
                        instance.function.delete()
                    
                    # Create a new (obligatory) 'Function' instance, with accompanying Argument instances
                    instance.function = Function.create(instance.functiondef, None, None, None, instance)
                    # Indicate that changes have been made and saving of 'instance' is needed
                    has_changed = True
            elif instance.feattype == "json":
                # We are uploading a function: get the JSON contents
                if 'file_source' in form.cleaned_data:
                    data_file = form.cleaned_data["file_source"]
                    oData = import_data_file(data_file, self.arErr)
                    # Check the top layer
                    if oData != None and 'type' in oData and oData['type'] == 'func':
                        # If the instance is not yet made, it needs to be saved before we continue
                        if instance.id == None:
                            instance.save()
                        # Make sure we change the type
                        instance.feattype = "func"
                        if instance.function != None:
                            # Remove the old function
                            instance.function.delete()
                        # Be sure to get the correct gateway
                        gateway = self.obj.gateway
                        # Create the functions for this CVAR
                        func_main = Function.create_from_list(oData['value'], gateway, "feat", None, None, instance)
                        # Double check if a function has been made
                        if func_main == None:
                            # No function made - return error
                            self.arErr.append("For some reason no feature-function was created")
                            return False                              
                        else:
                            # At least save the function
                            func_main.save()
                            instance.function = func_main
                            instance.functiondef = func_main.functiondef
                            # Indicate that changes have been made
                            has_changed = True
                            # Check the output type
                            func_out_type = func_main.get_outputtype()
                            if not arg_matches('str', func_out_type):
                                self.arErr.append("The output type of the imported function should be string (or int or bool)")

            # Retrieve the 'order' field
            if instance.order != form.cleaned_data['ORDER']:
                if form.cleaned_data['ORDER'] == None:
                    instance.order = len(instance.gateway.features.all())+1
                else:
                    instance.order = form.cleaned_data['ORDER']
                has_changed = True
        # Return the change-indicator to trigger saving
        return has_changed
          

class ResearchPart72(ResearchPart):
    MainModel = Feature
    template_name = 'seeker/research_part_72.html'
    # Provide a form that allows filling in the specifics of a function
    form_objects = [{'form': FunctionForm, 'prefix': 'function', 'readonly': False}]
    # The function that is provided contains a particular number of arguments
    ArgFormSet = inlineformset_factory(Function, Argument, 
                                          form=ArgumentForm, min_num=1, extra=0)
    formset_objects = [{'formsetClass': ArgFormSet,  'prefix': 'arg',  'readonly': False}]

    def get_instance(self, prefix):
        if prefix == 'function' or prefix == 'arg':
            # This returns the FUNCTION object we are linked to
            feature = self.obj
            if feature.function_id == None or feature.function == None:
                # Check the function definition
                if feature.functiondef_id == None:
                    # There is an error: we need to have a function definition here
                    return None
                # Make sure both changes are saved in one database-go
                with transaction.atomic():
                    # Create a new function 
                    function = Function.create(feature.functiondef, None, None, None, feature)
                    # Add a link to this function from the feature object
                    feature.function = function
                    # Make sure we save the feature object
                    feature.save()
            return feature.function

    def custom_init(self):
        """Make sure the ARGUMENT formset gets the correct number of arguments"""

        # Check if we have a Feature object
        if self.obj:
            # Check if the feature-object type is 'func' - functiewaarde
            if self.obj.feattype == "func":
                # Get the function definition
                functiondef = self.obj.functiondef
                if functiondef != None:
                    # Get the number of arguments
                    argnum = functiondef.argnum
                    # Adapt the minimum number of items in the argument formset
                    self.ArgFormSet = inlineformset_factory(Function, Argument, 
                                          form=ArgumentForm, min_num=argnum, extra=0)
                    self.formset_objects[0]['formsetClass'] = self.ArgFormSet

        return True

    def add_to_context(self, context):
        # Note: the self.obj is a Feature
        feature = self.obj

        if feature == None:
            currentowner = None
            context['research_id'] = None
            context['feature_this'] = None
            targettype = ""
        else:
            context['feature_this'] = feature
            gateway = feature.gateway
            currentowner = gateway.research.owner
            context['research_id'] = gateway.research.id
            targettype = gateway.research.targetType

            # Further action if this feature is of the 'func' type
            if feature.feattype == "func":

                # Need to specify the template for the function
                functiondef = feature.functiondef
                # Adapt the arguments for this form
                arg_defs = ArgumentDef.objects.filter(function=functiondef)

                # Calculate the initial queryset for 'gvar'
                qs_gvar = GlobalVariable.objects.filter(gateway=gateway)


                # Calculate the initial queryset for 'dvar'
                lstQ = []
                lstQ.append(Q(gateway=gateway))
                qs_dvar = VarDef.objects.filter(*lstQ).order_by('order')

                # - adapting the 'arg_formset' for the 'function' form (editable)
                fun_this = self.get_instance('function')
                if fun_this != None:
                    context['arg_formset'] = check_arguments(context['arg_formset'], functiondef, qs_gvar, None, qs_dvar, '73')
                else:
                    context['arg_formset'] = None

                # Get the list of functions for this feature here
                context['function_list'] = self.obj.get_functions()

        context['currentowner'] = currentowner
        # We also need to make the object_id available
        context['object_id'] = self.object_id
        context['targettype'] = targettype
        return context

    def before_save(self, prefix, request, instance=None, form=None):
        has_changed = False
        # When we save an ARG, we need to add a link to the Function it belongs to
        if prefix == 'arg':
            # The instance is the argument instance

            # Check the argtype of this argument: is it NOT a function any more?
            if instance.argtype != "func":
                # Then REmove all functions pointing to me
                func_child = instance.functionparent.first()
                if func_child != None:
                    func_child.delete()
                    has_changed = True
            else:
                # THis is a function: check if the function definition has not changed
                func_child = instance.functionparent.first()
                if func_child == None or instance.functiondef != func_child.functiondef:
                    # The function definition changed >> replace the child with a completely NEW version
                    # [1] remove the child (as well as anything attached to it)
                    if func_child != None:
                        func_child.delete()
                    # [2] Create a new version
                    func_child = Function.create(instance.functiondef, None, None, instance, instance.function.rootfeat)
                    # [3] Save it here (or that is done one level up)
                    func_child.save()
                    # Indicate changes have been made
                    has_changed = True
        elif prefix == 'function':
            if instance != None:
                if self.obj.function != instance:
                    # Link the function-instance to the  CVAR instance
                    self.obj.function = instance
                    # Save the adapted CVAR instance
                    self.obj.save()
                    # We have already saved the above, so 'has_changed' does not need to be touched
        # Save the related RESEARCH object
        self.obj.function.feature.gateway.research.save()
        # Return the change-indicator to trigger saving
        return has_changed


class ResearchPart73(ResearchPart):
    MainModel = Argument
    template_name = 'seeker/research_part_73.html'
    ArgFormSet = inlineformset_factory(Function, Argument, 
                                          form=ArgumentForm, min_num=1, extra=0)
    # Two forms:
    # - the 'parent' form is view-only and contains the argument we are supplying a function for
    # - the 'function' form is editable and contains the function for the argument 
    form_objects = [{'form': FunctionForm, 'prefix': 'parent', 'readonly': True},
                    {'form': FunctionForm, 'prefix': 'function', 'readonly': False}]
    # Two formsets:
    # - the 'arg'  formset belongs to the 'function' (see above)
    # - the 'parg' formset belongs to the 'parent'
    formset_objects = [{'formsetClass': ArgFormSet, 'prefix': 'arg', 'readonly': False},
                       {'formsetClass': ArgFormSet, 'prefix': 'parg', 'readonly': True}]

    def get_instance(self, prefix):
        # NOTE: For '73' the self.obj is an Argument!!

        if prefix == 'function' or prefix == 'arg':
            # This returns the EXISTING or NEW function object belonging to the argument
            qs = Function.objects.filter(parent=self.obj)
            # Does it exist?
            if qs.count() == 0:
                # It does not yet exist: create it
                # Make sure both changes are saved in one database-go
                with transaction.atomic():
                    # Create a new function 
                    function = Function(functiondef = self.obj.functiondef, 
                                        rootfeat = self.obj.function.rootfeat,
                                        parent = self.obj)
                    # Make sure the function instance gets saved
                    function.save()
            else:
                # It exists, so assign it
                function = qs[0]
            return function
        elif prefix == 'parent' or prefix == 'parg':
            # This returns the PARENT function object the argument belongs to
            return self.obj.function

    def custom_init(self):
        """Make sure the ARG formset gets the correct number of arguments"""

        # Check if we have a FUNCTION object
        fun_this = self.get_instance('function')
        if fun_this:
            # Get the function definition
            functiondef = fun_this.functiondef
            if functiondef != None:
                # Get the number of arguments
                argnum = functiondef.argnum
                # Adapt the minimum number of items in the argument formset
                self.ArgFormSet = inlineformset_factory(Function, Argument, 
                                        form=ArgumentForm, min_num=argnum, extra=0)
                # The 'arg' object is the one with index '0'
                self.formset_objects[0]['formsetClass'] = self.ArgFormSet

        return True

    def add_to_context(self, context):
        # NOTE: the instance (self.obj) for the '44' form is an ARGUMENT
        pfun_this = self.get_instance('parent')
        if pfun_this == None:
            currentowner = None
            context['research_id'] = None
            context['vardef_this'] = None
            targettype = ""
        else:
            # Get to the FEATURE instance
            feat = pfun_this.rootfeat
            gateway = feat.gateway
            currentowner = gateway.research.owner
            context['research_id'] = gateway.research.id
            context['vardef_this'] = feat.variable
            context['feat_this'] = feat
            targettype = gateway.research.targetType

            # Since this is a '73' form, we know this is a calculation

            # Calculate the initial queryset for 'gvar'
            #   (These are the variables available for this GATEWAY)
            qs_gvar = GlobalVariable.objects.filter(gateway=gateway)

            # Calculate the initial queryset for 'dvar'
            lstQ = []
            lstQ.append(Q(gateway=gateway))
            qs_dvar = VarDef.objects.filter(*lstQ).order_by('order')

            # - adapting the 'parg_formset' for the 'parent' form (view-only)
            context['parg_formset'] = check_arguments(context['parg_formset'], pfun_this.functiondef, qs_gvar, None, qs_dvar, '73')

            # - adapting the 'arg_formset' for the 'function' form (editable)
            fun_this = self.get_instance('function')
            if fun_this != None:
                context['arg_formset'] = check_arguments(context['arg_formset'], fun_this.functiondef, qs_gvar, None, qs_dvar, '73')
            else:
                context['arg_formset'] = None

            # Get a list of all ancestors
            context['anc_list'] = fun_this.get_ancestors()

        context['currentowner'] = currentowner
        # We also need to make the object_id available -- but is this the correct one?
        context['object_id'] = self.object_id
        context['targettype'] = targettype
        return context

    def before_save(self, prefix, request, instance=None, form=None):
        has_changed = False
        # When we save an ARG, we need to add a link to the Function it belongs to
        if prefix == 'arg':
            # The instance is the argument instance

            # Check the argtype of this argument: is it NOT a function any more?
            if instance.argtype != "func":
                # Then REmove all functions pointing to me
                func_child = instance.functionparent.first()
                if func_child != None:
                    func_child.delete()
                    has_changed = True
            else:
                # THis is a function: check if the function definition has not changed
                func_child = instance.functionparent.first()
                if func_child == None or instance.functiondef != func_child.functiondef:
                    # The function definition changed >> replace the child with a completely NEW version
                    # [1] remove the child
                    if func_child != None:
                        func_child.delete()
                    # [2] Create a new version
                    func_child = Function.create(instance.functiondef, None, None, instance, instance.function.rootfeat)
                    # [3] Save it here (or that is done one level up)
                    func_child.save()
                    # Indicate changes have been made
                    has_changed = True
        # Save the related RESEARCH object
        if has_changed:
            self.obj.function.rootfeat.gateway.research.save()
        # Return the change-indicator to trigger saving
        return has_changed


class ResearchPart8(ResearchPart):
    """Generic class to treat any function at any level, under Condition, Feature of Cvar"""

    MainModel = Function
    template_name = 'seeker/research_part_8.html'
    ArgFormSet = inlineformset_factory(Function, Argument, 
                                          form=ArgumentForm, min_num=1, extra=0)
    # One forms:
    # - the 'function' form is editable and contains the function for the argument 
    form_objects = [{'form': FunctionForm, 'prefix': 'function', 'readonly': False}]

    # One formsets:
    # - the 'arg'  formset belongs to the 'function' (see above)
    formset_objects = [{'formsetClass': ArgFormSet, 'prefix': 'arg', 'readonly': False}]

    def get_instance(self, prefix):
        # NOTE: For '8' the self.obj is a Function

        if prefix == 'function' or prefix == 'arg':
            return self.obj

    def custom_init(self):
        """Make sure the ARG formset gets the correct number of arguments"""

        # Check if we have a FUNCTION object
        fun_this = self.get_instance('function')
        if fun_this:
            # Get the function definition
            functiondef = fun_this.functiondef
            if functiondef != None:
                # Get the number of arguments
                argnum = functiondef.argnum
                # Adapt the minimum number of items in the argument formset
                self.ArgFormSet = inlineformset_factory(Function, Argument, 
                                        form=ArgumentForm, min_num=argnum, extra=0)
                # The 'arg' object is the one with index '0'
                self.formset_objects[0]['formsetClass'] = self.ArgFormSet

        return True

    def add_to_context(self, context):
        # NOTE: the instance (self.obj) for the '8' form is a FUNCTION
        fun_this = self.get_instance('function')
        if fun_this == None:
            currentowner = None
            context['research_id'] = None
            targettype = ""
        else:
            # Get to the GATEWAY instance
            if fun_this.root:
                gateway = fun_this.root.construction.gateway
            elif fun_this.rootfeat:
                gateway = fun_this.rootfeat.gateway
            elif fun_this.rootcond:
                gateway = fun_this.rootcond.gateway
            else:
                gateway = None
            currentowner = gateway.research.owner
            context['research_id'] = gateway.research.id
            targettype = gateway.research.targetType

            # Calculate the initial queryset for 'gvar'
            #   (These are the variables available for this GATEWAY)
            qs_gvar = GlobalVariable.objects.filter(gateway=gateway)

            # Calculate the initial queryset for 'dvar'
            lstQ = []
            lstQ.append(Q(gateway=gateway))
            if fun_this.root:
                # Need to restrict to what lies before me
                lstQ.append(Q(order__lt=fun_this.root.variable.order))
            else:
                # For Conditions and Features we may make use of *ALL* dvars
                pass
            qs_dvar = VarDef.objects.filter(*lstQ).order_by('order')

            # - adapting the 'arg_formset' for the 'function' form (editable)
            fun_this = self.get_instance('function')
            if fun_this != None:
                context['arg_formset'] = check_arguments(context['arg_formset'], fun_this.functiondef, qs_gvar, None, qs_dvar, '8')
            else:
                context['arg_formset'] = None

            # Indicate that EDIT buttons for the ARGUMENTS need to be hidden
            context['hide_buttons'] = "yes"

        context['currentowner'] = currentowner
        # We also need to make the object_id available -- but is this the correct one?
        context['object_id'] = self.object_id
        context['targettype'] = targettype
        return context

    def before_save(self, prefix, request, instance=None, form=None):
        has_changed = False

        # Find the rootobject (root, rootcond, rootfeat)
        func_this = instance
        gateway = None

        if prefix == "func":
            if func_this.root != None:
                gateway = func_this.root.construction.gateway
            elif func.rootfeat != None:
                gateway = func_this.rootfeat.gateway
            elif func.rootcond != None:
                gateway = func_this.rootcond.gateway

        elif prefix == 'arg':
            # When we save an ARG, we need to add a link to the Function it belongs to
            # The instance is the argument instance

            # Check the argtype of this argument: is it NOT a function any more?
            if instance.argtype != "func":
                # Then REmove all functions pointing to me
                func_child = instance.functionparent.first()
                if func_child != None:
                    func_child.delete()
                    has_changed = True
            else:
                # THis is a function: check if the function definition has not changed
                func_child = instance.functionparent.first()
                if func_child == None or instance.functiondef != func_child.functiondef:
                    # The function definition changed >> replace the child with a completely NEW version
                    # [1] remove the child
                    if func_child != None:
                        func_child.delete()
                    # [2] Create a new version
                    func_child = Function.create(instance.functiondef, instance.function.root, instance.function.rootcond, instance, instance.function.rootfeat)
                    # [3] Save it here (or that is done one level up)
                    func_child.save()
                    # Indicate changes have been made
                    has_changed = True

        # Save the related RESEARCH object
        if has_changed and gateway != None:
            gateway.research.save()
        # Return the change-indicator to trigger saving
        return has_changed


class ObjectCopyMixin:
    model = None
    data = {'status': 'ok', 'html': ''}       # Create data to be returned    

    def post(self, request, object_id=None):
        # Copying is only possible through a POST request
        obj = get_object_or_404(self.model, id=object_id)
        # Create a copy of the object
        kwargs = {'currentuser': request.user}
        copy_obj = obj.get_copy(**kwargs)
        sMsg = ""
        # Check result
        if copy_obj == None:
            sMsg = "There was a problem copying the object"
            self.data['html'] = sMsg
            self.data['status'] = "error" 
        else:
            # Save the new object
            copy_obj.save()
        # Return the information
        return JsonResponse(self.data)


class ResearchCopy(ObjectCopyMixin, View):
    """Copy one 'Research' object"""
    model = Research
    success_url = reverse_lazy('seeker_list')


class ObjectDeleteMixin:
    model = None
    data = {'status': 'ok', 'html': ''}       # Create data to be returned    

    def post(self, request, object_id=None):
        # Note: deleting is only possible through a POST request
        obj = get_object_or_404(self.model, id=object_id)
        try:
            obj.delete()
        except:
            lInfo = sys.exc_info()
            if len(lInfo) == 1:
                sMsg = lInfo[0]
            else:
                sMsg = "{}<br>{}".format(lInfo[0], lInfo[1])
            self.data['html'] = sMsg
            self.data['status'] = "error" 
        # Return the information
        return JsonResponse(self.data)


class ResearchDelete(ObjectDeleteMixin, View):
    """Delete one 'Research' object"""
    model = Research


class ResultDelete(ObjectDeleteMixin, View):
    """Delete one 'Basket' object and its associates"""

    model = Basket


class ResearchResultDetail(View):
    """Show details (all) of one result"""
    model = Basket
    template_name = "seeker/research_result.html"

    def get(self, request, object_id=None):
        # Initialize the context
        context = dict()
        # Fetch the basket_id
        obj = get_object_or_404(self.model, id=object_id)
        if obj != None:
            # Provide search URL and search name
            context['search_edit_url'] = reverse("seeker_edit", kwargs={"object_id": obj.research.id})
            context['search_name'] = obj.research.name
            context['original'] = obj.research
            context['intro_message'] = "Research project: <b>{}</b>".format(obj.research.name)
            # Get to the result object, a quantor
            quantor = Quantor.objects.filter(basket=obj).first()
            if quantor != None:
                context['quantor'] = quantor

        # Return the html page to be shown
        return render( request, self.template_name, context)


class KwicView(DetailView):
    """The main view that hosts the KWIC listview"""

    model = Basket
    template_name = 'seeker/kwic_main.html'
    basket = None
    kwic = None
    qc = 1        # Selected QC

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(KwicView, self).get_context_data(**kwargs)

        # Get parameters for the KWIC ordering and filtering
        initial = self.request.GET

        # EITHER: Pass on the form
        # context['form'] = self.form_class(instance=self.object)
        # OR: Specify the search form
        # search_form = KwicSearchForm(initial)
        # context['searchform'] = search_form

        status = ""
        if 'status' in initial:
            status = initial['status']
        context['status'] = status

        # Process the chosen QC (if any)
        if 'qc' in initial:
            self.qc = initial['qc']
        context['qc'] = self.qc

        # Add some elements to the context
        self.basket = self.object
        context['basket'] = self.basket
        context['object_id'] = self.basket.id
        research = self.basket.research
        context['search_edit_url'] = reverse("seeker_edit", kwargs={"object_id": research.id})
        context['search_name'] = research.name
        context['original'] = research
        context['intro_message'] = "Research project: <b>{}</b>".format(research.name)
        qc_list = Kwic.objects.filter(basket=self.basket).order_by("qc")
        if qc_list.count() == 0:
            # Create KWIC objects
            self.basket.create_kwic_objects()
        context['qc_list'] = Kwic.objects.filter(basket=self.basket).order_by("qc")

        # Return what we have
        return context


class KwicListView(View):
    """This listview manages showing the results that are gathered using /crpp/dbinfo"""

    model = Kwic
    arErr = []              # errors   
    template_name = 'seeker/kwic_list.html'
    page_function = "ru.cesar.seeker.kwic_page"
    form_div = "kwic_search_button"
    paginate_by = paginateEntries
    entrycount = 0
    basket = None       # Object: Basket
    qcline = 0          # Number
    page_obj = None     # Initializing the page object
    qs = None
    arFilter = [{'field': 'TextId', 'id': 'filter-text'},
                {'field': 'Cat', 'id': 'filter-cat'},
                {'field': 'SubType', 'id': 'filter-subtype'},
                {'field': 'Title', 'id': 'filter-title'},
                {'field': 'Genre', 'id': 'filter-genre'},
                {'field': 'Author', 'id': 'filter-author'},
                {'field': 'Date', 'id': 'filter-date'}]
    data = {'status': 'ok', 'html': '', 'statuscode': ''}       # Create data to be returned   

    def initialize(self, request):
        # Get the parameters
        if request.POST:
            self.qd = request.POST
        else:
            self.qd = request.GET

    def post(self, request, object_id=None):
        self.initialize(request)
        if self.checkAuthentication(request):
            # Define the context
            sStatusCode = "none"
            context = dict(object_id = object_id, statuscode=sStatusCode)
            # Get the BASKET object
            self.basket = get_object_or_404(Basket, id=object_id)
            # Get relevant information from the query dictionary
            qcNumber = self.qd.get('selected_qc')
            page = self.qd.get('page')
            if page == None:
                page = 1
            else:
                page = int(page)
            kwic_object = self.basket.kwiclines.filter(qc=qcNumber).first()

            # Filtering
            kwic_object.kwicfilters.all().delete()
            for item in self.arFilter:
                field = item['field']
                id = item['id']
                if self.qd.get(id) != "":
                    kwic_object.add_filter(field, self.qd.get(id))
            # Get the filter myself
            oFilter = kwic_object.get_filter()
            # Apply the filter
            kwic_object.apply_filter(oFilter)


            # Pagination
            hit_count = kwic_object.hitcount
            result_number_list = list(range(1, hit_count))
            paginator = Paginator(result_number_list, self.paginate_by)
            self.page_obj = paginator.page(page)

            # Set the number of items to be fetched
            iCount = self.paginate_by
            iStart = (page-1) * self.paginate_by
            # Fetch the data for this page
            oData = crpp_dbinfo(self.basket.research.owner.username, 
                                self.basket.research.name,
                                qcNumber, 
                                iStart,
                                iCount,
                                filter=oFilter,
                                sPart=self.basket.part.dir)
            if oData['commandstatus'] == "ok" and oData['status']['code'] == "completed":
                # Provide all the information needed to create the Html presentation of the data
                context['result_list'] = oData['Results']
                context['feature_list'] = oData['Features']

                # Put all the available filtering lists in here
                context['list_textid'] = []
                context['list_cat'] = []
                context['list_subtype'] = []
                context['list_title'] = []
                context['list_genre'] = []
                context['list_author'] = []

                # A list of filters that the user has specified in the past and saved
                context['list_filter'] = []

                # Add pagination information
                context['page_obj'] = self.page_obj
                context['page_function'] = self.page_function
                context['hitcount'] = hit_count

                # Information on this basket/kwic object
                context['formdiv'] = self.form_div

                # Make sure we have a list of any errors
                error_list = [str(item) for item in self.arErr]
                self.data['error_list'] = error_list
                self.data['errors'] = self.arErr
                if len(self.arErr) >0:
                    self.data['status'] = "error"

                # Add items to context
                context['error_list'] = error_list
                context['status'] = self.data['status']

                # Get the HTML response
                self.data['html'] = render_to_string(self.template_name, context, request)

            else:
                # Some error has occurred
                self.data['status'] = "error"
        else:
            self.data['html'] = "Please log in before continuing"

        # Return the information
        return JsonResponse(self.data)
      
    def get(self, request, object_id=None):
        self.initialize(request)
        # Fetch the basket_id
        self.basket = get_object_or_404(Basket, id=object_id)
        return super(KwicListView,self).get(self, request)

    def checkAuthentication(self,request):
        # first check for authentication
        if not request.user.is_authenticated:
            # Simply redirect to the home page
            self.data['html'] = "Please log in to work on a research project"
            return False
        else:
            return True

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(KwicListView, self).get_context_data(**kwargs)

        # Get parameters for the search
        initial = self.request.GET

        # Add some elements to the context
        research = self.basket.research
        context['search_edit_url'] = reverse("seeker_edit", kwargs={"object_id": research.id})
        context['search_name'] = research.name
        context['original'] = research
        context['intro_message'] = "Research project: <b>{}</b>".format(research.name)
        context['entrycount'] = self.entrycount
        context['basket'] = self.basket
        quantor = Quantor.objects.filter(basket=self.basket).first()
        context['qc_list'] = [idx for idx in range(1,quantor.qcNum+1)]

        # Return the calculated context
        return context

    def get_queryset(self):
        """Determine the queryset that is to be used"""

        # Get the parameters passed on with the GET request
        get = self.request.GET

        # Determine the selection
        lstQ = []
        lstQ.append(Q(basket=self.basket))
        if self.qcline > 0:
          lstQ.append(Q(qc=self.qcline))

        # Set and order the selection
        self.qs = Kwic.objects.filter(*lstQ)

        self.entrycount = self.qs.count()

        # Return the resulting filtered and sorted queryset
        return self.qs


class ResultDetailView(DetailView):
    """Show the results connected to one basket/quantor"""

    model = Basket
    template_name = 'seeker/result_details.html'
    basket = None     # Object: basket
    quantor = None    # Object: quantor


    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(ResultDetailView, self).get_context_data(**kwargs)

        # Get parameters for the search
        initial = self.request.GET

        # Find out where we are 
        self.basket = self.object
        self.quantor = Quantor.objects.filter(basket=self.basket).first()

        # Add some elements to the context
        research = self.basket.research
        context['object_id'] = self.object.id
        context['original'] = research
        context['quantor'] = self.quantor
        context['filters'] = self.basket.get_filters()
        # Check if this is the simple search
        if research.stype == STYPE_SIMPLE or research.name == SIMPLENAME:
            # This is the simple research project - special treatment
            context['search_edit_url'] = reverse('simple_details')
            context['search_name'] = "simple"
            context['intro_message'] = "Simple search"
        else:
            context['search_edit_url'] = reverse("seeker_edit", kwargs={"object_id": research.id})
            context['search_name'] = research.name
            context['intro_message'] = "Research project: <b>{}</b>".format(research.name)
        # BEWARE: perhaps quantor is empty?
        if self.quantor == None:
            context['qclines'] = []
        else:
            context['qclines'] = self.quantor.qclines.all()

        # Return the calculated context
        return context


class QuantorListView(ListView):
    """Show the search results connected to one quantor << Basket << Research"""

    model = Qsubinfo    # Each element consists of QsubInfo
    template_name = 'seeker/quantor_list.html'
    paginate_by = paginateEntries
    entrycount = 0
    qcTarget = 1
    basket = None       # Object: Basket
    quantor = None      # Object: Quantor
    qcline = None       # Object: QCline
    qs = None

    def get_qs(self):
        """Get the Texts that are selected"""
        if self.qs == None:
            # Get the Lemma PKs
            qs = self.get_queryset()
        else:
            qs = self.qs
        return qs

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super(QuantorListView, self).get_context_data(**kwargs)

        # Get parameters for the search
        initial = self.request.GET

        # Specify the search form
        search_form = QuantorSearchForm(initial)
        context['searchform'] = search_form

        # Add some elements to the context
        research = self.basket.research
        context['search_edit_url'] = reverse("seeker_edit", kwargs={"object_id": research.id})
        context['search_name'] = research.name
        context['original'] = research
        context['intro_message'] = "Research project: <b>{}</b>".format(research.name)
        context['quantor'] = self.quantor
        context['entrycount'] = self.entrycount
        context['qsubcats'] = self.qcline.qsubcats.all()

        # Return the calculated context
        return context

    def get(self, request, object_id=None):
        # Fetch the basket_id
        self.basket = get_object_or_404(Basket, id=object_id)
        return super(QuantorListView,self).get(self, request)

    def get_queryset(self):
        """Determine the Qsubinfo queryset"""

        # Get the parameters passed on with the GET request
        get = self.request.GET

        # Find out where we are 
        self.quantor = Quantor.objects.filter(basket=self.basket).first()
        self.qcline = self.quantor.qclines.filter(qc=self.qcTarget).first()

        # Determine the selection
        lstQ = []
        lstQ.append(Q(subcat__qcline=self.qcline))

        # Filter on subcat
        if 'subcategory' in get and get['subcategory'] != '':
            val = get['subcategory']
            # The subcat value is an id
            lstQ.append(Q(subcat=val))

        # Filter on text Name
        if 'textname' in get and get['textname'] != '':
            # Allow simple wildcard search
            val = adapt_search(get['textname'])
            lstQ.append(Q(text__fileName__iregex=val))

        # Filter on the minimum number of hits
        if 'minhits' in get and get['minhits'] != '':
            # Get the minimum number of hits
            val = get['minhits']
            try:
                iVal = int(val)-1
                lstQ.append(Q(count__gt=iVal))
            except:
                iDoNothing = 1

        # Set and order the selection
        self.qs = Qsubinfo.objects.filter(*lstQ).distinct().select_related().order_by(
            Lower('subcat__name'),
            Lower('text__fileName'))

        self.entrycount = self.qs.count()

        # Return the resulting filtered and sorted queryset
        return self.qs


class ResultPart1(ResearchPart):
    MainModel = Basket
    template_name = 'seeker/result_part_1.html'
    page_function = "ru.cesar.seeker.kwic_page"
    form_div = "result_part1_button"
    paginate_by = paginateEntries
    entrycount = 0
    qcTarget = 1
    bHideEmpty = False  # Hide empty documents
    basket = None       # Object: Basket
    quantor = None      # Object: Quantor
    qcline = None       # Object: QCline
    qs = None
    form_objects = []
    
    def custom_init(self):
        """Calculate stuff"""

        # Determine the QC line that has been passed on
        qc = self.qd['qc_select']
        if qc != None and qc != '':
            self.qcTarget = int(qc)
        # Set the basket
        self.basket = self.obj
        # Need to calculate the queryset already
        self.get_queryset('quantor')

        # Paging...
        page = self.qd.get('page')
        page = 1 if page == None else int(page)
        # Create a list [page_obj] that contains just these results
        paginator = Paginator(self.qs, self.paginate_by)
        self.page_obj = paginator.page(page)

    def add_to_context(self, context):
        # Specify the search form
        search_form = QuantorSearchForm(self.qd)
        context['searchform'] = search_form

        # Add some elements to the context
        context['quantor'] = self.quantor
        context['entrycount'] = self.entrycount
        context['qsubcats'] = self.qcline.qsubcats.all()
        context['filter_list'] = self.basket.get_filter_list(self.qcTarget)

        # Add pagination information
        context['object_list'] = self.page_obj
        context['page_obj'] = self.page_obj
        context['page_function'] = self.page_function
        context['hitcount'] = self.entrycount

        # Other required information
        context['formdiv'] = self.form_div

        # Return the calculated context
        return context

    def get_queryset(self, prefix):
        """Determine the Qsubinfo queryset"""

        # The parameters (POST or GET) are available through self.qd

        # Find out where we are 
        self.quantor = Quantor.objects.filter(basket=self.basket).first()
        self.qcline = self.quantor.qclines.filter(qc=self.qcTarget).first()

        # Determine the selection
        lstQ = []
        lstQ.append(Q(subcat__qcline=self.qcline))

        # Filter on subcat
        if 'subcategory' in self.qd and self.qd['subcategory'] != '':
            val = self.qd['subcategory']
            # The subcat value is an id
            lstQ.append(Q(subcat=val))

        # Filter on text Name
        if 'textname' in self.qd and self.qd['textname'] != '':
            # Allow simple wildcard search
            val = adapt_search(self.qd['textname'])
            lstQ.append(Q(text__fileName__iregex=val))

        # Filter on the minimum number of hits
        if 'minhits' in self.qd and self.qd['minhits'] != '':
            # Get the minimum number of hits
            val = self.qd['minhits']
            try:
                iVal = int(val)-1
                lstQ.append(Q(count__gt=iVal))
            except:
                iDoNothing = 1
        # Filter on zero hits
        elif 'hide_empty' in self.qd and self.qd['hide_empty'] == 'true':
            # make sure the count is larger than zero
            lstQ.append(Q(count__gt=0))

        # Set and order the selection
        self.qs = Qsubinfo.objects.filter(*lstQ).distinct().select_related().order_by(
            Lower('subcat__name'),
            Lower('text__fileName'))

        self.entrycount = self.qs.count()

        # Return the resulting filtered and sorted queryset
        return self.qs


class ResultPart14(ResearchPart):
    """Fetch all hits within a particular document"""
    MainModel = Qsubinfo
    template_name = 'seeker/result_part_14.html'
    kwic_object = None
    result_list = []
    feature_list = []
    form_objects = [{'form': KwicForm, 'prefix': 'kwic', 'readonly': True}]  

    def custom_init(self):
        """Calculate stuff"""

        # Determine the QC line that has been passed on
        if 'qc_select'in self.qd:
            qc = self.qd['qc_select']
        else:
            # Take a default value just in case
            qc = "1"
        if qc != None and qc != '':
            self.qcline = int(qc)

        # Get the Qsubinfo
        qsubinfo = self.obj
        # This leads to subcat and text
        self.text = qsubinfo.text
        self.subcat = qsubinfo.subcat.name
        # Get to the basket
        self.basket = qsubinfo.subcat.qcline.quantor.basket

        self.kwic_object = self.basket.kwiclines.filter(qc=self.qcline).first()
        if self.kwic_object == None:
            # there is no KWIC object for this one yet: create the KWIC objects
            self.basket.set_kwic(self.qcline)
            self.kwic_object = self.basket.kwiclines.filter(qc=self.qcline).first()

        # Validate
        if self.kwic_object == None:
            # Some error has occurred
            self.data['status'] = "error"
        else:
            # Create a filter for the selection
            # oFilter = dict(TextId=self.text.fileName+".gz")
            oFilter = { 'or_list': [{'TextId': self.text.fileName}, {'TextId': self.text.fileName+".gz"}]}
            iStart = 0
            iCount = 100
            page = 1
            # Fetch the data for this page
            oData = crpp_dbinfo(self.basket.research.owner.username, 
                                self.basket.research.name,
                                qc, 
                                iStart,
                                iCount,
                                filter=oFilter,
                                sPart=self.basket.part.dir)
            if oData['commandstatus'] == "ok" and oData['code'] == "completed":
                self.result_list = oData['Results']
                self.feature_list = oData['Features']
                # Also add the results as objects
                self.kwic_object.add_result_list(oData['Results'], page)
            else:
                # Some error has occurred
                self.data['status'] = "error"

    def add_to_context(self, context):
        # Add some elements to the context
        if self.kwic_object == None:
            context['kwic_id'] = None
        else:
            context['kwic_id'] = self.kwic_object.id
        # Provide all the information needed to create the Html presentation of the data
        context['result_list'] = self.result_list
        context['feature_list'] = self.feature_list


        # Return the adapted context
        return context


class ResultPart2(ResearchPart):
    MainModel = Basket
    template_name = 'seeker/result_part_2.html'
    page_function = "ru.cesar.seeker.kwic_page"
    form_div = "result_part2_button"
    paginate_by = paginateEntries
    kwic_object = None
    result_list = []
    feature_list = []
    form_objects = [{'form': KwicForm, 'prefix': 'kwic', 'readonly': True}]  

    def custom_init(self):
        """Calculate stuff"""

        # Determine the QC line that has been passed on
        if 'qc_select'in self.qd:
            qc = self.qd['qc_select']
        else:
            # Take a default value just in case
            qc = "1"
        if qc != None and qc != '':
            self.qcline = int(qc)
        # Set the basket
        self.basket = self.obj
        # Need to get the queryset already
        self.get_queryset('kwic')
        # Paging...
        page = self.qd.get('page')
        if page == None:
            page = 1
        else:
            page = int(page)
        self.kwic_object = self.basket.kwiclines.filter(qc=self.qcline).first()
        if self.kwic_object == None:
            # there is no KWIC object for this one yet: create the KWIC objects
            self.basket.set_kwic(self.qcline)
            self.kwic_object = self.basket.kwiclines.filter(qc=self.qcline).first()

        # Validate
        if self.kwic_object == None:
            # Some error has occurred
            self.data['status'] = "error"
        else:
            # Pagination
            self.hit_count = self.kwic_object.hitcount
            result_number_list = list(range(1, self.hit_count))
            paginator = Paginator(result_number_list, self.paginate_by)
            self.page_obj = paginator.page(page)

            # Get the filter myself
            oFilter = self.kwic_object.get_filter()
            # Check filter with what is available
            if self.kwic_object.has_results(oFilter, page):
                # Use the features and the results we already have
                self.result_list = [json.loads(item.result) for item in self.kwic_object.kwicresults.all()]
                self.feature_list = self.kwic_object.get_features()
            else:
                if not self.kwic_object.has_filter(oFilter):
                    # Apply the filter
                    self.kwic_object.apply_filter(oFilter)
                    # Now re-calculate the numbers
                    self.hit_count = self.kwic_object.hitcount
                    result_number_list = list(range(1, self.hit_count))
                    paginator = Paginator(result_number_list, self.paginate_by)
                    self.page_obj = paginator.page(page)

                # Set the number of items to be fetched
                iCount = self.paginate_by
                iStart = (page-1) * self.paginate_by
                # Fetch the data for this page
                oData = crpp_dbinfo(self.basket.research.owner.username, 
                                    self.basket.research.name,
                                    qc, 
                                    iStart,
                                    iCount,
                                    filter=oFilter,
                                    sPart=self.basket.part.dir)
                if oData['commandstatus'] == "ok" and oData['code'] == "completed":
                    self.result_list = oData['Results']
                    self.feature_list = oData['Features']
                    # Also add the results as objects
                    self.kwic_object.add_result_list(oData['Results'], page)
                else:
                    # Some error has occurred
                    self.data['status'] = "error"

    def add_to_context(self, context):
        # Add some elements to the context
        research = self.basket.research
        context['search_edit_url'] = reverse("seeker_edit", kwargs={"object_id": research.id})
        context['search_name'] = research.name
        context['original'] = research
        context['entrycount'] = self.entrycount
        context['basket'] = self.basket
        if self.kwic_object == None:
            context['kwic_id'] = None
        else:
            context['kwic_id'] = self.kwic_object.id
        # Provide all the information needed to create the Html presentation of the data
        context['result_list'] = self.result_list
        context['feature_list'] = self.feature_list

        # Put all the available filtering lists in here
        context['list_textid'] = []
        context['list_cat'] = []
        context['list_subtype'] = []
        context['list_title'] = []
        context['list_genre'] = []
        context['list_author'] = []

        # A list of filters that the user has specified in the past and saved
        context['list_filter'] = []

        if self.kwic_object == None:
            context['filter_list'] = []
        else:
            context['filter_list'] = self.kwic_object.kwicfilters.all()

        # Add pagination information
        context['page_obj'] = self.page_obj
        context['page_function'] = self.page_function
        context['hitcount'] = self.hit_count

        # Information on this basket/kwic object
        context['formdiv'] = self.form_div
        return context

    def get_queryset(self, prefix):
        """Determine the queryset that is to be used"""

        # Determine the selection
        lstQ = []
        lstQ.append(Q(basket=self.basket))
        if self.qcline > 0:
          lstQ.append(Q(qc=self.qcline))

        # Set and order the selection
        self.qs = Kwic.objects.filter(*lstQ)

        self.entrycount = self.qs.count()

        # Return the resulting filtered and sorted queryset
        return self.qs


class ResultPart3(ResearchPart):
    MainModel = Basket
    template_name = 'seeker/result_part_3.html'
    qcTarget = 1
    basket = None
    kwic = None
    FilterFormSet = inlineformset_factory(Kwic, KwicFilter, 
                                          form=KwicFilterForm, min_num=1, 
                                          extra=0, can_delete=True, can_order=False)
    formset_objects = [{'formsetClass': FilterFormSet, 'prefix': 'filter', 'readonly': False}]

    def custom_init(self):
        """Calculate stuff"""

        # Determine the QC line that has been passed on
        qc = self.qd['qc_select']
        if qc != None and qc != '':
            self.qcTarget = int(qc)
        # Set the basket-object
        self.basket = self.obj
        # Set the kwic object under which all filters are going to appear
        self.kwic = self.basket.get_kwic(self.qcTarget)

    def get_instance(self, prefix):
        if prefix == 'filter':
            return self.kwic

    def get_form_kwargs(self, prefix):
        if prefix == 'filter':
            return {"features":  self.kwic.get_features()}
        else:
            return None

    def add_to_context(self, context):
        if self.obj == None:
            currentowner = None
        else:
            currentowner = self.obj.research.owner
        context['currentowner'] = currentowner
        context['qc_select'] = self.qcTarget
        return context


class ResultPart4(ResearchPart):
    MainModel = Kwic
    template_name = 'seeker/result_part_4.html'

    def add_to_context(self, context):
        # find out which sentence has been identified by the user
        self.basket = self.obj.basket

        # Other information is also required: basket and so forth
        context['basket'] = self.basket
        context['kwic'] = self.obj
        context['kwic_id'] = self.obj.id
        context['quantor'] = Quantor.objects.filter(basket=self.basket).first()
        oResult = None
        if 'resid' in self.qd:
            iResId = int(self.qd['resid'])
            oResult = self.obj.get_result(iResId)
            # The RESULT object also contains all features and their values
            context['result'] = oResult
        if oResult == None:
            # What do we do if the result is empty??
            context['no_result'] = True
            return context
        # Check whether a 'back' identifis passed on
        if 'backid' in self.qd:
          sBackId = self.qd['backid']
        else:
          sBackId = 'result_container_2'
        context['backid'] = sBackId
        # Return the adapted context
        return context


class ResultPart5(ResearchPart):
    MainModel = Kwic
    template_name = 'seeker/result_part_5.html'
    do_dump = True

    def add_to_context(self, context):
        # find out which sentence has been identified by the user
        self.basket = self.obj.basket

        # Other information is also required: basket and so forth
        context['basket'] = self.basket
        context['kwic'] = self.obj
        context['kwic_id'] = self.obj.id
        context['quantor'] = Quantor.objects.filter(basket=self.basket).first()
        oResult = None
        if 'resid' in self.qd:
            iResId = int(self.qd['resid'])
            oResult = self.obj.get_result(iResId)
            # The RESULT object also contains all features and their values
            context['result'] = oResult
        # Check whether a 'back' identifis passed on
        if 'backid' in self.qd:
          sBackId = self.qd['backid']
        else:
          sBackId = 'result_container_2'
        context['backid'] = sBackId
        if oResult == None:
            # What do we do if the result is empty??
            context['no_result'] = True
            return context
        # Fetch information from the /crpp service
        options = {'userid': self.basket.research.owner.username,
                   'lng': self.basket.part.corpus.get_lng_display(),
                   'dir': self.basket.part.dir,
                   'ext': self.basket.get_format_display(),
                   'name': oResult['File'],
                   'locs': oResult['Locs'],
                   'locw': oResult['Locw'],
                   'type': 'syntax_tree'}

        # Retrieve the info for this sentence
        oInfo = get_crpp_sent_info(options)
        if oInfo != None and oInfo['status'] == "ok":
            # Make sure that 'object' sections are translated to proper JSON
            oSentInfo = oInfo['info']
            # Get the sentence tree and extract any top-node features from it
            treeSent = oSentInfo['allT']
            context['eng'] = ""
            if treeSent != None and treeSent['f'] != None:
                f = treeSent['f']
                if 'eng' in f and f['eng'] != "":
                    # We have an 'eng' translation: make it available
                    context['eng'] = f['eng']
            # Put the information into the DATA object
            self.data['sent_info'] = oSentInfo

            if self.do_dump:
                # Replace the 'allT' and 'hitT' with STRING json
                if 'allT' in oSentInfo: oSentInfo['allT'] = json.dumps(oSentInfo['allT'])
                if 'hitT' in oSentInfo: oSentInfo['hitT'] = json.dumps(oSentInfo['hitT'])
            context['sent_info'] = oSentInfo
        else:
            context['sent_info'] = None
        return context


class ResultPart6(ResultPart5):
    """The hierarchical table model uses the same information as the tree"""

    MainModel = Kwic
    template_name = 'seeker/result_part_6.html'
    do_dump = True


class ResultHitView(ResultPart5):
    action = "downloadhit"


class SeekerResultDownload(ResearchPart):
    MainModel = Basket
    template_name = "seeker/download_status.html"
    qcTarget = 1
    dtype = "csv"           # downloadtype
    basket = None
    action = "download"

    def custom_init(self):
        """Calculate stuff"""
        
        if 'qc_select' in self.qd:
            qc = self.qd['qc_select']
            if qc != None and qc != '':
                self.qcTarget = int(qc)
        else:
            qc = 1
        dt = self.qd['downloadtype']
        if dt != None and dt != '':
            self.dtype = dt
        # Set the basket-object
        self.basket = self.obj

    def add_to_context(self, context):
        # Provide search URL and search name
        context['search_edit_url'] = reverse("seeker_edit", kwargs={"object_id": self.basket.research.id})
        context['search_name'] = self.basket.research.name
        context['basket'] = self.basket
        return context


# ================================ Other views =========================================================

def research_oview(request, object_id=None):
    """Entry point to show a summary of a search specification"""
    
    # Initialisations
    template = 'seeker/research_oview.html'
    arErr = []      # Start without errors
    
    # Check if the user is authenticated
    if not request.user.is_authenticated:
        # Simply redirect to the home page
        return redirect('nlogin')

    # Get the instance of this research object
    obj = Research.objects.get(pk=object_id)
    title = "Specifications for project {}".format(obj.name)

    # Get a list of errors
    error_list = [str(item) for item in arErr]

    # Create the context
    context = dict(
        object_id = object_id,
        title=title,
        original=obj,
        gateway=obj.gateway,
        error_list=error_list
        )

    # Open the template that allows Editing an existing or Creating a new research project
    #   or editing the existing project
    return render(request, template, context)

def get_partlist(request):
    """Get a list of Part elements + first/last information"""

    # Longdale check
    longdale_user = user_is_ingroup(request, "longdale_user")
    # REtrieve the correct queryset, sorted on the correct levels
    if longdale_user:
        qs = [prt for prt in Part.objects.all().order_by('corpus__lng', 'corpus__name', 'name')]
    else:
        longdale = "Longdale"
        qs = [prt for prt in Part.objects.exclude(Q(name__istartswith=longdale)).order_by('corpus__lng', 'corpus__name', 'name')]
    # Start the output
    html = []
    # Initialize the variables whose changes are important
    lVars = ["corpus_lng", "corpus_name", "name"]
    lFuns = [Part.language, ["corpus", "name"], ["name"]]
    # Get a list of items containing 'first' and 'last' information
    lItem = get_item_list(lVars, lFuns, qs)
    # REturn this list
    return lItem
    
def research_edit(request, object_id=None):
    """Main entry point for the specification of a seeker research project"""

    # Initialisations
    template = 'seeker/research_details.html'
    arErr = []              # Start out with no errors
    partchoice = None       # Default value for part choice

    # Check if the user is authenticated
    if not request.user.is_authenticated:
        # Simply redirect to the home page
        return redirect('nlogin')

    # Get the 'obj' to this project (or 'None' if it is a new one)
    if object_id is None:
        obj = None
        intro_message = "Create a new project"
        intro_breadcrumb = "New Project"
        sTargetType = ""
        part_list = None
        search_type = None
        basket = None
    else:
        # Get the instance of this research object
        obj = Research.objects.get(pk=object_id)
        intro_message = "Research project: <b>{}</b>".format(obj.name)
        intro_breadcrumb = "[{}]".format(obj.name)
        sTargetType = obj.targetType
        # Since this has gone well, provide the context to the user with a list of corpora to explore
        part_list = get_partlist(request)
        # Also provide a set of search options
        # search_type = ['all', 'first n', 'random n']
        search_type = [ {'name': 'all', 'value': 'all'},
                        {'name': 'first n', 'value': 'first'},
                        {'name': 'random n', 'value': 'random'}]
        # Do we have a basket-id?
        if 'basket_id' in request.GET:
            basket_id = request.GET.get('basket_id')
            basket = Basket.objects.filter(id=basket_id).first()
            # There already is a search: find its associated part
            if basket != None:
                partchoice = basket.part.id
        else:
            # There already is a search: find its associated part
            basket = Basket.objects.filter(research=obj).order_by('-saved').first()
            if basket != None:
                partchoice = basket.part.id
            # Reset the basket, because it is not used in this call
            basket = None

    search_count = 1000

    # Get a list of errors
    error_list = [str(item) for item in arErr]

    # Create the context
    context = dict(
        object_id = object_id,
        original=obj,
        is_new=(obj==None),
        intro_message=intro_message,
        intro_breadcrumb=intro_breadcrumb,
        part_list=part_list,
        targettype=sTargetType,
        search_type = search_type,
        search_count = search_count,
        partchoice=partchoice,
        basket=basket,
        error_list=error_list
        )

    # Open the template that allows Editing an existing or Creating a new research project
    #   or editing the existing project
    return render(request, template, context)

def research_simple(request, pk=None):
    """Allow the user to specify a simple search"""

    # Initialisations
    template = "seeker/research_simple.html"
    arErr = []                  # Start without errors
    simplename = SIMPLENAME     # Name of the simple project
    oErr = ErrHandle()
    sTargetType = "w"
    # Get the correct owner
    owner = User.objects.filter(Q(username=request.user)).first()

    # Get a list of SIMPLE searches
    # qs_simple = Research.objects.filter(owner=owner, gateway__name="simple").exclude(name=SIMPLENAME)
    qs_simple = Research.objects.filter(owner=owner, stype=STYPE_SIMPLE).exclude(name=SIMPLENAME)

    # Get parameters
    qd = request.GET if request.method == "GET" else request.POST

    # Action depends on GET or POST
    if request.method == "GET": 

        # Check if the user is authenticated
        if owner == None or not request.user.is_authenticated:
            # Simply redirect to the home page
            return redirect('nlogin')

        # Initialisations
        partchoice = None
        obj = None

        # Initially the 'related' formset is empty
        related_formset = RelatedFormset(prefix='simplerel')


        # Get the correct research object
        if pk != None:
            obj = Research.objects.filter(id=pk).first()
        if pk == None or obj == None or obj.compact == None or obj.compact == "":
            # Get the 'simple' project of this owner
            lstQ = []
            lstQ.append(Q(owner=owner))
            lstQ.append(Q(name=simplename))
            obj = Research.objects.filter(*lstQ).first()
            if obj == None:
                gateway = Gateway(name="simple")
                gateway.save()
                obj = Research(name=simplename, purpose="simple", targetType=sTargetType,
                               gateway=gateway, owner=owner)
                obj.save()
            else:
                # There already is a search: find its associated part
                basket = Basket.objects.filter(research=obj).order_by('-saved').first()
                if basket != None:
                    partchoice = basket.part.id
            # Create a form based on this research object
            simpleform = SimpleSearchForm()
            simpleform.fields["targetType"].initial = obj.targetType
            simpleform.fields["searchwords"].initial = ""
            simpleform.fields["searchpos"].initial = ""
            simpleform.fields["searchlemma"].initial = ""
            simpleform.fields["searchexc"].initial = ""
            simpleform.fields["searchrel"].initial = ""
            simpleform.fields["searchcql"].initial = ""
            simpleform.fields["searchfcat"].initial = ""
            simpleform.fields["searchfval"].initial = ""
            simpleform.fields["baresimple"].initial = ""
            simpleform.fields["description"].initial = ""
            cns = obj.gateway.constructions.first()
            if cns != None and cns.search != None:
                svalue = cns.search.value
                if obj.targetType == "w":
                    # Simple word search
                    simpleform.fields["searchwords"].initial = svalue
                elif obj.targetType == "c":
                    # Constituent category search
                    simpleform.fields["searchpos"].initial = svalue
                elif obj.targetType == "q":     # Cesar CQL
                    simpleform.fields["searchcql"].initial = svalue
                else:
                    # This should be targettype "e" (extended)
                    # Extended search: at least 'lemma' or 'constituent', and possibly also 'word'
                    simpleform.fields["searchwords"].initial = svalue
                    simpleform.fields["searchpos"].initial = cns.search.category
                    simpleform.fields["searchexc"].initial = cns.search.exclude
                    simpleform.fields["searchlemma"].initial = cns.search.lemma

                if cns.search.related != None and cns.search.related != "":
                    # Treat 'related'
                    simpleform.fields["searchrel"].initial = cns.search.related
                    # Fill the 'related' formset with data from [related]
                    lSearchRelated = json.loads(cns.search.related)
                    related_formset = RelatedFormset(initial=lSearchRelated, prefix='simplerel')
                    lTowards = [("search", "Search Hit")]
                    i = 0
                    for form in related_formset:
                        form.fields['towards'].choices = lTowards
                        # Make sure 'name' is in there
                        if 'name' in lSearchRelated[i]:
                            sDvar = lSearchRelated[i]['name']
                            tThis = (sDvar, sDvar)
                            lTowards.append( tThis )
                        i += 1
        else:
            # Get the information from the simple search
            oSearch = json.loads(obj.compact)
            # Fill the simple search form with the information in oSearch
            simpleform = SimpleSearchForm()
            simpleform.fields["targetType"].initial = oSearch['targetType']
            simpleform.fields["searchwords"].initial = oSearch['searchwords']
            simpleform.fields["searchpos"].initial = oSearch['searchpos']
            simpleform.fields["searchlemma"].initial = oSearch['searchlemma']
            simpleform.fields["searchexc"].initial = oSearch['searchexc']
            simpleform.fields["searchcql"].initial = oSearch['searchcql']
            simpleform.fields["searchfcat"].initial = oSearch.get('searchfcat')
            simpleform.fields["searchfval"].initial = oSearch.get('searchfval')
            simpleform.fields["baresimple"].initial = oSearch['name']
            simpleform.fields["description"].initial = "" if "description" not in oSearch else oSearch['description']

            # TODO: calculate [searchrel] and [ltowards] from what is in oSearch
            # simpleform.fields["searchrel"].initial = oSearch['searchrel']

            # Get the list of related searches
            lSearchRelated = oSearch['related']
            # Make sure the string is saved correctly
            simpleform.fields["searchrel"].initial = json.dumps(lSearchRelated)
            # Create the correct formset
            related_formset = RelatedFormset(initial=lSearchRelated, prefix='simplerel')
            lTowards = [("search", "Search Hit")]
            i = 0
            for form in related_formset:
                form.fields['towards'].choices = lTowards
                # Make sure 'name' is in there
                if 'name' in lSearchRelated[i]:
                    sDvar = lSearchRelated[i]['name']
                    tThis = (sDvar, sDvar)
                    lTowards.append( tThis )
                i += 1

        # Determine whether the 'more' part should be shown or not
        show_more = "more"
        more_fields = ["searchpos", "searchlemma", "searchexc", "searchfcat", "searchfval"]
        for field in more_fields:
            if simpleform.fields[field].initial: 
                show_more = "less"
                break

        object_id = obj.id

        intro_message = "Make a simple search"
        intro_breadcrumb = "Simple"
        # Also provide a set of search options
        search_type = [ {'name': 'all', 'value': 'all'},
                        {'name': 'first n', 'value': 'first'},
                        {'name': 'random n', 'value': 'random'}]
        search_count = 1000

        # Provide a list of axis
        axis_list = []
        for rel in Relation.objects.filter(Q(type='axis')).order_by('name'):
            axis_list.append({"id": rel.id, "name": rel.name})

        # Get a list of errors
        error_list = [str(item) for item in arErr]

        # Create the context
        context = dict(
            object_id = object_id,
            original=obj,
            simpleform=simpleform,
            show_more=show_more,
            related_formset=related_formset,
            axis_list=axis_list,
            intro_message=intro_message,
            intro_breadcrumb=intro_breadcrumb,
            targettype=sTargetType,
            search_type = search_type,
            search_count = search_count,
            part_list=get_partlist(request),
            partchoice = partchoice,
            simple_list = qs_simple,
            error_list=error_list
            )
        # x = render_to_string(template, context, request=request)
        # Open the template that allows Editing an existing or Creating a new research project
        #   or editing the existing project
        return render(request, template, context)

    elif request.method == "POST":
        # The POST method is reserved for MODIFYING an existing simple search
        data = {'status': 'ok', 'html': ''}

        # Check if the user is authenticated
        if owner == None or not request.user.is_authenticated:
            # Simply redirect to the home page
            return redirect('nlogin')

        try:
            # Get the parameters
            qd = request.POST

            # See if a specific action is required
            bReady = False
            if 'action' in qd:
                action = qd['action']
                if action == "delete":
                    # Delete this item and then return the simple list view
                    obj = Research.objects.filter(id=pk).first()
                    if obj:
                        # Remove it
                        obj.delete()
                        # Now return the correct listview
                        data['afterdelurl'] = reverse('simple_list')
                        bReady = True

            # Are we ready?
            if not bReady:
                # Try to load the form
                simpleform = SimpleSearchForm(qd)
                if simpleform.is_valid():

                    # See if it has a different name now
                    simplename = simpleform.cleaned_data["baresimple"]
                    if simplename == "": simplename = SIMPLENAME

                    # Get the correct research project FOR THIS USER
                    lstQ = []
                    lstQ.append(Q(owner=owner))
                    lstQ.append(Q(name=simplename))
                    research = Research.objects.filter(*lstQ).first()
                    if research == None:
                        # There is no simple research project to adapt
                        data['status'] = 'error'
                        data['html'] = "research_simple cannot find a Research project that should be adapted"
                    else:
                        # Found the research
                        if modify_simple_search(research, qd, simpleform.cleaned_data):
                            # We are okay
                            data['html'] = "The simple project has been adapted"
                        else:
                            # Something went wrong
                            data['status'] = 'error'
                            data['html'] = "research_simple: modify_simple_search returns with an error"
                else:
                    # There are errors in the form
                    data['status'] = 'error'
                    data['html'] = "Simple form errors: {}".format(simpleform.errors)
            
        except:
            data['status'] = 'error'
            data['html'] = oErr.get_error_message()

        # What we return should be a JSON response
        return JsonResponse(data)

    else:
        return None

def get_related(qd):
    """Get the list of related"""

    sRelated = ""
    bResult = True

    # Adapt the 'RelatedForm' with the information from lTowards
    lTmp = json.loads(qd.get('ltowards', '[]'))
    lTowards = [(x['name'], x['value']) for x in lTmp if x['name'] != ""]

    # Get the 'simplerel' formset from the input
    formset = RelatedFormset(qd, prefix='simplerel')
    for form in formset:
        form.fields['towards'].choices = lTowards
    # if formset.is_valid():
    bResetRelated = False
    if formset != None:
        # Check for validness
        if formset.is_valid():
            # Everything okay, continue
            if formset.total_form_count() == 0:
                bResetRelated = True
            else:
                lRelated = []
                # Walk the forms in the formset
                for rel_form in formset:
                    if rel_form.is_valid():
                        # Process the information in this form
                        # oFields = rel_form.cleaned_data
                        oFields = copy.copy(rel_form.cleaned_data)
                        # lRelated.append({"name": oFields['name'], "cat": oFields['cat'], "raxis": oFields['raxis'], "towards": oFields['towards']})
                        lRelated.append(oFields)
                sRelated = json.dumps(lRelated)
        else:
            # One of the forms is not valid, got to do something about it
            bResult = False
            sRelated = formset.errors
    return bResult, sRelated

def modify_simple_search(research, qd, data = None):
    """Modify and save the research on the basis of the info we get"""

    # Sanity checking
    if research == None:
        return False

    oErr = ErrHandle()

    # Check if we have the data
    if data == None:
        # Clean the form and get the data
        form = SimpleSearchForm(qd)
        if form.is_valid():
            data = form.cleaned_data

    # Continue
    if data != None and "targetType" in data:
        # Make sure the targettype is set correctly
        research.targetType = qd.get("targetType", "w")
        # Make sure the stype is set correctly too
        research.stype = STYPE_SIMPLE
        #  Possibly adapt the name of this search
        simplename = SIMPLENAME
        if 'baresimple' in data and data['baresimple'] != "":
            simplename = data['baresimple']
        research.name = simplename
        # Save the adapted research
        research.save()
        
        # Look for value
        sValue = ""         # Word value
        sExclude = ""       # Category to be excluded
        sCategory = ""      # Extended: category
        sLemma = ""         # Extended: lemma
        sRelated = ""       # Extended: related constituent(s)
        lTowards = []       # List of towards options
        sFeatCat = ""       # Feature category
        sFeatVal = ""       # Feature value
        sCql = ""           # CQL
        if research.targetType == "w":              # Word-level
            sValue = qd.get("searchwords", "")
        elif research.targetType == "q":            # Cesar CQL
            sCql = qd.get("searchcql", "")
        elif research.targetType == "c":            # Construction level
            sValue = qd.get("searchpos", "")
            sExclude = qd.get("searchexc", "")
        else:                                       # Extended
            sValue = qd.get("searchwords", "")
            sCategory = qd.get("searchpos", "")
            sLemma = qd.get("searchlemma", "")
            sExclude = qd.get("searchexc", "")
            sFeatCat = qd.get("searchfcat", "")
            sFeatVal = qd.get("searchfval", "")
            # sRelated = qd.get("searchrel", "")      # Related constituent(s)

        # Adapt the 'RelatedForm' with the information from lTowards
        lTmp = json.loads(qd.get('ltowards', '[]'))
        lTowards = [(x['name'], x['value']) for x in lTmp if x['name'] != ""]

        # Get the 'simplerel' formset from the input
        formset = RelatedFormset(qd, prefix='simplerel')
        for form in formset:
            form.fields['towards'].choices = lTowards
        # if formset.is_valid():
        bResetRelated = False
        if formset != None:
            # Check for validness
            if formset.is_valid():
                # Everything okay, continue
                if formset.total_form_count() == 0:
                    bResetRelated = True
                else:
                    lRelated = []
                    # Walk the forms in the formset
                    for rel_form in formset:
                        if rel_form.is_valid():
                            # Process the information in this form
                            # oFields = rel_form.cleaned_data
                            oFields = copy.copy(rel_form.cleaned_data)
                            # lRelated.append({"name": oFields['name'], "cat": oFields['cat'], "raxis": oFields['raxis'], "towards": oFields['towards']})
                            lRelated.append(oFields)
                    sRelated = json.dumps(lRelated)
            else:
                # One of the forms is not valid, got to do something about it
                return False

        # Adapt the value in searchmain
        if sValue != "" or sCategory != "" or sLemma != "" or sExclude != "" or sCql != "" or sFeatVal != "" or bResetRelated:
            search = None
            cns = research.gateway.constructions.first()
            if cns == None or cns.search == None:
                # there is no construction yet: make a SearchMain and a Construction
                if research.targetType == "w":
                    search = SearchMain.create_item("word-group", sValue, 'groupmatches')
                    if sRelated == None or sRelated == "":
                        # This is simple 'w' without related constituents
                        search = SearchMain.create_item("word-group", sValue, 'groupmatches')
                    else:
                        # This is 'w' + related
                        search = SearchMain.create_item("word-group", sValue, 'groupmatches', related=sRelated)
                elif research.targetType == "c" or research.targetType == "e":
                    # This is targetType "c" (constituent) or "e" (extended)
                    search = SearchMain.create_item("const-group", sValue, 
                                                    'groupmatches', sExclude, sCategory, 
                                                    sLemma, sFeatCat, sFeatVal, sRelated)
                elif research.targetType == "q":
                    # CQL translation
                    search = SearchMain.create_item("const-group", sValue, 'groupmatches', cql = sCql)
                else:
                    # This may not happen
                    oErr.DoError("modify_simple_search: targettype not known")
                    return False

            # Make sure the correct SEARCH is set
            if search == None:
                search = cns.search
            # Possibly reset related
            if bResetRelated:
                # Yes: remove all related items
                search.related = ""
                search.save()
            # Make sure the value of the word(s)/constituent(s) is set correctly
            if research.targetType == "w":
                if sRelated == None or sRelated == "":
                    # This is simple 'w' without related constituents
                    search.adapt_item("word-group", sValue, 'groupmatches')
                else:
                    # This is 'w' + related
                    search.adapt_item("word-group", sValue, 'groupmatches', related=sRelated)
            elif research.targetType == "c" or research.targetType == "e":
                # Targettype "c" (constituent) or "e" (extended)
                search.adapt_item("const-group", sValue, 
                                  'groupmatches', sExclude, sCategory, 
                                  sLemma, sFeatCat, sFeatVal, sRelated)
            elif research.targetType == "q":
                # CQL translation
                search.adapt_item("const-group", sValue, 'groupmatches', cql = sCql)
            else:  
                # This may not happen
                oErr.DoError("modify_simple_search: targettype not known")
                return False

            # Make sure the construction is saved
            if cns == None:
                cns = Construction(name="simple", gateway=research.gateway, search=search)
                cns.save()


    # Return positively
    return True

def research_simple_save(request):
    """Given a simple search that has already run, save it under the name provided"""

    # Initialisations
    arErr = []
    simplename = SIMPLENAME     # Name of the simple project - this is by default the standard name
    data = {'status': 'error', 'html': ''}

    # This only works for action POST
    if request.method == "POST":

        # There should be a parameter: savename
        params = request.POST
        if 'savename' not in params:
            data['html'] = "Need parameter [savename]"
            return JsonResponse(data)
        sSaveName = params['savename']

        # Check if this has a particular name
        sSearchName = params.get("baresimple", "")
        if sSearchName != "": simplename = sSearchName

        # Get the correct owner
        owner = User.objects.filter(Q(username=request.user)).first()

        # Check if the user is authenticated
        if owner == None or not request.user.is_authenticated:
            # Simply redirect to the home page
            return redirect('nlogin')

        # Get the 'simple' project of this owner
        lstQ = []
        lstQ.append(Q(owner=owner))
        lstQ.append(Q(name=simplename))
        obj = Research.objects.filter(*lstQ).first()
        # Double check to see if the search has run
        if obj != None:
            # There already is a search: find its associated part
            basket = Basket.objects.filter(research=obj).order_by('-saved').first()
            if basket != None:
                # This search has run recently, so that is good
                object_id = obj.id
                # Get the JSON definition of the search 
                oProj = obj.get_json()
                if 'json_data' in oProj:
                    sData = oProj['json_data'].decode("utf-8")
                    oData = json.loads(sData)

                    # Since this is only called when a Simple research is converted...
                    oData['stype'] = 'p'    #plain
                    iStop = 1

                    # Create a new research based on the JSON definition
                    research = Research.read_data(owner.username, "", arErr, oData, sSaveName)
                    # Check if a new object has been made
                    if research != None:
                        # Indicate in the reply that all has gone well
                        data['status'] = "ok"
                        # Get the URL to edit it
                        data['editurl'] = reverse("seeker_edit", kwargs={"object_id": research.id})
                        data['name'] = sSaveName
                    else:
                        data['html'] = "Could not create a project with that name (does it exist already?)"
                else:
                    data['html'] = "Could not process JSON data of simple search"
            else:
                data['html'] = "First execute a simple search"
        else:
            data['html'] = "First define and execute a simple search"

    # Return the information
    return JsonResponse(data)

def get_targettype(sWord, sCat, sLemma, sFeatVal, sCql):
    """Determine the targettype based on the information passed on"""

    targetType = "e"    # Default: extended
    if sCql != "": 
        # Whenever CQL is specified, do a CQL search
        targetType = "q"
    elif sWord != "" and sLemma == "" and sCat == "" and sFeatVal == "":
        # THis is *ONLY* looking for a particular word
        targetType = "w"
    elif sWord == "" and sLemma == "" and sFeatVal == "" and sCat != "":
        # THis is *ONLY* constituent category searching!!!
        targetType = "c"
    return targetType

def read_simple_form(qd):
    """Read the forms in the research query data and transform them into a simple search object"""

    oSearch = {}
    bOkay = False
    msg = ""
    oErr = ErrHandle()
    try:
        # Read the form
        simpleform = SimpleSearchForm(qd)
        if simpleform.is_valid():
            # Get the data
            cleaned_data = simpleform.cleaned_data
            # Unpack the data
            baresimple = cleaned_data['baresimple']
            simplename = SIMPLENAME
            searchwords = cleaned_data['searchwords']
            searchcql = cleaned_data['searchcql']
            searchpos = cleaned_data['searchpos']
            searchexc = cleaned_data['searchexc']
            searchlemma = cleaned_data['searchlemma']
            searchfcat = cleaned_data['searchfcat']
            searchfval = cleaned_data['searchfval']
            targetType = cleaned_data['targetType']
            description = cleaned_data['description']
            overwrite = (qd.get("overwrite", "false") == "true")

            # Determine the targettype based on the data above
            # (i.e: overriding what has been determined previously...)
            targetType = get_targettype(searchwords, searchpos, searchlemma, searchfval, searchcql)

            if baresimple != "":
                # Get the search parameters
                oSearch = dict(name=baresimple, targetType=targetType, searchwords=searchwords, 
                               searchcql=searchcql, searchpos=searchpos, searchexc=searchexc, 
                               searchfcat=searchfcat, searchfval=searchfval,
                               searchlemma=searchlemma, description=description)
                # Get the related ones
                bOkay, sRelated = get_related(qd)
                if bOkay:
                    related = [] if sRelated == "" else json.loads(sRelated)
                    oSearch['related'] = related
                    bOkay = True
                else:
                    # Creating related goes wrong
                    oSearch['msg']  = sRelated
            else:
                # No name has been specified
                oSearch['msg'] = "When using 'Save as', a name is required"
        else:
            # form error
            oSearch['msg'] = "There are errors: {}".format(simpleform.errors)
    except:
        oSearch['msg'] = oErr.get_error_message()

    # Return our result
    return bOkay, oSearch

def research_simple_baresave(request):
    """Given a simple search *definition*, save this definition under the name provided"""

    # Initialisations
    arErr = []
    simplename = SIMPLENAME     # Name of the simple project
    data = {'status': 'error', 'html': 'nothing'}

    # Get the correct owner
    owner = User.objects.filter(Q(username=request.user)).first()

    # Check if the user is authenticated
    if owner == None or not request.user.is_authenticated:
        # Simply redirect to the home page
        return redirect('nlogin')        
            
    # This only works for action POST
    if request.method == "POST":

        # There should be a parameter: savename
        params = request.POST
        # overwrite = (params.get("overwrite", "false") == "true")
        overwrite = True

        # Read the form
        bOkay, oSearch = read_simple_form(params)
        if bOkay:
            baresimple = oSearch['name']
            # Now save [oSearch] with [baresimple]
            research, msg = Research.create_simple(oSearch, baresimple, owner, overwrite)

            if research == None:
                data['html'] = msg
            else:
                # Show all went well
                data['status'] = "ok"
                data['name'] = research.name
                data['view'] = reverse('simple_details', kwargs={'pk': research.id})
        else:
            if 'msg' in oSearch:
                data['html'] = oSearch['msg']
            else:
                data['html'] = "Could not read the simple form"

    # Return the information
    return JsonResponse(data)


@csrf_exempt
def import_json(request):
    """Import a JSON file that contains details of a (new) research project"""

    # Initialisations
    # NOTE: do ***not*** add a breakpoint until *AFTER* form.is_valid
    arErr = []
    error_list = []
    transactions = []
    data = {'status': 'ok', 'html': ''}
    template_name = 'seeker/import_status.html'
    obj = None
    data_file = ""
    bClean = False
    username = request.user.username

    # Check if the user is authenticated and if it is POST
    if not request.user.is_authenticated  or request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            # NOTE: from here a breakpoint may be inserted!
            print('valid form')
            # Get the contents of the imported file
            data_file = request.FILES['file_source']

            # Get the source file
            if data_file == None or data_file == "":
                arErr.append("No source file specified for the selected project")
            else:
                # 
                # Read the data for the selected project
                research = Research.read_data(username, data_file, arErr)

                # Determine a status code
                statuscode = "error" if research == None else "completed"
                if research == None:
                    arErr.append("There was an error. No research project has been created")

            # Get a list of errors
            error_list = [str(item) for item in arErr]

            # Create the context
            context = dict(
                statuscode=statuscode,
                research=research,
                error_list=error_list
                )

            if len(arErr) == 0:
                # Get the HTML response
                data['html'] = render_to_string(template_name, context, request)
            else:
                data['html'] = "Please log in before continuing"


        else:
            data['html'] = 'invalid form: {}'.format(form.errors)
            data['status'] = "error"
    else:
        data['html'] = 'Only use POST and make sure you are logged in'
        data['status'] = "error"
 
    # Return the information
    return JsonResponse(data)

